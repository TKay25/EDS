import uuid
import os
import numpy as np
from mysql.connector import Error
from flask import Flask, request, jsonify, session, render_template, redirect, url_for, send_file,flash, make_response, after_this_request
from datetime import datetime, timedelta
import pandas as pd
from xhtml2pdf import pisa
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import matplotlib.pyplot as plt
import seaborn as sns
import psycopg2
from psycopg2 import sql
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font
from werkzeug.utils import secure_filename
import matplotlib.pyplot as plt
import io
import base64
import json
import requests
import pdfkit
from weasyprint import HTML
import re
from paynow import Paynow
import time
import random




today_date = datetime.now().strftime('%d %B %Y')
applied_date = datetime.now().strftime('%Y-%m-%d')

app = Flask(__name__)
app.secret_key = 'your_secret_key'  
app.secret_key = '011235'
app.permanent_session_lifetime = timedelta(minutes=30)
user_sessions = {}

external_database_url = "postgresql://lmsdatabase_8ag3_user:6WD9lOnHkiU7utlUUjT88m4XgEYQMTLb@dpg-ctp9h0aj1k6c739h9di0-a.oregon-postgres.render.com/lmsdatabase_8ag3"
database = 'lmsdatabase_8ag3'

connection = psycopg2.connect(external_database_url)

cursor = connection.cursor()

# WhatsApp API Credentials (Replace with your actual credentials)
#ACCESS_TOKEN = "EAATESj1oB5YBOyIVfVPEAIZAZA7sgPboDN36Wa2Or11uZCBEZCVWaNAZB0exkYYG6gcIdiYbvPCST9tKjS54ib1NqXbNg7UvJYaZCIZAjxgTBQwvyoWE8cZCMgje1wkrUyb335TMwNwYSTA3rNwppRZAeQGt3M7s5x15nZCbZBtEfZBtSIu3p7ZCHOcF0pMTuLgjQreLz2QZDZD"
#PHONE_NUMBER_ID = "558392750697195"
#VERIFY_TOKEN = "521035180620700"
#WHATSAPP_API_URL = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"



create_table_query_comp_creation = f"""
CREATE TABLE IF NOT EXISTS companyreg (
    compid SERIAL PRIMARY KEY,
    companyname VARCHAR (100),
    datecreated date
);
"""
cursor.execute(create_table_query_comp_creation)
connection.commit()

create_table_query = f"""
CREATE TABLE IF NOT EXISTS cagwatick (
    id SERIAL PRIMARY KEY,
    idwanumber INT,
    route VARCHAR (100),
    time VARCHAR (100),
    paymethod VARCHAR (100),    
    fare VARCHAR (100),
    ecocashnum INT,
    pollurl VARCHAR (100),
    status VARCHAR (100),
    datebought date 
);
"""
cursor.execute(create_table_query)
connection.commit()



create_table_query = f"""
CREATE TABLE IF NOT EXISTS whatsapptempapplication (
    id SERIAL PRIMARY KEY,
    empidwa INT,
    leavetypewa VARCHAR (100),
    startdate date,
    enddate date
);
"""
cursor.execute(create_table_query)
connection.commit()

cursor.execute("""
    ALTER TABLE whatsapptempapplication
    ADD COLUMN IF NOT EXISTS companynamewa VARCHAR(255);
""")

connection.commit()
print(f"column added to Table whatsapptempapplication successfully!")

##################### client test ####################################################################################################

@app.errorhandler(400)
def bad_request(e):
    return jsonify({'status': 'error', 'message': str(e)}), 400

@app.route("/webhook", methods=["GET", "POST"])
def webhook():
    if request.method == "GET":
        print("ouch")
        if request.args.get("hub.verify_token") == VERIFY_TOKEN:
            return request.args.get("hub.challenge")
        return "Verification failed", 403

    if request.method == "POST":
        global today_date
        data = request.get_json()

        #global ACCESS_TOKEN
        #global PHONE_NUMBER_ID

        try:
            # Navigate the JSON structure to get the display_phone_number
            display_phone_number = data["entry"][0]["changes"][0]["value"]["metadata"]["display_phone_number"]

            # Example condition: check if it's a specific number
            if display_phone_number == "15556291389":
                print(display_phone_number)

                VERIFY_TOKENcc = "1412803596375322"
                ACCESS_TOKEN = "EAAUppTRo5q4BPATlxuMt4ZANFhgbyrtQI7iB1bR5FAI7K5Rv9yolg1OEwgt5J8xRJKKkTc2F9lHutvNcDXyHPEZAoGEuMQlv1THfAGRuTtZBEzmwbJG04f1sLxEAUFze09rHvmtuqa50ccT6ik2nm7cfcMOI8vn6id1PZBId5fMDf2WNZASQFIBIZBX6UIyTr3vVkaaTvIwO1ZB1ZAnQS6LUMtC6b14MZBeisR6XvHIvZBSSooWwZDZD"
                PHONE_NUMBER_IDcc = "618334968023252"

                def send_whatsapp_messagecc(to, text, buttons=None):
                    """Function to send a WhatsApp message using Meta API, with optional buttons."""
                    headers = {
                        "Authorization": f"Bearer {ACCESS_TOKEN}",
                        "Content-Type": "application/json"
                    }

                    # If buttons are provided, send an interactive message
                    if buttons:
                        data = {
                            "messaging_product": "whatsapp",
                            "to": to,
                            "type": "interactive",
                            "interactive": {
                                "type": "button",
                                "body": {"text": text},
                                "action": {
                                    "buttons": buttons
                                }
                            }
                        }
                    else:
                        # Send a normal text message
                        data = {
                            "messaging_product": "whatsapp",
                            "to": to,
                            "type": "text",
                            "text": {"body": text}
                        }

                    response = requests.post(f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_IDcc}/messages", headers=headers, json=data)
                    
                    # Debugging logs
                    print("✅ Sending message to:", to)
                    print("📩 Message body:", text)
                    print("📡 WhatsApp API Response Status:", response.status_code)

                    try:
                        response_json = response.json()
                        print("📝 WhatsApp API Response Data:", response_json)
                    except Exception as e:
                        print("❌ Error parsing response JSON:", e)

                    return response.json()

                def send_whatsapp_list_messagecc(recipient, text, list_title, sections):
                    headers = {
                        "Authorization": f"Bearer {ACCESS_TOKEN}",
                        "Content-Type": "application/json"
                    }
                    
                    payload = {
                        "messaging_product": "whatsapp",
                        "recipient_type": "individual",
                        "to": recipient,
                        "type": "interactive",
                        "interactive": {
                            "type": "list",
                            "header": {
                                "type": "text",
                                "text": list_title
                            },
                            "body": {
                                "text": text
                            },
                            "action": {
                                "button": "Select Option",
                                "sections": sections
                            }
                        }
                    }
                    
                    response = requests.post(
                        f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_IDcc}/messages",
                        headers=headers,
                        json=payload
                    )
                    
                    print("List message response:", response.json())
                    return response
                

                try:

                    if data and "entry" in data:
                        for entry in data["entry"]:
                            for change in entry["changes"]:
                                if "messages" in change["value"]:
                                    for message in change["value"]["messages"]:

                                        conversation_id = str(uuid.uuid4())
                                        session['conversation_id'] = conversation_id
                                    

                                        sender_id = message["from"]
                                        sender_number = sender_id[-9:]
                                        print(f"📱 Conversation {conversation_id}: Sender's WhatsApp number: {sender_number}")
                                        session['client'] = str(sender_number)

                                        external_database_url = "postgresql://lmsdatabase_8ag3_user:6WD9lOnHkiU7utlUUjT88m4XgEYQMTLb@dpg-ctp9h0aj1k6c739h9di0-a.oregon-postgres.render.com/lmsdatabase_8ag3"


                                        try:
                                            connection = psycopg2.connect(external_database_url)
                                            cursor = connection.cursor()   

                                            if message.get("type") == "interactive":
                                                interactive = message.get("interactive", {})


                                                if interactive.get("type") == "list_reply":
                                                    selected_option = interactive.get("list_reply", {}).get("id")
                                                    print(f"📋 User selected: {selected_option}")

                                                    if selected_option == "book_ticket":

                                                        url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_IDcc}/messages"
                                                        headers = {
                                                            "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                            "Content-Type": "application/json"
                                                        }

                                                        payload = {
                                                            "messaging_product": "whatsapp",
                                                            "to": sender_id,
                                                            "type": "interactive",
                                                            "interactive": {
                                                                "type": "list",
                                                                "header": {
                                                                    "type": "text",
                                                                    "text": "🚍 CAG TOURS MENU"
                                                                },
                                                                "body": {
                                                                    "text": (
                                                                        "Okay. Kindly select the route of travel for which you want to book a ticket on the menu below. ⬇️"
                                                                    )
                                                                },
                                                                "action": {
                                                                    "button": "ROUTE SELECTION",
                                                                    "sections": [
                                                                        {
                                                                            "title": "📦ROUTES",
                                                                            "rows": [
                                                                                {"id": "HreByo", "title": "Harare to Bulawayo", "description": "BUS FARE: USD 15"},
                                                                                {"id": "HreCheg", "title": "Harare to Chegutu", "description": "BUS FARE: USD 3"},
                                                                                {"id": "HreKad", "title": "Harare to Kadoma", "description": "BUS FARE: USD 5"},
                                                                                {"id": "HreKwek", "title": "Harare to Kwekwe", "description": "BUS FARE: USD 8"},
                                                                                {"id": "HreGwe", "title": "Harare to Gweru", "description": "BUS FARE: USD 10"},
                                                                                {"id": "ByoHre", "title": "Bulawayo to Harare", "description": "BUS FARE: USD 15"},
                                                                            ]
                                                                        }
                                                                    ]
                                                                }
                                                            }
                                                        }



                                                        # Send the request to WhatsApp
                                                        response = requests.post(url, headers=headers, json=payload)

                                                        # Optional: Print result for debugging
                                                        print(response.status_code)
                                                        print(response.text)
                                                
                                                    elif selected_option == "HreByo" or selected_option == "ByoHre" or selected_option == "HreCheg" or selected_option == "HreKad" or selected_option == "HreKwek" or selected_option == "HreGwe"  :

                                                        if selected_option == "HreByo":

                                                            route = "Harare to Bulawayo"
                                                            amount = "15"

                                                        elif selected_option == "ByoHre":

                                                            route = "Bulawayo to Harare"
                                                            amount = "15"

                                                        elif selected_option == "HreCheg":

                                                            route = "Harare to Chegutu"
                                                            amount = "3"

                                                        elif selected_option == "HreKad":

                                                            route = "Harare to Kadoma"
                                                            amount = "5"

                                                        elif selected_option == "HreKwe":

                                                            route = "Harare to Kwekwe"
                                                            amount = "8"

                                                        elif selected_option == "HreGwe":

                                                            route = "Harare to Gweru"
                                                            amount = "10"

                                                        cursor.execute("SELECT status FROM cagwatick WHERE idwanumber = %s", (sender_id[-9:],))
                                                        rows = cursor.fetchall()

                                                        if rows:
                                                            # Step 2: Check if any row has empty or NULL status
                                                            has_empty_status = any(row[0] in (None, '') for row in rows)

                                                            if not has_empty_status:
                                                                # No empty status found, safe to insert a new row
                                                                cursor.execute("""
                                                                    INSERT INTO cagwatick (idwanumber, route, fare)
                                                                    VALUES (%s, %s, %s)
                                                                """, (sender_id[-9:], route, amount))

                                                                connection.commit()

                                                            else:
                                                                print("Not inserting: an existing row with empty status found for this sender_id.")
                                                        else:
                                                            # sender_id does not exist at all — insert new
                                                            cursor.execute("""
                                                                INSERT INTO cagwatick (idwanumber, route, fare)
                                                                VALUES (%s, %s, %s)
                                                            """, (sender_id[-9:], route, amount))

                                                            connection.commit()


                                                        url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_IDcc}/messages"
                                                        headers = {
                                                            "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                            "Content-Type": "application/json"
                                                        }

                                                        payload = {
                                                            "messaging_product": "whatsapp",
                                                            "to": sender_id,
                                                            "type": "interactive",
                                                            "interactive": {
                                                                "type": "list",
                                                                "header": {
                                                                    "type": "text",
                                                                    "text": "🚍 DEPARTURE TIME"
                                                                },
                                                                "body": {
                                                                    "text": (
                                                                        "Okay. Kindly select the departure time from Harare for which you want to book a ticket on the menu below. ⬇️"
                                                                    )
                                                                },
                                                                "action": {
                                                                    "button": "DEPARTURE TIME",
                                                                    "sections": [
                                                                        {
                                                                            "title": "DEPARTURE TIME",
                                                                            "rows": [
                                                                                {"id": "8am", "title": "8 am"},
                                                                                {"id": "9am", "title": "9 am"},
                                                                                {"id": "2pm", "title": "2 pm"},

                                                                            ]
                                                                        }
                                                                    ]
                                                                }
                                                            }
                                                        }



                                                        # Send the request to WhatsApp
                                                        response = requests.post(url, headers=headers, json=payload)

                                                        # Optional: Print result for debugging
                                                        print(response.status_code)
                                                        print(response.text)


                                                    elif selected_option == "8am" or selected_option == "9am" or  selected_option == "2pm":

                                                        cursor.execute("""
                                                            SELECT id FROM cagwatick
                                                            WHERE idwanumber = %s
                                                            ORDER BY id DESC
                                                            LIMIT 1
                                                        """, (sender_id[-9:],))
                                                        result = cursor.fetchone()

                                                        if result:
                                                            highest_id = result[0]
                                                            cursor.execute("""
                                                                UPDATE cagwatick
                                                                SET time = %s
                                                                WHERE id = %s
                                                            """, (selected_option, highest_id))

                                                            connection.commit()

                                                        else:
                                                            print("No row found for this sender_id.")


                                                        url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_IDcc}/messages"
                                                        headers = {
                                                            "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                            "Content-Type": "application/json"
                                                        }

                                                        payload = {
                                                            "messaging_product": "whatsapp",
                                                            "to": sender_id,
                                                            "type": "interactive",
                                                            "interactive": {
                                                                "type": "list",
                                                                "header": {
                                                                    "type": "text",
                                                                    "text": "🚍 PAYMENT METHODS"
                                                                },
                                                                "body": {
                                                                    "text": (
                                                                        "Okay. Kindly select the payment method that you would like to use to book a ticket on the menu below. ⬇️"
                                                                    )
                                                                },
                                                                "action": {
                                                                    "button": "PAYMENT METHODS",
                                                                    "sections": [
                                                                        {
                                                                            "title": "PAYMENT METHODS",
                                                                            "rows": [
                                                                                {"id": "ecocash", "title": "EcoCash"},
                                                                                {"id": "onemoney", "title": "OneMoney"},
                                                                            ]
                                                                        }
                                                                    ]
                                                                }
                                                            }
                                                        }



                                                        # Send the request to WhatsApp
                                                        response = requests.post(url, headers=headers, json=payload)

                                                        # Optional: Print result for debugging
                                                        print(response.status_code)
                                                        print(response.text)


                                                    elif selected_option == "ecocash":


                                                        cursor.execute("""
                                                            SELECT id FROM cagwatick
                                                            WHERE idwanumber = %s
                                                            ORDER BY id DESC
                                                            LIMIT 1
                                                        """, (sender_id[-9:],))
                                                        result = cursor.fetchone()

                                                        if result:
                                                            highest_id = result[0]
                                                            cursor.execute("""
                                                                UPDATE cagwatick
                                                                SET paymethod = %s
                                                                WHERE id = %s
                                                            """, (selected_option, highest_id))

                                                            connection.commit()

                                                        else:
                                                            print("No row found for this sender_id.")



                                                        url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_IDcc}/messages"
                                                        headers = {
                                                            "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                            "Content-Type": "application/json"
                                                        }

                                                        payload = {
                                                            "messaging_product": "whatsapp",
                                                            "to": sender_id,
                                                            "type": "interactive",
                                                            "interactive": {
                                                                "type": "list",
                                                                "header": {
                                                                    "type": "text",
                                                                    "text": "🚍 CAG TOURS MAIN MENU"
                                                                },
                                                                "body": {
                                                                    "text": (
                                                                        "Ok. Kindly provide the EcoCash number that you would like to use to pay. \n\n eg `0777111234`"
                                                                    )
                                                                },
                                                                "action": {
                                                                    "button": "📋 CAG TOURS MENU",
                                                                    "sections": [
                                                                        {
                                                                            "title": "📦 CAG TOURS SERVICES",
                                                                            "rows": [
                                                                                {
                                                                                    "id": "book_ticket",
                                                                                    "title": "Book a Ticket",
                                                                                    "description": "Reserve your seat instantly"
                                                                                },
                                                                                {
                                                                                    "id": "routes",
                                                                                    "title": "View Routes",
                                                                                    "description": "Get info regarding our travel routes"
                                                                                },
                                                                                {
                                                                                    "id": "parcel_delivery",
                                                                                    "title": "Parcel Delivery",
                                                                                    "description": "Send or collect packages"
                                                                                },
                                                                                {
                                                                                    "id": "find_stop",
                                                                                    "title": "Find Bus Stop",
                                                                                    "description": "Locate nearest pick-up point"
                                                                                },
                                                                                {
                                                                                    "id": "promotions",
                                                                                    "title": "Promotions & Offers",
                                                                                    "description": "Current discounts & deals"
                                                                                }
                                                                            ]
                                                                        },
                                                                        {
                                                                            "title": "🚌 ABOUT CAG TOURS",
                                                                            "rows": [
                                                                                {
                                                                                    "id": "know_more",
                                                                                    "title": "Know More",
                                                                                    "description": "Our story, mission & travel experience"
                                                                                },
                                                                                {
                                                                                    "id": "why_choose",
                                                                                    "title": "Why Choose Us",
                                                                                    "description": "Luxury, safety & comfort explained"
                                                                                }
                                                                            ]
                                                                        },
                                                                        {
                                                                            "title": "🛎 CUSTOMER SERVICE",
                                                                            "rows": [
                                                                                {
                                                                                    "id": "faqs",
                                                                                    "title": "❓ FAQs",
                                                                                    "description": "Get answers to common questions"
                                                                                },
                                                                                {
                                                                                    "id": "policies",
                                                                                    "title": "Travel Policies",
                                                                                    "description": "Baggage rules, safety, refunds"
                                                                                },
                                                                                {
                                                                                    "id": "get_help",
                                                                                    "title": "Get Help",
                                                                                    "description": "Talk to a support agent now"
                                                                                }
                                                                            ]
                                                                        }
                                                                    ]
                                                                }
                                                            }
                                                        }



                                                        # Send the request to WhatsApp
                                                        response = requests.post(url, headers=headers, json=payload)

                                                        # Optional: Print result for debugging
                                                        print(response.status_code)
                                                        print(response.text)


                                                    elif selected_option == "faqs":
                                                        button_id_leave_type = str(selected_option)

                                                        sections = [
                                                            {
                                                                "title": "FAQs",
                                                                "rows": [
                                                                    {"id": "Routes", "title": "routes"},
                                                                    {"id": "BusTypes", "title": "Bus Types"},
                                                                    {"id": "Privatehires", "title": "Do you do private hires?"},
                                                                    {"id": "Sunday", "title": "Do you work on Sundays"},
                                                                ]
                                                            }
                                                        ]

                                                        send_whatsapp_list_messagecc(
                                                            sender_id, 
                                                            "Ok. Select a FAQ for more info...", 
                                                            "CAG TOURS FAQs",
                                                            sections) 
                                                        
                                                    elif selected_option == "Fares":
                                                        button_id_leave_type = str(selected_option)

                                                        table_message = (
                                                                
                                                                "*🚌 Bus Fares*\n\n"
                                                                "```"
                                                                "Cities/Towns         | Fare\n"
                                                                "-------------------- |-----\n"
                                                                "Harare ↔️ Bulawayo   | $15\n"
                                                                "Harare ↔️ Kariba     | $14\n"
                                                                "Harare ↔️ Vic Falls  | $25\n"
                                                                "Chitungwiza ↔️ Mutare| $10\n"
                                                                "Harare ↔️ Gokwe      | $15"
                                                                "```"
                                                            )

                                                        payload = {
                                                            "messaging_product": "whatsapp",
                                                            "to": sender_id,
                                                            "type": "text",
                                                            "text": {
                                                                "body": table_message,
                                                                "preview_url": False
                                                            }
                                                        }

                                                        headers = {
                                                            "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                            "Content-Type": "application/json"
                                                        }

                                                        url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_IDcc}/messages"

                                                        response = requests.post(url, headers=headers, json=payload)

                                                        sections = [
                                                            {
                                                                "title": "Leave Type Options",
                                                                "rows": [
                                                                    {"id": "Book", "title": "Book A Bus Ticket"},
                                                                    {"id": "View", "title": "View Route & Times"},
                                                                    {"id": "Contact", "title": "Contact Support"},
                                                                    {"id": "FAQs", "title": "FAQs"},
                                                                ]
                                                            }
                                                        ]

                                                        send_whatsapp_list_messagecc(
                                                            sender_id, 
                                                            f"Kindly select an option for enquiry.", 
                                                            "CAG TOURS Options",
                                                            sections) 


                                                    elif selected_option == "routes":
                                                        button_id_leave_type = str(selected_option)

                                                        url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_IDcc}/messages"
                                                        headers = {
                                                            "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                            "Content-Type": "application/json"
                                                        }

                                                        payload = {
                                                            "messaging_product": "whatsapp",
                                                            "to": sender_id,
                                                            "type": "interactive",
                                                            "interactive": {
                                                                "type": "list",
                                                                "header": { "type": "text", "text": "🚌 CAG TOURS Schedule" },
                                                                "body": { "text": "Select a destination to view bus times:" },
                                                                "action": {
                                                                    "button": "View Routes",
                                                                    "sections": [{
                                                                        "title": "Available Routes",
                                                                        "rows": [
                                                                            { "id": "vbulawayo", "title": "Harare → Bulawayo", "description": "US$15" },
                                                                            { "id": "vmutare", "title": "Chitungwiza → Mutare", "description": "6:00 AM" },
                                                                            { "id": "vkariba", "title": "Harare → Kariba", "description": "US$14" },
                                                                            { "id": "vvictoria", "title": "Harare → Victoria Falls", "description": "US$25" },
                                                                            { "id": "vgokwe", "title": "Harare → Gokwe", "description": "10 daily buses" },
                                                                            { "id": "vkaroi", "title": "Harare → Karoi / Magunje", "description": "7 departures" },
                                                                            { "id": "vhonde", "title": "Harare → Honde Valley", "description": "10 departures" },
                                                                            { "id": "vchirundu", "title": "Harare → Chirundu", "description": "9:00 AM" },
                                                                            { "id": "vmukumbura", "title": "Harare → Mukumbura", "description": "9 departures" },
                                                                            { "id": "vall", "title": "📄 View Full Schedule PDF", "description": "Download file" }
                                                                        ]
                                                                    }]
                                                                }
                                                            }
                                                        }

                                                        requests.post(url, headers=headers, json=payload)
                                                                                                            


                                                    elif selected_option == "Contact":
                                                        button_id_leave_type = str(selected_option)

                                                        send_whatsapp_messagecc(sender_id, "✅ Okay. A Customer Representative has been notified to assit you. They will contact you shortly.")

                                                    elif selected_option == "vkariba":
                                                        button_id_leave_type = str(selected_option)

                                                        sections = [
                                                            {
                                                                "title": "Leave Type Options",
                                                                "rows": [
                                                                    {"id": "Book", "title": "Book A Bus Ticket"},
                                                                    {"id": "View", "title": "View Route & Times"},
                                                                    {"id": "Contact", "title": "Contact Support"},
                                                                    {"id": "FAQs", "title": "FAQs"},
                                                                    {"id": "Download", "title": "Download Brochure"},
                                                                ]
                                                            }
                                                        ]

                                                        send_whatsapp_list_messagecc(
                                                            sender_id, 
                                                            f"Departures from Mbare Musika Rank. \n\n Departure Times: 7:00AM, 8:30AM, 10:00AM, 12:30PM, 2:30PM, 8:00PM.\n\n Kindly select an option for enquiry.", 
                                                            "CAG TOURS Options",
                                                            sections) 
                                                        

                                                    elif selected_option == "vbulawayo":
                                                        button_id_leave_type = str(selected_option)

                                                        sections = [
                                                            {
                                                                "title": "Leave Type Options",
                                                                "rows": [
                                                                    {"id": "Book", "title": "Book A Bus Ticket"},
                                                                    {"id": "View", "title": "View Route & Times"},
                                                                    {"id": "Contact", "title": "Contact Support"},
                                                                    {"id": "FAQs", "title": "FAQs"},
                                                                    {"id": "Download", "title": "Download Brochure"},
                                                                ]
                                                            }
                                                        ]

                                                        send_whatsapp_list_messagecc(
                                                            sender_id, 
                                                            f"Departures from Harare Showgrounds (CAG House). \n\n Departure Times: 8:00AM, 9:00AM, 2:00PM.\n\n Kindly select an option for enquiry.", 
                                                            "CAG TOURS Options",
                                                            sections) 

                                                    elif selected_option == "vmutare":
                                                        button_id_leave_type = str(selected_option)

                                                        sections = [
                                                            {
                                                                "title": "Leave Type Options",
                                                                "rows": [
                                                                    {"id": "Book", "title": "Book A Bus Ticket"},
                                                                    {"id": "View", "title": "View Route & Times"},
                                                                    {"id": "Contact", "title": "Contact Support"},
                                                                    {"id": "FAQs", "title": "FAQs"},
                                                                    {"id": "Download", "title": "Download Brochure"},
                                                                ]
                                                            }
                                                        ]

                                                        send_whatsapp_list_messagecc(
                                                            sender_id, 
                                                            f"Departures from Harare Showgrounds (CAG House). \n\n Departure Times: 8:00AM, 9:00AM, 2:00PM.\n\n Kindly select an option for enquiry.", 
                                                            "CAG TOURS Options",
                                                            sections)  


                                            elif message.get("type") == "text":

                                                text = message.get("text", {}).get("body", "").lower()
                                                print(f"📨 Message from {sender_id}: {text}")
                                                
                                                print("yearrrrrrrrrrrrrrrrrrrrrrrrrrrssrsrsrsrsrs")

                                                digits_only = ''.join(filter(str.isdigit, text.replace(" ", "")))

                                                if len(digits_only) > 9:


                                                    cursor.execute("""
                                                        SELECT id FROM cagwatick
                                                        WHERE idwanumber = %s
                                                        ORDER BY id DESC
                                                        LIMIT 1
                                                    """, (sender_id[-9:],))
                                                    result = cursor.fetchone()

                                                    if result:
                                                        highest_id = result[0]
                                                        cursor.execute("""
                                                            UPDATE cagwatick
                                                            SET ecocashnum = %s
                                                            WHERE id = %s
                                                        """, (digits_only, highest_id))

                                                        connection.commit()

                                                    else:
                                                        print("No row found for this sender_id.")

                                                    print("Message contains more than 9 digits after removing spaces")
                                                    # You can now process it as needed, e.g., assume it's an ID number or phone number

                                                    print("yeah")

                                                    try:
                                                    
                                                        paynow = Paynow('20625',
                                                                        'f6559511-ab13-45b0-b75b-07b36890f6a6',
                                                                        'https://eds-dfym.onrender.com/paynow/return',
                                                                        'https://eds-dfym.onrender.com/paynow/result/update'
                                                                        )
                                                        
                                                        print(paynow)

                                                        payment = paynow.create_payment('Order', 'takudzwazvaks@gmail.com')

                                                        payment.add('Payment for stuff', 0.01)


                                                        cursor.execute("""
                                                            SELECT id, idwanumber, route, time, paymethod, fare, ecocashnum FROM cagwatick
                                                            WHERE idwanumber = %s
                                                            ORDER BY id DESC
                                                            LIMIT 1
                                                        """, (sender_id[-9:],))
                                                        result = cursor.fetchone()

                                                        if result:
                                                            highest_id = result[0]

                                                            send_whatsapp_messagecc(
                                                                sender_id, 
                                                                f"We are initiating your ticket for route `{result[2]}` on bus departing at `{result[3]}`.\n\n You will receive a USSD prompt on `{result[6]}` shortly to provide your EcoCah PIN to process your USD {result[5]} bus fare payment."
                                                            ) 

                                                        else:
                                                            print("No row found for this sender_id.")



                                                        response = paynow.send_mobile(payment, digits_only, 'ecocash')

                                                        print("pending")

                                    

                                                        if(response.success):

                                                            print('success')
                                                            poll_url = response.poll_url

                                                            print("Poll Url: ", poll_url)


                                                            cursor.execute("""
                                                                SELECT id, idwanumber, route, time, paymethod, fare, ecocashnum FROM cagwatick
                                                                WHERE idwanumber = %s
                                                                ORDER BY id DESC
                                                                LIMIT 1
                                                            """, (sender_id[-9:],))
                                                            result = cursor.fetchone()

                                                            if result:
                                                                highest_id = result[0]
                                                                cursor.execute("""
                                                                    UPDATE cagwatick
                                                                    SET pollurl = %s
                                                                    WHERE id = %s
                                                                """, (poll_url, highest_id))

                                                                connection.commit()

                                                            else:
                                                                print("No row found for this sender_id.")

                                                            status = paynow.check_transaction_status(poll_url)

                                                            time.sleep(20)

                                                            print("Payment Status: ", status.status)


                                                        return 'OK', 200


                                                    
                                                    except Exception as e:
                                                        print(e)

                                                    return 'OK', 200










                                                else:


                                                    url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_IDcc}/messages"
                                                    headers = {
                                                        "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                        "Content-Type": "application/json"
                                                    }

                                                    payload = {
                                                        "messaging_product": "whatsapp",
                                                        "to": sender_id,
                                                        "type": "interactive",
                                                        "interactive": {
                                                            "type": "list",
                                                            "header": {
                                                                "type": "text",
                                                                "text": "🚍 CAG TOURS MAIN MENU"
                                                            },
                                                            "body": {
                                                                "text": (
                                                                    "Welcome aboard! 👋\n\n"
                                                                    "Explore our available routes, services, and customer support options.\n"
                                                                    "Tap *OPEN MENU* below to get started. ⬇️"
                                                                )
                                                            },
                                                            "action": {
                                                                "button": "📋 CAG TOURS MENU",
                                                                "sections": [
                                                                    {
                                                                        "title": "📦 CAG TOURS SERVICES",
                                                                        "rows": [
                                                                            {
                                                                                "id": "book_ticket",
                                                                                "title": "Book a Ticket",
                                                                                "description": "Reserve your seat instantly"
                                                                            },
                                                                            {
                                                                                "id": "routes",
                                                                                "title": "View Routes",
                                                                                "description": "Get info regarding our travel routes"
                                                                            },
                                                                            {
                                                                                "id": "parcel_delivery",
                                                                                "title": "Parcel Delivery",
                                                                                "description": "Send or collect packages"
                                                                            },
                                                                            {
                                                                                "id": "find_stop",
                                                                                "title": "Find Bus Stop",
                                                                                "description": "Locate nearest pick-up point"
                                                                            },
                                                                            {
                                                                                "id": "promotions",
                                                                                "title": "Promotions & Offers",
                                                                                "description": "Current discounts & deals"
                                                                            }
                                                                        ]
                                                                    },
                                                                    {
                                                                        "title": "🚌 CAG TOURS",
                                                                        "rows": [
                                                                            {
                                                                                "id": "know_more",
                                                                                "title": "Know More",
                                                                                "description": "Our story, mission & travel experience"
                                                                            },
                                                                            {
                                                                                "id": "why_choose",
                                                                                "title": "Why Choose Us",
                                                                                "description": "Luxury, safety & comfort explained"
                                                                            }
                                                                        ]
                                                                    },
                                                                    {
                                                                        "title": "🛎 CUSTOMER SERVICE",
                                                                        "rows": [
                                                                            {
                                                                                "id": "faqs",
                                                                                "title": "❓ FAQs",
                                                                                "description": "Get answers to common questions"
                                                                            },
                                                                            {
                                                                                "id": "policies",
                                                                                "title": "Travel Policies",
                                                                                "description": "Baggage rules, safety, refunds"
                                                                            },
                                                                            {
                                                                                "id": "get_help",
                                                                                "title": "Get Help",
                                                                                "description": "Talk to a support agent now"
                                                                            }
                                                                        ]
                                                                    }
                                                                ]
                                                            }
                                                        }
                                                    }



                                                    # Send the request to WhatsApp
                                                    response = requests.post(url, headers=headers, json=payload)

                                                    # Optional: Print result for debugging
                                                    print(response.status_code)
                                                    print(response.text)


                                        finally:
                                            if connection:
                                                print('DONE')

                except Exception as e:
                    print(e)

                return "OK", 200








################### end of test #########################





#####################################

            elif display_phone_number == "263789339777":
                ACCESS_TOKEN = "EAATESj1oB5YBPIzFCv7ulvosr2S2ZAiWBJrFp7bti6L0ZCWS2AOz5dUABlJ6q16a4hRwEXdq5vZAP5tp4rGXfOQ2sx0hg1EOwMpL002eqUrygbPc3jkY8FPOzR7c6tMvKJxT3XxXP8Qp9U1n30MIMVcNy9JUCZB8UyIwaAZBAjf2U32TVTwSBJlSeHoNYrGH0dwZDZD"
                PHONE_NUMBER_ID = "756962384159644"
                VERIFY_TOKEN = "2644686099068373"
                WHATSAPP_API_URL = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
                power = "Alluire Marketing Agency"
                bot = "Alluire"

            elif display_phone_number == "263772860855":
                ACCESS_TOKEN = "EAATESj1oB5YBPOHzgE9i7DOSQGWJARySZABPAutZBQKBZCOhMfZAOtweOol9LhkR1dIfzMhplJMN9ECLvANqOYR0io2q5twXiog8iU9kV5M48T8NHJVcWDtjWQCMxXuCXNwEAJOw3F2UCBE5iEGECFSeYOe68BAxPwnDzADPTS2vuZAbfxYAEA0cacrWY0qbl5wZDZD"
                PHONE_NUMBER_ID = "709678028897732"
                VERIFY_TOKEN = "1122583909768342"
                WHATSAPP_API_URL = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
                power = "Echelon Equipment Pvt Ltd"
                bot = "EDS"

            def send_whatsapp_message(to, text, buttons=None):

                print("send mess initialised")

                """Function to send a WhatsApp message using Meta API, with optional buttons."""
                headers = {
                    "Authorization": f"Bearer {ACCESS_TOKEN}",
                    "Content-Type": "application/json"
                }

                # If buttons are provided, send an interactive message
                if buttons:
                    data = {
                        "messaging_product": "whatsapp",
                        "to": to,
                        "type": "interactive",
                        "interactive": {
                            "type": "button",
                            "body": {"text": text},
                            "action": {
                                "buttons": buttons
                            }
                        }
                    }
                else:
                    # Send a normal text message
                    print("inside else")

                    data = {
                        "messaging_product": "whatsapp",
                        "to": to,
                        "type": "text",
                        "text": {"body": text}
                    }
                
                # Debugging logs
                print("✅ Sending message to:", to)
                print("📩 Message body:", text)

                """try:
                    response_json = response.json()
                    print("📝 WhatsApp API Response Data:", response_json)
                except Exception as e:
                    print("❌ Error parsing response JSON:", e)"""

                try:
                    print("trying in def")

                    response = requests.post(WHATSAPP_API_URL, headers=headers, json=data)
                    print("📡 WhatsApp API Response Status:", response.status_code)
                    print("📡 WhatsApp API Response Text:", response.text)

                    response.raise_for_status()  # will throw if not 2xx

                    print("✅ Message sent successfully.")
                    print("📝 Response JSON:", response.json())

                    print("done trying")
                    return response.json()
                
                except requests.exceptions.RequestException as e:
                    print("❌ WhatsApp API Error:", e)
                    return {"error": str(e)}


            def send_whatsapp_list_message(recipient, text, list_title, sections):

                headers = {
                    "Authorization": f"Bearer {ACCESS_TOKEN}",
                    "Content-Type": "application/json"
                }
                
                payload = {
                    "messaging_product": "whatsapp",
                    "recipient_type": "individual",
                    "to": recipient,
                    "type": "interactive",
                    "interactive": {
                        "type": "list",
                        "header": {
                            "type": "text",
                            "text": list_title
                        },
                        "body": {
                            "text": text
                        },
                        "action": {
                            "button": "Select Option",
                            "sections": sections
                        }
                    }
                }
                
                response = requests.post(WHATSAPP_API_URL,headers=headers,json=payload)

                print("✅ Sending message to:", recipient)
                print("📩 Message body:", text)
                print("📡 WhatsApp API Response Status:", response.status_code)  

                print("List message response:", response.json())
                return response


            print("📥 Full incoming data:", json.dumps(data, indent=2))

            if data and "entry" in data:
                for entry in data["entry"]:
                    for change in entry["changes"]:
                        if "messages" in change["value"]:
                            for message in change["value"]["messages"]:

                                conversation_id = str(uuid.uuid4())
                                session['conversation_id'] = conversation_id
                            

                                sender_id = message["from"]
                                sender_number = sender_id[-9:]
                                print(f"📱 Conversation {conversation_id}: Sender's WhatsApp number: {sender_number}")
                                session['client'] = str(sender_number)

                                external_database_url = "postgresql://lmsdatabase_8ag3_user:6WD9lOnHkiU7utlUUjT88m4XgEYQMTLb@dpg-ctp9h0aj1k6c739h9di0-a.oregon-postgres.render.com/lmsdatabase_8ag3"

                                try:
                                    connection = psycopg2.connect(external_database_url)
                                    cursor = connection.cursor()

                                    cursor.execute("""
                                        SELECT DISTINCT table_name
                                        FROM information_schema.columns
                                        WHERE table_schema = 'public'
                                        AND column_name = 'password'
                                    """)
                                    tables = cursor.fetchall()

                                    found = False

                                    for table in tables:
                                        table_name = table[0]  

                                        print(table_name)
                                        print(sender_number)
                                        
                                        query = f"""
                                            SELECT * FROM {table_name}
                                            WHERE whatsapp::TEXT LIKE %s
                                        """
                                        cursor.execute(query, (f"%{sender_number}",))
                                        result = cursor.fetchone()

                                        print(result)

                                        if result:

                                            found = True 

                                            id_user = result[0]  
                                            first_name = result[1]  
                                            last_name = result[2]  
                                            whatsapp_foc_8 = f"0{result[3]}"
                                            email_foc_8 = result[4]
                                            address_foc_8 = result[5]
                                            role_foc_8 = result[8]
                                            days_days_balance = result[13]
                                            company_reg = table_name[:-4]  

                                            print(id_user)
                                            print(first_name)
                                            print(last_name)
                                            print(role_foc_8)

                                            print(f"Credentials found in table: {table_name}")
                                            first_column_value = result[0]
                                            print(f"First column value: {first_column_value}")

                                            break  
                                            
                                    if not found:
                                        send_whatsapp_message(
                                            sender_id, 
                                            "Oops, you are not registered. Kindly get in touch with your leave administrator for assistance."
                                        )

                                        return jsonify({"status": "received"}), 200

                                finally:
                                    if connection:
                                        print('DONE')

                                if role_foc_8 == "Ordinary User":

                                    table_namexxxx = company_reg + "main"        

                                    query = f"SELECT id FROM {table_namexxxx} WHERE leaveapproverid = {str(id_user)};"
                                    cursor.execute(query)
                                    rows = cursor.fetchall()

                                    df_employeesempapp = pd.DataFrame(rows, columns=["id"])

                                    if len(df_employeesempapp) == 0:

                                        try:
                                        
                                            if message.get("type") == "interactive":
                                                interactive = message.get("interactive", {})


                                                if interactive.get("type") == "list_reply":
                                                    selected_option = interactive.get("list_reply", {}).get("id")
                                                    print(f"📋 User selected: {selected_option}")
                                                    button_id = ""

                                                elif interactive.get("type") == "button_reply":
                                                    button_id = interactive.get("button_reply", {}).get("id")
                                                    print(f"🔘 Button clicked: {button_id}")
                                                    selected_option = ""

                                                if selected_option in ["Annual","Sick","Study","Parental", "Bereavement","Other"] :
                                                    button_id_leave_type = str(selected_option)

                                                    cursor.execute("""
                                                        DELETE FROM whatsapptempapplication
                                                        WHERE empidwa = %s
                                                    """, (str(id_user),))  
                                                    
                                                    connection.commit()

                                                    cursor.execute(f"""
                                                        INSERT INTO whatsapptempapplication (empidwa, leavetypewa, companynamewa)
                                                        VALUES (%s, %s, %s)
                                                    """, (id_user, button_id_leave_type, company_reg))

                                                    connection.commit()

                                                    send_whatsapp_message(
                                                        sender_id, 
                                                        f"Ok. When would you like your {selected_option} Leave to start {first_name}?\n\n"
                                                        "Please enter your response using the format: 👇🏻\n"
                                                        "`start 24 january 2025`"
                                                    )

                                                    continue

                                                elif "reminder" in button_id.lower():

                                                    app_id = button_id.split("_")[1]
                                                    print(app_id)

                                                    try:
                                                    
                                                        print ("eissssssssshhhhhhhhhhhhhhhhhhhhhhhhhhhh")

                                                        table_name = company_reg + 'main'
                                                        company_name = company_reg.replace("_", " ").title()
                                                        table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                        table_name_apps_approved = f"{company_reg}appsapproved"


                                                        if not app_id:
                                                            print("none on appid")

                                                            return jsonify({"message": "Application ID is missing."}), 400

                                                        print(table_name_apps_pending_approval)

                                                        query = f"SELECT * FROM {table_name_apps_pending_approval} WHERE appid = %s;"
                                                        cursor.execute(query, (app_id,))
                                                        result = cursor.fetchone()
                                                        app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf, statuspre = result
                                                        print("chiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
                                                        print(employee_number)
                                                        print(approver_name)

                                                        try:

                                                            query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation, department FROM {table_name};"
                                                            cursor.execute(query)
                                                            rows = cursor.fetchall()

                                                            df_employees = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Address", "Role","Leave Approver Name","Leave Approver ID","Leave Approver Email", "Leave Approver WhatsAapp", "Leave Days Balance","Days Accumulated per Month", "Department"])
                                                            print(df_employees)
                                                            userdf = df_employees[df_employees['id'] == int(np.int64(employee_number))].reset_index()
                                                            print("yeaarrrrr")
                                                            print(userdf)
                                                            firstname = userdf.iat[0,2]
                                                            surname = userdf.iat[0,3]
                                                            whatsapp = userdf.iat[0,4]
                                                            address = userdf.iat[0,6]
                                                            email = userdf.iat[0,5]
                                                            fullnamedisp = firstname + ' ' + surname
                                                            leaveapprovername = userdf.iat[0,8]
                                                            leaveapproverid = userdf.iat[0,9]
                                                            leaveapproveremail = userdf.iat[0, 10]
                                                            leaveapproverwhatsapp = userdf.iat[0,11]
                                                            role = userdf.iat[0,7]
                                                            leavedaysbalance = userdf.iat[0,12]
                                                            department = userdf.iat[0,14] 
                                                            print('check')

                                                            departmentdf = df_employees[df_employees['Department'] == department].reset_index()
                                                            numberindepartment = len(departmentdf)
                                                            
                                                            startdatex = pd.Timestamp(start_date)
                                                            enddatex = pd.Timestamp(end_date)

                                                            leave_dates = pd.date_range(startdatex, enddatex)

                                                            query = f"""
                                                                SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate,
                                                                    leaveenddate, leavedaysappliedfor, approvalstatus, statusdate,
                                                                    leavedaysbalancebf, department
                                                                FROM {table_name_apps_approved}
                                                                WHERE department = %s;
                                                            """
                                                            cursor.execute(query, (department,))
                                                            rows = cursor.fetchall()

                                                            df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf","department"]) 
                                                            df_employeesappsapprovedcheck["leavestartdate"] = pd.to_datetime(df_employeesappsapprovedcheck["leavestartdate"])
                                                            df_employeesappsapprovedcheck["leaveenddate"] = pd.to_datetime(df_employeesappsapprovedcheck["leaveenddate"])
            
                                                            df_employeesappsapprovedcheck.dropna(subset=["leavestartdate", "leaveenddate"], inplace=True)
                                                            # Create daily impact report
                                                            impact_report = []

                                                            for date in leave_dates:

                                                                date = pd.Timestamp(date)

                                                                print(type(date))  # Should be pandas._libs.tslibs.timestamps.Timestamp or datetime.datetime
                                                                print(df_employeesappsapprovedcheck.dtypes)  # Check all datetime columns

                                                                on_leave = ((df_employeesappsapprovedcheck["leavestartdate"] <= date) & (df_employeesappsapprovedcheck["leaveenddate"] >= date)).sum()
                                                                remaining = numberindepartment - on_leave - 1  # subtract 1 for the new leave
                                                                impact_report.append({
                                                                    "date": date,  # <=== Keep as datetime, don't convert to string
                                                                    "on leave": on_leave + 1,
                                                                    "employees remaining": remaining
                                                                })

                                                            # Convert to DataFrame for display
                                                            impact_df = pd.DataFrame(impact_report)
                                                            print("IMPAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACT")
                                                            print(impact_df)
                                                            print(numberindepartment)

                                                            impact_df["date"] = pd.to_datetime(impact_df["date"], format="%Y-%m-%d")
                                                            impact_df = impact_df[impact_df["date"].dt.weekday != 6].copy()

                                                            change = (impact_df[["on leave", "employees remaining"]] != impact_df[["on leave", "employees remaining"]].shift()).any(axis=1)
                                                            change.iloc[0] = True  # ensure the first row starts a group
                                                            impact_df["group"] = change.cumsum()

                                                            statements = []
                                                            for _, group_df in impact_df.groupby("group"):
                                                                start = group_df["date"].iloc[0].strftime("%d %B %Y")
                                                                end = group_df["date"].iloc[-1].strftime("%d %B %Y")
                                                                on_leave = group_df["on leave"].iloc[0]
                                                                remaining = group_df["employees remaining"].iloc[0]
                                                                
                                                                if start == end:
                                                                    statements.append(f"On {start}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                                                                else:
                                                                    statements.append(f"From {start} to {end}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                                                                    # Combine all statements into a single variable
                                                            final_summary = "\n".join(statements)
                                                            # Print output
                                                            for s in statements:
                                                                print(s)

                                                        except Exception as e:
                                                            print(e)


                                                        sections = [
                                                            {
                                                                "title": "User Options",
                                                                "rows": [
                                                                    {"id": "Apply", "title": "Apply for Leave"},
                                                                    {"id": "Track", "title": "Track My Application"},
                                                                    {"id": "Checkbal", "title": "Check Days Balance"},
                                                                    {"id": "myhist", "title": "My Applications History"},
                                                                    {"id": "Myinfo", "title": "My Info"}
                                                                ]
                                                            }
                                                        ]


                                                        send_whatsapp_list_message(
                                                            sender_id, 
                                                            f"Hey {first_name}. A reminder has been sent to {approver_name} on {approver_whatsapp} to decide on your `{leave_days}-day {leave_type} leave` running from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}` that you applied for on `{date_applied.strftime('%d %B %Y')}`✅! \n Select an option below to continue 👇",
                                                        "User Options",
                                                        sections)

                                                        if approver_whatsapp:

                                                            try:

                                                                buttons = [
                                                                    {"type": "reply", "reply": {"id": f"Approve5appwa_{app_id}", "title": "Approve"}},
                                                                    {"type": "reply", "reply": {"id": f"Disapproveappwa_{app_id}", "title": "Disapprove"}},
                                                                ]
                                                                send_whatsapp_message(
                                                                    f"263{approver_whatsapp}", 
                                                                    f"Hey {approver_name}! 😊. A gentle reminder, you have a new `{leave_type}` Leave Application from `{first_name} {surname}` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}`.\n\n" 
                                                                    f"If you approve this leave application, {final_summary}\n\n"  
                                                                    f"Select an option below to either approve or disapprove the application."         
                                                                    , 
                                                                    buttons
                                                                )

                                                            except Exception as e:
                                                                print(e)



                                                    except Exception as e:
                                                        print(e)
                                                        return jsonify({"message": "Error approving leave application.", "error": str(e)}), 500
                                                


                                                elif selected_option == "Myinfo":

                                                    companyxx = company_reg.replace("_"," ").title()

                                                    try:

                                                        table_name = f"{company_reg}main"

                                                        query = f"SELECT id, firstname, surname, whatsapp, address, email, role, department, currentleavedaysbalance, monthlyaccumulation, leaveapprovername, leaveapproverwhatsapp, leaveapproveremail FROM {table_name} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        row = cursor.fetchone()

                                                        if row:

                                                            columns = ["ID", "First Name", "Surname", "WhatsApp", "Address", "Email", 
                                                                    "Role", "Department", "Leave Days", "Monthly Accrual", 
                                                                    "Approver", "Approver WhatsApp", "Approver Email"]

                                                            message_text = "*📄 Employee Details:*\n\n"
                                                            for col, val in zip(columns, row):
                                                                message_text += f"*{col}:* {val}\n"

                                                            sections = [
                                                                {
                                                                    "title": "User Options",
                                                                    "rows": [
                                                                        {"id": "Editname", "title": "Edit My Name"},
                                                                        {"id": "Editwhatsapp", "title": "Change My WhatsApp #"},
                                                                        {"id": "Editemail", "title": "Change My Email"},
                                                                        {"id": "Editwebpass", "title": "Change Web Password"},
                                                                        {"id": "Editaddress", "title": "Edit My Address"},
                                                                        {"id": "MyInfo", "title": "My Info"},
                                                                        {"id": "Menu", "title": "Main Menu"}
                                                                    ]
                                                                }
                                                            ]

                                                            send_whatsapp_list_message(
                                                                sender_id, 
                                                                f"Hey there {first_name}!\n Your information in {companyxx}'s Leave Management System is as follows;\n\n {message_text}", 
                                                            "User Options",
                                                            sections)

                                                    except Exception as e:

                                                        print(e)

                                                        send_whatsapp_message(f"+263710910052", f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully! Error; {e}")                      

                                                elif selected_option == "Menu" or button_id == "Menu" :

                                                    companyxx = company_reg.replace("_"," ").title()
                                                    
                                                    sections = [
                                                        {
                                                            "title": "User Options",
                                                            "rows": [
                                                                {"id": "Apply", "title": "Apply for Leave"},
                                                                {"id": "Track", "title": "Track My Application"},
                                                                {"id": "Checkbal", "title": "Check Days Balance"},
                                                                {"id": "myhist", "title": "My Applications History"},
                                                                {"id": "Myinfo", "title": "My Info"}
                                                            ]
                                                        }
                                                    ]


                                                    send_whatsapp_list_message(
                                                        sender_id, 
                                                        f"Hello {first_name} {last_name} from {companyxx}!\n\n {bot} LMS Bot Here 😎. How can I assist you?", 
                                                    "User Options",
                                                    sections)

                                                elif selected_option == "Track" or button_id == "Track":

                                                    table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                    table_name_apps_approved = f"{company_reg}appsapproved"
                                                    table_name_apps_declined = f"{company_reg}appsdeclined"
                                                    table_name_apps_cancelled = f"{company_reg}appscancelled"


                                                    query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leaveapproverwhatsapp  FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                    cursor.execute(query)
                                                    rows = cursor.fetchall()

                                                    df_employeesappspendingcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor", "leaveapproverwhatsapp"])    

                                                    if len(df_employeesappspendingcheck) == 0:

                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate, leavedaysbalancebf  FROM {table_name_apps_approved} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()
                                                        df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf"]) 

                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate, leavedaysbalancebf  FROM {table_name_apps_declined} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()
                                                        df_employeesappsdeclinedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf"])  
                                
                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate, leavedaysbalancebf  FROM {table_name_apps_cancelled} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()
                                                        df_employeesappscancelledcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf"])
                                
                                                        all_approved_declined = df_employeesappsapprovedcheck._append(df_employeesappsdeclinedcheck)
                                                        all_approved_declined_cancelled = all_approved_declined._append(df_employeesappscancelledcheck)

                                                        all_approved_declined_cancelled = all_approved_declined_cancelled.sort_values(by="appid", ascending=False)  

                                                        if len(all_approved_declined_cancelled) > 0:

                                                            print(f" hhhhhhhhhhhhhhhhhhhh  {all_approved_declined_cancelled.iat[0,8] }")

                                                            if all_approved_declined_cancelled.iat[0,8] == "Approved":

                                                                buttons = [
                                                                    {"type": "reply", "reply": {"id": "Revoke", "title": "Revoke Application"}},
                                                                    {"type": "reply", "reply": {"id": "myhist", "title": "Download My History"}},
                                                                    {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                                ]
                                                                send_whatsapp_message(
                                                                    sender_id, 
                                                                    f"Hey {first_name}, your recent `{all_approved_declined_cancelled.iat[0,2]}` Leave Application `[ID - {all_approved_declined_cancelled.iat[0,0]}]` that you applied for on `{all_approved_declined_cancelled.iat[0,4].strftime('%d %B %Y')}` for `{all_approved_declined_cancelled.iat[0,7]} days` from `{all_approved_declined_cancelled.iat[0,5].strftime('%d %B %Y')}` to `{all_approved_declined_cancelled.iat[0,6].strftime('%d %B %Y')}` was {all_approved_declined_cancelled.iat[0,8]}✅ by `{all_approved_declined_cancelled.iat[0,3].title()}` on `{all_approved_declined_cancelled.iat[0,9].strftime('%d %B %Y')}`." 
                                                                )


                                                                def generate_leave_pdf():
                                                                    app = {
                                                                        'company_logo': 44,
                                                                        'company_name': company_reg.replace("_"," ").title(),
                                                                        'employee_name': f"{first_name} {last_name}",
                                                                        'leave_type': all_approved_declined_cancelled.iat[0,2],
                                                                        'generated_on': today_date,
                                                                        'date_applied': all_approved_declined_cancelled.iat[0,4].strftime('%d %B %Y'),
                                                                        'approver_name': all_approved_declined_cancelled.iat[0,3].title(),
                                                                        'reference_number': all_approved_declined_cancelled.iat[0,0],
                                                                        'approved_date': all_approved_declined_cancelled.iat[0,9].strftime('%d %B %Y'),
                                                                        'new_balance': all_approved_declined_cancelled.iat[0,10],
                                                                        'start_date':  all_approved_declined_cancelled.iat[0,5].strftime('%d %B %Y'),
                                                                        'end_date':  all_approved_declined_cancelled.iat[0,6].strftime('%d %B %Y'),
                                                                        'days_requested':  all_approved_declined_cancelled.iat[0,7], 
                                                                        'address': address_foc_8, 
                                                                        'whatsapp': whatsapp_foc_8, 
                                                                        'email': email_foc_8, 
                                                                        'status': 'Approved',
                                                                        'power': power,
                                                                    }

                                                                    html_out = render_template("leave_pdf_template.html", app=app)
                                                                    
                                                                    # ✅ Return as bytes instead of saving to file
                                                                    pdf_bytes = HTML(string=html_out).write_pdf()
                                                                    return pdf_bytes

                                                                


                                                                def upload_pdf_to_whatsapp(pdf_bytes):
                                                                    compxxy = company_reg.replace("_"," ").title()
                                                                    filename=f"leave_application_{all_approved_declined_cancelled.iat[0,0]}_{first_name}_{last_name}_{compxxy}.pdf"
                                                                
                                                                    url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/media"
                                                                    headers = {
                                                                        "Authorization": f"Bearer {ACCESS_TOKEN}"
                                                                    }

                                                                    files = {
                                                                        "file": (filename, io.BytesIO(pdf_bytes), "application/pdf"),
                                                                        "type": (None, "application/pdf"),
                                                                        "messaging_product": (None, "whatsapp")
                                                                    }

                                                                    response = requests.post(url, headers=headers, files=files)
                                                                    print("📥 Full incoming data:", response.text)  # Good for debugging
                                                                    response.raise_for_status()
                                                                    return response.json()["id"]

                                                                                                                
                                                                def send_whatsapp_pdf_by_media_id(recipient_number, media_id):
                                                                    compxxy = company_reg.replace("_"," ").title()
                                                                    filename=f"leave_application_{all_approved_declined_cancelled.iat[0,0]}_{first_name}_{last_name}_{compxxy}.pdf"
                                                                    url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                                    headers = {
                                                                        "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                                        "Content-Type": "application/json"
                                                                    }
                                                                    payload = {
                                                                        "messaging_product": "whatsapp",
                                                                        "to": recipient_number,
                                                                        "type": "document",
                                                                        "document": {
                                                                            "id": media_id,            # Media ID from upload step
                                                                            "filename": filename       # Desired file name on recipient's phone
                                                                        }
                                                                    }

                                                                    response = requests.post(url, headers=headers, json=payload)
                                                                    response.raise_for_status()
                                                                    return response.json()


                                                                pdf_path = generate_leave_pdf()
                                                                media_id = upload_pdf_to_whatsapp(pdf_path)
                                                                send_whatsapp_pdf_by_media_id(sender_id, media_id)

                                                                send_whatsapp_message(
                                                                    sender_id,
                                                                    "Select an option below to continue 👇",
                                                                    buttons
                                                                )

                                                            elif all_approved_declined_cancelled.iat[0,8] == "Disapproved":
 
                                                                app_id = all_approved_declined_cancelled.iat[0,0]

                                                                buttons = [
                                                                    {"type": "reply", "reply": {"id": f"Resubapp_{app_id}", "title": "ReSubmit Application"}},
                                                                    {"type": "reply", "reply": {"id": "myhist", "title": "Download My History"}},
                                                                    {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                                ]
                                                                send_whatsapp_message(
                                                                    sender_id, 
                                                                    f"Hey {first_name}, your recent `{all_approved_declined_cancelled.iat[0,2]}` Leave Application `[ID - {all_approved_declined_cancelled.iat[0,0]}]` that you applied for on `{all_approved_declined_cancelled.iat[0,4].strftime('%d %B %Y')}` for `{all_approved_declined_cancelled.iat[0,7]} days` from `{all_approved_declined_cancelled.iat[0,5].strftime('%d %B %Y')}` to `{all_approved_declined_cancelled.iat[0,6].strftime('%d %B %Y')}` was {all_approved_declined_cancelled.iat[0,8]}❌ by `{all_approved_declined_cancelled.iat[0,3].title()}` on `{all_approved_declined_cancelled.iat[0,9].strftime('%d %B %Y')}`.",
                                                                    buttons 
                                                                )

                                                            elif all_approved_declined_cancelled.iat[0,8] == "Cancelled":

                                                                app_id = all_approved_declined_cancelled.iat[0,0]

                                                                buttons = [
                                                                    {"type": "reply", "reply": {"id": f"Resubapp_{app_id}", "title": "ReSubmit Application"}},
                                                                    {"type": "reply", "reply": {"id": "myhist", "title": "Download My History"}},
                                                                    {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                                ]
                                                                send_whatsapp_message(
                                                                    sender_id, 
                                                                    f"Hey {first_name}, on `{all_approved_declined_cancelled.iat[0,9].strftime('%d %B %Y')}` you Cancelled ⛔ your recent `{all_approved_declined_cancelled.iat[0,2]} Leave Application [ID - {all_approved_declined_cancelled.iat[0,0]}]` that you applied for on `{all_approved_declined_cancelled.iat[0,4].strftime('%d %B %Y')}` for `{all_approved_declined_cancelled.iat[0,7]} days` from `{all_approved_declined_cancelled.iat[0,5].strftime('%d %B %Y')}` to `{all_approved_declined_cancelled.iat[0,6].strftime('%d %B %Y')}`.",
                                                                    buttons 
                                                                )

                                                        else:

                                                            buttons = [
                                                                {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                                                                {"type": "reply", "reply": {"id": "Checkbal", "title": "Check Days Balance"}},
                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                            ]
                                                            companyxx = company_reg.replace("_"," ").title()
                                                            send_whatsapp_message(
                                                                sender_id, 
                                                                f"Hello {first_name} {last_name} from {companyxx}!\n\n You have not applied for any leave days yet.", 
                                                                buttons
                                                            )


                                                    elif len(df_employeesappspendingcheck) > 0:

                                                        app_idx = df_employeesappspendingcheck.iat[0,0]

                                                        buttons = [
                                                            {"type": "reply", "reply": {"id": f"Reminder_{app_idx}", "title": "Remind Approver"}},
                                                            {"type": "reply", "reply": {"id": "Cancelapp", "title": "Cancel Pending App"}},
                                                            {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                        ]
                                                        approoooover = df_employeesappspendingcheck.iat[0,3].title()
                                                        send_whatsapp_message(
                                                            sender_id, 
                                                            f"Hey {first_name}, your recent `{df_employeesappspendingcheck.iat[0,2]}` Leave Application `[ID - {df_employeesappspendingcheck.iat[0,0]}]` applied on `{df_employeesappspendingcheck.iat[0,4].strftime('%d %B %Y')}` for `{df_employeesappspendingcheck.iat[0,7]} days from {df_employeesappspendingcheck.iat[0,5].strftime('%d %B %Y')} to {df_employeesappspendingcheck.iat[0,6].strftime('%d %B %Y')}` is still pending approval from `{approoooover}`.\n\n" 
                                                            f"Select an option below to either remind `{approoooover}` to approve your pending leave application or you can cancel the pending application to submit a new leave application."         
                                                            , 
                                                            buttons
                                                        )

                                                elif selected_option == "Checkbal" or button_id == "Checkbal" :

                                                    buttons = [
                                                    {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                                                    {"type": "reply", "reply": {"id": "Track", "title": "Track Application"}},
                                                    {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                    ]

                                                    send_whatsapp_message(
                                                        sender_id, 
                                                        f"Hey {first_name}, your current available leave days balance is `{days_days_balance} days`.\n\n"
                                                        "Select an option below to continue 👇",
                                                        buttons
                                                    )

                                                elif selected_option == "myhist" or  button_id == "myhist":

                                                    try:

                                                        print(id_user)
                                                        companyxx = company_reg.replace("_"," ").title()

                                                        table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                        table_name_apps_approved = f"{company_reg}appsapproved"
                                                        table_name_apps_declined = f"{company_reg}appsdeclined"
                                                        table_name_apps_cancelled = f"{company_reg}appscancelled"

                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, approvalstatus, TO_CHAR(statusdate, 'FMDD Month YYYY') AS statusdate  FROM {table_name_apps_approved} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()
                                                        df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid", "id","leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate","approvalstatus","statusdate"]) 

                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, approvalstatus, TO_CHAR(statusdate, 'FMDD Month YYYY') AS statusdate   FROM {table_name_apps_declined} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()
                                                        df_employeesappsdeclinedcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate","approvalstatus","statusdate"])  
                                
                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, approvalstatus, TO_CHAR(statusdate, 'FMDD Month YYYY') AS statusdate  FROM {table_name_apps_cancelled} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()
                                                        df_employeesappscancelledcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate","approvalstatus","statusdate"])

                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, approvalstatus FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()
                                                        df_employeesappspenpendingcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate" ,"approvalstatus"])
                                                        df_employeesappspenpendingcheck["statusdate"] = ""


                                                        all_approved_declined = df_employeesappsapprovedcheck._append(df_employeesappsdeclinedcheck)
                                                        all_approved_declined_cancelled = all_approved_declined._append(df_employeesappscancelledcheck)
                                                        all_approved_declined_cancelled_pending = all_approved_declined_cancelled._append(df_employeesappspenpendingcheck)

                                                        all_approved_declined_cancelled_pending["dateapplied"] = pd.to_datetime(all_approved_declined_cancelled_pending["dateapplied"], errors='coerce')

                                                        all_approved_declined_cancelled_pending = all_approved_declined_cancelled_pending.sort_values(by="dateapplied", ascending=False)

                                                        print("hist hist hist")
                                                        all_approved_declined_cancelled_pending.drop('id', axis=1, inplace=True)
                                                        all_approved_declined_cancelled_pending["dateapplied"] = all_approved_declined_cancelled_pending["dateapplied"].dt.strftime("%-d %B %Y")

                                                        print(all_approved_declined_cancelled_pending)

                                                        def generate_leave_hist_pdf():
                                                            app = {
                                                                'company_name': company_reg.replace("_", " ").title(),
                                                                'employee_name': f"{first_name} {last_name}",
                                                                'generated_on': today_date,
                                                                'power': power
                                                            }

                                                            table_hist_html = all_approved_declined_cancelled_pending.to_html(index=False, classes='data', border=0, justify='center',escape=False)

                                                            html_out = render_template("leave_applications_history.html", app=app, table_hist_html=table_hist_html)
                                                            pdf_bytes = HTML(string=html_out).write_pdf()
                                                            return pdf_bytes

                                                        def upload_pdf_to_whatsapp(pdf_bytes):
                                                            compxxy = company_reg.replace("_"," ").title()
                                                            filename=f"{first_name}_{last_name}_{compxxy}_leave_applications_history.pdf"
                                                        
                                                            url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/media"
                                                            headers = {
                                                                "Authorization": f"Bearer {ACCESS_TOKEN}"
                                                            }

                                                            files = {
                                                                "file": (filename, io.BytesIO(pdf_bytes), "application/pdf"),
                                                                "type": (None, "application/pdf"),
                                                                "messaging_product": (None, "whatsapp")
                                                            }

                                                            response = requests.post(url, headers=headers, files=files)
                                                            print("📥 Full incoming data:", response.text)  # Good for debugging
                                                            response.raise_for_status()
                                                            return response.json()["id"]

                                                                                                        
                                                        def send_whatsapp_pdf_by_media_id(recipient_number, media_id):
                                                            compxxy = company_reg.replace("_"," ").title()
                                                            filename=f"{first_name}_{last_name}_{compxxy}_leave_applications_history.pdf"
                                                            url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                            headers = {
                                                                "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                                "Content-Type": "application/json"
                                                            }
                                                            payload = {
                                                                "messaging_product": "whatsapp",
                                                                "to": recipient_number,
                                                                "type": "document",
                                                                "document": {
                                                                    "id": media_id,            # Media ID from upload step
                                                                    "filename": filename       # Desired file name on recipient's phone
                                                                }
                                                            }

                                                            response = requests.post(url, headers=headers, json=payload)
                                                            response.raise_for_status()
                                                            return response.json()


                                                        pdf_path = generate_leave_hist_pdf()
                                                        media_id = upload_pdf_to_whatsapp(pdf_path)

                                                        appscountnum = len(all_approved_declined_cancelled_pending)

                                                        if appscountnum > 0:

                                                            send_whatsapp_pdf_by_media_id(sender_id, media_id)

                                                            buttons = [
                                                                {"type": "reply", "reply": {"id": "Analytics", "title": "Analytics & Insights"}},
                                                                {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                            ]
                                                            send_whatsapp_message(
                                                                sender_id, 
                                                                f"Hey {first_name} {last_name} from {companyxx}! You may go ahead and download your leave applications history file attached here 😎.", 
                                                                buttons
                                                            )

                                                        else:

                                                            buttons = [
                                                                {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                                                                {"type": "reply", "reply": {"id": "Checkbal", "title": "Check Days Balance"}},
                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                            ]

                                                            send_whatsapp_message(
                                                                sender_id, 
                                                                f"Hello {first_name} {last_name} from {companyxx}!\n\n You have not applied for any leave days yet.", 
                                                                buttons
                                                            )

                                                    except Exception as e:

                                                        send_whatsapp_message(f"+263710910052", f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully! Error; {e}")                      

                                                elif selected_option == "Apply" or button_id == "Apply":

                                                    table_name_apps_pending_approval = f"{company_reg}appspendingapproval"

                                                    query = f"SELECT  appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor  FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                    cursor.execute(query)
                                                    rows = cursor.fetchall()

                                                    df_employeesappspendingcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor"])    

                                                    if len(df_employeesappspendingcheck) == 0:

                                                        sections = [
                                                            {
                                                                "title": "Leave Type Options",
                                                                "rows": [
                                                                    {"id": "Annual", "title": "Annual Leave"},
                                                                    {"id": "Sick", "title": "Sick Leave"},
                                                                    {"id": "Study", "title": "Study Leave"},
                                                                    {"id": "Bereavement", "title": "Bereavement Leave"},
                                                                    {"id": "Parental", "title": "Parental Leave"},
                                                                    {"id": "Other", "title": "Other"},
                                                                ]
                                                            }
                                                        ]

                                                        send_whatsapp_list_message(
                                                            sender_id, 
                                                            f"{first_name}, kindly select the type of Leave that you are applying for.", 
                                                            "Leave Type Options",
                                                            sections) 

                                                    elif len(df_employeesappspendingcheck) > 0:

                                                        app_idx = df_employeesappspendingcheck.iat[0,0]

                                                        buttons = [
                                                            {"type": "reply", "reply": {"id": f"Reminder_{app_idx}", "title": "Remind Approver"}},
                                                            {"type": "reply", "reply": {"id": "Cancelapp", "title": "Cancel Pending App"}},
                                                            {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                        ]
                                                        send_whatsapp_message(
                                                            sender_id, 
                                                            f"Oops! 🥲. Sorry {first_name}, you cannot apply for leave whilst you have another leave application which is still pending approval.\n\n" 
                                                            f"Your `{df_employeesappspendingcheck.iat[0,2]}` Leave Application `[ID - {df_employeesappspendingcheck.iat[0,0]}]` applied on `{df_employeesappspendingcheck.iat[0,4].strftime('%d %B %Y')}` for `{df_employeesappspendingcheck.iat[0,7]} days from {df_employeesappspendingcheck.iat[0,5].strftime('%d %B %Y')} to {df_employeesappspendingcheck.iat[0,6].strftime('%d %B %Y')}` is still pending approval from {df_employeesappspendingcheck.iat[0,3]}.\n\n" 
                                                            f"Select an option below to either remind the approver to approved your pending application or you can cancel the pending application to submit a new leave application."         
                                                            , 
                                                            buttons
                                                        )
                                                
                                                elif button_id == "Submitapp":
                                        
                                                    try:

                                                        table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                        table_name_apps_approved = f"{company_reg}appsapproved"
                                                        companyxx = company_reg.replace("_", " ").title()

                                                        cursor.execute("""
                                                            SELECT id ,empidwa, leavetypewa, startdate, enddate FROM whatsapptempapplication
                                                            WHERE empidwa = %s
                                                        """, (id_user,))
                                                
                                                        result = cursor.fetchone()

                                                        appid = result[0]
                                                        id_user = result[1]
                                                        leavetype = result[2]
                                                        startdate = result[3]
                                                        enddate = result[4]
                                                        table_name = f"{company_reg}main"
                                                        
                                                        query = f"SELECT id FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()

                                                        df_employeesappspendingcheck = pd.DataFrame(rows, columns=["id"])    

                                                        if len(df_employeesappspendingcheck) == 0:

                                                            query = f"""SELECT appid, id, leavestartdate, leaveenddate FROM {table_name_apps_approved} WHERE id = %s AND leavestartdate <= %s AND leaveenddate >= %s"""

                                                            cursor.execute(query, (id_user, enddate, startdate))
                                                            results = cursor.fetchall()

                                                            # Process results
                                                            if results:
                                                                print("Overlapping records found:")

                                                                try:

                                                                    overlap_messages = []

                                                                    for row in results:

                                                                        formatted_date_start = row[2].strftime("%d %B %Y")
                                                                        formatted_date_end = row[3].strftime("%d %B %Y")

                                                                        overlap_messages.append(f"appID: {row[0]}, Starting Date: {formatted_date_start}, Ending Date: {formatted_date_end}")

                                                                    # Combine into one single string (newline-separated)
                                                                    overlap_info = "\n".join(overlap_messages)

                                                                    buttons = [
                                                                        {"type": "reply", "reply": {"id": f"Apply", "title": "Restart Application"}},
                                                                        {"type": "reply", "reply": {"id": f"ApplyRevoke", "title": "Revoke Conflictn App"}},
                                                                        {"type": "reply", "reply": {"id": f"Menu", "title": "Main Menu"}},
                                                                    ]

                                                                    send_whatsapp_message(sender_id, f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully!\n\n"
                                                                        f"One of your previously approved leave applications include days within the period that you are currently applying for.\n\n Leave App; {overlap_info}.\n\n Either restart your application with different dates from these, or apply that this conflicting approved Leave application be Revoked.",
                                                                        buttons
                                                                        )
                                                                
                                                                except Exception as e:

                                                                    send_whatsapp_message(f"+263710910052", f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully! Error; {e}")                      
                                                            
                                                            else:

                                                                print("No Overlapping records found:")

                                                                try:

                                                                    table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                                    table_name_apps_approved = f"{company_reg}appsapproved"

                                                                    query = f"SELECT id FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                                    cursor.execute(query)
                                                                    rows = cursor.fetchall()

                                                                    df_employeesappspendingcheck = pd.DataFrame(rows, columns=["id"])   

                                                                    print("1212") 

                                                                    if len(df_employeesappspendingcheck) == 0:

                                                                        cursor.execute("""
                                                                            SELECT id ,empidwa, leavetypewa, startdate, enddate FROM whatsapptempapplication
                                                                            WHERE empidwa = %s
                                                                        """, (id_user,))
                                                                
                                                                        result = cursor.fetchone()

                                                                        print("select successful")

                                                                        appid = result[0]
                                                                        leavetype = result[2]
                                                                        startdate = result[3]
                                                                        enddate = result[4]
                                                                        table_name = f"{company_reg}main"

                                                                        if isinstance(startdate, str):
                                                                            startdate = datetime.datetime.strptime(startdate, "%Y-%m-%d").date()
                                                                        if isinstance(enddate, str):
                                                                            enddate = datetime.datetime.strptime(enddate, "%Y-%m-%d").date()

                                                                        business_days = 0
                                                                        current_date = startdate

                                                                        while current_date <= enddate:
                                                                            if current_date.weekday() != 6:  # 0=Mon, 1=Tue, ..., 4=Fri
                                                                                business_days += 1
                                                                            current_date += timedelta(days=1)  # Use timedelta directly

                                                                        query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation, department FROM {table_name};"
                                                                        cursor.execute(query)
                                                                        rows = cursor.fetchall()

                                                                        df_employees = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Address", "Role","Leave Approver Name","Leave Approver ID","Leave Approver Email", "Leave Approver WhatsAapp", "Leave Days Balance","Days Accumulated per Month", "Department"])
                                                                        print(df_employees)
                                                                        userdf = df_employees[df_employees['id'] == int(np.int64(id_user))].reset_index()
                                                                        print("yeaarrrrr")
                                                                        print(userdf)
                                                                        firstname = userdf.iat[0,2]
                                                                        surname = userdf.iat[0,3]
                                                                        whatsapp = userdf.iat[0,4]
                                                                        address = userdf.iat[0,6]
                                                                        email = userdf.iat[0,5]
                                                                        fullnamedisp = firstname + ' ' + surname
                                                                        leaveapprovername = userdf.iat[0,8]
                                                                        leaveapproverid = userdf.iat[0,9]
                                                                        leaveapproveremail = userdf.iat[0, 10]
                                                                        leaveapproverwhatsapp = userdf.iat[0,11]
                                                                        role = userdf.iat[0,7]
                                                                        leavedaysbalance = userdf.iat[0,12]
                                                                        department = userdf.iat[0,14] 
                                                                        print('check')

                                                                        departmentdf = df_employees[df_employees['Department'] == department].reset_index()
                                                                        numberindepartment = len(departmentdf)
                                                                        
                                                                        startdatex = pd.Timestamp(startdate)
                                                                        enddatex = pd.Timestamp(enddate)

                                                                        leave_dates = pd.date_range(startdatex, enddatex)

                                                                        query = f"""
                                                                            SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate,
                                                                                leaveenddate, leavedaysappliedfor, approvalstatus, statusdate,
                                                                                leavedaysbalancebf, department
                                                                            FROM {table_name_apps_approved}
                                                                            WHERE department = %s;
                                                                        """
                                                                        cursor.execute(query, (department,))
                                                                        rows = cursor.fetchall()

                                                                        df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf","department"]) 
                                                                        df_employeesappsapprovedcheck["leavestartdate"] = pd.to_datetime(df_employeesappsapprovedcheck["leavestartdate"])
                                                                        df_employeesappsapprovedcheck["leaveenddate"] = pd.to_datetime(df_employeesappsapprovedcheck["leaveenddate"])
                        
                                                                        df_employeesappsapprovedcheck.dropna(subset=["leavestartdate", "leaveenddate"], inplace=True)
                                                                        # Create daily impact report
                                                                        impact_report = []

                                                                        for date in leave_dates:

                                                                            date = pd.Timestamp(date)

                                                                            print(type(date))  # Should be pandas._libs.tslibs.timestamps.Timestamp or datetime.datetime
                                                                            print(df_employeesappsapprovedcheck.dtypes)  # Check all datetime columns

                                                                            on_leave = ((df_employeesappsapprovedcheck["leavestartdate"] <= date) & (df_employeesappsapprovedcheck["leaveenddate"] >= date)).sum()
                                                                            remaining = numberindepartment - on_leave - 1  # subtract 1 for the new leave
                                                                            impact_report.append({
                                                                                "date": date,  # <=== Keep as datetime, don't convert to string
                                                                                "on leave": on_leave + 1,
                                                                                "employees remaining": remaining
                                                                            })

                                                                        # Convert to DataFrame for display
                                                                        impact_df = pd.DataFrame(impact_report)
                                                                        print("IMPAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACT")
                                                                        print(impact_df)
                                                                        print(numberindepartment)

                                                                        impact_df["date"] = pd.to_datetime(impact_df["date"], format="%Y-%m-%d")
                                                                        impact_df = impact_df[impact_df["date"].dt.weekday != 6].copy()

                                                                        change = (impact_df[["on leave", "employees remaining"]] != impact_df[["on leave", "employees remaining"]].shift()).any(axis=1)
                                                                        change.iloc[0] = True  # ensure the first row starts a group
                                                                        impact_df["group"] = change.cumsum()

                                                                        statements = []
                                                                        for _, group_df in impact_df.groupby("group"):
                                                                            start = group_df["date"].iloc[0].strftime("%d %B %Y")
                                                                            end = group_df["date"].iloc[-1].strftime("%d %B %Y")
                                                                            on_leave = group_df["on leave"].iloc[0]
                                                                            remaining = group_df["employees remaining"].iloc[0]
                                                                            
                                                                            if start == end:
                                                                                statements.append(f"On {start}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                                                                            else:
                                                                                statements.append(f"From {start} to {end}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                                                                                # Combine all statements into a single variable
                                                                        final_summary = "\n".join(statements)
                                                                        # Print output
                                                                        for s in statements:
                                                                            print(s)

                                                                        if leavetype == "Annual":

                                                                            leavedaysbalancebf = float(leavedaysbalance) - float(business_days)

                                                                        else:

                                                                            leavedaysbalancebf = float(leavedaysbalance)

                                                                        if leavedaysbalancebf >= 0:



                                                                            status = "Pending"

                                                                            insert_query = f"""
                                                                            INSERT INTO {table_name_apps_pending_approval} (id, firstname, surname, department, leavetype, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus)
                                                                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                                                            """
                                                                            cursor.execute(insert_query, (int(np.int64(id_user)), first_name, last_name, department, leavetype, leaveapprovername, int(np.int64(leaveapproverid)), leaveapproveremail, int(np.int64(leaveapproverwhatsapp)), float(np.float64(leavedaysbalance)), today_date, startdate, enddate, float(np.int64(business_days)), float(np.float64(leavedaysbalancebf)), status))
                                                                            connection.commit()

                                                                            query = f"SELECT appid FROM {table_name_apps_pending_approval};"
                                                                            cursor.execute(query)
                                                                            rows = cursor.fetchall()

                                                                            df_employees = pd.DataFrame(rows, columns=["id"])
                                                                            leaveappid = df_employees.iat[0,0]
                                                                            companyxx = company_reg.replace("_"," ").title()
                                                                            approovvver = leaveapprovername.title()

                                                                            buttons = [
                                                                            {"type": "reply", "reply": {"id": "Track", "title": "Track Application"}},
                                                                            {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                                            ]

                                                                            send_whatsapp_message(sender_id, f"✅ Great News {first_name} from {companyxx}! \n\n Your `{leavetype} Leave Application` for `{business_days} days` from `{startdate.strftime('%d %B %Y')}` to `{enddate.strftime('%d %B %Y')}` has been submitted successfully!\n\n"
                                                                                f"Your Leave Application ID is `{leaveappid}`.\n\n"
                                                                                f"A Notification has been sent to `{approovvver}`  on `+263{leaveapproverwhatsapp}` to decide on  your application.\n\n", 
                                                                                buttons)
                                                                            
                                                                            if leaveapproverwhatsapp:
                                
                                                                                buttons = [
                                                                                    {"type": "reply", "reply": {"id": f"Approve5appwa_{leaveappid}", "title": "Approve"}},
                                                                                    {"type": "reply", "reply": {"id": f"Disapproveappwa_{leaveappid}", "title": "Disapprove"}},
                                                                                ]
                                                                                send_whatsapp_message(
                                                                                    f"263{leaveapproverwhatsapp}", 
                                                                                    f"Hey {approovvver}! 😊. New `{leavetype}` Leave Application from `{first_name} {surname}` for `{business_days} days` from `{startdate.strftime('%d %B %Y')}` to `{enddate.strftime('%d %B %Y')}`.\n\n" 
                                                                                    f"If you approve this leave application, {final_summary}\n\n"  
                                                                                    f"Select an option below to either approve or disapprove the application."         
                                                                                    , 
                                                                                    buttons
                                                                                )

                                                                        else:


                                                                            buttons = [
                                                                                {"type": "reply", "reply": {"id": f"Apply", "title": "Restart Application"}},
                                                                                {"type": "reply", "reply": {"id": f"Checkbal", "title": "Check Days Balance"}},
                                                                                {"type": "reply", "reply": {"id": f"Menu", "title": "Main Menu"}},
                                                                            ]

                                                                            send_whatsapp_message(sender_id, f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully!\n\n"
                                                                                f"You only have *{leavedaysbalance}* days available for leave but you are applying for *{business_days}*.\n\n You can restart your application and apply for leave such that the days between your leave start date and end date do not exceed your available balance of *{leavedaysbalance}* days.",
                                                                                buttons
                                                                                )

                                                                    else:
                                                                        print("leave app submission failed")
                                                                except Error as e:
                                                                    print(f"Ooops, submission failed, {e}")

                                                    except ValueError as e:
                                                        send_whatsapp_message(
                                                            sender_id,
                                                            f"{e}, ❌ No, incorrect message format. Please use:\n"
                                                            "`end 24 january 2025`\n"
                                                            "Example: `end 15 march 2024`"
                                                        )

                                                elif button_id == "Resubapp" :

                                                    table_name_apps_cancelled = f"{company_reg}appscancelled"
                                                    table_name_apps_pending_approval = f"{company_reg}appspendingapproval"

                                                    query = f"SELECT appid, id, firstname, surname, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail , leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf FROM {table_name_apps_pending_approval} WHERE id = %s;"
                                                    cursor.execute(query, (id_user,))
                                                    result = cursor.fetchone()
                                                    if result:
                                                        (app_id, employee_number, first_name, surname, leave_type,  leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf) = result
                                                    
                                                        try:
                                                                status = "Cancelled"
                                                                statusdate = today_date
                                                            
                                                                insert_query = f"""
                                                                INSERT INTO {table_name_apps_cancelled} 
                                                                (appid, id, firstname, surname, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus, statusdate)
                                                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                                                """

                                                                cursor.execute(insert_query, (
                                                                    app_id, employee_number, first_name, surname, leave_type, leave_specify, approver_name, 
                                                                    approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                                                                    end_date, leave_days, leavedaysbalancebf, status, statusdate
                                                                ))
                                                                
                                                                connection.commit()
                                                                print("Insert successful!")

                                                        except Exception as e:
                                                            print("Error inserting data:", e)

                                                        # SQL query to delete or mark the leave as canceled
                                                        query = f"""DELETE FROM {table_name_apps_pending_approval} WHERE appid = %s"""
                                                        cursor.execute(query, (app_id,))
                                                        connection.commit()                                       

                                                        companyxx = company_reg.replace("_", " ").title()
                                                        buttons = [
                                                            {"type": "reply", "reply": {"id": f"Resubapp_{app_id}", "title": "ReSubmit Application"}},
                                                            {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                                                            {"type": "reply", "reply": {"id": "Checkbal", "title": "Check Days Balance"}},
                                                        ]

                                                        send_whatsapp_message(sender_id, f"Hey {first_name} from {companyxx}! \n\n Your `{leave_type} Leave Application` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}` has been Cancelled successfully✅!\n\n"
                                                            "Select an option below to continue 👇",
                                                            buttons
                                                        )                                          
                                                    
                                                    else:
                                                        print("No record found for the user.")

                                                elif button_id == "Cancelapp" :

                                                    table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                    table_name_apps_cancelled = f"{company_reg}appscancelled"

                                                    query = f"SELECT appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail , leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf FROM {table_name_apps_pending_approval} WHERE id = %s;"
                                                    cursor.execute(query, (id_user,))
                                                    result = cursor.fetchone()
                                                    if result:
                                                        (app_id, employee_number, first_name, surname, department, leave_type,  leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf) = result
                                                    
                                                        try:
                                                                status = "Cancelled"
                                                                statusdate = today_date
                                                            
                                                                insert_query = f"""
                                                                INSERT INTO {table_name_apps_cancelled} 
                                                                (appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus, statusdate)
                                                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                                                """

                                                                cursor.execute(insert_query, (
                                                                    app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, 
                                                                    approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                                                                    end_date, leave_days, leavedaysbalancebf, status, statusdate
                                                                ))
                                                                
                                                                connection.commit()
                                                                print("Insert successful!")

                                                        except Exception as e:
                                                            print("Error inserting data:", e)

                                                        # SQL query to delete or mark the leave as canceled
                                                        query = f"""DELETE FROM {table_name_apps_pending_approval} WHERE appid = %s"""
                                                        cursor.execute(query, (app_id,))
                                                        connection.commit()                                       

                                                        companyxx = company_reg.replace("_", " ").title()
                                                        buttons = [
                                                            {"type": "reply", "reply": {"id": f"Resubapp_{app_id}", "title": "ReSubmit Application"}},
                                                            {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                                                            {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                        ]

                                                        send_whatsapp_message(sender_id, f"Hey {first_name} from {companyxx}! \n\n Your `{leave_type} Leave Application` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}` has been Cancelled successfully✅!\n\n"
                                                            "Select an option below to continue 👇",
                                                            buttons
                                                        )                                          
                                                    
                                                    else:
                                                        print("No record found for the user.")

                                            else:

                                                text = message.get("text", {}).get("body", "").lower()
                                                print(f"📨 Message from {sender_id}: {text}")
                                                
                                                print("yearrrrrrrrrrrrrrrrrrrrrrrrrrrssrsrsrsrsrs")

                                                print(role_foc_8)
                                                    
                                                companyxx = company_reg.replace("_", " ").title()

                                                if "hello" in text.lower():

                                                    sections = [
                                                        {
                                                            "title": "User Options",
                                                            "rows": [
                                                                {"id": "Apply", "title": "Apply for Leave"},
                                                                {"id": "Track", "title": "Track My Application"},
                                                                {"id": "Checkbal", "title": "Check Days Balance"},
                                                                {"id": "myhist", "title": "My Applications History"},
                                                                {"id": "Myinfo", "title": "My Info"}
                                                            ]
                                                        }
                                                    ]


                                                    send_whatsapp_list_message(
                                                        sender_id, 
                                                        f"Hello {first_name} {last_name} from {companyxx}!\n\n {bot} LMS Bot Here 😎. How can I assist you?", 
                                                    "User Options",
                                                    sections)

                                                elif "email" in text.lower():

                                                    table_name = company_reg + "main"
                                                    table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                    table_name_apps_cancelled = f"{company_reg}appscancelled"
                                                    table_name_apps_approved = f"{company_reg}appsapproved"
                                                    table_name_apps_declined = f"{company_reg}appsdeclined"
                                                    table_name_apps_revoked = f"{company_reg}appsrevoked"

                                                    # Regex to extract email after the word "email"
                                                    match = re.search(r"email\s+([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)", text.strip(), re.IGNORECASE)

                                                    if not match:
                                                        raise ValueError("Invalid Email format")

                                                        sections = [
                                                            {
                                                                "title": "User Options",
                                                                "rows": [
                                                                    {"id": "Editname", "title": "Edit My Name"},
                                                                    {"id": "Editwhatsapp", "title": "Change My WhatsApp #"},
                                                                    {"id": "Editemail", "title": "Change My Email"},
                                                                    {"id": "Editwebpass", "title": "Change Web Password"},
                                                                    {"id": "Editaddress", "title": "Edit My Address"},
                                                                    {"id": "MyInfo", "title": "My Info"},
                                                                    {"id": "Menu", "title": "Main Menu"}
                                                                ]
                                                            }
                                                        ]

                                                        send_whatsapp_list_message(
                                                            sender_id, 
                                                            f"Hey {first_name}!\n You have provided an invalid Email Address. Enter your new email address starting with the word `email` as shown below 👇. \n\n `email epsilon@gmail.com`", 
                                                        "User Options",
                                                        sections)                


                                                    email = match.group(1)
                                                    print("Extracted email:", email)

                                                    try:

                                                        query = f"UPDATE {table_name} SET email = %s WHERE id = %s;"
                                                        cursor.execute(query, (email, id_user))
                                                        connection.commit()

                                                        query = f"UPDATE {table_name_apps_pending_approval} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                        cursor.execute(query, (email, id_user))
                                                        connection.commit()
                                            
                                                        query = f"UPDATE {table_name_apps_cancelled} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                        cursor.execute(query, (email, id_user))
                                                        connection.commit()

                                                        query = f"UPDATE {table_name_apps_approved} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                        cursor.execute(query, (email, id_user))
                                                        connection.commit()

                                                        query = f"UPDATE {table_name_apps_declined} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                        cursor.execute(query, (email, id_user))
                                                        connection.commit()

                                                        query = f"UPDATE {table_name_apps_revoked} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                        cursor.execute(query, (email, id_user))
                                                        connection.commit()

                                                        sections = [
                                                            {
                                                                "title": "User Options",
                                                                "rows": [
                                                                    {"id": "Editname", "title": "Edit My Name"},
                                                                    {"id": "Editwhatsapp", "title": "Change My WhatsApp #"},
                                                                    {"id": "Editemail", "title": "Change My Email"},
                                                                    {"id": "Editwebpass", "title": "Change Web Password"},
                                                                    {"id": "Editaddress", "title": "Edit My Address"},
                                                                    {"id": "MyInfo", "title": "My Info"},
                                                                    {"id": "Menu", "title": "Main Menu"}
                                                                ]
                                                            }
                                                        ]

                                                        send_whatsapp_list_message(
                                                            sender_id, 
                                                            f"Hey {first_name}!\n Email Address Successfully Changed to `{email}`. Select an option below to proceed 👇.", 
                                                        "User Options",
                                                        sections)  

                                                    except Exception as e:
                                                        print(e)

                                                elif "start" in text.lower():
                                                    try:
                                                        # Match: "start 20 july 2025"
                                                        match = re.match(r"start\s+(\d{1,2}\s+[a-zA-Z]+\s+\d{4})", text.strip(), re.IGNORECASE)
                                                        if not match:
                                                            raise ValueError("Invalid format")

                                                        date_part = match.group(1)
                                                        parsed_date = datetime.strptime(date_part, "%d %B %Y")  # Will raise ValueError if invalid

                                                        # ✅ Now it's safe to update the DB
                                                        cursor.execute("""
                                                            UPDATE whatsapptempapplication
                                                            SET startdate = %s
                                                            WHERE empidwa = %s
                                                        """, (date_part, id_user))
                                                        connection.commit()

                                                        cursor.execute("""
                                                            SELECT empidwa, leavetypewa FROM whatsapptempapplication
                                                            WHERE empidwa = %s
                                                        """, (id_user,))
                                                        result = cursor.fetchone()
                                                        leavetypewa = result[1] if result else "your"

                                                        send_whatsapp_message(sender_id,
                                                            f"✅ Got it! Start date saved.\n\nNow enter your last day on {leavetypewa} leave like this:\n"
                                                            "`end 28 July 2025`"
                                                        )

                                                    except ValueError:
                                                        send_whatsapp_message(
                                                            sender_id,
                                                            f"❌ Invalid start date message format, {first_name}. Please use the date format givem below 👇:\n"
                                                            "`start 24 january 2025`\n\n"
                                                            "Example: `start 15 march 2024`"
                                                        )

                                                    except Exception as e:
                                                        import traceback
                                                        print("🔴 Unexpected error:", e)
                                                        traceback.print_exc()

                                                        try:
                                                            send_whatsapp_message(
                                                                sender_id,
                                                                "⚠️ Something went wrong while processing your start date. Please try again or contact support."
                                                            )
                                                        except Exception as send_err:
                                                            print("🔴 Failed to send WhatsApp error message:", send_err)


                                                elif "end" in text.lower():

                                                    try:
                                                        # ✅ Match "end 24 january 2025"
                                                        match = re.match(r"end\s+(\d{1,2}\s+[a-zA-Z]+\s+\d{4})", text.strip(), re.IGNORECASE)
                                                        if not match:
                                                            raise ValueError("Invalid end date format.")

                                                        date_part = match.group(1)
                                                        parsed_end_date = datetime.strptime(date_part, "%d %B %Y").date()  # Will raise ValueError if invalid

                                                        # ✅ Update DB now that it's valid
                                                        cursor.execute("""
                                                            UPDATE whatsapptempapplication
                                                            SET enddate = %s
                                                            WHERE empidwa = %s
                                                        """, (date_part, id_user))
                                                        connection.commit()

                                                        # ✅ Fetch full leave application
                                                        cursor.execute("""
                                                            SELECT id, empidwa, leavetypewa, startdate, enddate FROM whatsapptempapplication
                                                            WHERE empidwa = %s
                                                        """, (id_user,))
                                                        result = cursor.fetchone()

                                                        if not result:
                                                            raise Exception("No leave record found.")

                                                        appid = result[0]
                                                        leavetype = result[2]
                                                        startdate = result[3]
                                                        enddate = result[4]

                                                        # ✅ Ensure both dates are datetime.date objects
                                                        if isinstance(startdate, str):
                                                            startdate = datetime.strptime(startdate, "%Y-%m-%d").date()
                                                        if isinstance(enddate, str):
                                                            enddate = datetime.strptime(enddate, "%Y-%m-%d").date()

                                                        # ✅ Calculate business days
                                                        business_days = 0
                                                        current_date = startdate
                                                        while current_date <= enddate:
                                                            if current_date.weekday() != 6:  # Weekday: Mon-Fri
                                                                business_days += 1
                                                            current_date += timedelta(days=1)

                                                        # ✅ Ask user to confirm submission
                                                        buttons = [
                                                            {"type": "reply", "reply": {"id": "Submitapp", "title": "Yes, Submit"}},
                                                            {"type": "reply", "reply": {"id": "Dontsubmit", "title": "No"}}
                                                        ]
                                                        send_whatsapp_message(
                                                            sender_id,
                                                            f"📝 Do you wish to submit your `{business_days}-day {leavetype} Leave Application` from "
                                                            f"`{startdate.strftime('%d %B %Y')}` to `{enddate.strftime('%d %B %Y')}`, {first_name}?",
                                                            buttons
                                                        )

                                                    except ValueError:
                                                        send_whatsapp_message(
                                                            sender_id,
                                                            f"❌ Invalid end date message format, {first_name}. Please use the date format givem below 👇:\n"
                                                            "`end 24 january 2025`\n\n"
                                                            "Example: `end 28 march 2024`"
                                                        )

                                                    except Exception as e:
                                                        import traceback
                                                        print("🔴 ERROR during end date processing:", e)
                                                        traceback.print_exc()
                                                        try:
                                                            send_whatsapp_message(
                                                                sender_id,
                                                                "⚠️ Something went wrong while processing your end date. Please try again or contact support."
                                                            )
                                                        except Exception as send_err:
                                                            print("🔴 Failed to send error message via WhatsApp:", send_err)
                                                            
                                                else:
                                                    send_whatsapp_message(
                                                        sender_id, 
                                                        f"{bot} LMS Bot Here 😎. Say 'hello' to start!"
                                                    )

                                        except Exception as e:
                                            print(e)

################## ORDINARY USER APPROVER

                                    elif len(df_employeesempapp) > 0:

                                        print("uuuuuuuuuuuuuuuuuuuuuuuuuuuuuuuuuuuuuuuuuuuuuu")
                                    

                                        if message.get("type") == "interactive":
                                            interactive = message.get("interactive", {})

                                            if interactive.get("type") == "list_reply":

                                                selected_option = interactive.get("list_reply", {}).get("id")
                                                print(f"📋 User selected: {selected_option}")
                                                button_id = ""

                                            elif interactive.get("type") == "button_reply":

                                                button_id = interactive.get("button_reply", {}).get("id")
                                                print(f"🔘 Button clicked: {button_id}")
                                                selected_option = ""

                                            if selected_option in ["Annual","Sick","Study","Parental", "Bereavement","Other"] :

                                                button_id_leave_type = str(selected_option)

                                                cursor.execute("""
                                                    DELETE FROM whatsapptempapplication
                                                    WHERE empidwa = %s
                                                """, (str(id_user),))  
                                                
                                                connection.commit()

                                                cursor.execute(f"""
                                                    INSERT INTO whatsapptempapplication (empidwa, leavetypewa, companynamewa)
                                                    VALUES (%s, %s, %s)
                                                """, (id_user, button_id_leave_type, company_reg))

                                                connection.commit()

                                                send_whatsapp_message(
                                                    sender_id, 
                                                    f"Ok. When would you like your {selected_option} Leave to start {first_name}?\n\n"
                                                    "Please enter your response using the format: 👇🏻\n"
                                                    "`start 24 january 2025`"
                                                )

                                                continue

                                            elif selected_option == "Apply" or button_id == "Apply":

                                                table_name_apps_pending_approval = f"{company_reg}appspendingapproval"

                                                query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor  FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                cursor.execute(query)
                                                rows = cursor.fetchall()

                                                df_employeesappspendingcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor"])    

                                                if len(df_employeesappspendingcheck) == 0:

                                                        sections = [
                                                            {
                                                                "title": "Leave Type Options",
                                                                "rows": [
                                                                    {"id": "Annual", "title": "Annual Leave"},
                                                                    {"id": "Sick", "title": "Sick Leave"},
                                                                    {"id": "Study", "title": "Study Leave"},
                                                                    {"id": "Bereavement", "title": "Bereavement Leave"},
                                                                    {"id": "Parental", "title": "Parental Leave"},
                                                                    {"id": "Other", "title": "Other"},
                                                                ]
                                                            }
                                                        ]

                                                        send_whatsapp_list_message(
                                                            sender_id, 
                                                            f"{first_name}, kindly select the type of Leave that you are applying for.", 
                                                            "Leave Type Options",
                                                            sections) 

                                                elif len(df_employeesappspendingcheck) > 0:

                                                    app_idx = df_employeesappspendingcheck.iat[0,0]

                                                    buttons = [
                                                        {"type": "reply", "reply": {"id": f"Reminder_{app_idx}", "title": "Remind Approver"}},
                                                        {"type": "reply", "reply": {"id": "Cancelapp", "title": "Cancel Pending App"}},
                                                        {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                    ]
                                                    send_whatsapp_message(
                                                        sender_id, 
                                                        f"Oops! 🥲. Sorry {first_name}, you cannot apply for leave whilst you have another leave application which is still pending approval.\n\n" 
                                                        f"Your `{df_employeesappspendingcheck.iat[0,2]}` Leave Application `[ID - {df_employeesappspendingcheck.iat[0,0]}]` applied on `{df_employeesappspendingcheck.iat[0,4].strftime('%d %B %Y')}` for `{df_employeesappspendingcheck.iat[0,7]} days from {df_employeesappspendingcheck.iat[0,5].strftime('%d %B %Y')} to {df_employeesappspendingcheck.iat[0,6].strftime('%d %B %Y')}` is still pending approval from {df_employeesappspendingcheck.iat[0,3]}.\n\n" 
                                                        f"Select an option below to either remind the approver to approved your pending application or you can cancel the pending application to submit a new leave application."         
                                                        , 
                                                        buttons
                                                    )

                                            elif selected_option == "Pending" or button_id == "Pending":

                                                table_name_apps_pending_approval = f"{company_reg}appspendingapproval"

                                                query = f"SELECT id, leavetype, firstname, surname, leaveapprovername, leaveapproverid, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, appid  FROM {table_name_apps_pending_approval} WHERE leaveapproverid = {str(id_user)};"
                                                cursor.execute(query)
                                                rows = cursor.fetchall()

                                                df_employeesappspendingcheck = pd.DataFrame(rows, columns=["id", "leavetype", "firstname", "surname", "leaveapprovername", "leaveapproverid", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor", "appid"])    
                                                df_employeesappspendingcheck = df_employeesappspendingcheck.sort_values(by=df_employeesappspendingcheck.columns[10], ascending=False)

                                                if len(df_employeesappspendingcheck) == 0:

                                                    companyxx = company_reg.replace("_", " ").title()
                                                    sections = [
                                                        {
                                                            "title": "Approver Options",
                                                            "rows": [
                                                                {"id": "Apply", "title": "Apply for Leave"},
                                                                {"id": "Track", "title": "Track My Application"},
                                                                {"id": "Checkbal", "title": "Check Days Balance"},
                                                                {"id": "Pending", "title": "Apps Pending My Approval"},
                                                                {"id": "myhist", "title": "My Applications History"},
                                                                {"id": "Myinfo", "title": "My Info"}     
                                                            ]
                                                        }
                                                    ]
    
                                                    send_whatsapp_list_message(
                                                        sender_id, 
                                                        f"{first_name}, there are currently no leave applications that are pending your approval.", 
                                                    "Approver Options",
                                                    sections) 

                                                elif len(df_employeesappspendingcheck) > 0:

                                                    firstnameemp2 = df_employeesappspendingcheck.iat[0,2]
                                                    appid = df_employeesappspendingcheck.iat[0,10]
                                                    surnameemp2 = df_employeesappspendingcheck.iat[0,3]
                                                    leave_type2 = df_employeesappspendingcheck.iat[0,1]
                                                    days = df_employeesappspendingcheck.iat[0,9]
                                                    date_applied2 = df_employeesappspendingcheck.iat[0,6]
                                                    start_date2 = df_employeesappspendingcheck.iat[0,7]
                                                    end_date2 = df_employeesappspendingcheck.iat[0,8]

                                                    buttons = [
                                                        {"type": "reply", "reply": {"id": f"Approve5appwa_{appid}", "title": "Approve"}},
                                                        {"type": "reply", "reply": {"id": f"Disapproveappwa_{appid}", "title": "Disapprove"}},
                                                    ]

                                                    send_whatsapp_message(
                                                        sender_id, 
                                                        f"Hey {first_name}, {firstnameemp2} {surnameemp2}'s {days} day {leave_type2} Leave Application, applied on {date_applied2.strftime('%d %B %Y')} and running from {start_date2.strftime('%d %B %Y')} to {end_date2.strftime('%d %B %Y')} is pending your Approval.\n\n" 
                                                        "Select an option below to either approve or disapprove this leave application.", 
                                                        buttons
                                                    )

                                            elif selected_option == "myhist" or button_id == "myhist":

                                                try:

                                                    print(id_user)
                                                    companyxx = company_reg.replace("_"," ").title()

                                                    table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                    table_name_apps_approved = f"{company_reg}appsapproved"
                                                    table_name_apps_declined = f"{company_reg}appsdeclined"
                                                    table_name_apps_cancelled = f"{company_reg}appscancelled"

                                                    query = f"SELECT appid, id, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, approvalstatus, TO_CHAR(statusdate, 'FMDD Month YYYY') AS statusdate  FROM {table_name_apps_approved} WHERE id = {str(id_user)};"
                                                    cursor.execute(query)
                                                    rows = cursor.fetchall()
                                                    df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid", "id","leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate","approvalstatus","statusdate"]) 

                                                    query = f"SELECT appid, id, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, approvalstatus, TO_CHAR(statusdate, 'FMDD Month YYYY') AS statusdate   FROM {table_name_apps_declined} WHERE id = {str(id_user)};"
                                                    cursor.execute(query)
                                                    rows = cursor.fetchall()
                                                    df_employeesappsdeclinedcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate","approvalstatus","statusdate"])  
                            
                                                    query = f"SELECT appid, id, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, approvalstatus, TO_CHAR(statusdate, 'FMDD Month YYYY') AS statusdate  FROM {table_name_apps_cancelled} WHERE id = {str(id_user)};"
                                                    cursor.execute(query)
                                                    rows = cursor.fetchall()
                                                    df_employeesappscancelledcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate","approvalstatus","statusdate"])

                                                    query = f"SELECT appid, id, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, approvalstatus FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                    cursor.execute(query)
                                                    rows = cursor.fetchall()
                                                    df_employeesappspenpendingcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate" ,"approvalstatus"])
                                                    df_employeesappspenpendingcheck["statusdate"] = ""


                                                    all_approved_declined = df_employeesappsapprovedcheck._append(df_employeesappsdeclinedcheck)
                                                    all_approved_declined_cancelled = all_approved_declined._append(df_employeesappscancelledcheck)
                                                    all_approved_declined_cancelled_pending = all_approved_declined_cancelled._append(df_employeesappspenpendingcheck)

                                                    all_approved_declined_cancelled_pending["dateapplied"] = pd.to_datetime(all_approved_declined_cancelled_pending["dateapplied"], errors='coerce')

                                                    all_approved_declined_cancelled_pending = all_approved_declined_cancelled_pending.sort_values(by="dateapplied", ascending=False)

                                                    print("hist hist hist")
                                                    all_approved_declined_cancelled_pending.drop('id', axis=1, inplace=True)
                                                    all_approved_declined_cancelled_pending["dateapplied"] = all_approved_declined_cancelled_pending["dateapplied"].dt.strftime("%-d %B %Y")

                                                    print(all_approved_declined_cancelled_pending)
                                                
                                                    def generate_leave_hist_pdf():
                                                        app = {
                                                            'company_name': company_reg.replace("_", " ").title(),
                                                            'employee_name': f"{first_name} {last_name}",
                                                            'generated_on': today_date,
                                                            'power':power,
                                                        }

                                                        table_hist_html = all_approved_declined_cancelled_pending.to_html(index=False, classes='data', border=0, justify='center',escape=False)

                                                        html_out = render_template("leave_applications_history.html", app=app, table_hist_html=table_hist_html)
                                                        pdf_bytes = HTML(string=html_out).write_pdf()
                                                        return pdf_bytes

                                                    def upload_pdf_to_whatsapp(pdf_bytes):
                                                        compxxy = company_reg.replace("_"," ").title()
                                                        filename=f"{first_name}_{last_name}_{compxxy}_leave_applications_history.pdf"
                                                    
                                                        url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/media"
                                                        headers = {
                                                            "Authorization": f"Bearer {ACCESS_TOKEN}"
                                                        }

                                                        files = {
                                                            "file": (filename, io.BytesIO(pdf_bytes), "application/pdf"),
                                                            "type": (None, "application/pdf"),
                                                            "messaging_product": (None, "whatsapp")
                                                        }

                                                        response = requests.post(url, headers=headers, files=files)
                                                        print("📥 Full incoming data:", response.text)  # Good for debugging
                                                        response.raise_for_status()
                                                        return response.json()["id"]

                                                                                                    
                                                    def send_whatsapp_pdf_by_media_id(recipient_number, media_id):
                                                        compxxy = company_reg.replace("_"," ").title()
                                                        filename=f"{first_name}_{last_name}_{compxxy}_leave_applications_history.pdf"
                                                        url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                        headers = {
                                                            "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                            "Content-Type": "application/json"
                                                        }
                                                        payload = {
                                                            "messaging_product": "whatsapp",
                                                            "to": recipient_number,
                                                            "type": "document",
                                                            "document": {
                                                                "id": media_id,            # Media ID from upload step
                                                                "filename": filename       # Desired file name on recipient's phone
                                                            }
                                                        }

                                                        response = requests.post(url, headers=headers, json=payload)
                                                        response.raise_for_status()
                                                        return response.json()


                                                    pdf_path = generate_leave_hist_pdf()
                                                    media_id = upload_pdf_to_whatsapp(pdf_path)

                                                    appscountnum = len(all_approved_declined_cancelled_pending)

                                                    if appscountnum > 0:

                                                        send_whatsapp_pdf_by_media_id(sender_id, media_id)

                                                        buttons = [
                                                            {"type": "reply", "reply": {"id": "Analytics", "title": "Analytics & Insights"}},
                                                            {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                                                            {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                        ]
                                                        send_whatsapp_message(
                                                            sender_id, 
                                                            f"Hey {first_name} {last_name} from {companyxx}! You may go ahead and download your leave applications history file attached here 😎.", 
                                                            buttons
                                                        )

                                                    else:

                                                        buttons = [
                                                            {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                                                            {"type": "reply", "reply": {"id": "Checkbal", "title": "Check Days Balance"}},
                                                            {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                        ]

                                                        send_whatsapp_message(
                                                            sender_id, 
                                                            f"Hello {first_name} {last_name} from {companyxx}!\n\n You have not applied for any leave days yet.", 
                                                            buttons
                                                        )



                                                except Exception as e:

                                                    send_whatsapp_message(f"+263710910052", f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully! Error; {e}")                      

                                            elif selected_option == "Myinfo":

                                                companyxx = company_reg.replace("_"," ").title()

                                                try:

                                                    table_name = f"{company_reg}main"

                                                    query = f"SELECT id, firstname, surname, whatsapp, address, email, role, department, currentleavedaysbalance, monthlyaccumulation, leaveapprovername, leaveapproverwhatsapp, leaveapproveremail FROM {table_name} WHERE id = {str(id_user)};"
                                                    cursor.execute(query)
                                                    row = cursor.fetchone()

                                                    if row:

                                                        columns = ["ID", "First Name", "Surname", "WhatsApp", "Address", "Email", 
                                                                "Role", "Department", "Leave Days", "Monthly Accrual", 
                                                                "Approver", "Approver WhatsApp", "Approver Email"]

                                                        message_text = "*📄 Employee Details:*\n\n"
                                                        for col, val in zip(columns, row):
                                                            message_text += f"*{col}:* {val}\n"

                                                        sections = [
                                                            {
                                                                "title": "User Options",
                                                                "rows": [
                                                                    {"id": "Editname", "title": "Edit My Name"},
                                                                    {"id": "Editwhatsapp", "title": "Change My WhatsApp #"},
                                                                    {"id": "Editemail", "title": "Change My Email"},
                                                                    {"id": "Editwebpass", "title": "Change Web Password"},
                                                                    {"id": "Editaddress", "title": "Edit My Address"},
                                                                    {"id": "MyInfo", "title": "My Info"},
                                                                    {"id": "Menu", "title": "Main Menu"}
                                                                ]
                                                            }
                                                        ]

                                                        send_whatsapp_list_message(
                                                            sender_id, 
                                                            f"Hey there {first_name}!\n Your information in {companyxx}'s Leave Management System is as follows;\n\n {message_text}", 
                                                        "User Options",
                                                        sections)



                                                except Exception as e:

                                                    print(e)

                                                    send_whatsapp_message(f"+263710910052", f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully! Error; {e}")                      

                                            elif button_id == "Track" or selected_option == "Track":

                                                table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                table_name_apps_approved = f"{company_reg}appsapproved"
                                                table_name_apps_declined = f"{company_reg}appsdeclined"
                                                table_name_apps_cancelled = f"{company_reg}appscancelled"


                                                query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leaveapproverwhatsapp, appid  FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                cursor.execute(query)
                                                rows = cursor.fetchall()

                                                df_employeesappspendingcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor", "leaveapproverwhatsapp","appid"])    

                                                if len(df_employeesappspendingcheck) == 0:

                                                    query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate, leavedaysbalancebf  FROM {table_name_apps_approved} WHERE id = {str(id_user)};"
                                                    cursor.execute(query)
                                                    rows = cursor.fetchall()
                                                    df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf"]) 

                                                    query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate, leavedaysbalancebf  FROM {table_name_apps_declined} WHERE id = {str(id_user)};"
                                                    cursor.execute(query)
                                                    rows = cursor.fetchall()
                                                    df_employeesappsdeclinedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf"])  
                            
                                                    query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate, leavedaysbalancebf  FROM {table_name_apps_cancelled} WHERE id = {str(id_user)};"
                                                    cursor.execute(query)
                                                    rows = cursor.fetchall()
                                                    df_employeesappscancelledcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf"])
                            
                                                    all_approved_declined = df_employeesappsapprovedcheck._append(df_employeesappsdeclinedcheck)
                                                    all_approved_declined_cancelled = all_approved_declined._append(df_employeesappscancelledcheck)
                                                    all_approved_declined_cancelled = all_approved_declined_cancelled.sort_values(by="appid", ascending=False) 

                                                    if len(all_approved_declined_cancelled) > 0:


                                                        print(f" hhhhhhhhhhhhhhhhhhhh  {all_approved_declined_cancelled.iat[0,8] }")

                                                        if all_approved_declined_cancelled.iat[0,8] == "Approved":

                                                            buttons = [
                                                                {"type": "reply", "reply": {"id": "Revoke", "title": "Revoke Application"}},
                                                                {"type": "reply", "reply": {"id": "myhist", "title": "Download My History"}},
                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                            ]
                                                            send_whatsapp_message(
                                                                sender_id, 
                                                                f"Hey {first_name}, your recent `{all_approved_declined_cancelled.iat[0,2]}` Leave Application `[ID - {all_approved_declined_cancelled.iat[0,0]}]` that you applied for on `{all_approved_declined_cancelled.iat[0,4].strftime('%d %B %Y')}` for `{all_approved_declined_cancelled.iat[0,7]} days` from `{all_approved_declined_cancelled.iat[0,5].strftime('%d %B %Y')}` to `{all_approved_declined_cancelled.iat[0,6].strftime('%d %B %Y')}` was {all_approved_declined_cancelled.iat[0,8]}✅ by `{all_approved_declined_cancelled.iat[0,3].title()}` on `{all_approved_declined_cancelled.iat[0,9].strftime('%d %B %Y')}`." 
                                                            )


                                                            def generate_leave_pdf():
                                                                app = {
                                                                    'company_logo': 44,
                                                                    'company_name': company_reg.replace("_"," ").title(),
                                                                    'employee_name': f"{first_name} {last_name}",
                                                                    'leave_type': all_approved_declined_cancelled.iat[0,2],
                                                                    'generated_on': today_date,
                                                                    'date_applied': all_approved_declined_cancelled.iat[0,4].strftime('%d %B %Y'),
                                                                    'approver_name': all_approved_declined_cancelled.iat[0,3].title(),
                                                                    'reference_number': all_approved_declined_cancelled.iat[0,0],
                                                                    'approved_date': all_approved_declined_cancelled.iat[0,9].strftime('%d %B %Y'),
                                                                    'new_balance': all_approved_declined_cancelled.iat[0,10],
                                                                    'start_date':  all_approved_declined_cancelled.iat[0,5].strftime('%d %B %Y'),
                                                                    'end_date':  all_approved_declined_cancelled.iat[0,6].strftime('%d %B %Y'),
                                                                    'days_requested':  all_approved_declined_cancelled.iat[0,7], 
                                                                    'address': address_foc_8, 
                                                                    'whatsapp': whatsapp_foc_8, 
                                                                    'email': email_foc_8, 
                                                                    'status': 'Approved',
                                                                    'power': power,
                                                                }

                                                                html_out = render_template("leave_pdf_template.html", app=app)
                                                                
                                                                # ✅ Return as bytes instead of saving to file
                                                                pdf_bytes = HTML(string=html_out).write_pdf()
                                                                return pdf_bytes


                                                            def upload_pdf_to_whatsapp(pdf_bytes):
                                                                compxxy = company_reg.replace("_"," ").title()
                                                                filename=f"leave_application_{all_approved_declined_cancelled.iat[0,0]}_{first_name}_{last_name}_{compxxy}.pdf"
                                                            
                                                                url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/media"
                                                                headers = {
                                                                    "Authorization": f"Bearer {ACCESS_TOKEN}"
                                                                }

                                                                files = {
                                                                    "file": (filename, io.BytesIO(pdf_bytes), "application/pdf"),
                                                                    "type": (None, "application/pdf"),
                                                                    "messaging_product": (None, "whatsapp")
                                                                }

                                                                response = requests.post(url, headers=headers, files=files)
                                                                print("📥 Full incoming data:", response.text)  # Good for debugging
                                                                response.raise_for_status()
                                                                return response.json()["id"]

                                                                                                            
                                                            def send_whatsapp_pdf_by_media_id(recipient_number, media_id):
                                                                compxxy = company_reg.replace("_"," ").title()
                                                                filename=f"leave_application_{all_approved_declined_cancelled.iat[0,0]}_{first_name}_{last_name}_{compxxy}.pdf"
                                                                url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                                headers = {
                                                                    "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                                    "Content-Type": "application/json"
                                                                }
                                                                payload = {
                                                                    "messaging_product": "whatsapp",
                                                                    "to": recipient_number,
                                                                    "type": "document",
                                                                    "document": {
                                                                        "id": media_id,            # Media ID from upload step
                                                                        "filename": filename       # Desired file name on recipient's phone
                                                                    }
                                                                }

                                                                response = requests.post(url, headers=headers, json=payload)
                                                                response.raise_for_status()
                                                                return response.json()


                                                            pdf_path = generate_leave_pdf()
                                                            media_id = upload_pdf_to_whatsapp(pdf_path)
                                                            send_whatsapp_pdf_by_media_id(sender_id, media_id)

                                                            send_whatsapp_message(
                                                                sender_id,
                                                                "Select an option below to continue 👇",
                                                                buttons
                                                            )

                                                        elif all_approved_declined_cancelled.iat[0,8] == "Disapproved":

                                                            app_id = all_approved_declined_cancelled.iat[0,0] 

                                                            buttons = [
                                                                {"type": "reply", "reply": {"id": f"Resubapp_{app_id}", "title": "ReSubmit Application"}},
                                                                {"type": "reply", "reply": {"id": "myhist", "title": "Download My History"}},
                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                            ]
                                                            send_whatsapp_message(
                                                                sender_id, 
                                                                f"Hey {first_name}, your recent `{all_approved_declined_cancelled.iat[0,2]}` Leave Application `[ID - {all_approved_declined_cancelled.iat[0,9]}]` that you applied for on `{all_approved_declined_cancelled.iat[0,4].strftime('%d %B %Y')}` for `{all_approved_declined_cancelled.iat[0,7]} days` from `{all_approved_declined_cancelled.iat[0,5].strftime('%d %B %Y')}` to `{all_approved_declined_cancelled.iat[0,6].strftime('%d %B %Y')}` was {all_approved_declined_cancelled.iat[0,8]}❌ by `{all_approved_declined_cancelled.iat[0,3].title()}` on `{all_approved_declined_cancelled.iat[0,9].strftime('%d %B %Y')}`.",
                                                                buttons 
                                                            )

                                                        elif all_approved_declined_cancelled.iat[0,8] == "Cancelled":

                                                            app_id = all_approved_declined_cancelled.iat[0,0] 

                                                            buttons = [
                                                                {"type": "reply", "reply": {"id": f"Resubapp_{app_id}", "title": "ReSubmit Application"}},
                                                                {"type": "reply", "reply": {"id": "myhist", "title": "Download My History"}},
                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                            ]
                                                            send_whatsapp_message(
                                                                sender_id, 
                                                                f"Hey {first_name}, on `{all_approved_declined_cancelled.iat[0,9].strftime('%d %B %Y')}` you Cancelled ⛔ your recent `{all_approved_declined_cancelled.iat[0,2]} Leave Application [ID - {all_approved_declined_cancelled.iat[0,0]}]` that you applied for on `{all_approved_declined_cancelled.iat[0,4].strftime('%d %B %Y')}` for `{all_approved_declined_cancelled.iat[0,7]} days` from `{all_approved_declined_cancelled.iat[0,5].strftime('%d %B %Y')}` to `{all_approved_declined_cancelled.iat[0,6].strftime('%d %B %Y')}`.",
                                                                buttons 
                                                            )

                                                    else:


                                                        buttons = [
                                                            {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                                                            {"type": "reply", "reply": {"id": "Checkbal", "title": "Check Days Balance"}},
                                                            {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                        ]
                                                        companyxx = company_reg.replace("_"," ").title()

                                                        send_whatsapp_message(
                                                            sender_id, 
                                                            f"Hello {first_name} {last_name}, LMS Leave Applications Approver from {companyxx}!\n\n You have not applied for any leave days yet.", 
                                                            buttons 
                                                        )

                                                elif len(df_employeesappspendingcheck) > 0:

                                                    app_idx = df_employeesappspendingcheck.iat[0,0]

                                                    buttons = [
                                                        {"type": "reply", "reply": {"id": f"Reminder_{app_idx}", "title": "Remind Approver"}},
                                                        {"type": "reply", "reply": {"id": "Cancelapp", "title": "Cancel Pending App"}},
                                                        {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                    ]
                                                    approoooover = df_employeesappspendingcheck.iat[0,3].title()
                                                    send_whatsapp_message(
                                                        sender_id, 
                                                        f"Hey {first_name}, your recent `{df_employeesappspendingcheck.iat[0,2]}` Leave Application `[ID - {df_employeesappspendingcheck.iat[0,0]}]` applied on `{df_employeesappspendingcheck.iat[0,4].strftime('%d %B %Y')}` for `{df_employeesappspendingcheck.iat[0,7]} days from {df_employeesappspendingcheck.iat[0,5].strftime('%d %B %Y')} to {df_employeesappspendingcheck.iat[0,6].strftime('%d %B %Y')}` is still pending approval from `{approoooover}`.\n\n" 
                                                        f"Select an option below to either remind `{approoooover}` to approve your pending leave application or you can cancel the pending application to submit a new leave application."         
                                                        , 
                                                        buttons
                                                    )
                                            
                                            elif button_id == "Menu" or selected_option == "Menu":

                                                companyxx = company_reg.replace("_", " ").title()

                                                sections = [
                                                    {
                                                        "title": "Approver Options",
                                                        "rows": [
                                                            {"id": "Apply", "title": "Apply for Leave"},
                                                            {"id": "Track", "title": "Track My Application"},
                                                            {"id": "Checkbal", "title": "Check Days Balance"},
                                                            {"id": "Pending", "title": "Apps Pending My Approval"},
                                                            {"id": "myhist", "title": "My Applications History"},
                                                            {"id": "Myinfo", "title": "My Info"}
                                                        ]
                                                    }
                                                ]


                                                send_whatsapp_list_message(
                                                    sender_id, 
                                                    f"Hello {first_name} {last_name}, LMS Leave Applications Approver from {companyxx}!\n\n {bot} LMS Bot Here 😎. How can I assist you?", 
                                                "User Options",
                                                sections)

                                            elif button_id == "Submitapp":
                                    
                                                try:
                                                    
                                                    table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                    table_name_apps_approved = f"{company_reg}appsapproved"
                                                    companyxx = company_reg.replace("_", " ").title()

                                                    cursor.execute("""
                                                        SELECT id ,empidwa, leavetypewa, startdate, enddate FROM whatsapptempapplication
                                                        WHERE empidwa = %s
                                                    """, (id_user,))
                                            
                                                    result = cursor.fetchone()

                                                    appid = result[0]
                                                    id_user = result[1]
                                                    leavetype = result[2]
                                                    startdate = result[3]
                                                    enddate = result[4]
                                                    table_name = f"{company_reg}main"
                                                    
                                                    query = f"SELECT id FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                    cursor.execute(query)
                                                    rows = cursor.fetchall()

                                                    df_employeesappspendingcheck = pd.DataFrame(rows, columns=["id"])    

                                                    if len(df_employeesappspendingcheck) == 0:

                                                        query = f"""SELECT appid, id, leavestartdate, leaveenddate FROM {table_name_apps_approved} WHERE id = %s AND leavestartdate <= %s AND leaveenddate >= %s"""

                                                        cursor.execute(query, (id_user, enddate, startdate))
                                                        results = cursor.fetchall()

                                                        # Process results
                                                        if results:
                                                            print("Overlapping records found:")

                                                            try:

                                                                overlap_messages = []

                                                                for row in results:

                                                                    formatted_date_start = row[2].strftime("%d %B %Y")
                                                                    formatted_date_end = row[3].strftime("%d %B %Y")

                                                                    overlap_messages.append(f"appID: {row[0]}, Starting Date: {formatted_date_start}, Ending Date: {formatted_date_end}")

                                                                # Combine into one single string (newline-separated)
                                                                overlap_info = "\n".join(overlap_messages)

                                                                buttons = [
                                                                    {"type": "reply", "reply": {"id": f"Apply", "title": "Restart Application"}},
                                                                    {"type": "reply", "reply": {"id": f"ApplyRevoke", "title": "Revoke Conflictn App"}},
                                                                    {"type": "reply", "reply": {"id": f"Menu", "title": "Main Menu"}},
                                                                ]

                                                                send_whatsapp_message(sender_id, f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully!\n\n"
                                                                    f"One of your previously approved leave applications include days within the period that you are currently applying for.\n\n Leave App; {overlap_info}.\n\n Either restart your application with different dates from these, or apply that this conflicting approved Leave application be Revoked.",
                                                                    buttons
                                                                    )
                                                            
                                                            except Exception as e:

                                                                send_whatsapp_message(f"+263710910052", f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully! Error; {e}")                      
                                                        
                                                        else:

                                                            print("No Overlapping records found:")

                                                            table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                            table_name_apps_approved = f"{company_reg}appsapproved"

                                                            query = f"SELECT id FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                            cursor.execute(query)
                                                            rows = cursor.fetchall()

                                                            df_employeesappspendingcheck = pd.DataFrame(rows, columns=["id"])    

                                                            if len(df_employeesappspendingcheck) == 0:

                                                                cursor.execute("""
                                                                    SELECT id ,empidwa, leavetypewa, startdate, enddate FROM whatsapptempapplication
                                                                    WHERE empidwa = %s
                                                                """, (id_user,))
                                                        
                                                                result = cursor.fetchone()

                                                                appid = result[0]
                                                                leavetype = result[2]
                                                                startdate = result[3]
                                                                enddate = result[4]
                                                                table_name = f"{company_reg}main"

                                                                if isinstance(startdate, str):
                                                                    startdate = datetime.datetime.strptime(startdate, "%Y-%m-%d").date()
                                                                if isinstance(enddate, str):
                                                                    enddate = datetime.datetime.strptime(enddate, "%Y-%m-%d").date()

                                                                business_days = 0
                                                                current_date = startdate

                                                                while current_date <= enddate:
                                                                    if current_date.weekday() != 6:  # 0=Mon, 1=Tue, ..., 4=Fri
                                                                        business_days += 1
                                                                    current_date += timedelta(days=1)  # Use timedelta directly

                                                                query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation, department FROM {table_name};"
                                                                cursor.execute(query)
                                                                rows = cursor.fetchall()

                                                                df_employees = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Address", "Role","Leave Approver Name","Leave Approver ID","Leave Approver Email", "Leave Approver WhatsAapp", "Leave Days Balance","Days Accumulated per Month", "Department"])
                                                                print(df_employees)
                                                                userdf = df_employees[df_employees['id'] == int(np.int64(id_user))].reset_index()
                                                                print("yeaarrrrr")
                                                                print(userdf)
                                                                firstname = userdf.iat[0,2]
                                                                surname = userdf.iat[0,3]
                                                                whatsapp = userdf.iat[0,4]
                                                                address = userdf.iat[0,6]
                                                                email = userdf.iat[0,5]
                                                                fullnamedisp = firstname + ' ' + surname
                                                                leaveapprovername = userdf.iat[0,8]
                                                                leaveapproverid = userdf.iat[0,9]
                                                                leaveapproveremail = userdf.iat[0, 10]
                                                                leaveapproverwhatsapp = userdf.iat[0,11]
                                                                role = userdf.iat[0,7]
                                                                leavedaysbalance = userdf.iat[0,12]
                                                                department = userdf.iat[0,14] 
                                                                print('check')

                                                                departmentdf = df_employees[df_employees['Department'] == department].reset_index()
                                                                numberindepartment = len(departmentdf)
                                                                
                                                                startdatex = pd.Timestamp(startdate)
                                                                enddatex = pd.Timestamp(enddate)

                                                                leave_dates = pd.date_range(startdatex, enddatex)

                                                                query = f"""
                                                                    SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate,
                                                                        leaveenddate, leavedaysappliedfor, approvalstatus, statusdate,
                                                                        leavedaysbalancebf, department
                                                                    FROM {table_name_apps_approved}
                                                                    WHERE department = %s;
                                                                """
                                                                cursor.execute(query, (department,))
                                                                rows = cursor.fetchall()

                                                                df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf","department"]) 
                                                                df_employeesappsapprovedcheck["leavestartdate"] = pd.to_datetime(df_employeesappsapprovedcheck["leavestartdate"])
                                                                df_employeesappsapprovedcheck["leaveenddate"] = pd.to_datetime(df_employeesappsapprovedcheck["leaveenddate"])
                
                                                                df_employeesappsapprovedcheck.dropna(subset=["leavestartdate", "leaveenddate"], inplace=True)
                                                                # Create daily impact report
                                                                impact_report = []

                                                                for date in leave_dates:

                                                                    date = pd.Timestamp(date)

                                                                    print(type(date))  # Should be pandas._libs.tslibs.timestamps.Timestamp or datetime.datetime
                                                                    print(df_employeesappsapprovedcheck.dtypes)  # Check all datetime columns

                                                                    on_leave = ((df_employeesappsapprovedcheck["leavestartdate"] <= date) & (df_employeesappsapprovedcheck["leaveenddate"] >= date)).sum()
                                                                    remaining = numberindepartment - on_leave - 1  # subtract 1 for the new leave
                                                                    impact_report.append({
                                                                        "date": date,  # <=== Keep as datetime, don't convert to string
                                                                        "on leave": on_leave + 1,
                                                                        "employees remaining": remaining
                                                                    })

                                                                # Convert to DataFrame for display
                                                                impact_df = pd.DataFrame(impact_report)
                                                                print("IMPAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACT")
                                                                print(impact_df)
                                                                print(numberindepartment)

                                                                impact_df["date"] = pd.to_datetime(impact_df["date"], format="%Y-%m-%d")
                                                                impact_df = impact_df[impact_df["date"].dt.weekday != 6].copy()

                                                                change = (impact_df[["on leave", "employees remaining"]] != impact_df[["on leave", "employees remaining"]].shift()).any(axis=1)
                                                                change.iloc[0] = True  # ensure the first row starts a group
                                                                impact_df["group"] = change.cumsum()

                                                                statements = []
                                                                for _, group_df in impact_df.groupby("group"):
                                                                    start = group_df["date"].iloc[0].strftime("%d %B %Y")
                                                                    end = group_df["date"].iloc[-1].strftime("%d %B %Y")
                                                                    on_leave = group_df["on leave"].iloc[0]
                                                                    remaining = group_df["employees remaining"].iloc[0]
                                                                    
                                                                    if start == end:
                                                                        statements.append(f"On {start}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                                                                    else:
                                                                        statements.append(f"From {start} to {end}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                                                                        # Combine all statements into a single variable
                                                                final_summary = "\n".join(statements)
                                                                # Print output
                                                                for s in statements:
                                                                    print(s)

                                                                if leavetype == "Annual":

                                                                    leavedaysbalancebf = float(leavedaysbalance) - float(business_days)

                                                                else:

                                                                    leavedaysbalancebf = float(leavedaysbalance)

                                                                if leavedaysbalancebf >= 0:

                                                                    status = "Pending"

                                                                    insert_query = f"""
                                                                    INSERT INTO {table_name_apps_pending_approval} (id, firstname, surname, department, leavetype, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus)
                                                                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                                                    """
                                                                    cursor.execute(insert_query, (int(np.int64(id_user)), first_name, last_name, department, leavetype, leaveapprovername, int(np.int64(leaveapproverid)), leaveapproveremail, int(np.int64(leaveapproverwhatsapp)), float(np.float64(leavedaysbalance)), today_date, startdate, enddate, float(np.int64(business_days)), float(np.float64(leavedaysbalancebf)), status))
                                                                    connection.commit()

                                                                    query = f"SELECT appid FROM {table_name_apps_pending_approval};"
                                                                    cursor.execute(query)
                                                                    rows = cursor.fetchall()

                                                                    df_employees = pd.DataFrame(rows, columns=["id"])
                                                                    leaveappid = df_employees.iat[0,0]
                                                                    companyxx = company_reg.replace("_"," ").title()
                                                                    approovvver = leaveapprovername.title()

                                                                    buttons = [
                                                                    {"type": "reply", "reply": {"id": "Track", "title": "Track Application"}},
                                                                    {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                                    ]

                                                                    send_whatsapp_message(sender_id, f"✅ Great News {first_name} from {companyxx}! \n\n Your `{leavetype} Leave Application` for `{business_days} days` from `{startdate.strftime('%d %B %Y')}` to `{enddate.strftime('%d %B %Y')}` has been submitted successfully!\n\n"
                                                                        f"Your Leave Application ID is `{leaveappid}`.\n\n"
                                                                        f"A Notification has been sent to `{approovvver}`  on `+263{leaveapproverwhatsapp}` to decide on  your application.",
                                                                        buttons)
                                                                    
                                                                    if leaveapproverwhatsapp:
                        
                                                                        buttons = [
                                                                            {"type": "reply", "reply": {"id": f"Approve5appwa_{leaveappid}", "title": "Approve"}},
                                                                            {"type": "reply", "reply": {"id": f"Disapproveappwa_{leaveappid}", "title": "Disapprove"}},
                                                                        ]
                                                                        send_whatsapp_message(
                                                                            f"263{leaveapproverwhatsapp}", 
                                                                            f"Hey {approovvver}! 😊. New `{leavetype}` Leave Application from `{first_name} {surname}` for `{business_days} days` from `{startdate.strftime('%d %B %Y')}` to `{enddate.strftime('%d %B %Y')}`.\n\n" 
                                                                            f"If you approve this leave application, {final_summary}\n\n"  
                                                                            f"Select an option below to either approve or disapprove the application."         
                                                                            , 
                                                                            buttons
                                                                        )

                                                                else:

                                                                    buttons = [
                                                                        {"type": "reply", "reply": {"id": f"Apply", "title": "Restart Application"}},
                                                                        {"type": "reply", "reply": {"id": f"Checkbal", "title": "Check Days Balance"}},
                                                                        {"type": "reply", "reply": {"id": f"Menu", "title": "Main Menu"}},
                                                                    ]

                                                                    send_whatsapp_message(sender_id, f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully!\n\n"
                                                                        f"You only have *{leavedaysbalance}* days available for leave but you are applying for *{business_days}*.\n\n You can restart your application and apply for leave such that the days between your leave start date and end date do not exceed your available balance of *{leavedaysbalance}* days.",
                                                                        buttons
                                                                        )

                                                            else:
                                                                print("leave app submission failed")

                                                except ValueError as e:
                                                    send_whatsapp_message(
                                                        sender_id,
                                                        f"{e}, ❌ No, incorrect message format. Please use:\n"
                                                        "`end 24 january 2025`\n"
                                                        "Example: `end 15 march 2024`"
                                                    )

                                            elif button_id == "Checkbal" or selected_option == "Checkbal":

                                                buttons = [
                                                {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                                                {"type": "reply", "reply": {"id": "Track", "title": "Track Application"}},
                                                {"type": "reply", "reply": {"id": f"Menu", "title": "Main Menu"}},

                                                ]

                                                send_whatsapp_message(
                                                    sender_id, 
                                                    f"Hey {first_name}, your current available leave days balance is `{days_days_balance} days`.\n\n"
                                                    "Select an option below to continue 👇",
                                                    buttons
                                                )

                                            elif button_id == "Resubapp" :

                                                table_name_apps_cancelled = f"{company_reg}appscancelled"
                                                table_name_apps_pending_approval = f"{company_reg}appspendingapproval"

                                                query = f"SELECT appid, id, firstname, surname, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail , leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf FROM {table_name_apps_cancelled} WHERE id = %s;"
                                                cursor.execute(query, (id_user,))
                                                result = cursor.fetchone()
                                                if result:
                                                    (app_id, employee_number, first_name, surname, leave_type,  leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf) = result
                                                
                                                    try:
                                                            status = "Pending"
                                                            statusdate = today_date
                                                        
                                                            insert_query = f"""
                                                            INSERT INTO {table_name_apps_pending_approval} 
                                                            (appid, id, firstname, surname, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus, statusdate)
                                                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                                            """

                                                            cursor.execute(insert_query, (
                                                                app_id, employee_number, first_name, surname, leave_type, leave_specify, approver_name, 
                                                                approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                                                                end_date, leave_days, leavedaysbalancebf, status, statusdate
                                                            ))
                                                            
                                                            connection.commit()
                                                            print("Insert successful!")

                                                    except Exception as e:
                                                        print("Error inserting data:", e)

                                                    # SQL query to delete or mark the leave as canceled
                                                    query = f"""DELETE FROM {table_name_apps_cancelled} WHERE appid = %s"""
                                                    cursor.execute(query, (app_id,))
                                                    connection.commit()                                       

                                                    companyxx = company_reg.replace("_", " ").title()
                                                    buttons = [
                                                        {"type": "reply", "reply": {"id": "Track", "title": "Track Application"}},
                                                        {"type": "reply", "reply": {"id": "Checkbal", "title": "Check Days Balance"}},
                                                        {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                    ]

                                                    send_whatsapp_message(sender_id, f"Hey {first_name} from {companyxx}! \n\n Your `{leave_type} Leave Application` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}` has been Cancelled successfully✅!\n\n"
                                                        "Select an option below to continue 👇",
                                                        buttons
                                                    )                                          
                                                
                                                else:
                                                    print("No record found for the user.")

                                            elif button_id == "Cancelapp" :

                                                table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                table_name_apps_cancelled = f"{company_reg}appscancelled"

                                                query = f"SELECT appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail , leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf FROM {table_name_apps_pending_approval} WHERE id = %s;"
                                                cursor.execute(query, (id_user,))
                                                result = cursor.fetchone()
                                                if result:
                                                    (app_id, employee_number, first_name, surname, department, leave_type,  leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf) = result
                                                
                                                    try:
                                                            status = "Cancelled"
                                                            statusdate = today_date
                                                        
                                                            insert_query = f"""
                                                            INSERT INTO {table_name_apps_cancelled} 
                                                            (appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus, statusdate)
                                                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                                            """

                                                            cursor.execute(insert_query, (
                                                                app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, 
                                                                approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                                                                end_date, leave_days, leavedaysbalancebf, status, statusdate
                                                            ))
                                                            
                                                            connection.commit()
                                                            print("Insert successful!")

                                                    except Exception as e:
                                                        print("Error inserting data:", e)

                                                    # SQL query to delete or mark the leave as canceled
                                                    query = f"""DELETE FROM {table_name_apps_pending_approval} WHERE appid = %s"""
                                                    cursor.execute(query, (app_id,))
                                                    connection.commit()                                       

                                                    companyxx = company_reg.replace("_", " ").title()
                                                    buttons = [
                                                        {"type": "reply", "reply": {"id": f"Resubapp_{app_id}", "title": "ReSubmit Application"}},
                                                        {"type": "reply", "reply": {"id": "Checkbal", "title": "Check Days Balance"}},
                                                        {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                    ]

                                                    send_whatsapp_message(sender_id, f"Hey {first_name} from {companyxx}! \n\n Your `{leave_type} Leave Application` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}` has been Cancelled successfully✅!\n\n"
                                                        "Select an option below to continue 👇",
                                                        buttons
                                                    )                                          
                                                
                                                else:
                                                    print("No record found for the user.")

                                            elif "reminder" in button_id.lower():

                                                app_id = button_id.split("_")[1]
                                                print(app_id)

                                                try:
                                                
                                                    print ("eissssssssshhhhhhhhhhhhhhhhhhhhhhhhhhhh")

                                                    table_name = company_reg + 'main'
                                                    company_name = company_reg.replace("_", " ").title()
                                                    table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                    table_name_apps_approved = f"{company_reg}appsapproved"


                                                    if not app_id:
                                                        print("none on appid")

                                                        return jsonify({"message": "Application ID is missing."}), 400

                                                    print(table_name_apps_pending_approval)

                                                    query = f"SELECT * FROM {table_name_apps_pending_approval} WHERE appid = %s;"
                                                    cursor.execute(query, (app_id,))
                                                    result = cursor.fetchone()
                                                    app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf, statuspre = result
                                                    print("chiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
                                                    print(employee_number)
                                                    print(approver_name)

                                                    try:

                                                        query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation, department FROM {table_name};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()

                                                        df_employees = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Address", "Role","Leave Approver Name","Leave Approver ID","Leave Approver Email", "Leave Approver WhatsAapp", "Leave Days Balance","Days Accumulated per Month", "Department"])
                                                        print(df_employees)
                                                        userdf = df_employees[df_employees['id'] == int(np.int64(employee_number))].reset_index()
                                                        print("yeaarrrrr")
                                                        print(userdf)
                                                        firstname = userdf.iat[0,2]
                                                        surname = userdf.iat[0,3]
                                                        whatsapp = userdf.iat[0,4]
                                                        address = userdf.iat[0,6]
                                                        email = userdf.iat[0,5]
                                                        fullnamedisp = firstname + ' ' + surname
                                                        leaveapprovername = userdf.iat[0,8]
                                                        leaveapproverid = userdf.iat[0,9]
                                                        leaveapproveremail = userdf.iat[0, 10]
                                                        leaveapproverwhatsapp = userdf.iat[0,11]
                                                        role = userdf.iat[0,7]
                                                        leavedaysbalance = userdf.iat[0,12]
                                                        department = userdf.iat[0,14] 
                                                        print('check')

                                                        departmentdf = df_employees[df_employees['Department'] == department].reset_index()
                                                        numberindepartment = len(departmentdf)
                                                        
                                                        startdatex = pd.Timestamp(start_date)
                                                        enddatex = pd.Timestamp(end_date)

                                                        leave_dates = pd.date_range(startdatex, enddatex)

                                                        query = f"""
                                                            SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate,
                                                                leaveenddate, leavedaysappliedfor, approvalstatus, statusdate,
                                                                leavedaysbalancebf, department
                                                            FROM {table_name_apps_approved}
                                                            WHERE department = %s;
                                                        """
                                                        cursor.execute(query, (department,))
                                                        rows = cursor.fetchall()

                                                        df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf","department"]) 
                                                        df_employeesappsapprovedcheck["leavestartdate"] = pd.to_datetime(df_employeesappsapprovedcheck["leavestartdate"])
                                                        df_employeesappsapprovedcheck["leaveenddate"] = pd.to_datetime(df_employeesappsapprovedcheck["leaveenddate"])
        
                                                        df_employeesappsapprovedcheck.dropna(subset=["leavestartdate", "leaveenddate"], inplace=True)
                                                        # Create daily impact report
                                                        impact_report = []

                                                        for date in leave_dates:

                                                            date = pd.Timestamp(date)

                                                            print(type(date))  # Should be pandas._libs.tslibs.timestamps.Timestamp or datetime.datetime
                                                            print(df_employeesappsapprovedcheck.dtypes)  # Check all datetime columns

                                                            on_leave = ((df_employeesappsapprovedcheck["leavestartdate"] <= date) & (df_employeesappsapprovedcheck["leaveenddate"] >= date)).sum()
                                                            remaining = numberindepartment - on_leave - 1  # subtract 1 for the new leave
                                                            impact_report.append({
                                                                "date": date,  # <=== Keep as datetime, don't convert to string
                                                                "on leave": on_leave + 1,
                                                                "employees remaining": remaining
                                                            })

                                                        # Convert to DataFrame for display
                                                        impact_df = pd.DataFrame(impact_report)
                                                        print("IMPAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACT")
                                                        print(impact_df)
                                                        print(numberindepartment)

                                                        impact_df["date"] = pd.to_datetime(impact_df["date"], format="%Y-%m-%d")
                                                        impact_df = impact_df[impact_df["date"].dt.weekday != 6].copy()

                                                        change = (impact_df[["on leave", "employees remaining"]] != impact_df[["on leave", "employees remaining"]].shift()).any(axis=1)
                                                        change.iloc[0] = True  # ensure the first row starts a group
                                                        impact_df["group"] = change.cumsum()

                                                        statements = []
                                                        for _, group_df in impact_df.groupby("group"):
                                                            start = group_df["date"].iloc[0].strftime("%d %B %Y")
                                                            end = group_df["date"].iloc[-1].strftime("%d %B %Y")
                                                            on_leave = group_df["on leave"].iloc[0]
                                                            remaining = group_df["employees remaining"].iloc[0]
                                                            
                                                            if start == end:
                                                                statements.append(f"On {start}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                                                            else:
                                                                statements.append(f"From {start} to {end}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                                                                # Combine all statements into a single variable
                                                        final_summary = "\n".join(statements)
                                                        # Print output
                                                        for s in statements:
                                                            print(s)

                                                    except Exception as e:
                                                        print(e)


                                                    sections = [
                                                        {
                                                            "title": "Approver Options",
                                                            "rows": [
                                                                {"id": "Apply", "title": "Apply for Leave"},
                                                                {"id": "Track", "title": "Track My Application"},
                                                                {"id": "Checkbal", "title": "Check Days Balance"},
                                                                {"id": "Pending", "title": "Apps Pending My Approval"},
                                                                {"id": "myhist", "title": "My Applications History"},
                                                                {"id": "Myinfo", "title": "My Info"}
                                                            ]
                                                        }
                                                    ]


                                                    send_whatsapp_list_message(
                                                        sender_id, 
                                                        f"Hey {first_name}. A reminder has been sent to {approver_name} on {approver_whatsapp} to decide on your `{leave_days}-day {leave_type} leave` running from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}` that you applied for on `{date_applied.strftime('%d %B %Y')}`✅! \n Select an option below to continue 👇",
                                                    "User Options",
                                                    sections)

                                                    if approver_whatsapp:

                                                        try:

                                                            buttons = [
                                                                {"type": "reply", "reply": {"id": f"Approve5appwa_{app_id}", "title": "Approve"}},
                                                                {"type": "reply", "reply": {"id": f"Disapproveappwa_{app_id}", "title": "Disapprove"}},
                                                            ]
                                                            send_whatsapp_message(
                                                                f"263{approver_whatsapp}", 
                                                                f"Hey {approver_name}! 😊. A gentle reminder, you have a new `{leave_type}` Leave Application from `{first_name} {surname}` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}`.\n\n" 
                                                                f"If you approve this leave application, {final_summary}\n\n"  
                                                                f"Select an option below to either approve or disapprove the application."         
                                                                , 
                                                                buttons
                                                            )

                                                        except Exception as e:
                                                            print(e)


                                                except Exception as e:
                                                    print(e)
                                                    return jsonify({"message": "Error sending reminder.", "error": str(e)}), 500
                                                
                                            elif "appwa" in button_id.lower():

                                                app_id = button_id.split("_")[1]
                                                print(app_id)

                                                if "approve5" in button_id.lower():

                                                    try:
                                                    
                                                        print ("eissssssssshhhhhhhhhhhhhhhhhhhhhhhhhhhh")

                                                        table_name = company_reg + 'main'
                                                        company_name = company_reg.replace("_", " ").title()
                                                        table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                        table_name_apps_approved = f"{company_reg}appsapproved"

                                                        if not app_id:
                                                            print("none on appid")

                                                            return jsonify({"message": "Application ID is missing."}), 400

                                                        status = "Approved"
                                                        statusdate = today_date
                                                        print("bababababababababa")
                                                        print(table_name_apps_pending_approval)

                                                        query = f"SELECT * FROM {table_name_apps_pending_approval} WHERE appid = %s;"
                                                        cursor.execute(query, (app_id,))
                                                        result = cursor.fetchone()
                                                        app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf, statuspre = result
                                                        print("chiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
                                                        print(employee_number)
                                                        print(approver_name)

                                                        try:
                                                            insert_query = f"""
                                                            INSERT INTO {table_name_apps_approved} 
                                                            (appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus, statusdate)
                                                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                                            """
                                                            
                                                            cursor.execute(insert_query, (
                                                                app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, 
                                                                approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                                                                end_date, leave_days, leavedaysbalancebf, status, statusdate
                                                            ))
                                                            
                                                            connection.commit()
                                                            print("Insert successful!")

                                                            query = f"UPDATE {table_name} SET currentleavedaysbalance = %s WHERE id = %s;"
                                                            cursor.execute(query, (leavedaysbalancebf, employee_number))
                                                            connection.commit()

                                                        except Exception as e:
                                                            print("Error inserting data:", e)

                                                        query = f"""DELETE FROM {table_name_apps_pending_approval} WHERE appid = %s"""
                                                        cursor.execute(query, (app_id,))
                                                        connection.commit()

                                                        query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation, department FROM {table_name};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()

                                                        df_employees = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Address", "Role","Leave Approver Name","Leave Approver ID","Leave Approver Email", "Leave Approver WhatsAapp", "Leave Days Balance","Days Accumulated per Month", "Department"])
                                                        print(df_employees)
                                                        userdf = df_employees[df_employees['id'] == int(np.int64(employee_number))].reset_index()
                                                        print("yeaarrrrr")
                                                        print(userdf)
                                                        firstname = userdf.iat[0,2].title()
                                                        surname = userdf.iat[0,3].title()
                                                        whatsappemp = userdf.iat[0,4]
                                                        email = userdf.iat[0,5]
                                                        address = userdf.iat[0,6]
                                                        companyxx = company_name.replace("_", " ").title()
                                                        app_namexx = approver_name.title()

                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate, leavedaysbalancebf  FROM {table_name_apps_approved} WHERE id = {str(employee_number)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()
                                                        df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf"]) 

                                                        df_employeesappsapprovedcheck = df_employeesappsapprovedcheck.sort_values(by="appid", ascending=False)  

                                                        def generate_leave_pdf():
                                                            app = {
                                                                'company_logo': 44,
                                                                'company_name': companyxx,
                                                                'employee_name': f"{first_name} {surname}",
                                                                'leave_type': leave_type,
                                                                'generated_on': today_date,
                                                                'department': department,
                                                                'date_applied': df_employeesappsapprovedcheck.iat[0,4].strftime('%d %B %Y'),
                                                                'approver_name': df_employeesappsapprovedcheck.iat[0,3].title(),
                                                                'reference_number': df_employeesappsapprovedcheck.iat[0,0],
                                                                'approved_date': df_employeesappsapprovedcheck.iat[0,9].strftime('%d %B %Y'),
                                                                'new_balance': df_employeesappsapprovedcheck.iat[0,10],
                                                                'start_date':  df_employeesappsapprovedcheck.iat[0,5].strftime('%d %B %Y'),
                                                                'end_date':  df_employeesappsapprovedcheck.iat[0,6].strftime('%d %B %Y'),
                                                                'days_requested':  df_employeesappsapprovedcheck.iat[0,7], 
                                                                'address': address, 
                                                                'whatsapp': f"+263{whatsappemp}", 
                                                                'email': email, 
                                                                'status': 'Approved',
                                                                'power': power,
                                                            }

                                                            html_out = render_template("leave_pdf_template.html", app=app)
                                                            
                                                            # ✅ Return as bytes instead of saving to file
                                                            pdf_bytes = HTML(string=html_out).write_pdf()
                                                            return pdf_bytes

                                                        

                                                        def upload_pdf_to_whatsapp(pdf_bytes):
                                                            filename=f"leave_application_{df_employeesappsapprovedcheck.iat[0,0]}_{first_name}_{surname}_{companyxx}.pdf"
                                                        
                                                            url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/media"
                                                            headers = {
                                                                "Authorization": f"Bearer {ACCESS_TOKEN}"
                                                            }

                                                            files = {
                                                                "file": (filename, io.BytesIO(pdf_bytes), "application/pdf"),
                                                                "type": (None, "application/pdf"),
                                                                "messaging_product": (None, "whatsapp")
                                                            }

                                                            response = requests.post(url, headers=headers, files=files)
                                                            print("📥 Full incoming data:", response.text)  # Good for debugging
                                                            response.raise_for_status()
                                                            return response.json()["id"]

                                                                                                        
                                                        def send_whatsapp_pdf_by_media_id(recipient_number, media_id):
                                                            filename=f"leave_application_{df_employeesappsapprovedcheck.iat[0,0]}_{first_name}_{surname}_{companyxx}.pdf"
                                                            url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                            headers = {
                                                                "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                                "Content-Type": "application/json"
                                                            }
                                                            payload = {
                                                                "messaging_product": "whatsapp",
                                                                "to": recipient_number,
                                                                "type": "document",
                                                                "document": {
                                                                    "id": media_id,            # Media ID from upload step
                                                                    "filename": filename       # Desired file name on recipient's phone
                                                                }
                                                            }

                                                            response = requests.post(url, headers=headers, json=payload)
                                                            response.raise_for_status()
                                                            return response.json()


                                                        pdf_path = generate_leave_pdf()
                                                        media_id = upload_pdf_to_whatsapp(pdf_path)

                                                        buttonsapproval = [
                                                            {"type": "reply", "reply": {"id": "Revoke", "title": "Revoke Approval"}},
                                                            {"type": "reply", "reply": {"id": "Pending", "title": "Pending My Approval"}},
                                                            {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                        ]

                                                        send_whatsapp_message(sender_id, f"✅ Great News {approver_name} from {companyxx}! \n\n You have successfully approved `{first_name} {surname}`'s  `{leave_days} day` `{leave_type} Leave Application` running from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}`✅!")
                                                        send_whatsapp_pdf_by_media_id(sender_id, media_id)
                                                        send_whatsapp_message(
                                                            sender_id,
                                                            "Select an option below to continue 👇, or Type `Hello` to view all Approver Options",
                                                            buttonsapproval
                                                        )

                                                        if whatsappemp:

                                                            buttons = [
                                                                {"type": "reply", "reply": {"id": "Revoke", "title": "Revoke Application"}},
                                                                {"type": "reply", "reply": {"id": "myhist", "title": "Download My History"}},
                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                            ]

                                                            send_whatsapp_message(f"263{whatsappemp}", f"✅ Great News {first_name} {surname} from {companyxx}! \n\n Your `{leave_type} Leave Application` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}`, has been Approved ✅ by `{app_namexx}`!")
                                                            send_whatsapp_pdf_by_media_id(f"263{whatsappemp}", media_id)
                                                            send_whatsapp_message(
                                                                f"263{whatsappemp}",
                                                                "Select an option below to continue 👇, or Type `Hello` to view all your User Options",
                                                                buttons
                                                            )

                                                    except Exception as e:
                                                        print(e)
                                                        return jsonify({"message": "Error approving leave application.", "error": str(e)}), 500


                                                elif "disapprove" in button_id.lower():

                                                    print("disapproved")

                                                    try:
                                                    
                                                        print ("eissssssssshhhhhhhhhhhhhhhhhhhhhhhhhhhh")

                                                        table_name = company_reg + 'main'
                                                        company_name = company_reg.replace("_", " ").title()
                                                        table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                        table_name_apps_approved = f"{company_reg}appsapproved"
                                                        table_name_apps_declined = f"{company_reg}appsdeclined"


                                                        if not app_id:
                                                            print("none on appid")

                                                            return jsonify({"message": "Application ID is missing."}), 400

                                                        status = "Disapproved"
                                                        statusdate = today_date
                                                        print("bababababababababa")
                                                        print(table_name_apps_pending_approval)

                                                        query = f"SELECT * FROM {table_name_apps_pending_approval} WHERE appid = %s;"
                                                        cursor.execute(query, (app_id,))
                                                        result = cursor.fetchone()
                                                        app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf, statuspre = result
                                                        print("chiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
                                                        print(employee_number)
                                                        print(approver_name)

                                                        try:
                                                            insert_query = f"""
                                                            INSERT INTO {table_name_apps_declined} 
                                                            (appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus, statusdate)
                                                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                                            """
                                                            
                                                            cursor.execute(insert_query, (
                                                                app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, 
                                                                approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                                                                end_date, leave_days, leavedaysbalancebf, status, statusdate
                                                            ))
                                                            
                                                            connection.commit()
                                                            print("Insert successful!")

                                                        except Exception as e:
                                                            print("Error inserting data:", e)

                                                        query = f"""DELETE FROM {table_name_apps_pending_approval} WHERE appid = %s"""
                                                        cursor.execute(query, (app_id,))
                                                        connection.commit()

                                                        query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation, department FROM {table_name};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()

                                                        df_employees = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Address", "Role","Leave Approver Name","Leave Approver ID","Leave Approver Email", "Leave Approver WhatsAapp", "Leave Days Balance","Days Accumulated per Month", "Department"])
                                                        print(df_employees)
                                                        userdf = df_employees[df_employees['id'] == int(np.int64(employee_number))].reset_index()
                                                        print("yeaarrrrr")
                                                        print(userdf)
                                                        firstname = userdf.iat[0,2].title()
                                                        surname = userdf.iat[0,3].title()
                                                        whatsappemp = userdf.iat[0,4]
                                                        email = userdf.iat[0,5]
                                                        address = userdf.iat[0,6]
                                                        companyxx = company_name.replace("_", " ").title()
                                                        app_namexx = approver_name.title()

                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate, leavedaysbalancebf  FROM {table_name_apps_approved} WHERE id = {str(employee_number)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()
                                                        df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf"]) 

                                                        df_employeesappsapprovedcheck["dateapplied"] = pd.to_datetime(df_employeesappsapprovedcheck["dateapplied"], errors='coerce')
                                                        df_employeesappsapprovedcheck = df_employeesappsapprovedcheck.sort_values(by="dateapplied", ascending=False)
                                                        

                                                        buttonsapproval = [
                                                            {"type": "reply", "reply": {"id": "Revokedis", "title": "Revoke Disapproval"}},
                                                            {"type": "reply", "reply": {"id": "Pending", "title": "Apps Pending My Approval"}},
                                                            {"type": "reply", "reply": {"id": "Main", "title": "Main Menu"}}
                                                        ]

                                                        send_whatsapp_message(sender_id, f"✅ Hey {approver_name} from {companyxx}! \n\n You have successfully disapproved `{first_name} {surname}`'s  `{leave_days} day` `{leave_type} Leave Application` running from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}`✅!")
                                                        send_whatsapp_message(
                                                            sender_id,
                                                            "Select an option below to continue 👇y, or Type `Hello` to view all Approver options",
                                                            buttonsapproval
                                                        )

                                                        if whatsappemp:

                                                            buttons = [
                                                                {"type": "reply", "reply": {"id": "Reapply", "title": "Resubmit Application"}},
                                                                {"type": "reply", "reply": {"id": "myhist", "title": "Download My History"}},
                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                            ]

                                                            send_whatsapp_message(f"263{whatsappemp}", f"✅ Oops, {first_name} {surname} from {companyxx}! \n\n Your `{leave_type} Leave Application` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}`, has been disapproved ❌ by `{app_namexx}`!")
                                                            send_whatsapp_message(
                                                                f"263{whatsappemp}",
                                                                "Select an option below to continue 👇",
                                                                buttons
                                                            )


                                                    except Exception as e:
                                                        print(e)
                                                        return jsonify({"message": "Error approving leave application.", "error": str(e)}), 500

                                        else:

                                            text = message.get("text", {}).get("body", "").lower()
                                            print(f"📨 Message from {sender_id}: {text}")
                                            
                                            print("yearrrrrrrrrrrrrrrrrrrrrrrrrrrssrsrsrsrsrs")

                                            print(role_foc_8)
                                            companyxx = company_reg.replace("_", " ").title()

                                            if "hello" in text.lower():

                                                sections = [
                                                    {
                                                        "title": "Approver Options",
                                                        "rows": [
                                                            {"id": "Apply", "title": "Apply for Leave"},
                                                            {"id": "Track", "title": "Track My Application"},
                                                            {"id": "Checkbal", "title": "Check Days Balance"},
                                                            {"id": "Pending", "title": "Apps Pending My Approval"},
                                                            {"id": "myhist", "title": "My Applications History"},
                                                            {"id": "Myinfo", "title": "My Info"}
                                                        ]
                                                    }
                                                ]


                                                send_whatsapp_list_message(
                                                    sender_id, 
                                                    f"Hello {first_name} {last_name}, LMS Leave Applications Approver from {companyxx}!\n\n {bot} LMS Bot Here 😎. How can I assist you?", 
                                                "User Options",
                                                sections)

                                            elif "email" in text.lower():

                                                table_name = company_reg + "main"
                                                table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                table_name_apps_cancelled = f"{company_reg}appscancelled"
                                                table_name_apps_approved = f"{company_reg}appsapproved"
                                                table_name_apps_declined = f"{company_reg}appsdeclined"
                                                table_name_apps_revoked = f"{company_reg}appsrevoked"

                                                # Regex to extract email after the word "email"
                                                match = re.search(r"email\s+([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)", text.strip(), re.IGNORECASE)

                                                if not match:
                                                    raise ValueError("Invalid Email format")

                                                    sections = [
                                                        {
                                                            "title": "User Options",
                                                            "rows": [
                                                                {"id": "Editname", "title": "Edit My Name"},
                                                                {"id": "Editwhatsapp", "title": "Change My WhatsApp #"},
                                                                {"id": "Editemail", "title": "Change My Email"},
                                                                {"id": "Editwebpass", "title": "Change Web Password"},
                                                                {"id": "Editaddress", "title": "Edit My Address"},
                                                                {"id": "MyInfo", "title": "My Info"},
                                                                {"id": "Menu", "title": "Main Menu"}
                                                            ]
                                                        }
                                                    ]

                                                    send_whatsapp_list_message(
                                                        sender_id, 
                                                        f"Hey {first_name}!\n You have provided an invalid Email Address. Enter your new email address starting with the word `email` as shown below 👇. \n\n `email epsilon@gmail.com`", 
                                                    "User Options",
                                                    sections)                


                                                email = match.group(1)
                                                print("Extracted email:", email)

                                                try:

                                                    query = f"UPDATE {table_name} SET email = %s WHERE id = %s;"
                                                    cursor.execute(query, (email, id_user))
                                                    connection.commit()

                                                    query = f"UPDATE {table_name_apps_pending_approval} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                    cursor.execute(query, (email, id_user))
                                                    connection.commit()
                                        
                                                    query = f"UPDATE {table_name_apps_cancelled} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                    cursor.execute(query, (email, id_user))
                                                    connection.commit()

                                                    query = f"UPDATE {table_name_apps_approved} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                    cursor.execute(query, (email, id_user))
                                                    connection.commit()

                                                    query = f"UPDATE {table_name_apps_declined} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                    cursor.execute(query, (email, id_user))
                                                    connection.commit()

                                                    query = f"UPDATE {table_name_apps_revoked} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                    cursor.execute(query, (email, id_user))
                                                    connection.commit()

                                                    sections = [
                                                        {
                                                            "title": "User Options",
                                                            "rows": [
                                                                {"id": "Editname", "title": "Edit My Name"},
                                                                {"id": "Editwhatsapp", "title": "Change My WhatsApp #"},
                                                                {"id": "Editemail", "title": "Change My Email"},
                                                                {"id": "Editwebpass", "title": "Change Web Password"},
                                                                {"id": "Editaddress", "title": "Edit My Address"},
                                                                {"id": "MyInfo", "title": "My Info"},
                                                                {"id": "Menu", "title": "Main Menu"}
                                                            ]
                                                        }
                                                    ]

                                                    send_whatsapp_list_message(
                                                        sender_id, 
                                                        f"Hey {first_name}!\n Email Address Successfully Changed to `{email}`. Select an option below to proceed 👇.", 
                                                    "User Options",
                                                    sections)  

                                                except Exception as e:
                                                    print(e)


                                            elif "start" in text.lower():
                                                
                                                try:
                                                    # Match: "start 20 july 2025"
                                                    match = re.match(r"start\s+(\d{1,2}\s+[a-zA-Z]+\s+\d{4})", text.strip(), re.IGNORECASE)
                                                    if not match:
                                                        raise ValueError("Invalid format")

                                                    date_part = match.group(1)
                                                    parsed_date = datetime.strptime(date_part, "%d %B %Y")  # Will raise ValueError if invalid

                                                    # ✅ Now it's safe to update the DB
                                                    cursor.execute("""
                                                        UPDATE whatsapptempapplication
                                                        SET startdate = %s
                                                        WHERE empidwa = %s
                                                    """, (date_part, id_user))
                                                    connection.commit()

                                                    cursor.execute("""
                                                        SELECT empidwa, leavetypewa FROM whatsapptempapplication
                                                        WHERE empidwa = %s
                                                    """, (id_user,))
                                                    result = cursor.fetchone()
                                                    leavetypewa = result[1] if result else "your"

                                                    send_whatsapp_message(sender_id,
                                                        f"✅ Got it! Start date saved.\n\nNow enter your last day on {leavetypewa} leave like this:\n"
                                                        "`end 28 July 2025`"
                                                    )

                                                except ValueError:
                                                    send_whatsapp_message(
                                                        sender_id,
                                                        f"❌ Invalid start date message format, {first_name}. Please use the date format givem below 👇:\n"
                                                        "`start 24 january 2025`\n\n"
                                                        "Example: `start 15 march 2024`"
                                                    )

                                                except Exception as e:
                                                    import traceback
                                                    print("🔴 Unexpected error:", e)
                                                    traceback.print_exc()

                                                    try:
                                                        send_whatsapp_message(
                                                            sender_id,
                                                            "⚠️ Something went wrong while processing your start date. Please try again or contact support."
                                                        )
                                                    except Exception as send_err:
                                                        print("🔴 Failed to send WhatsApp error message:", send_err)

                                            elif "end" in text.lower():

                                                try:
                                                    # ✅ Match "end 24 january 2025"
                                                    match = re.match(r"end\s+(\d{1,2}\s+[a-zA-Z]+\s+\d{4})", text.strip(), re.IGNORECASE)
                                                    if not match:
                                                        raise ValueError("Invalid end date format.")

                                                    date_part = match.group(1)
                                                    parsed_end_date = datetime.strptime(date_part, "%d %B %Y").date()  # Will raise ValueError if invalid

                                                    # ✅ Update DB now that it's valid
                                                    cursor.execute("""
                                                        UPDATE whatsapptempapplication
                                                        SET enddate = %s
                                                        WHERE empidwa = %s
                                                    """, (date_part, id_user))
                                                    connection.commit()

                                                    # ✅ Fetch full leave application
                                                    cursor.execute("""
                                                        SELECT id, empidwa, leavetypewa, startdate, enddate FROM whatsapptempapplication
                                                        WHERE empidwa = %s
                                                    """, (id_user,))
                                                    result = cursor.fetchone()

                                                    if not result:
                                                        raise Exception("No leave record found.")

                                                    appid = result[0]
                                                    leavetype = result[2]
                                                    startdate = result[3]
                                                    enddate = result[4]

                                                    # ✅ Ensure both dates are datetime.date objects
                                                    if isinstance(startdate, str):
                                                        startdate = datetime.strptime(startdate, "%Y-%m-%d").date()
                                                    if isinstance(enddate, str):
                                                        enddate = datetime.strptime(enddate, "%Y-%m-%d").date()

                                                    # ✅ Calculate business days
                                                    business_days = 0
                                                    current_date = startdate
                                                    while current_date <= enddate:
                                                        if current_date.weekday() != 6:  # Weekday: Mon-Fri
                                                            business_days += 1
                                                        current_date += timedelta(days=1)

                                                    # ✅ Ask user to confirm submission
                                                    buttons = [
                                                        {"type": "reply", "reply": {"id": "Submitapp", "title": "Yes, Submit"}},
                                                        {"type": "reply", "reply": {"id": "Dontsubmit", "title": "No"}}
                                                    ]
                                                    send_whatsapp_message(
                                                        sender_id,
                                                        f"📝 Do you wish to submit your `{business_days}-day {leavetype} Leave Application` from "
                                                        f"`{startdate.strftime('%d %B %Y')}` to `{enddate.strftime('%d %B %Y')}`, {first_name}?",
                                                        buttons
                                                    )

                                                except ValueError:
                                                    send_whatsapp_message(
                                                        sender_id,
                                                        f"❌ Invalid end date message format, {first_name}. Please use the date format givem below 👇:\n"
                                                        "`end 24 january 2025`\n\n"
                                                        "Example: `end 28 march 2024`"
                                                    )

                                                except Exception as e:
                                                    import traceback
                                                    print("🔴 ERROR during end date processing:", e)
                                                    traceback.print_exc()
                                                    try:
                                                        send_whatsapp_message(
                                                            sender_id,
                                                            "⚠️ Something went wrong while processing your end date. Please try again or contact support."
                                                        )
                                                    except Exception as send_err:
                                                        print("🔴 Failed to send error message via WhatsApp:", send_err)
                                                        
                                            else:
                                                send_whatsapp_message(
                                                    sender_id, 
                                                    f"{bot} LMS Bot Here 😎. Say 'hello' to start!"
                                                )

                                elif role_foc_8 == "Administrator":

                                    try:

                                        table_namexxxx = company_reg + "main"        

                                        query = f"SELECT id FROM {table_namexxxx} WHERE leaveapproverid = {str(id_user)};"
                                        cursor.execute(query)
                                        rows = cursor.fetchall()

                                        df_employeesempapp = pd.DataFrame(rows, columns=["id"])

                                        if len(df_employeesempapp) == 0:

                                            try:

                                                if message.get("type") == "interactive":
                                                    interactive = message.get("interactive", {})


                                                    if interactive.get("type") == "button_reply":
                                                        button_id = interactive.get("button_reply", {}).get("id")
                                                        print(f"🔘 Button clicked: {button_id}")
                                                        selected_option = ""

                                                    elif interactive.get("type") == "list_reply":

                                                        selected_option = interactive.get("list_reply", {}).get("id")
                                                        print(f"📋 User selected: {selected_option}")
                                                        button_id = ""


                                                        
                                                    if button_id == "Track" or selected_option == "Track":

                                                        table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                        table_name_apps_approved = f"{company_reg}appsapproved"
                                                        table_name_apps_declined = f"{company_reg}appsdeclined"
                                                        table_name_apps_cancelled = f"{company_reg}appscancelled"


                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leaveapproverwhatsapp  FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()

                                                        df_employeesappspendingcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor", "leaveapproverwhatsapp"])    

                                                        if len(df_employeesappspendingcheck) == 0:

                                                            query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate, leavedaysbalancebf  FROM {table_name_apps_approved} WHERE id = {str(id_user)};"
                                                            cursor.execute(query)
                                                            rows = cursor.fetchall()
                                                            df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf"]) 

                                                            query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate, leavedaysbalancebf  FROM {table_name_apps_declined} WHERE id = {str(id_user)};"
                                                            cursor.execute(query)
                                                            rows = cursor.fetchall()
                                                            df_employeesappsdeclinedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf"])  
                                    
                                                            query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate, leavedaysbalancebf  FROM {table_name_apps_cancelled} WHERE id = {str(id_user)};"
                                                            cursor.execute(query)
                                                            rows = cursor.fetchall()
                                                            df_employeesappscancelledcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf"])
                                    
                                                            all_approved_declined = df_employeesappsapprovedcheck._append(df_employeesappsdeclinedcheck)
                                                            all_approved_declined_cancelled = all_approved_declined._append(df_employeesappscancelledcheck)
                                                            all_approved_declined_cancelled = all_approved_declined_cancelled.sort_values(by="appid", ascending=False)

                                                            if len(all_approved_declined_cancelled) > 0:
        
                                                                print(f" hhhhhhhhhhhhhhhhhhhh  {all_approved_declined_cancelled.iat[0,8] }")

                                                                if all_approved_declined_cancelled.iat[0,8] == "Approved":

                                                                    buttons = [
                                                                        {"type": "reply", "reply": {"id": "Revoke", "title": "Revoke Application"}},
                                                                        {"type": "reply", "reply": {"id": "myhist", "title": "Download My History"}},
                                                                        {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                                    ]
                                                                    send_whatsapp_message(
                                                                        sender_id, 
                                                                        f"Hey {first_name}, your recent `{all_approved_declined_cancelled.iat[0,2]}` Leave Application `[ID - {all_approved_declined_cancelled.iat[0,0]}]` that you applied for on `{all_approved_declined_cancelled.iat[0,4].strftime('%d %B %Y')}` for `{all_approved_declined_cancelled.iat[0,7]} days` from `{all_approved_declined_cancelled.iat[0,5].strftime('%d %B %Y')}` to `{all_approved_declined_cancelled.iat[0,6].strftime('%d %B %Y')}` was {all_approved_declined_cancelled.iat[0,8]}✅ by `{all_approved_declined_cancelled.iat[0,3].title()}` on `{all_approved_declined_cancelled.iat[0,9].strftime('%d %B %Y')}`." 
                                                                    )


                                                                    def generate_leave_pdf():
                                                                        app = {
                                                                            'company_logo': 44,
                                                                            'company_name': company_reg.replace("_"," ").title(),
                                                                            'employee_name': f"{first_name} {last_name}",
                                                                            'leave_type': all_approved_declined_cancelled.iat[0,2],
                                                                            'generated_on': today_date,
                                                                            'date_applied': all_approved_declined_cancelled.iat[0,4].strftime('%d %B %Y'),
                                                                            'approver_name': all_approved_declined_cancelled.iat[0,3].title(),
                                                                            'reference_number': all_approved_declined_cancelled.iat[0,0],
                                                                            'approved_date': all_approved_declined_cancelled.iat[0,9].strftime('%d %B %Y'),
                                                                            'new_balance': all_approved_declined_cancelled.iat[0,10],
                                                                            'start_date':  all_approved_declined_cancelled.iat[0,5].strftime('%d %B %Y'),
                                                                            'end_date':  all_approved_declined_cancelled.iat[0,6].strftime('%d %B %Y'),
                                                                            'days_requested':  all_approved_declined_cancelled.iat[0,7], 
                                                                            'address': address_foc_8, 
                                                                            'whatsapp': whatsapp_foc_8, 
                                                                            'email': email_foc_8, 
                                                                            'status': 'Approved',
                                                                            'power': power,
                                                                        }

                                                                        html_out = render_template("leave_pdf_template.html", app=app)
                                                                        
                                                                        # ✅ Return as bytes instead of saving to file
                                                                        pdf_bytes = HTML(string=html_out).write_pdf()
                                                                        return pdf_bytes

                                                                

                                                                    def upload_pdf_to_whatsapp(pdf_bytes):
                                                                        compxxy = company_reg.replace("_"," ").title()
                                                                        filename=f"leave_application_{all_approved_declined_cancelled.iat[0,0]}_{first_name}_{last_name}_{compxxy}.pdf"
                                                                    
                                                                        url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/media"
                                                                        headers = {
                                                                            "Authorization": f"Bearer {ACCESS_TOKEN}"
                                                                        }

                                                                        files = {
                                                                            "file": (filename, io.BytesIO(pdf_bytes), "application/pdf"),
                                                                            "type": (None, "application/pdf"),
                                                                            "messaging_product": (None, "whatsapp")
                                                                        }

                                                                        response = requests.post(url, headers=headers, files=files)
                                                                        print("📥 Full incoming data:", response.text)  # Good for debugging
                                                                        response.raise_for_status()
                                                                        return response.json()["id"]

                                                                                                                    
                                                                    def send_whatsapp_pdf_by_media_id(recipient_number, media_id):
                                                                        compxxy = company_reg.replace("_"," ").title()
                                                                        filename=f"leave_application_{all_approved_declined_cancelled.iat[0,0]}_{first_name}_{last_name}_{compxxy}.pdf"
                                                                        url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                                        headers = {
                                                                            "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                                            "Content-Type": "application/json"
                                                                        }
                                                                        payload = {
                                                                            "messaging_product": "whatsapp",
                                                                            "to": recipient_number,
                                                                            "type": "document",
                                                                            "document": {
                                                                                "id": media_id,            # Media ID from upload step
                                                                                "filename": filename       # Desired file name on recipient's phone
                                                                            }
                                                                        }

                                                                        response = requests.post(url, headers=headers, json=payload)
                                                                        response.raise_for_status()
                                                                        return response.json()


                                                                    pdf_path = generate_leave_pdf()
                                                                    media_id = upload_pdf_to_whatsapp(pdf_path)
                                                                    send_whatsapp_pdf_by_media_id(sender_id, media_id)

                                                                    send_whatsapp_message(
                                                                        sender_id,
                                                                        "Select an option below to continue 👇",
                                                                        buttons
                                                                    )

                                                                elif all_approved_declined_cancelled.iat[0,8] == "Disapproved":

                                                                    app_id = all_approved_declined_cancelled.iat[0,0]

                                                                    buttons = [
                                                                        {"type": "reply", "reply": {"id": f"Resubapp_{app_id}", "title": "ReSubmit Application"}},
                                                                        {"type": "reply", "reply": {"id": "myhist", "title": "Download My History"}},
                                                                        {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                                    ]
                                                                    send_whatsapp_message(
                                                                        sender_id, 
                                                                        f"Hey {first_name}, your recent `{all_approved_declined_cancelled.iat[0,2]}` Leave Application `[ID - {all_approved_declined_cancelled.iat[0,0]}]` that you applied for on `{all_approved_declined_cancelled.iat[0,4].strftime('%d %B %Y')}` for `{all_approved_declined_cancelled.iat[0,7]} days` from `{all_approved_declined_cancelled.iat[0,5].strftime('%d %B %Y')}` to `{all_approved_declined_cancelled.iat[0,6].strftime('%d %B %Y')}` was {all_approved_declined_cancelled.iat[0,8]}❌ by `{all_approved_declined_cancelled.iat[0,3].title()}` on `{all_approved_declined_cancelled.iat[0,9].strftime('%d %B %Y')}`.",
                                                                        buttons 
                                                                    )

                                                                elif all_approved_declined_cancelled.iat[0,8] == "Cancelled":

                                                                    buttons = [
                                                                        {"type": "reply", "reply": {"id": f"Resubapp_{app_id}", "title": "ReSubmit Application"}},
                                                                        {"type": "reply", "reply": {"id": "myhist", "title": "Download My History"}},
                                                                        {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                                    ]
                                                                    send_whatsapp_message(
                                                                        sender_id, 
                                                                        f"Hey {first_name}, on `{all_approved_declined_cancelled.iat[0,9].strftime('%d %B %Y')}` you Cancelled ⛔ your recent `{all_approved_declined_cancelled.iat[0,2]} Leave Application [ID - {all_approved_declined_cancelled.iat[0,0]}]` that you applied for on `{all_approved_declined_cancelled.iat[0,4].strftime('%d %B %Y')}` for `{all_approved_declined_cancelled.iat[0,7]} days` from `{all_approved_declined_cancelled.iat[0,5].strftime('%d %B %Y')}` to `{all_approved_declined_cancelled.iat[0,6].strftime('%d %B %Y')}`.",
                                                                        buttons 
                                                                    )
                                                            
                                                            else:


                                                                sections = [
                                                                    {
                                                                        "title": "Administrator Options",
                                                                        "rows": [
                                                                            {"id": "Apply", "title": "Apply for Leave"},
                                                                            {"id": "Track", "title": "Track My Application"},
                                                                            {"id": "Checkbal", "title": "Check Days Balance"},
                                                                            {"id": "myhist", "title": "My Applications History"},
                                                                            {"id": "Myinfo", "title": "My Info"},
                                                                            {"id": "Empmgt", "title": "Employee Management"},
                                                                            {"id": "Analyticscomp", "title": "Analytics & Insights"},
                                                                            {"id": "Company", "title": "Company Profile"}
                                                                        ]
                                                                    }
                                                                ]
                                                                companyxx = company_reg.replace("_"," ").title()


                                                                send_whatsapp_list_message(
                                                                    sender_id, 
                                                                    f"Hello {first_name} {last_name}, LMS Leave Applications Approver from {companyxx}!\n\n You have not applied for any leave days yet.", 
                                                                    "Administrator Options",
                                                                    sections
                                                                )


                                                        elif len(df_employeesappspendingcheck) > 0:

                                                            app_idx = df_employeesappspendingcheck.iat[0,0]

                                                            buttons = [
                                                                {"type": "reply", "reply": {"id": f"Reminder_{app_idx}", "title": "Remind Approver"}},
                                                                {"type": "reply", "reply": {"id": "Cancelapp", "title": "Cancel Pending App"}},
                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                            ]
                                                            approoooover = df_employeesappspendingcheck.iat[0,3].title()
                                                            send_whatsapp_message(
                                                                sender_id, 
                                                                f"Hey {first_name}, your recent `{df_employeesappspendingcheck.iat[0,2]}` Leave Application `[ID - {df_employeesappspendingcheck.iat[0,0]}]` applied on `{df_employeesappspendingcheck.iat[0,4].strftime('%d %B %Y')}` for `{df_employeesappspendingcheck.iat[0,7]} days from {df_employeesappspendingcheck.iat[0,5].strftime('%d %B %Y')} to {df_employeesappspendingcheck.iat[0,6].strftime('%d %B %Y')}` is still pending approval from `{approoooover}`.\n\n" 
                                                                f"Select an option below to either remind `{approoooover}` to approve your pending leave application or you can cancel the pending application to submit a new leave application."         
                                                                , 
                                                                buttons
                                                            )

                                                        else:

                                                            sections = [
                                                                {
                                                                    "title": "Administrator Options",
                                                                    "rows": [
                                                                        {"id": "Apply", "title": "Apply for Leave"},
                                                                        {"id": "Track", "title": "Track My Application"},
                                                                        {"id": "Checkbal", "title": "Check Days Balance"},
                                                                        {"id": "myhist", "title": "My Applications History"},
                                                                        {"id": "Myinfo", "title": "My Info"},
                                                                        {"id": "Empmgt", "title": "Employee Management"},
                                                                        {"id": "Analyticscomp", "title": "Analytics & Insights"},
                                                                        {"id": "Company", "title": "Company Profile"}
                                                                    ]
                                                                }
                                                            ]

                                                            send_whatsapp_list_message(
                                                                sender_id, 
                                                                f"Hello {first_name} {last_name}, LMS Leave Applications Approver from {companyxx}!\n\n You have not applied for any leave days yet.", 
                                                            "Administrator Options",
                                                            sections)

                                                    elif "reminder" in button_id.lower():

                                                        app_id = button_id.split("_")[1]
                                                        print(app_id)

                                                        try:
                                                        
                                                            print ("eissssssssshhhhhhhhhhhhhhhhhhhhhhhhhhhh")

                                                            table_name = company_reg + 'main'
                                                            company_name = company_reg.replace("_", " ").title()
                                                            table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                            table_name_apps_approved = f"{company_reg}appsapproved"


                                                            if not app_id:
                                                                print("none on appid")

                                                                return jsonify({"message": "Application ID is missing."}), 400

                                                            print(table_name_apps_pending_approval)

                                                            query = f"SELECT * FROM {table_name_apps_pending_approval} WHERE appid = %s;"
                                                            cursor.execute(query, (app_id,))
                                                            result = cursor.fetchone()
                                                            app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf, statuspre = result
                                                            print("chiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
                                                            print(employee_number)
                                                            print(approver_name)

                                                            try:

                                                                query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation, department FROM {table_name};"
                                                                cursor.execute(query)
                                                                rows = cursor.fetchall()

                                                                df_employees = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Address", "Role","Leave Approver Name","Leave Approver ID","Leave Approver Email", "Leave Approver WhatsAapp", "Leave Days Balance","Days Accumulated per Month", "Department"])
                                                                print(df_employees)
                                                                userdf = df_employees[df_employees['id'] == int(np.int64(employee_number))].reset_index()
                                                                print("yeaarrrrr")
                                                                print(userdf)
                                                                firstname = userdf.iat[0,2]
                                                                surname = userdf.iat[0,3]
                                                                whatsapp = userdf.iat[0,4]
                                                                address = userdf.iat[0,6]
                                                                email = userdf.iat[0,5]
                                                                fullnamedisp = firstname + ' ' + surname
                                                                leaveapprovername = userdf.iat[0,8]
                                                                leaveapproverid = userdf.iat[0,9]
                                                                leaveapproveremail = userdf.iat[0, 10]
                                                                leaveapproverwhatsapp = userdf.iat[0,11]
                                                                role = userdf.iat[0,7]
                                                                leavedaysbalance = userdf.iat[0,12]
                                                                department = userdf.iat[0,14] 
                                                                print('check')

                                                                departmentdf = df_employees[df_employees['Department'] == department].reset_index()
                                                                numberindepartment = len(departmentdf)
                                                                
                                                                startdatex = pd.Timestamp(start_date)
                                                                enddatex = pd.Timestamp(end_date)

                                                                leave_dates = pd.date_range(startdatex, enddatex)

                                                                query = f"""
                                                                    SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate,
                                                                        leaveenddate, leavedaysappliedfor, approvalstatus, statusdate,
                                                                        leavedaysbalancebf, department
                                                                    FROM {table_name_apps_approved}
                                                                    WHERE department = %s;
                                                                """
                                                                cursor.execute(query, (department,))
                                                                rows = cursor.fetchall()

                                                                df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf","department"]) 
                                                                df_employeesappsapprovedcheck["leavestartdate"] = pd.to_datetime(df_employeesappsapprovedcheck["leavestartdate"])
                                                                df_employeesappsapprovedcheck["leaveenddate"] = pd.to_datetime(df_employeesappsapprovedcheck["leaveenddate"])
                
                                                                df_employeesappsapprovedcheck.dropna(subset=["leavestartdate", "leaveenddate"], inplace=True)
                                                                # Create daily impact report
                                                                impact_report = []

                                                                for date in leave_dates:

                                                                    date = pd.Timestamp(date)

                                                                    print(type(date))  # Should be pandas._libs.tslibs.timestamps.Timestamp or datetime.datetime
                                                                    print(df_employeesappsapprovedcheck.dtypes)  # Check all datetime columns

                                                                    on_leave = ((df_employeesappsapprovedcheck["leavestartdate"] <= date) & (df_employeesappsapprovedcheck["leaveenddate"] >= date)).sum()
                                                                    remaining = numberindepartment - on_leave - 1  # subtract 1 for the new leave
                                                                    impact_report.append({
                                                                        "date": date,  # <=== Keep as datetime, don't convert to string
                                                                        "on leave": on_leave + 1,
                                                                        "employees remaining": remaining
                                                                    })

                                                                # Convert to DataFrame for display
                                                                impact_df = pd.DataFrame(impact_report)
                                                                print("IMPAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACT")
                                                                print(impact_df)
                                                                print(numberindepartment)

                                                                impact_df["date"] = pd.to_datetime(impact_df["date"], format="%Y-%m-%d")
                                                                impact_df = impact_df[impact_df["date"].dt.weekday != 6].copy()

                                                                change = (impact_df[["on leave", "employees remaining"]] != impact_df[["on leave", "employees remaining"]].shift()).any(axis=1)
                                                                change.iloc[0] = True  # ensure the first row starts a group
                                                                impact_df["group"] = change.cumsum()

                                                                statements = []
                                                                for _, group_df in impact_df.groupby("group"):
                                                                    start = group_df["date"].iloc[0].strftime("%d %B %Y")
                                                                    end = group_df["date"].iloc[-1].strftime("%d %B %Y")
                                                                    on_leave = group_df["on leave"].iloc[0]
                                                                    remaining = group_df["employees remaining"].iloc[0]
                                                                    
                                                                    if start == end:
                                                                        statements.append(f"On {start}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                                                                    else:
                                                                        statements.append(f"From {start} to {end}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                                                                        # Combine all statements into a single variable
                                                                final_summary = "\n".join(statements)
                                                                # Print output
                                                                for s in statements:
                                                                    print(s)

                                                            except Exception as e:
                                                                print(e)


                                                            sections = [
                                                                {
                                                                    "title": "Administrator Options",
                                                                    "rows": [
                                                                        {"id": "Apply", "title": "Apply for Leave"},
                                                                        {"id": "Track", "title": "Track My Application"},
                                                                        {"id": "Checkbal", "title": "Check Days Balance"},
                                                                        {"id": "myhist", "title": "My Applications History"},
                                                                        {"id": "Myinfo", "title": "My Info"},
                                                                        {"id": "Empmgt", "title": "Employee Management"},
                                                                        {"id": "Analyticscomp", "title": "Analytics & Insights"},
                                                                        {"id": "Company", "title": "Company Profile"}
                                                                    ]
                                                                }
                                                            ]
                                                            
                                                            send_whatsapp_list_message(
                                                                sender_id,
                                                                f"Hey {first_name}. A reminder has been sent to {approver_name} on {approver_whatsapp} to decide on your `{leave_days}-day {leave_type} leave` running from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}` that you applied for on `{date_applied.strftime('%d %B %Y')}`✅! \n Select an option below to continue 👇",
                                                                "Administrator Options",
                                                                sections
                                                            )

                                                            if approver_whatsapp:

                                                                try:

                                                                    buttons = [
                                                                        {"type": "reply", "reply": {"id": f"Approve5appwa_{app_id}", "title": "Approve"}},
                                                                        {"type": "reply", "reply": {"id": f"Disapproveappwa_{app_id}", "title": "Disapprove"}},
                                                                    ]
                                                                    send_whatsapp_message(
                                                                        f"263{approver_whatsapp}", 
                                                                        f"Hey {approver_name}! 😊. A gentle reminder, you have a new `{leave_type}` Leave Application from `{first_name} {surname}` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}`.\n\n" 
                                                                        f"If you approve this leave application, {final_summary}\n\n"  
                                                                        f"Select an option below to either approve or disapprove the application."         
                                                                        , 
                                                                        buttons
                                                                    )

                                                                except Exception as e:
                                                                    print(e)



                                                        except Exception as e:
                                                            print(e)
                                                            return jsonify({"message": "Error approving leave application.", "error": str(e)}), 500
                                                    
                                                    elif selected_option == "Myinfo":

                                                        companyxx = company_reg.replace("_"," ").title()

                                                        try:

                                                            table_name = f"{company_reg}main"

                                                            query = f"SELECT id, firstname, surname, whatsapp, address, email, role, department, currentleavedaysbalance, monthlyaccumulation, leaveapprovername, leaveapproverwhatsapp, leaveapproveremail FROM {table_name} WHERE id = {str(id_user)};"
                                                            cursor.execute(query)
                                                            row = cursor.fetchone()

                                                            if row:

                                                                columns = ["ID", "First Name", "Surname", "WhatsApp", "Address", "Email", 
                                                                        "Role", "Department", "Leave Days", "Monthly Accrual", 
                                                                        "Approver", "Approver WhatsApp", "Approver Email"]

                                                                message_text = "*📄 Employee Details:*\n\n"
                                                                for col, val in zip(columns, row):
                                                                    message_text += f"*{col}:* {val}\n"

                                                                sections = [
                                                                    {
                                                                        "title": "User Options",
                                                                        "rows": [
                                                                            {"id": "Editname", "title": "Edit My Name"},
                                                                            {"id": "Editwhatsapp", "title": "Change My WhatsApp #"},
                                                                            {"id": "Editemail", "title": "Change My Email"},
                                                                            {"id": "Editwebpass", "title": "Change Web Password"},
                                                                            {"id": "Editaddress", "title": "Edit My Address"},
                                                                            {"id": "MyInfo", "title": "My Info"},
                                                                            {"id": "Menu", "title": "Main Menu"}
                                                                        ]
                                                                    }
                                                                ]

                                                                send_whatsapp_list_message(
                                                                    sender_id, 
                                                                    f"Hey there {first_name}!\n Your information in {companyxx}'s Leave Management System is as follows;\n\n {message_text}", 
                                                                "User Options",
                                                                sections)



                                                        except Exception as e:

                                                            print(e)

                                                            send_whatsapp_message(f"+263710910052", f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully! Error; {e}")                      

                                                    elif button_id == "myhist" or selected_option == "myhist":

                                                        try:

                                                            print(id_user)
                                                            companyxx = company_reg.replace("_"," ").title()

                                                            table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                            table_name_apps_approved = f"{company_reg}appsapproved"
                                                            table_name_apps_declined = f"{company_reg}appsdeclined"
                                                            table_name_apps_cancelled = f"{company_reg}appscancelled"

                                                            query = f"SELECT appid, id, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, approvalstatus, TO_CHAR(statusdate, 'FMDD Month YYYY') AS statusdate  FROM {table_name_apps_approved} WHERE id = {str(id_user)};"
                                                            cursor.execute(query)
                                                            rows = cursor.fetchall()
                                                            df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid", "id","leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate","approvalstatus","statusdate"]) 

                                                            query = f"SELECT appid, id, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, approvalstatus, TO_CHAR(statusdate, 'FMDD Month YYYY') AS statusdate   FROM {table_name_apps_declined} WHERE id = {str(id_user)};"
                                                            cursor.execute(query)
                                                            rows = cursor.fetchall()
                                                            df_employeesappsdeclinedcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate","approvalstatus","statusdate"])  
                                    
                                                            query = f"SELECT appid, id, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, approvalstatus, TO_CHAR(statusdate, 'FMDD Month YYYY') AS statusdate  FROM {table_name_apps_cancelled} WHERE id = {str(id_user)};"
                                                            cursor.execute(query)
                                                            rows = cursor.fetchall()
                                                            df_employeesappscancelledcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate","approvalstatus","statusdate"])

                                                            query = f"SELECT appid, id, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, approvalstatus FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                            cursor.execute(query)
                                                            rows = cursor.fetchall()
                                                            df_employeesappspenpendingcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate" ,"approvalstatus"])
                                                            df_employeesappspenpendingcheck["statusdate"] = ""


                                                            all_approved_declined = df_employeesappsapprovedcheck._append(df_employeesappsdeclinedcheck)
                                                            all_approved_declined_cancelled = all_approved_declined._append(df_employeesappscancelledcheck)
                                                            all_approved_declined_cancelled_pending = all_approved_declined_cancelled._append(df_employeesappspenpendingcheck)

                                                            all_approved_declined_cancelled_pending["dateapplied"] = pd.to_datetime(all_approved_declined_cancelled_pending["dateapplied"], errors='coerce')

                                                            all_approved_declined_cancelled_pending = all_approved_declined_cancelled_pending.sort_values(by="dateapplied", ascending=False)

                                                            print("hist hist hist")
                                                            all_approved_declined_cancelled_pending.drop('id', axis=1, inplace=True)
                                                            all_approved_declined_cancelled_pending["dateapplied"] = all_approved_declined_cancelled_pending["dateapplied"].dt.strftime("%-d %B %Y")

                                                            print(all_approved_declined_cancelled_pending)
                                                        
                                                            def generate_leave_hist_pdf():
                                                                app = {
                                                                    'company_name': company_reg.replace("_", " ").title(),
                                                                    'employee_name': f"{first_name} {last_name}",
                                                                    'generated_on': today_date,
                                                                    'power': power,
                                                                }

                                                                table_hist_html = all_approved_declined_cancelled_pending.to_html(index=False, classes='data', border=0, justify='center',escape=False)

                                                                html_out = render_template("leave_applications_history.html", app=app, table_hist_html=table_hist_html)
                                                                pdf_bytes = HTML(string=html_out).write_pdf()
                                                                return pdf_bytes

                                                            def upload_pdf_to_whatsapp(pdf_bytes):
                                                                compxxy = company_reg.replace("_"," ").title()
                                                                filename=f"{first_name}_{last_name}_{compxxy}_leave_applications_history.pdf"
                                                            
                                                                url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/media"
                                                                headers = {
                                                                    "Authorization": f"Bearer {ACCESS_TOKEN}"
                                                                }

                                                                files = {
                                                                    "file": (filename, io.BytesIO(pdf_bytes), "application/pdf"),
                                                                    "type": (None, "application/pdf"),
                                                                    "messaging_product": (None, "whatsapp")
                                                                }

                                                                response = requests.post(url, headers=headers, files=files)
                                                                print("📥 Full incoming data:", response.text)  # Good for debugging
                                                                response.raise_for_status()
                                                                return response.json()["id"]

                                                                                                            
                                                            def send_whatsapp_pdf_by_media_id(recipient_number, media_id):
                                                                compxxy = company_reg.replace("_"," ").title()
                                                                filename=f"{first_name}_{last_name}_{compxxy}_leave_applications_history.pdf"
                                                                url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                                headers = {
                                                                    "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                                    "Content-Type": "application/json"
                                                                }
                                                                payload = {
                                                                    "messaging_product": "whatsapp",
                                                                    "to": recipient_number,
                                                                    "type": "document",
                                                                    "document": {
                                                                        "id": media_id,            # Media ID from upload step
                                                                        "filename": filename       # Desired file name on recipient's phone
                                                                    }
                                                                }

                                                                response = requests.post(url, headers=headers, json=payload)
                                                                response.raise_for_status()
                                                                return response.json()


                                                            pdf_path = generate_leave_hist_pdf()
                                                            media_id = upload_pdf_to_whatsapp(pdf_path)

                                                            appscountnum = len(all_approved_declined_cancelled_pending)

                                                            if appscountnum > 0:

                                                                send_whatsapp_pdf_by_media_id(sender_id, media_id)

                                                                buttons = [
                                                                    {"type": "reply", "reply": {"id": "Analytics", "title": "Analytics & Insights"}},
                                                                    {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                                                                    {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                                ]
                                                                send_whatsapp_message(
                                                                    sender_id, 
                                                                    f"Hey {first_name} {last_name} from {companyxx}! You may go ahead and download your leave applications history file attached here 😎.", 
                                                                    buttons
                                                                )

                                                            else:

                                                                buttons = [
                                                                    {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                                                                    {"type": "reply", "reply": {"id": "Checkbal", "title": "Check Days Balance"}},
                                                                    {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                                ]

                                                                send_whatsapp_message(
                                                                    sender_id, 
                                                                    f"Hello {first_name} {last_name} from {companyxx}!\n\n You have not applied for any leave days yet.", 
                                                                    buttons
                                                                )


                                                        except Exception as e:

                                                            send_whatsapp_message(f"+263710910052", f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully! Error; {e}")                      
                                                    
                                                    elif selected_option == "Analyticscomp":

                                                        companyxx = company_reg.replace("_"," ").title()

                                                        buttons = [
                                                            {"type": "reply", "reply": {"id": "Book", "title": "Extract Leave Book"}},
                                                            {"type": "reply", "reply": {"id": "Summarycomp", "title": "Get Insights"}},
                                                            {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                        ]
                                                        send_whatsapp_message(
                                                            sender_id,
                                                            f"Hey {first_name}, select an option below to proceed.",
                                                            buttons
                                                        )

                                                    elif button_id == "Menu" or selected_option == "Menu":

                                                        companyxx = company_reg.replace("_"," ").title()
                                                        
                                                        sections = [
                                                            {
                                                                "title": "Administrator Options",
                                                                "rows": [
                                                                    {"id": "Apply", "title": "Apply for Leave"},
                                                                    {"id": "Track", "title": "Track My Application"},
                                                                    {"id": "Checkbal", "title": "Check Days Balance"},
                                                                    {"id": "myhist", "title": "My Applications History"},
                                                                    {"id": "Myinfo", "title": "My Info"},
                                                                    {"id": "Empmgt", "title": "Employee Management"},
                                                                    {"id": "Analyticscomp", "title": "Analytics & Insights"},
                                                                    {"id": "Company", "title": "Company Profile"}
                                                                ]
                                                            }
                                                        ]
                                                        
                                                        send_whatsapp_list_message(
                                                            sender_id,
                                                            f"Hello {first_name} {last_name}, LMS Administrator from {companyxx}!\n\n {bot} LMS Bot Here 😎. How can I assist you?",
                                                            "Administrator Options",
                                                            sections
                                                        )

                                                    elif button_id == "Apply" or selected_option == "Apply":

                                                        table_name_apps_pending_approval = f"{company_reg}appspendingapproval"

                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor  FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()

                                                        df_employeesappspendingcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor"])    

                                                        if len(df_employeesappspendingcheck) == 0:

                                                            sections = [
                                                                {
                                                                    "title": "Leave Type Options",
                                                                    "rows": [
                                                                        {"id": "Annual", "title": "Annual Leave"},
                                                                        {"id": "Sick", "title": "Sick Leave"},
                                                                        {"id": "Study", "title": "Study Leave"},
                                                                        {"id": "Bereavement", "title": "Bereavement Leave"},
                                                                        {"id": "Parental", "title": "Parental Leave"},
                                                                        {"id": "Other", "title": "Other"},
                                                                    ]
                                                                }
                                                            ]

                                                            send_whatsapp_list_message(
                                                                sender_id, 
                                                                f"{first_name}, kindly select the type of Leave that you are applying for.", 
                                                                "Leave Type Options",
                                                                sections) 

                                                        elif len(df_employeesappspendingcheck) > 0:

                                                            app_idx = df_employeesappspendingcheck.iat[0,0]

                                                            buttons = [
                                                                {"type": "reply", "reply": {"id": f"Reminder_{app_idx}", "title": "Remind Approver"}},
                                                                {"type": "reply", "reply": {"id": "Cancelapp", "title": "Cancel Pending App"}},
                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                            ]
                                                            send_whatsapp_message(
                                                                sender_id, 
                                                                f"Oops! 🥲. Sorry {first_name}, you cannot apply for leave whilst you have another leave application which is still pending approval.\n\n" 
                                                                f"Your `{df_employeesappspendingcheck.iat[0,2]}` Leave Application `[ID - {df_employeesappspendingcheck.iat[0,0]}]` applied on `{df_employeesappspendingcheck.iat[0,4].strftime('%d %B %Y')}` for `{df_employeesappspendingcheck.iat[0,7]} days from {df_employeesappspendingcheck.iat[0,5].strftime('%d %B %Y')} to {df_employeesappspendingcheck.iat[0,6].strftime('%d %B %Y')}` is still pending approval from {df_employeesappspendingcheck.iat[0,3]}.\n\n" 
                                                                f"Select an option below to either remind the approver to approved your pending application or you can cancel the pending application to submit a new leave application."         
                                                                , 
                                                                buttons
                                                            )

                                                    elif button_id == "Submitapp":
                                            
                                                        try:
                                                            table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                            table_name_apps_approved = f"{company_reg}appsapproved"
                                                            companyxx = company_reg.replace("_", " ").title()

                                                            cursor.execute("""
                                                                SELECT id ,empidwa, leavetypewa, startdate, enddate FROM whatsapptempapplication
                                                                WHERE empidwa = %s
                                                            """, (id_user,))
                                                    
                                                            result = cursor.fetchone()

                                                            appid = result[0]
                                                            id_user = result[1]
                                                            leavetype = result[2]
                                                            startdate = result[3]
                                                            enddate = result[4]
                                                            table_name = f"{company_reg}main"
                                                            
                                                            query = f"SELECT id FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                            cursor.execute(query)
                                                            rows = cursor.fetchall()

                                                            df_employeesappspendingcheck = pd.DataFrame(rows, columns=["id"])    

                                                            if len(df_employeesappspendingcheck) == 0:

                                                                query = f"""SELECT appid, id, leavestartdate, leaveenddate FROM {table_name_apps_approved} WHERE id = %s AND leavestartdate <= %s AND leaveenddate >= %s"""

                                                                cursor.execute(query, (id_user, enddate, startdate))
                                                                results = cursor.fetchall()

                                                                # Process results
                                                                if results:
                                                                    print("Overlapping records found:")

                                                                    try:

                                                                        overlap_messages = []

                                                                        for row in results:

                                                                            formatted_date_start = row[2].strftime("%d %B %Y")
                                                                            formatted_date_end = row[3].strftime("%d %B %Y")

                                                                            overlap_messages.append(f"appID: {row[0]}, Starting Date: {formatted_date_start}, Ending Date: {formatted_date_end}")

                                                                        # Combine into one single string (newline-separated)
                                                                        overlap_info = "\n".join(overlap_messages)

                                                                        buttons = [
                                                                            {"type": "reply", "reply": {"id": f"Apply", "title": "Restart Application"}},
                                                                            {"type": "reply", "reply": {"id": f"ApplyRevoke", "title": "Revoke Conflictn App"}},
                                                                            {"type": "reply", "reply": {"id": f"Menu", "title": "Main Menu"}},
                                                                        ]

                                                                        send_whatsapp_message(sender_id, f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully!\n\n"
                                                                            f"One of your previously approved leave applications include days within the period that you are currently applying for.\n\n Leave App; {overlap_info}.\n\n Either restart your application with different dates from these, or apply that this conflicting approved Leave application be Revoked.",
                                                                            buttons
                                                                            )
                                                                    
                                                                    except Exception as e:

                                                                        send_whatsapp_message(f"+263710910052", f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully! Error; {e}")                      
                                                                
                                                                else:

                                                                    print("No Overlapping records found:")

                                                                    table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                                    table_name_apps_approved = f"{company_reg}appsapproved"

                                                                    query = f"SELECT id FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                                    cursor.execute(query)
                                                                    rows = cursor.fetchall()

                                                                    df_employeesappspendingcheck = pd.DataFrame(rows, columns=["id"])    

                                                                    if len(df_employeesappspendingcheck) == 0:

                                                                        cursor.execute("""
                                                                            SELECT id ,empidwa, leavetypewa, startdate, enddate FROM whatsapptempapplication
                                                                            WHERE empidwa = %s
                                                                        """, (id_user,))
                                                                
                                                                        result = cursor.fetchone()

                                                                        appid = result[0]
                                                                        leavetype = result[2]
                                                                        startdate = result[3]
                                                                        enddate = result[4]
                                                                        table_name = f"{company_reg}main"

                                                                        if isinstance(startdate, str):
                                                                            startdate = datetime.datetime.strptime(startdate, "%Y-%m-%d").date()
                                                                        if isinstance(enddate, str):
                                                                            enddate = datetime.datetime.strptime(enddate, "%Y-%m-%d").date()

                                                                        business_days = 0
                                                                        current_date = startdate

                                                                        while current_date <= enddate:
                                                                            if current_date.weekday() != 6:  # 0=Mon, 1=Tue, ..., 4=Fri
                                                                                business_days += 1
                                                                            current_date += timedelta(days=1)  # Use timedelta directly

                                                                        query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation, department FROM {table_name};"
                                                                        cursor.execute(query)
                                                                        rows = cursor.fetchall()

                                                                        df_employees = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Address", "Role","Leave Approver Name","Leave Approver ID","Leave Approver Email", "Leave Approver WhatsAapp", "Leave Days Balance","Days Accumulated per Month", "Department"])
                                                                        print(df_employees)
                                                                        userdf = df_employees[df_employees['id'] == int(np.int64(id_user))].reset_index()
                                                                        print("yeaarrrrr")
                                                                        print(userdf)
                                                                        firstname = userdf.iat[0,2]
                                                                        surname = userdf.iat[0,3]
                                                                        whatsapp = userdf.iat[0,4]
                                                                        address = userdf.iat[0,6]
                                                                        email = userdf.iat[0,5]
                                                                        fullnamedisp = firstname + ' ' + surname
                                                                        leaveapprovername = userdf.iat[0,8]
                                                                        leaveapproverid = userdf.iat[0,9]
                                                                        leaveapproveremail = userdf.iat[0, 10]
                                                                        leaveapproverwhatsapp = userdf.iat[0,11]
                                                                        role = userdf.iat[0,7]
                                                                        leavedaysbalance = userdf.iat[0,12]
                                                                        department = userdf.iat[0,14] 
                                                                        print('check')

                                                                        departmentdf = df_employees[df_employees['Department'] == department].reset_index()
                                                                        numberindepartment = len(departmentdf)
                                                                        
                                                                        startdatex = pd.Timestamp(startdate)
                                                                        enddatex = pd.Timestamp(enddate)

                                                                        leave_dates = pd.date_range(startdatex, enddatex)

                                                                        query = f"""
                                                                            SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate,
                                                                                leaveenddate, leavedaysappliedfor, approvalstatus, statusdate,
                                                                                leavedaysbalancebf, department
                                                                            FROM {table_name_apps_approved}
                                                                            WHERE department = %s;
                                                                        """
                                                                        cursor.execute(query, (department,))
                                                                        rows = cursor.fetchall()

                                                                        df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf","department"]) 
                                                                        df_employeesappsapprovedcheck["leavestartdate"] = pd.to_datetime(df_employeesappsapprovedcheck["leavestartdate"])
                                                                        df_employeesappsapprovedcheck["leaveenddate"] = pd.to_datetime(df_employeesappsapprovedcheck["leaveenddate"])
                        
                                                                        df_employeesappsapprovedcheck.dropna(subset=["leavestartdate", "leaveenddate"], inplace=True)
                                                                        # Create daily impact report
                                                                        impact_report = []

                                                                        for date in leave_dates:

                                                                            date = pd.Timestamp(date)

                                                                            print(type(date))  # Should be pandas._libs.tslibs.timestamps.Timestamp or datetime.datetime
                                                                            print(df_employeesappsapprovedcheck.dtypes)  # Check all datetime columns

                                                                            on_leave = ((df_employeesappsapprovedcheck["leavestartdate"] <= date) & (df_employeesappsapprovedcheck["leaveenddate"] >= date)).sum()
                                                                            remaining = numberindepartment - on_leave - 1  # subtract 1 for the new leave
                                                                            impact_report.append({
                                                                                "date": date,  # <=== Keep as datetime, don't convert to string
                                                                                "on leave": on_leave + 1,
                                                                                "employees remaining": remaining
                                                                            })

                                                                        # Convert to DataFrame for display
                                                                        impact_df = pd.DataFrame(impact_report)
                                                                        print("IMPAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACT")
                                                                        print(impact_df)
                                                                        print(numberindepartment)

                                                                        impact_df["date"] = pd.to_datetime(impact_df["date"], format="%Y-%m-%d")
                                                                        impact_df = impact_df[impact_df["date"].dt.weekday != 6].copy()

                                                                        change = (impact_df[["on leave", "employees remaining"]] != impact_df[["on leave", "employees remaining"]].shift()).any(axis=1)
                                                                        change.iloc[0] = True  # ensure the first row starts a group
                                                                        impact_df["group"] = change.cumsum()

                                                                        statements = []
                                                                        for _, group_df in impact_df.groupby("group"):
                                                                            start = group_df["date"].iloc[0].strftime("%d %B %Y")
                                                                            end = group_df["date"].iloc[-1].strftime("%d %B %Y")
                                                                            on_leave = group_df["on leave"].iloc[0]
                                                                            remaining = group_df["employees remaining"].iloc[0]
                                                                            
                                                                            if start == end:
                                                                                statements.append(f"On {start}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                                                                            else:
                                                                                statements.append(f"From {start} to {end}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                                                                                # Combine all statements into a single variable
                                                                        final_summary = "\n".join(statements)
                                                                        # Print output
                                                                        for s in statements:
                                                                            print(s)

                                                                        if leavetype == "Annual":

                                                                            leavedaysbalancebf = float(leavedaysbalance) - float(business_days)

                                                                        else:

                                                                            leavedaysbalancebf = float(leavedaysbalance)


                                                                        if leavedaysbalancebf >= 0:

                                                                            status = "Pending"

                                                                            insert_query = f"""
                                                                            INSERT INTO {table_name_apps_pending_approval} (id, firstname, surname, department, leavetype, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus)
                                                                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                                                            """
                                                                            cursor.execute(insert_query, (int(np.int64(id_user)), first_name, last_name, department, leavetype, leaveapprovername, int(np.int64(leaveapproverid)), leaveapproveremail, int(np.int64(leaveapproverwhatsapp)), float(np.float64(leavedaysbalance)), today_date, startdate, enddate, float(np.int64(business_days)), float(np.float64(leavedaysbalancebf)), status))
                                                                            connection.commit()

                                                                            query = f"SELECT appid FROM {table_name_apps_pending_approval};"
                                                                            cursor.execute(query)
                                                                            rows = cursor.fetchall()

                                                                            df_employees = pd.DataFrame(rows, columns=["id"])
                                                                            leaveappid = df_employees.iat[0,0]
                                                                            companyxx = company_reg.replace("_"," ").title()
                                                                            approovvver = leaveapprovername.title()

                                                                            buttons = [
                                                                                {"type": "reply", "reply": {"id": "Track", "title": "Track Application"}},
                                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                                                ]

                                                                            send_whatsapp_message(sender_id, f"✅ Great News {first_name} from {companyxx}! \n\n Your `{leavetype} Leave Application` for `{business_days} days` from `{startdate.strftime('%d %B %Y')}` to `{enddate.strftime('%d %B %Y')}` has been submitted successfully!\n\n"
                                                                                f"Your Leave Application ID is `{leaveappid}`.\n\n"
                                                                                f"A Notification has been sent to `{approovvver}`  on `+263{leaveapproverwhatsapp}` to decide on  your application.",
                                                                                buttons)
                                                                            
                                                                            if leaveapproverwhatsapp:
                                
                                                                                buttons = [
                                                                                    {"type": "reply", "reply": {"id": f"Approve5appwa_{leaveappid}", "title": "Approve"}},
                                                                                    {"type": "reply", "reply": {"id": f"Disapproveappwa_{leaveappid}", "title": "Disapprove"}},
                                                                                ]
                                                                                send_whatsapp_message(
                                                                                    f"263{leaveapproverwhatsapp}", 
                                                                                    f"Hey {approovvver}! 😊. New `{leavetype}` Leave Application from `{first_name} {surname}` for `{business_days} days` from `{startdate.strftime('%d %B %Y')}` to `{enddate.strftime('%d %B %Y')}`.\n\n" 
                                                                                    f"If you approve this leave application, {final_summary}\n\n"  
                                                                                    f"Select an option below to either approve or disapprove the application."         
                                                                                    , 
                                                                                    buttons
                                                                                )

                                                                        else:

                                                                            buttons = [
                                                                                {"type": "reply", "reply": {"id": f"Apply", "title": "Restart Application"}},
                                                                                {"type": "reply", "reply": {"id": f"Checkbal", "title": "Check Days Balance"}},
                                                                                {"type": "reply", "reply": {"id": f"Menu", "title": "Main Menu"}},
                                                                            ]

                                                                            send_whatsapp_message(sender_id, f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully!\n\n"
                                                                                f"You only have *{leavedaysbalance}* days available for leave but you are applying for *{business_days}*.\n\n You can restart your application and apply for leave such that the days between your leave start date and end date do not exceed your available balance of *{leavedaysbalance}* days.",
                                                                                buttons
                                                                                )


                                                                    else:
                                                                        print("leave app submission failed")

                                                        except ValueError as e:
                                                            send_whatsapp_message(
                                                                sender_id,
                                                                f"{e}, ❌ No, incorrect message format. Please use:\n"
                                                                "`end 24 january 2025`\n"
                                                                "Example: `end 15 march 2024`"
                                                            )

                                                    elif button_id == "Resubapp" :

                                                        table_name_apps_cancelled = f"{company_reg}appscancelled"
                                                        table_name_apps_pending_approval = f"{company_reg}appspendingapproval"

                                                        query = f"SELECT appid, id, firstname, surname, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail , leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, department FROM {table_name_apps_cancelled} WHERE id = %s;"
                                                        cursor.execute(query, (id_user,))
                                                        result = cursor.fetchall()

                                                        if result:

                                                            df_employees = pd.DataFrame(result)
                                                            df_employees = df_employees.sort_values(by=df_employees.columns[0], ascending=False)
                                                            print(df_employees)
                                                                    
                                                            try:

                                                                status = "Pending"
                                                                app_id = int(np.int64(df_employees.iat[0,0]))
                                                                employee_number = int(np.int64(df_employees.iat[0,1]))
                                                                first_name = df_employees.iat[0,2]
                                                                surname = df_employees.iat[0,3]
                                                                leave_type = df_employees.iat[0,4]
                                                                leave_specify = df_employees.iat[0,5]
                                                                approver_name = df_employees.iat[0,6]
                                                                approver_id =  int(np.int64(df_employees.iat[0,7]))
                                                                approver_email =  df_employees.iat[0,8]
                                                                approver_whatsapp =  int(np.int64(df_employees.iat[0,9]))
                                                                leave_days_balance =  float(np.float64(df_employees.iat[0,10]))
                                                                date_applied = df_employees.iat[0,11]
                                                                start_date = df_employees.iat[0,12]
                                                                end_date = df_employees.iat[0,13]
                                                                leave_days =  float(np.int64(df_employees.iat[0,14]))
                                                                leavedaysbalancebf =  float(np.float64(df_employees.iat[0,15]))
                                                                department = df_employees.iat[0,16]
                                                                insert_query = f"""
                                                                INSERT INTO {table_name_apps_pending_approval} 
                                                                (appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus)
                                                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                                                """

                                                                cursor.execute(insert_query, (
                                                                    app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, 
                                                                    approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                                                                    end_date, leave_days, leavedaysbalancebf, status
                                                                ))
                                                                
                                                                connection.commit()
                                                                print("Insert successful!")

                                                            except Exception as e:
                                                                print("Error inserting data:", e)

                                                            # SQL query to delete or mark the leave as canceled
                                                            query = f"""DELETE FROM {table_name_apps_cancelled} WHERE appid = %s"""
                                                            cursor.execute(query, (app_id,))
                                                            connection.commit()                                       

                                                            companyxx = company_reg.replace("_", " ").title()
                                                            sections = [
                                                                {
                                                                    "title": "Administrator Options",
                                                                    "rows": [
                                                                        {"id": "Apply", "title": "Apply for Leave"},
                                                                        {"id": "Track", "title": "Track My Application"},
                                                                        {"id": "Checkbal", "title": "Check Days Balance"},
                                                                        {"id": "myhist", "title": "My Applications History"},
                                                                        {"id": "Myinfo", "title": "My Info"},
                                                                        {"id": "Empmgt", "title": "Employee Management"},
                                                                        {"id": "Analyticscomp", "title": "Analytics & Insights"},
                                                                        {"id": "Company", "title": "Company Profile"}
                                                                    ]
                                                                }
                                                            ]

                                                            send_whatsapp_list_message(sender_id, f"Hey {first_name} from {companyxx}! \n\n Your `{leave_type} Leave Application` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}` has been Re-Submitted for approval successfully✅!",
                                                            "Administrator Options",
                                                            sections)                                          
                                                        
                                                        else:
                                                            print("No record found for the user.")


                                                    elif button_id == "Summarycomp":

                                                        try:

                                                            table_name = f"{company_reg}main"
                                                            table_name_apps_approved = f"{company_reg}appsapproved"
                                                            companyxx = company_reg.replace("_", " ").title()

                                                            query = f"""SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate,leavedaysbalancebf, department FROM {table_name_apps_approved}"""
                                                            cursor.execute(query)
                                                            rowsxxyy = cursor.fetchall()

                                                            df_apps_approved_lineg = pd.DataFrame(rowsxxyy, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate","leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf","department"])

                                                            startdate_lineg = datetime.today().date()
                                                            enddate_lineg = startdate_lineg + timedelta(days=30)


                                                            query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation, department FROM {table_name};"
                                                            cursor.execute(query)
                                                            rowsxxzz = cursor.fetchall()

                                                            df_employees_lineg = pd.DataFrame(rowsxxzz, columns=["id","firstname", "surname", "whatsapp","email", "address", "role","leaveapprovername","leaveapproverid","leaveapproveremail", "leaveapproverwhatsapp", "currentleavedaysbalance","monthlyaccumulation","department"])
                                                        
                                                            df_employees = df_employees_lineg
                                                            df_leaves = df_apps_approved_lineg

                                                            today = datetime.today().date()
                                                            next_30_days = today + timedelta(days=30)

                                                            # Total employees per department
                                                            total_by_dept = df_employees.groupby('department').size().to_dict()

                                                            # Ensure leave dates are datetime.date
                                                            df_leaves['leavestartdate'] = pd.to_datetime(df_leaves['leavestartdate']).dt.date
                                                            df_leaves['leaveenddate'] = pd.to_datetime(df_leaves['leaveenddate']).dt.date

                                                            result = {}
                                                            all_dates = pd.date_range(start=today, end=next_30_days).date

                                                            for dept, total_employees in total_by_dept.items():
                                                                result[dept] = []
                                                                df_dept_leaves = df_leaves[df_leaves['department'] == dept]

                                                                for date in all_dates:
                                                                    # Count employees on leave on this date
                                                                    on_leave = df_dept_leaves[
                                                                        (df_dept_leaves['leavestartdate'] <= date) & (df_dept_leaves['leaveenddate'] >= date)
                                                                    ].shape[0]

                                                                    remaining = ((total_employees - on_leave)/total_employees) * 100

                                                                    result[dept].append({
                                                                        "date": date.strftime("%Y-%m-%d"),
                                                                        "remaining": remaining
                                                                    })
                                                            print("CURRENT FIIIIIIIXXX")
                                                            print(result)

                                                            result2 = {}

                                                            for dept, total_employees in total_by_dept.items():
                                                                result2[dept] = []
                                                                df_dept_leaves = df_leaves[df_leaves['department'] == dept]

                                                                for date in all_dates:
                                                                    # Count employees on leave on this date
                                                                    on_leave = df_dept_leaves[
                                                                        (df_dept_leaves['leavestartdate'] <= date) & (df_dept_leaves['leaveenddate'] >= date)
                                                                    ].shape[0]

                                                                    remaining = total_employees - on_leave

                                                                    result2[dept].append({
                                                                        "date": date.strftime("%Y-%m-%d"),
                                                                        "remaining": remaining
                                                                    })
                                                            print("CURRENT FIIIIIIIXXX")
                                                            print(result2)
                                                        





                                                            def generate_graph_image_bytes(result_dict):
                                                                plt.figure(figsize=(12, 6))

                                                                for dept, values in result_dict.items():
                                                                    dates = [entry['date'] for entry in values]
                                                                    percentages = [entry['remaining'] for entry in values]
                                                                    plt.plot(dates, percentages, label=dept)

                                                                plt.xlabel("Date")
                                                                plt.ylabel("Remaining % Available Employees")
                                                                plt.title("Departmental Leave Trend")
                                                                plt.xticks(rotation=45)
                                                                plt.legend()
                                                                plt.tight_layout()

                                                                img_buffer = io.BytesIO()
                                                                plt.savefig(img_buffer, format='png')
                                                                img_buffer.seek(0)
                                                                return img_buffer

                                                            def generate_graph_image_bytes_bar(result_dict):
                                                                plt.figure(figsize=(12, 6))

                                                                # Extract all unique dates
                                                                all_dates = sorted({entry['date'] for values in result_dict.values() for entry in values})
                                                                x = np.arange(len(all_dates))  # the label locations

                                                                total_departments = len(result_dict)
                                                                width = 0.8 / total_departments  # total width shared among bars

                                                                for i, (dept, values) in enumerate(result_dict.items()):
                                                                    # Create a dict of date -> percentage for each dept (fill missing dates with 0)
                                                                    date_to_val = {entry['date']: entry['remaining'] for entry in values}
                                                                    percentages = [date_to_val.get(date, 0) for date in all_dates]
                                                                    plt.bar(x + i * width, percentages, width=width, label=dept)

                                                                plt.xlabel("Date")
                                                                plt.ylabel("Remaining % Available Employees")
                                                                plt.title("Departmental Leave Trend (Bar Chart)")
                                                                plt.xticks(x + width * (total_departments - 1) / 2, all_dates, rotation=45)
                                                                plt.legend()
                                                                plt.tight_layout()

                                                                img_buffer = io.BytesIO()
                                                                plt.savefig(img_buffer, format='png')
                                                                img_buffer.seek(0)
                                                                return img_buffer


                                                            def upload_image_to_whatsapp(img_buffer):
                                                                filename=f"{companyxx} insights {today_date}.png"
                                                            
                                                                url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/media"
                                                                headers = {
                                                                    "Authorization": f"Bearer {ACCESS_TOKEN}"
                                                                }

                                                                files = {
                                                                    "file": (filename, img_buffer.read(), "image/png"),
                                                                    "type": (None, "image/png"),
                                                                    "messaging_product": (None, "whatsapp")
                                                                }

                                                                response = requests.post(url, headers=headers, files=files)
                                                                print("📥 Full incoming data:", response.text)  # Good for debugging
                                                                response.raise_for_status()
                                                                return response.json()["id"]

                                                                                                            
                                                            def send_whatsapp_image_by_media_id(recipient_number, media_id):
                                                                filename=f"{companyxx} insights {today_date}.png"
                                                                url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                                headers = {
                                                                    "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                                    "Content-Type": "application/json"
                                                                }
                                                                payload = {
                                                                    "messaging_product": "whatsapp",
                                                                    "to": recipient_number,
                                                                    "type": "image",
                                                                    "image": {
                                                                        "id": media_id,
                                                                        "caption": "📊 % Employee Occupancy per Department +30-day"
                                                                    }
                                                                }

                                                                response = requests.post(url, headers=headers, json=payload)
                                                                response.raise_for_status()
                                                                return response.json()


                                                            def send_whatsapp_image_by_media_id2(recipient_number, media_id):
                                                                filename=f"{companyxx} insights {today_date}.png"
                                                                url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                                headers = {
                                                                    "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                                    "Content-Type": "application/json"
                                                                }
                                                                payload = {
                                                                    "messaging_product": "whatsapp",
                                                                    "to": recipient_number,
                                                                    "type": "image",
                                                                    "image": {
                                                                        "id": media_id,
                                                                        "caption": "📊 Number of Employees Remaining at work per Department +30-day"
                                                                    }
                                                                }

                                                                response = requests.post(url, headers=headers, json=payload)
                                                                response.raise_for_status()
                                                                return response.json()
                                                            






                                                            img_bytes = generate_graph_image_bytes_bar(result)
                                                            media_id = upload_image_to_whatsapp(img_bytes)
                                                            send_whatsapp_image_by_media_id(sender_id, media_id)

                                                            img_bytes2 = generate_graph_image_bytes(result2)
                                                            media_id2 = upload_image_to_whatsapp(img_bytes2)
                                                            send_whatsapp_image_by_media_id2(sender_id, media_id2)

                                                            avg_availability = {}
                                                            for dept, values in result.items():
                                                                percentages = [entry['remaining'] for entry in values]
                                                                avg_availability[dept] = sum(percentages) / len(percentages)

                                                            lowest_availability = {}
                                                            for dept, values in result.items():
                                                                lowest_entry = min(values, key=lambda x: x['remaining'])
                                                                lowest_availability[dept] = (lowest_entry['date'], lowest_entry['remaining'])

                                                            total_leave_days = {}
                                                            for dept, values in result.items():
                                                                total_leave_days[dept] = sum(100 - entry['remaining'] for entry in values)


                                                            msg = "📊 Departmental Leave Insights for Next 30 Days:\n\n"

                                                            msg += "Average Availability:\n"
                                                            for dept, avg in avg_availability.items():
                                                                msg += f"- {dept}: {avg:.1f}%\n"

                                                            msg += "\nLowest Availability Dates:\n"
                                                            for dept, (date, perc) in lowest_availability.items():
                                                                msg += f"- {dept}: {date} ({perc:.1f}%)\n"

                                                            msg += "\nTotal Planned Leave Days:\n"
                                                            for dept, days in total_leave_days.items():
                                                                msg += f"- {dept}: {days:.0f} days\n"

                                                            buttonsapproval = [
                                                                {"type": "reply", "reply": {"id": "Book", "title": "Extract Leave Book"}},
                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                            ]

                                                            send_whatsapp_message(
                                                                sender_id,
                                                                f" {msg} \n\n Select an option below to continue 👇.",
                                                                buttonsapproval
                                                            )




                                                        except Exception as e:

                                                            print(e)










                                                    elif button_id == "Cancelapp" :

                                                        table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                        table_name_apps_cancelled = f"{company_reg}appscancelled"

                                                        query = f"SELECT appid, id, firstname, surname, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail , leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, department FROM {table_name_apps_pending_approval} WHERE id = %s;"
                                                        cursor.execute(query, (id_user,))
                                                        result = cursor.fetchone()
                                                        if result:
                                                            (app_id, employee_number, first_name, surname, leave_type,  leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf, department) = result
                                                        
                                                            try:
                                                                    status = "Cancelled"
                                                                    statusdate = today_date
                                                                
                                                                    insert_query = f"""
                                                                    INSERT INTO {table_name_apps_cancelled} 
                                                                    (appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus, statusdate)
                                                                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                                                    """

                                                                    cursor.execute(insert_query, (
                                                                        app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, 
                                                                        approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                                                                        end_date, leave_days, leavedaysbalancebf, status, statusdate
                                                                    ))
                                                                    
                                                                    connection.commit()
                                                                    print("Insert successful!")

                                                            except Exception as e:
                                                                print("Error inserting data:", e)

                                                            # SQL query to delete or mark the leave as canceled
                                                            query = f"""DELETE FROM {table_name_apps_pending_approval} WHERE appid = %s"""
                                                            cursor.execute(query, (app_id,))
                                                            connection.commit()                                       

                                                            companyxx = company_reg.replace("_", " ").title()
                                                            buttons = [
                                                                {"type": "reply", "reply": {"id": f"Resubapp_{app_id}", "title": "ReSubmit Application"}},
                                                                {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                                                                {"type": "reply", "reply": {"id": "Main", "title": "Main Menu"}},
                                                            ]

                                                            send_whatsapp_message(sender_id, f"Hey {first_name} from {companyxx}! \n\n Your `{leave_type} Leave Application` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}` has been Cancelled successfully✅!\n\n"
                                                                "Select an option below to continue 👇",
                                                                buttons
                                                            )                                          
                                                        
                                                        else:
                                                            print("No record found for the user.")

                                                    elif button_id == "Addemp":

                                                        buttons = [
                                                        {"type": "reply", "reply": {"id": "Addone", "title": "Manual Addition"}},
                                                        {"type": "reply", "reply": {"id": "Bulkadd", "title": "Bulk Addition"}},
                                                        {"type": "reply", "reply": {"id": f"Menu", "title": "Main Menu"}},
                                                        ]

                                                        send_whatsapp_message(
                                                            sender_id, 
                                                            f"Hey {first_name}, select an option below on how you want to add employees to your company's Leave Management System.",
                                                            buttons
                                                        )

                                                    elif button_id == "Bulkadd":

                                                        buttons = [
                                                        {"type": "reply", "reply": {"id": "Uptemp", "title": "Upload Template"}},
                                                        {"type": "reply", "reply": {"id": "Downtemp", "title": "Download Template"}},
                                                        {"type": "reply", "reply": {"id": f"Menu", "title": "Main Menu"}},
                                                        ]

                                                        send_whatsapp_message(
                                                            sender_id, 
                                                            f"Hey {first_name}, select an option below on whether you want to upload an Excel temaplate already filled with employee details or you want to download the template to fill with Employee details.",
                                                            buttons
                                                        )

                                                    elif button_id == "Uptemp":

                                                        buttons = [
                                                        {"type": "reply", "reply": {"id": "Downtemp", "title": "Download Template"}},
                                                        {"type": "reply", "reply": {"id": f"Menu", "title": "Main Menu"}},
                                                        ]

                                                        send_whatsapp_message(
                                                            sender_id,
                                                            "📁 Please attach 📎 the filled Excel template file now by sending it directly here.",
                                                            buttons
                                                        )

                                                    elif button_id == "Downtemp":

                                                        companyxx = company_reg.replace("_", " ").title()

                                                        wb = openpyxl.Workbook()
                                                        ws = wb.active
                                                        ws.title = "Employee Details"

                                                        # Add headers
                                                        headers = [
                                                            "FirstName", "Surname", "WhatsApp", "Email", 
                                                            "Role", "Department", 
                                                            "Current Leave Days Balance", "Monthly Leave Days Accumulation"
                                                        ]
                                                        ws.append(headers)

                                                        # Style headers
                                                        dark_blue = "003366"
                                                        white = "FFFFFF"
                                                        for col in range(1, len(headers) + 1):
                                                            cell = ws.cell(row=1, column=col)
                                                            cell.fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
                                                            cell.font = Font(color=white, bold=True)

                                                        # Role dropdown (inline, short list)
                                                        role_dv = DataValidation(type="list", formula1='"Administrator,Ordinary User"', allow_blank=False)
                                                        ws.add_data_validation(role_dv)
                                                        for row in range(2, 500):
                                                            role_dv.add(ws[f"E{row}"])

                                                        # Department options (long list) - write to column Z
                                                        departments = [
                                                            "Human Resources and Administration",
                                                            "Finance and Accounting",
                                                            "Sales and Marketing",
                                                            "Sales and Distribution",
                                                            "Workshop and Maintenance",
                                                            "Operations and Production",
                                                            "Procurement and Purchasing",
                                                            "Customer Service and Support",
                                                            "IT and Digital Infrastructure",
                                                            "Risk Management",
                                                            "Legal and Compliance",
                                                            "Health, Safety and Environment",
                                                            "Research, Analytics and Reporting"
                                                        ]

                                                        for i, dept in enumerate(departments, start=1):
                                                            ws[f"R{i}"] = dept

                                                        # Department dropdown using cell range
                                                        dept_dv = DataValidation(type="list", formula1="=$R$1:$R$13", allow_blank=False)
                                                        ws.add_data_validation(dept_dv)
                                                        for row in range(2, 500):
                                                            dept_dv.add(ws[f"F{row}"])

                                                        # Hide the reference column
                                                        ws.column_dimensions['R'].hidden = True

                                                        # Save workbook to memory stream
                                                        output = io.BytesIO()
                                                        wb.save(output)
                                                        output.seek(0)

                                                        def send_whatsapp_excel_by_media_id(recipient_number, media_id, company_reg, reference_number=None, caption=None):
                                                            """Sends an Excel file via WhatsApp using the uploaded media ID"""                                                            
                                                            filename = f"LMS Employee Addition Template {companyxx}.xlsx"
                                                            
                                                            url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                            headers = {
                                                                "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                                "Content-Type": "application/json"
                                                            }
                                                            
                                                            payload = {
                                                                "messaging_product": "whatsapp",
                                                                "to": recipient_number,
                                                                "type": "document",
                                                                "document": {
                                                                    "id": media_id,
                                                                    "filename": filename
                                                                }
                                                            }
                                                            
                                                            if caption:
                                                                payload["document"]["caption"] = caption

                                                            response = requests.post(url, headers=headers, json=payload)
                                                            response.raise_for_status()
                                                            return response.json()

                                                        def upload_excel_to_whatsapp(excel_bytes, company_reg, reference_number=None):
                                                            compxxy = company_reg.replace("_"," ").title()
                                                            filename = f"LMS Employee Addition Template {companyxx}.xlsx"

                                                            url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/media"
                                                            headers = {
                                                                "Authorization": f"Bearer {ACCESS_TOKEN}"
                                                            }

                                                            files = {
                                                                "file": (filename, io.BytesIO(excel_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                                                                "type": (None, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                                                                "messaging_product": (None, "whatsapp")
                                                            }

                                                            response = requests.post(url, headers=headers, files=files)
                                                            print("📊 Excel upload response:", response.text)
                                                            response.raise_for_status()

                                                            return response.json()["id"]

                                                        try:
                                                            # Get the Excel bytes
                                                            excel_bytes = output.getvalue()

                                                            # Upload Excel to WhatsApp and get media ID
                                                            media_id = upload_excel_to_whatsapp(
                                                                excel_bytes=excel_bytes,
                                                                company_reg=table_name
                                                            )

                                                            # Send Excel to user
                                                            send_whatsapp_excel_by_media_id(
                                                                recipient_number=sender_id,
                                                                media_id=media_id,
                                                                company_reg=table_name,
                                                                caption=f"Employee Registration Template - {companyxx}"
                                                            )

                                                            # Confirmation message with button
                                                            buttons = [
                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                            ]
                                                            send_whatsapp_message(
                                                                sender_id,
                                                                f"Your *Employee Registration Template* for *{companyxx}* is attached 📎.\n\nYou may fill it in (*Kindly use Microsoft Excel, NOT Google Sheets*) and upload it when ready.",
                                                                buttons
                                                            )

                                                        except Exception as e:
                                                            print("Error sending employee template:", str(e))
                                                            send_whatsapp_message(
                                                                sender_id,
                                                                f"Sorry, we couldn't send the Employee Template right now.\nError: {e}"
                                                            )

                                                    elif selected_option == "Editemail":
                                                    
                                                        companyxx = company_reg.replace("_"," ").title()



                                                        sections = [
                                                            {
                                                                "title": "User Options",
                                                                "rows": [
                                                                    {"id": "Editname", "title": "Edit My Name"},
                                                                    {"id": "Editwhatsapp", "title": "Change My WhatsApp #"},
                                                                    {"id": "Editemail", "title": "Change My Email"},
                                                                    {"id": "Editwebpass", "title": "Change Web Password"},
                                                                    {"id": "Editaddress", "title": "Edit My Address"},
                                                                    {"id": "MyInfo", "title": "My Info"},
                                                                    {"id": "Menu", "title": "Main Menu"}
                                                                ]
                                                            }
                                                        ]

                                                        send_whatsapp_list_message(
                                                            sender_id, 
                                                            f"Hey {first_name}!\n Enter your new email address starting with the word `email` as shown below 👇. \n\n `email epsilon@gmail.com`", 
                                                        "User Options",
                                                        sections)

                                                    elif selected_option == "Empmgt":

                                                        sections = [
                                                            {
                                                            "title": "Administrator Options",
                                                            "rows": [
                                                                    {"id": "Addrememp", "title": "Add or Remove Employees"},
                                                                    {"id": "RoleApprover", "title": "Change Role or Approver"},
                                                                    {"id": "DepBalAcc", "title": "Edit Department or Days"},
                                                                    {"id": "Menu", "title": "Main Menu"}
                                                                ]
                                                            }
                                                        ]
                                                        companyxx = company_reg.replace("_"," ").title()


                                                        send_whatsapp_list_message(
                                                            sender_id, 
                                                            f"Hello {first_name} {last_name}, LMS Leave Applications Approver from {companyxx}!\n\n Select an Employee Management option to continue.", 
                                                            "Administrator Options",
                                                            sections
                                                        )

                                                    elif selected_option in ["Annual","Sick","Study","Parental", "Bereavement","Other"] :
                                                    
                                                        button_id_leave_type = str(selected_option)

                                                        cursor.execute("""
                                                            DELETE FROM whatsapptempapplication
                                                            WHERE empidwa = %s
                                                        """, (id_user,))  
                                                        
                                                        connection.commit()

                                                        cursor.execute(f"""
                                                            INSERT INTO whatsapptempapplication (empidwa, leavetypewa, companynamewa)
                                                            VALUES (%s, %s, %s)
                                                        """, (id_user, button_id_leave_type, company_reg))

                                                        connection.commit()

                                                        send_whatsapp_message(
                                                            sender_id, 
                                                            f"Ok. When would you like your {selected_option} Leave to start {first_name}?\n\n"
                                                            "Please enter your response using the format: 👇🏻\n"
                                                            "`start 24 january 2025`"
                                                        )

                                                        continue
                                                        
                                                    elif selected_option == "RoleApprover":

                                                        sections = [
                                                            {
                                                                "title": "Role & Approver Options",
                                                                "rows": [
                                                                    {"id": "Changerole", "title": "Edit Employee Role"},
                                                                    {"id": "Changeappr", "title": "Edit Employee Approver"},
                                                                    {"id": "RoleApprover", "title": "Role & Approver Schedule"},
                                                                    {"id": "Menu", "title": "Main Menu"}
                                                                ]
                                                            }
                                                        ]

                                                        send_whatsapp_list_message(
                                                            sender_id, 
                                                            f"Hey {first_name}, kindly select an option below.",
                                                            "Role & Approver Options",
                                                            sections
                                                        )
                                                        
                                                    elif selected_option == "Checkbal" or button_id == "Checkbal":

                                                        sections = [
                                                            {
                                                                "title": "Administrator Options",
                                                                "rows": [
                                                                    {"id": "Apply", "title": "Apply for Leave"},
                                                                    {"id": "Track", "title": "Track My Application"},
                                                                    {"id": "Checkbal", "title": "Check Days Balance"},
                                                                    {"id": "myhist", "title": "My Applications History"},
                                                                    {"id": "Myinfo", "title": "My Info"},
                                                                    {"id": "Empmgt", "title": "Employee Management"},
                                                                    {"id": "Analyticscomp", "title": "Analytics & Insights"},
                                                                    {"id": "Company", "title": "Company Profile"}
                                                                ]
                                                            }
                                                        ]

                                                        send_whatsapp_list_message(
                                                            sender_id, 
                                                            f"Hey {first_name}, your current available leave days balance is `{days_days_balance} days`.",
                                                            "Administrator Options",
                                                            sections
                                                        )

                                                    elif selected_option == "Addrememp":

                                                        buttons = [
                                                        {"type": "reply", "reply": {"id": "Addemp", "title": "Add Employees"}},
                                                        {"type": "reply", "reply": {"id": "Rememp", "title": "Remove Employees"}},
                                                        {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                        ]

                                                        send_whatsapp_message(
                                                            sender_id, 
                                                            f"Hey `{first_name}`, select an option below on how you want to add or remove employees to or from your company's Leave Management System.",
                                                            buttons
                                                        )
                                                                                    
                                                    elif selected_option == "Book" or button_id == "Book":

                                                        table_name = f"{company_reg}main"
                                                        appsapproved = f"{company_reg}appsapproved"
                                                        companyxx = company_reg.replace("_", " ").title()

                                                        query = f"SELECT id, firstname, surname, whatsapp, email, address ,role,currentleavedaysbalance, monthlyaccumulation, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp  FROM {table_name};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()

                                                        df_employees = pd.DataFrame(rows, columns=["ID","First Name", "Surname", "WhatsApp","Email", "Address", "Role","Leave Days Balance","Days Accumulated per Month","Leave Approver Name", "Leave Approver ID", "Leave Approver Email", "Leave Approver WhatsaApp"])
                                                        df_employees = df_employees.sort_values(by="ID", ascending=True)

                                                        query = f"SELECT appid, id, firstname, surname, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD-Month-YYYY') AS dateapplied,  TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate,   TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, leavedaysappliedfor,   TO_CHAR(statusdate, 'FMDD Month YYYY') AS statusdate, leavedaysbalancebf  FROM {appsapproved};"
                                                        cursor.execute(query)
                                                        rows2 = cursor.fetchall()

                                                        df_apps = pd.DataFrame(rows2, columns=["AppID","Emp ID", "First Name", "Surname", "Leave Type","Leave Approver Name", "Date Applied", "Leave Start Date", "Leave End Date","Leave Days Applied for","Date Approved","Leave Days Balance"])
                                                        df_apps = df_apps.sort_values(by="AppID", ascending=False)




                                                        print(df_employees)


                                                        df_apps['Leave Start Date'] = pd.to_datetime(df_apps['Leave Start Date'])
                                                        df_apps['Leave End Date'] = pd.to_datetime(df_apps['Leave End Date'])

                                                        # Function to expand dates and exclude Sundays
                                                        def expand_leave_days(row):
                                                            dates = pd.date_range(row['Leave Start Date'], row['Leave End Date'], freq='D')
                                                            # Exclude Sundays (weekday=6)
                                                            dates = [d for d in dates if d.weekday() != 6]
                                                            return dates

                                                        # Apply the function and explode the DataFrame
                                                        df_apps['Leave Dates'] = df_apps.apply(expand_leave_days, axis=1)
                                                        df_exploded = df_apps.explode('Leave Dates')

                                                        # Extract month and year for grouping
                                                        df_exploded['Month'] = df_exploded['Leave Dates'].dt.to_period('M')

                                                        # Group by Employee and Month
                                                        result = df_exploded.groupby(['Emp ID', 'First Name', 'Surname', 'Month']).size().reset_index(name='Leave Days Taken')

                                                        # Pivot to MoM format (months as columns)
                                                        mom_leave = result.pivot_table(
                                                            index=['Emp ID', 'First Name', 'Surname'],
                                                            columns='Month',
                                                            values='Leave Days Taken',
                                                            fill_value=0
                                                        ).reset_index()

                                                        # Rename columns for clarity
                                                        mom_leave.columns.name = None
                                                        mom_leave.columns = ['Emp ID', 'First Name', 'Surname'] + [f"{col.strftime('%b-%Y')}" for col in mom_leave.columns[3:]]

                                                        print(mom_leave)

                                                        def upload_excel_to_whatsapp(excel_bytes, company_reg, first_name, last_name, reference_number=None):
                                                            """Uploads an Excel file to WhatsApp servers and returns the media ID"""
                                                            compxxy = company_reg.replace("_"," ").title()
                                                            
                                                            filename = f"leave_records_{compxxy}.xlsx"
                                                            
                                                            url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/media"
                                                            headers = {
                                                                "Authorization": f"Bearer {ACCESS_TOKEN}"
                                                            }

                                                            files = {
                                                                "file": (filename, io.BytesIO(excel_bytes), 
                                                                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                                                                "type": (None, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                                                                "messaging_product": (None, "whatsapp")
                                                            }

                                                            response = requests.post(url, headers=headers, files=files)
                                                            print("📊 Excel upload response:", response.text)  # Debugging
                                                            response.raise_for_status()
                                                            return response.json()["id"]

                                                        def send_whatsapp_excel_by_media_id(recipient_number, media_id, company_reg, first_name, last_name, reference_number=None, caption=None):
                                                            """Sends an Excel file via WhatsApp using the uploaded media ID"""
                                                            compxxy = company_reg.replace("_"," ").title()
                                                            
                                                            filename = f"leave_records_{compxxy}.xlsx"
                                                            
                                                            url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                            headers = {
                                                                "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                                "Content-Type": "application/json"
                                                            }
                                                            
                                                            payload = {
                                                                "messaging_product": "whatsapp",
                                                                "to": recipient_number,
                                                                "type": "document",
                                                                "document": {
                                                                    "id": media_id,
                                                                    "filename": filename
                                                                }
                                                            }
                                                            
                                                            if caption:
                                                                payload["document"]["caption"] = caption

                                                            response = requests.post(url, headers=headers, json=payload)
                                                            response.raise_for_status()
                                                            return response.json()

                                                        output = io.BytesIO()
                                                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                                            df_employees.to_excel(writer, index=False, sheet_name=f'LMS Book {today_date}')
                                                            df_apps.to_excel(writer, index=False, sheet_name=f'All Approved')
                                                            mom_leave.to_excel(writer, index=False, sheet_name=f'Month on Month')

                                                        output.seek(0)
                                                        excel_bytes = output.getvalue()
                                                        
                                                        try:
                                                            media_id = upload_excel_to_whatsapp(
                                                                excel_bytes=excel_bytes,
                                                                company_reg=company_reg,
                                                                first_name=first_name,
                                                                last_name=last_name
                                                            )
                                                            
                                                            send_whatsapp_excel_by_media_id(
                                                                recipient_number=sender_id,
                                                                media_id=media_id,
                                                                company_reg=company_reg,
                                                                first_name=first_name,
                                                                last_name=last_name,
                                                                caption=f"{companyxx} Employee Leave Records as of {today_date}"
                                                            )
                                                            
                                                            buttons = [
                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                            ]
                                                            send_whatsapp_message(
                                                                sender_id, 
                                                                f"Hey there {first_name} {last_name}! You may go ahead and download Leave Book for {companyxx} attached here 😎.", 
                                                                buttons
                                                            )

                                                        except Exception as e:
                                                            print(f"Error sending Excel file: {str(e)}")
                                                            send_whatsapp_message(
                                                                sender_id,
                                                                f"Sorry {first_name}, we encountered an error preparing your document -- {e}. Please try again later."
                                                            )
                                                        
                                                elif message.get("type") == "text":
                                                    text = message.get("text", {}).get("body", "").lower()
                                                    print(f"📨 Message from {sender_id}: {text}")
                                                    
                                                    if "hello" in text.lower():
                                                        companyxx = company_reg.replace("_"," ").title()
                                                        
                                                        sections = [
                                                            {
                                                                "title": "Administrator Options",
                                                                "rows": [
                                                                    {"id": "Apply", "title": "Apply for Leave"},
                                                                    {"id": "Track", "title": "Track My Application"},
                                                                    {"id": "Checkbal", "title": "Check Days Balance"},
                                                                    {"id": "myhist", "title": "My Applications History"},
                                                                    {"id": "Myinfo", "title": "My Info"},
                                                                    {"id": "Empmgt", "title": "Employee Management"},
                                                                    {"id": "Analyticscomp", "title": "Analytics & Insights"},
                                                                    {"id": "Company", "title": "Company Profile"}
                                                                ]
                                                            }
                                                        ]
                                                        
                                                        send_whatsapp_list_message(
                                                            sender_id,
                                                            f"Hello {first_name} {last_name}, LMS Administrator from {companyxx}!\n\n {bot} LMS Bot Here 😎. How can I assist you?",
                                                            "Administrator Options",
                                                            sections
                                                        )

                                                    elif "email" in text.lower():

                                                        table_name = company_reg + "main"
                                                        table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                        table_name_apps_cancelled = f"{company_reg}appscancelled"
                                                        table_name_apps_approved = f"{company_reg}appsapproved"
                                                        table_name_apps_declined = f"{company_reg}appsdeclined"
                                                        table_name_apps_revoked = f"{company_reg}appsrevoked"

                                                        # Regex to extract email after the word "email"
                                                        match = re.search(r"email\s+([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)", text.strip(), re.IGNORECASE)

                                                        if not match:
                                                            raise ValueError("Invalid Email format")

                                                            sections = [
                                                                {
                                                                    "title": "User Options",
                                                                    "rows": [
                                                                        {"id": "Editname", "title": "Edit My Name"},
                                                                        {"id": "Editwhatsapp", "title": "Change My WhatsApp #"},
                                                                        {"id": "Editemail", "title": "Change My Email"},
                                                                        {"id": "Editwebpass", "title": "Change Web Password"},
                                                                        {"id": "Editaddress", "title": "Edit My Address"},
                                                                        {"id": "MyInfo", "title": "My Info"},
                                                                        {"id": "Menu", "title": "Main Menu"}
                                                                    ]
                                                                }
                                                            ]

                                                            send_whatsapp_list_message(
                                                                sender_id, 
                                                                f"Hey {first_name}!\n You have provided an invalid Email Address. Enter your new email address starting with the word `email` as shown below 👇. \n\n `email epsilon@gmail.com`", 
                                                            "User Options",
                                                            sections)                


                                                        email = match.group(1)
                                                        print("Extracted email:", email)

                                                        try:

                                                            query = f"UPDATE {table_name} SET email = %s WHERE id = %s;"
                                                            cursor.execute(query, (email, id_user))
                                                            connection.commit()

                                                            query = f"UPDATE {table_name_apps_pending_approval} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                            cursor.execute(query, (email, id_user))
                                                            connection.commit()
                                                
                                                            query = f"UPDATE {table_name_apps_cancelled} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                            cursor.execute(query, (email, id_user))
                                                            connection.commit()

                                                            query = f"UPDATE {table_name_apps_approved} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                            cursor.execute(query, (email, id_user))
                                                            connection.commit()

                                                            query = f"UPDATE {table_name_apps_declined} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                            cursor.execute(query, (email, id_user))
                                                            connection.commit()

                                                            query = f"UPDATE {table_name_apps_revoked} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                            cursor.execute(query, (email, id_user))
                                                            connection.commit()

                                                            sections = [
                                                                {
                                                                    "title": "User Options",
                                                                    "rows": [
                                                                        {"id": "Editname", "title": "Edit My Name"},
                                                                        {"id": "Editwhatsapp", "title": "Change My WhatsApp #"},
                                                                        {"id": "Editemail", "title": "Change My Email"},
                                                                        {"id": "Editwebpass", "title": "Change Web Password"},
                                                                        {"id": "Editaddress", "title": "Edit My Address"},
                                                                        {"id": "MyInfo", "title": "My Info"},
                                                                        {"id": "Menu", "title": "Main Menu"}
                                                                    ]
                                                                }
                                                            ]

                                                            send_whatsapp_list_message(
                                                                sender_id, 
                                                                f"Hey {first_name}!\n Email Address Successfully Changed to `{email}`. Select an option below to proceed 👇.", 
                                                            "User Options",
                                                            sections)  

                                                        except Exception as e:
                                                            print(e)

                                                    elif "start" in text.lower():

                                                        try:
                                                            # Match: "start 20 july 2025"
                                                            match = re.match(r"start\s+(\d{1,2}\s+[a-zA-Z]+\s+\d{4})", text.strip(), re.IGNORECASE)
                                                            if not match:
                                                                raise ValueError("Invalid format")

                                                            date_part = match.group(1)
                                                            parsed_date = datetime.strptime(date_part, "%d %B %Y")  # Will raise ValueError if invalid

                                                            # ✅ Now it's safe to update the DB
                                                            cursor.execute("""
                                                                UPDATE whatsapptempapplication
                                                                SET startdate = %s
                                                                WHERE empidwa = %s
                                                            """, (date_part, id_user))
                                                            connection.commit()

                                                            cursor.execute("""
                                                                SELECT empidwa, leavetypewa FROM whatsapptempapplication
                                                                WHERE empidwa = %s
                                                            """, (id_user,))
                                                            result = cursor.fetchone()
                                                            leavetypewa = result[1] if result else "your"

                                                            send_whatsapp_message(sender_id,
                                                                f"✅ Got it! Start date saved.\n\nNow enter your last day on {leavetypewa} leave like this:\n"
                                                                "`end 28 July 2025`"
                                                            )

                                                        except ValueError:
                                                            send_whatsapp_message(
                                                                sender_id,
                                                                f"❌ Invalid start date message format, {first_name}. Please use the date format givem below 👇:\n"
                                                                "`start 24 january 2025`\n\n"
                                                                "Example: `start 15 march 2024`"
                                                            )

                                                        except Exception as e:
                                                            import traceback
                                                            print("🔴 Unexpected error:", e)
                                                            traceback.print_exc()

                                                            try:
                                                                send_whatsapp_message(
                                                                    sender_id,
                                                                    "⚠️ Something went wrong while processing your start date. Please try again or contact support."
                                                                )
                                                            except Exception as send_err:
                                                                print("🔴 Failed to send WhatsApp error message:", send_err)

                                                    elif "end" in text.lower():

                                                        try:
                                                            # ✅ Match "end 24 january 2025"
                                                            match = re.match(r"end\s+(\d{1,2}\s+[a-zA-Z]+\s+\d{4})", text.strip(), re.IGNORECASE)
                                                            if not match:
                                                                raise ValueError("Invalid end date format.")

                                                            date_part = match.group(1)
                                                            parsed_end_date = datetime.strptime(date_part, "%d %B %Y").date()  # Will raise ValueError if invalid

                                                            # ✅ Update DB now that it's valid
                                                            cursor.execute("""
                                                                UPDATE whatsapptempapplication
                                                                SET enddate = %s
                                                                WHERE empidwa = %s
                                                            """, (date_part, id_user))
                                                            connection.commit()

                                                            # ✅ Fetch full leave application
                                                            cursor.execute("""
                                                                SELECT id, empidwa, leavetypewa, startdate, enddate FROM whatsapptempapplication
                                                                WHERE empidwa = %s
                                                            """, (id_user,))
                                                            result = cursor.fetchone()

                                                            if not result:
                                                                raise Exception("No leave record found.")

                                                            appid = result[0]
                                                            leavetype = result[2]
                                                            startdate = result[3]
                                                            enddate = result[4]

                                                            # ✅ Ensure both dates are datetime.date objects
                                                            if isinstance(startdate, str):
                                                                startdate = datetime.strptime(startdate, "%Y-%m-%d").date()
                                                            if isinstance(enddate, str):
                                                                enddate = datetime.strptime(enddate, "%Y-%m-%d").date()

                                                            # ✅ Calculate business days
                                                            business_days = 0
                                                            current_date = startdate
                                                            while current_date <= enddate:
                                                                if current_date.weekday() != 6:  # Weekday: Mon-Fri
                                                                    business_days += 1
                                                                current_date += timedelta(days=1)

                                                            # ✅ Ask user to confirm submission
                                                            buttons = [
                                                                {"type": "reply", "reply": {"id": "Submitapp", "title": "Yes, Submit"}},
                                                                {"type": "reply", "reply": {"id": "Dontsubmit", "title": "No"}}
                                                            ]
                                                            send_whatsapp_message(
                                                                sender_id,
                                                                f"📝 Do you wish to submit your `{business_days}-day {leavetype} Leave Application` from "
                                                                f"`{startdate.strftime('%d %B %Y')}` to `{enddate.strftime('%d %B %Y')}`, {first_name}?",
                                                                buttons
                                                            )

                                                        except ValueError:
                                                            send_whatsapp_message(
                                                                sender_id,
                                                                f"❌ Invalid end date message format, {first_name}. Please use the date format givem below 👇:\n"
                                                                "`end 24 january 2025`\n\n"
                                                                "Example: `end 28 march 2024`"
                                                            )

                                                        except Exception as e:
                                                            import traceback
                                                            print("🔴 ERROR during end date processing:", e)
                                                            traceback.print_exc()
                                                            try:
                                                                send_whatsapp_message(
                                                                    sender_id,
                                                                    "⚠️ Something went wrong while processing your end date. Please try again or contact support."
                                                                )
                                                            except Exception as send_err:
                                                                print("🔴 Failed to send error message via WhatsApp:", send_err)
                                                                
                                                    else:
                                                        send_whatsapp_message(
                                                            sender_id, 
                                                            f"{bot} LMS Bot Here 😎. Say 'hello' to start!"
                                                        )

                                                elif message.get("type") == "document":
                                                    mime_type = message["document"]["mime_type"]
                                                    filename = message["document"]["filename"]
                                                    file_id = message["document"]["id"]

                                                    def download_whatsapp_media(media_id):
                                                        media_url = f"https://graph.facebook.com/v19.0/{media_id}"
                                                        headers = {
                                                            "Authorization": f"Bearer {ACCESS_TOKEN}"
                                                        }

                                                        # Get the actual download URL
                                                        res = requests.get(media_url, headers=headers)
                                                        res.raise_for_status()
                                                        download_url = res.json()["url"]

                                                        # Download file content
                                                        file_response = requests.get(download_url, headers=headers)
                                                        file_response.raise_for_status()
                                                        return file_response.content


                                                    if mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or filename.endswith(".xlsx"):
                                                        try:
                                                            file_bytes = download_whatsapp_media(file_id)

                                                            excel_file = io.BytesIO(file_bytes)

                                                            df = pd.read_excel(excel_file)

                                                            print("yoooooooooooooooooooh upload!!!")
                                                            print(df)

                                                            send_whatsapp_message(sender_id, f"✅ Excel received. It contains {len(df)} employees to add.")

                                                        except Exception as e:
                                                            send_whatsapp_message(sender_id, f"❌ Error reading Excel file: {str(e)}")
                                                    else:
                                                        send_whatsapp_message(sender_id, "⚠️ Unsupported file type. Please upload a valid `.xlsx` file.")

                                            except Exception as e:
                                                print("error here")
                                                print(e)

                                        elif len(df_employeesempapp) > 0:

                                            if message.get("type") == "interactive":
                                                interactive = message.get("interactive", {})


                                                if interactive.get("type") == "button_reply":

                                                    button_id = interactive.get("button_reply", {}).get("id")
                                                    print(f"🔘 Button clicked: {button_id}")
                                                    selected_option = ""
                                            
                                                elif interactive.get("type") == "list_reply":
                                                    selected_option = interactive.get("list_reply", {}).get("id")
                                                    print(f"📋 User selected: {selected_option}")
                                                    button_id = ""

                                                if selected_option == "Empmgt":

                                                    sections = [
                                                        {
                                                        "title": "Administrator Options",
                                                        "rows": [
                                                                {"id": "Addrememp", "title": "Add or Remove Employees"},
                                                                {"id": "RoleApprover", "title": "Change Role or Approver"},
                                                                {"id": "DepBalAcc", "title": "Edit Department or Days"},
                                                                {"id": "Menu", "title": "Main Menu"}
                                                            ]
                                                        }
                                                    ]
                                                    companyxx = company_reg.replace("_"," ").title()


                                                    send_whatsapp_list_message(
                                                        sender_id, 
                                                        f"Hello {first_name} {last_name}, LMS Leave Applications Approver from {companyxx}!\n\n Select an Employee Management option to continue.", 
                                                        "Administrator Options",
                                                        sections
                                                    )


                                                elif button_id == "Summarycomp":

                                                    try:

                                                        table_name = f"{company_reg}main"
                                                        table_name_apps_approved = f"{company_reg}appsapproved"
                                                        companyxx = company_reg.replace("_", " ").title()

                                                        query = f"""SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate,leavedaysbalancebf, department FROM {table_name_apps_approved}"""
                                                        cursor.execute(query)
                                                        rowsxxyy = cursor.fetchall()

                                                        df_apps_approved_lineg = pd.DataFrame(rowsxxyy, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate","leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf","department"])

                                                        startdate_lineg = datetime.today().date()
                                                        enddate_lineg = startdate_lineg + timedelta(days=30)


                                                        query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation, department FROM {table_name};"
                                                        cursor.execute(query)
                                                        rowsxxzz = cursor.fetchall()

                                                        df_employees_lineg = pd.DataFrame(rowsxxzz, columns=["id","firstname", "surname", "whatsapp","email", "address", "role","leaveapprovername","leaveapproverid","leaveapproveremail", "leaveapproverwhatsapp", "currentleavedaysbalance","monthlyaccumulation","department"])
                                                    
                                                        df_employees = df_employees_lineg
                                                        df_leaves = df_apps_approved_lineg

                                                        today = datetime.today().date()
                                                        next_30_days = today + timedelta(days=30)

                                                        # Total employees per department
                                                        total_by_dept = df_employees.groupby('department').size().to_dict()

                                                        # Ensure leave dates are datetime.date
                                                        df_leaves['leavestartdate'] = pd.to_datetime(df_leaves['leavestartdate']).dt.date
                                                        df_leaves['leaveenddate'] = pd.to_datetime(df_leaves['leaveenddate']).dt.date

                                                        result = {}
                                                        all_dates = pd.date_range(start=today, end=next_30_days).date

                                                        for dept, total_employees in total_by_dept.items():
                                                            result[dept] = []
                                                            df_dept_leaves = df_leaves[df_leaves['department'] == dept]

                                                            for date in all_dates:
                                                                # Count employees on leave on this date
                                                                on_leave = df_dept_leaves[
                                                                    (df_dept_leaves['leavestartdate'] <= date) & (df_dept_leaves['leaveenddate'] >= date)
                                                                ].shape[0]

                                                                remaining = ((total_employees - on_leave)/total_employees) * 100

                                                                result[dept].append({
                                                                    "date": date.strftime("%Y-%m-%d"),
                                                                    "remaining": remaining
                                                                })
                                                        print("CURRENT FIIIIIIIXXX")
                                                        print(result)

                                                        result2 = {}

                                                        for dept, total_employees in total_by_dept.items():
                                                            result2[dept] = []
                                                            df_dept_leaves = df_leaves[df_leaves['department'] == dept]

                                                            for date in all_dates:
                                                                # Count employees on leave on this date
                                                                on_leave = df_dept_leaves[
                                                                    (df_dept_leaves['leavestartdate'] <= date) & (df_dept_leaves['leaveenddate'] >= date)
                                                                ].shape[0]

                                                                remaining = total_employees - on_leave

                                                                result2[dept].append({
                                                                    "date": date.strftime("%Y-%m-%d"),
                                                                    "remaining": remaining
                                                                })
                                                        print("CURRENT FIIIIIIIXXX")
                                                        print(result2)
                                                    





                                                        def generate_graph_image_bytes(result_dict):
                                                            plt.figure(figsize=(12, 6))

                                                            for dept, values in result_dict.items():
                                                                dates = [entry['date'] for entry in values]
                                                                percentages = [entry['remaining'] for entry in values]
                                                                plt.plot(dates, percentages, label=dept)

                                                            plt.xlabel("Date")
                                                            plt.ylabel("Remaining % Available Employees")
                                                            plt.title("Departmental Leave Trend")
                                                            plt.xticks(rotation=45)
                                                            plt.legend()
                                                            plt.tight_layout()

                                                            img_buffer = io.BytesIO()
                                                            plt.savefig(img_buffer, format='png')
                                                            img_buffer.seek(0)
                                                            return img_buffer

                                                        def generate_graph_image_bytes_bar(result_dict):
                                                            plt.figure(figsize=(12, 6))

                                                            # Extract all unique dates
                                                            all_dates = sorted({entry['date'] for values in result_dict.values() for entry in values})
                                                            x = np.arange(len(all_dates))  # the label locations

                                                            total_departments = len(result_dict)
                                                            width = 0.8 / total_departments  # total width shared among bars

                                                            for i, (dept, values) in enumerate(result_dict.items()):
                                                                # Create a dict of date -> percentage for each dept (fill missing dates with 0)
                                                                date_to_val = {entry['date']: entry['remaining'] for entry in values}
                                                                percentages = [date_to_val.get(date, 0) for date in all_dates]
                                                                plt.bar(x + i * width, percentages, width=width, label=dept)

                                                            plt.xlabel("Date")
                                                            plt.ylabel("Remaining % Available Employees")
                                                            plt.title("Departmental Leave Trend (Bar Chart)")
                                                            plt.xticks(x + width * (total_departments - 1) / 2, all_dates, rotation=45)
                                                            plt.legend()
                                                            plt.tight_layout()

                                                            img_buffer = io.BytesIO()
                                                            plt.savefig(img_buffer, format='png')
                                                            img_buffer.seek(0)
                                                            return img_buffer


                                                        def upload_image_to_whatsapp(img_buffer):
                                                            filename=f"{companyxx} insights {today_date}.png"
                                                        
                                                            url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/media"
                                                            headers = {
                                                                "Authorization": f"Bearer {ACCESS_TOKEN}"
                                                            }

                                                            files = {
                                                                "file": (filename, img_buffer.read(), "image/png"),
                                                                "type": (None, "image/png"),
                                                                "messaging_product": (None, "whatsapp")
                                                            }

                                                            response = requests.post(url, headers=headers, files=files)
                                                            print("📥 Full incoming data:", response.text)  # Good for debugging
                                                            response.raise_for_status()
                                                            return response.json()["id"]

                                                                                                        
                                                        def send_whatsapp_image_by_media_id(recipient_number, media_id):
                                                            filename=f"{companyxx} insights {today_date}.png"
                                                            url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                            headers = {
                                                                "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                                "Content-Type": "application/json"
                                                            }
                                                            payload = {
                                                                "messaging_product": "whatsapp",
                                                                "to": recipient_number,
                                                                "type": "image",
                                                                "image": {
                                                                    "id": media_id,
                                                                    "caption": "📊 % Employee Occupancy per Department +30-day"
                                                                }
                                                            }

                                                            response = requests.post(url, headers=headers, json=payload)
                                                            response.raise_for_status()
                                                            return response.json()


                                                        def send_whatsapp_image_by_media_id2(recipient_number, media_id):
                                                            filename=f"{companyxx} insights {today_date}.png"
                                                            url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                            headers = {
                                                                "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                                "Content-Type": "application/json"
                                                            }
                                                            payload = {
                                                                "messaging_product": "whatsapp",
                                                                "to": recipient_number,
                                                                "type": "image",
                                                                "image": {
                                                                    "id": media_id,
                                                                    "caption": "📊 Number of Employees Remaining at work per Department +30-day"
                                                                }
                                                            }

                                                            response = requests.post(url, headers=headers, json=payload)
                                                            response.raise_for_status()
                                                            return response.json()
                                                        






                                                        img_bytes = generate_graph_image_bytes_bar(result)
                                                        media_id = upload_image_to_whatsapp(img_bytes)
                                                        send_whatsapp_image_by_media_id(sender_id, media_id)

                                                        img_bytes2 = generate_graph_image_bytes(result2)
                                                        media_id2 = upload_image_to_whatsapp(img_bytes2)
                                                        send_whatsapp_image_by_media_id2(sender_id, media_id2)

                                                        avg_availability = {}
                                                        for dept, values in result.items():
                                                            percentages = [entry['remaining'] for entry in values]
                                                            avg_availability[dept] = sum(percentages) / len(percentages)

                                                        lowest_availability = {}
                                                        for dept, values in result.items():
                                                            lowest_entry = min(values, key=lambda x: x['remaining'])
                                                            lowest_availability[dept] = (lowest_entry['date'], lowest_entry['remaining'])

                                                        total_leave_days = {}
                                                        for dept, values in result.items():
                                                            total_leave_days[dept] = sum(100 - entry['remaining'] for entry in values)


                                                        msg = "📊 Departmental Leave Insights for Next 30 Days:\n\n"

                                                        msg += "Average Availability:\n"
                                                        for dept, avg in avg_availability.items():
                                                            msg += f"- {dept}: {avg:.1f}%\n"

                                                        msg += "\nLowest Availability Dates:\n"
                                                        for dept, (date, perc) in lowest_availability.items():
                                                            msg += f"- {dept}: {date} ({perc:.1f}%)\n"

                                                        msg += "\nTotal Planned Leave Days:\n"
                                                        for dept, days in total_leave_days.items():
                                                            msg += f"- {dept}: {days:.0f} days\n"

                                                        buttonsapproval = [
                                                            {"type": "reply", "reply": {"id": "Book", "title": "Extract Leave Book"}},
                                                            {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                        ]

                                                        send_whatsapp_message(
                                                            sender_id,
                                                            f" {msg} \n\n Select an option below to continue 👇.",
                                                            buttonsapproval
                                                        )




                                                    except Exception as e:

                                                        print(e)

                                                    
                                                elif button_id == "Track" or selected_option == "Track":

                                                    table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                    table_name_apps_approved = f"{company_reg}appsapproved"
                                                    table_name_apps_declined = f"{company_reg}appsdeclined"
                                                    table_name_apps_cancelled = f"{company_reg}appscancelled"


                                                    query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leaveapproverwhatsapp, appid, department  FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                    cursor.execute(query)
                                                    rows = cursor.fetchall()

                                                    df_employeesappspendingcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor", "leaveapproverwhatsapp", "appid", "department"])    

                                                    if len(df_employeesappspendingcheck) == 0:

                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate, leavedaysbalancebf, department  FROM {table_name_apps_approved} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()
                                                        df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf", "department"]) 

                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate, leavedaysbalancebf, department  FROM {table_name_apps_declined} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()
                                                        df_employeesappsdeclinedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf", "department"])  
                                
                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate, leavedaysbalancebf, department  FROM {table_name_apps_cancelled} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()
                                                        df_employeesappscancelledcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf", "department"])
                                
                                                        all_approved_declined = df_employeesappsapprovedcheck._append(df_employeesappsdeclinedcheck)
                                                        all_approved_declined_cancelled = all_approved_declined._append(df_employeesappscancelledcheck)
                                                        all_approved_declined_cancelled = all_approved_declined_cancelled.sort_values(by="appid", ascending=False)  


                                                        if len(all_approved_declined_cancelled) > 0:

                                                            print(f" hhhhhhhhhhhhhhhhhhhh  {all_approved_declined_cancelled.iat[0,8] }")

                                                            if all_approved_declined_cancelled.iat[0,8] == "Approved":

                                                                buttons = [
                                                                    {"type": "reply", "reply": {"id": "Revoke", "title": "Revoke Application"}},
                                                                    {"type": "reply", "reply": {"id": "myhist", "title": "Download My History"}},
                                                                    {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                                ]
                                                                send_whatsapp_message(
                                                                    sender_id, 
                                                                    f"Hey {first_name}, your recent `{all_approved_declined_cancelled.iat[0,2]}` Leave Application `[ID - {all_approved_declined_cancelled.iat[0,0]}]` that you applied for on `{all_approved_declined_cancelled.iat[0,4].strftime('%d %B %Y')}` for `{all_approved_declined_cancelled.iat[0,7]} days` from `{all_approved_declined_cancelled.iat[0,5].strftime('%d %B %Y')}` to `{all_approved_declined_cancelled.iat[0,6].strftime('%d %B %Y')}` was {all_approved_declined_cancelled.iat[0,8]}✅ by `{all_approved_declined_cancelled.iat[0,3].title()}` on `{all_approved_declined_cancelled.iat[0,9].strftime('%d %B %Y')}`." 
                                                                )


                                                                def generate_leave_pdf():
                                                                    app = {
                                                                        'company_logo': 44,
                                                                        'company_name': company_reg.replace("_"," ").title(),
                                                                        'employee_name': f"{first_name} {last_name}",
                                                                        'leave_type': all_approved_declined_cancelled.iat[0,2],
                                                                        'generated_on': today_date,
                                                                        'date_applied': all_approved_declined_cancelled.iat[0,4].strftime('%d %B %Y'),
                                                                        'approver_name': all_approved_declined_cancelled.iat[0,3].title(),
                                                                        'reference_number': all_approved_declined_cancelled.iat[0,0],
                                                                        'approved_date': all_approved_declined_cancelled.iat[0,9].strftime('%d %B %Y'),
                                                                        'new_balance': all_approved_declined_cancelled.iat[0,10],
                                                                        'start_date':  all_approved_declined_cancelled.iat[0,5].strftime('%d %B %Y'),
                                                                        'end_date':  all_approved_declined_cancelled.iat[0,6].strftime('%d %B %Y'),
                                                                        'days_requested':  all_approved_declined_cancelled.iat[0,7], 
                                                                        'department':  all_approved_declined_cancelled.iat[0,11], 
                                                                        'address': address_foc_8, 
                                                                        'whatsapp': whatsapp_foc_8, 
                                                                        'email': email_foc_8, 
                                                                        'status': 'Approved',
                                                                        'power': power,
                                                                    }

                                                                    html_out = render_template("leave_pdf_template.html", app=app)
                                                                    
                                                                    # ✅ Return as bytes instead of saving to file
                                                                    pdf_bytes = HTML(string=html_out).write_pdf()
                                                                    return pdf_bytes

                                                                

                                                                def upload_pdf_to_whatsapp(pdf_bytes):
                                                                    compxxy = company_reg.replace("_"," ").title()
                                                                    filename=f"leave_application_{all_approved_declined_cancelled.iat[0,0]}_{first_name}_{last_name}_{compxxy}.pdf"
                                                                
                                                                    url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/media"
                                                                    headers = {
                                                                        "Authorization": f"Bearer {ACCESS_TOKEN}"
                                                                    }

                                                                    files = {
                                                                        "file": (filename, io.BytesIO(pdf_bytes), "application/pdf"),
                                                                        "type": (None, "application/pdf"),
                                                                        "messaging_product": (None, "whatsapp")
                                                                    }

                                                                    response = requests.post(url, headers=headers, files=files)
                                                                    print("📥 Full incoming data:", response.text)  # Good for debugging
                                                                    response.raise_for_status()
                                                                    return response.json()["id"]

                                                                                                                
                                                                def send_whatsapp_pdf_by_media_id(recipient_number, media_id):
                                                                    compxxy = company_reg.replace("_"," ").title()
                                                                    filename=f"leave_application_{all_approved_declined_cancelled.iat[0,0]}_{first_name}_{last_name}_{compxxy}.pdf"
                                                                    url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                                    headers = {
                                                                        "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                                        "Content-Type": "application/json"
                                                                    }
                                                                    payload = {
                                                                        "messaging_product": "whatsapp",
                                                                        "to": recipient_number,
                                                                        "type": "document",
                                                                        "document": {
                                                                            "id": media_id,            # Media ID from upload step
                                                                            "filename": filename       # Desired file name on recipient's phone
                                                                        }
                                                                    }

                                                                    response = requests.post(url, headers=headers, json=payload)
                                                                    response.raise_for_status()
                                                                    return response.json()


                                                                pdf_path = generate_leave_pdf()
                                                                media_id = upload_pdf_to_whatsapp(pdf_path)
                                                                send_whatsapp_pdf_by_media_id(sender_id, media_id)

                                                                send_whatsapp_message(
                                                                    sender_id,
                                                                    "Select an option below to continue 👇",
                                                                    buttons
                                                                )

                                                            elif all_approved_declined_cancelled.iat[0,8] == "Disapproved":

                                                                app_id = all_approved_declined_cancelled.iat[0,0]

                                                                buttons = [
                                                                    {"type": "reply", "reply": {"id": f"Resubapp_{app_id}", "title": "ReSubmit Application"}},
                                                                    {"type": "reply", "reply": {"id": "myhist", "title": "Download My History"}},
                                                                    {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                                ]
                                                                send_whatsapp_message(
                                                                    sender_id, 
                                                                    f"Hey {first_name}, your recent `{all_approved_declined_cancelled.iat[0,2]}` Leave Application `[ID - {all_approved_declined_cancelled.iat[0,0]}]` that you applied for on `{all_approved_declined_cancelled.iat[0,4].strftime('%d %B %Y')}` for `{all_approved_declined_cancelled.iat[0,7]} days` from `{all_approved_declined_cancelled.iat[0,5].strftime('%d %B %Y')}` to `{all_approved_declined_cancelled.iat[0,6].strftime('%d %B %Y')}` was {all_approved_declined_cancelled.iat[0,8]}❌ by `{all_approved_declined_cancelled.iat[0,3].title()}` on `{all_approved_declined_cancelled.iat[0,9].strftime('%d %B %Y')}`.",
                                                                    buttons 
                                                                )

                                                            elif all_approved_declined_cancelled.iat[0,8] == "Cancelled":

                                                                app_id = all_approved_declined_cancelled.iat[0,0]

                                                                buttons = [
                                                                    {"type": "reply", "reply": {"id": f"Resubapp_{app_id}", "title": "ReSubmit Application"}},
                                                                    {"type": "reply", "reply": {"id": "myhist", "title": "Download My History"}},
                                                                    {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                                ]
                                                                send_whatsapp_message(
                                                                    sender_id, 
                                                                    f"Hey {first_name}, on `{all_approved_declined_cancelled.iat[0,9].strftime('%d %B %Y')}` you Cancelled ⛔ your recent `{all_approved_declined_cancelled.iat[0,2]} Leave Application [ID - {all_approved_declined_cancelled.iat[0,0]}]` that you applied for on `{all_approved_declined_cancelled.iat[0,4].strftime('%d %B %Y')}` for `{all_approved_declined_cancelled.iat[0,7]} days` from `{all_approved_declined_cancelled.iat[0,5].strftime('%d %B %Y')}` to `{all_approved_declined_cancelled.iat[0,6].strftime('%d %B %Y')}`.",
                                                                    buttons 
                                                                )

                                                        else:

                                                            sections = [
                                                                {
                                                                    "title": "MY PROFILE",
                                                                    "rows": [
                                                                        {"id": "Apply", "title": "Apply for Leave"},
                                                                        {"id": "Track", "title": "Track My Application"},
                                                                        {"id": "Checkbal", "title": "Check Days Balance"},
                                                                        {"id": "myhist", "title": "My Applications History"},
                                                                        {"id": "Myinfo", "title": "My Info"},
                                                                        {"id": "Pending", "title": "Apps Pending My Approval"},
                                                                    ]
                                                                },
                                                                {
                                                                    "title": "ADMINISTRATION",
                                                                    "rows": [
                                                                        {"id": "Empmgt", "title": "Employee Management"},
                                                                        {"id": "Analyticscomp", "title": "Analytics & Insights"},
                                                                        {"id": "Company", "title": "Company Profile"},
                                                                    ]
                                                                }
                                                            ]

                                                            companyxx = company_reg.replace("_"," ").title()

                                                            send_whatsapp_list_message(
                                                                sender_id, 
                                                                f"Hello {first_name} {last_name}, LMS Leave Applications Approver from {companyxx}!\n\n You have not applied for any leave days yet.", 
                                                                "Administrator Options",
                                                                sections
                                                            )

                                                    elif len(df_employeesappspendingcheck) > 0:
                                                        
                                                        app_idx = df_employeesappspendingcheck.iat[0,0]

                                                        buttons = [
                                                            {"type": "reply", "reply": {"id": f"Reminder_{app_idx}", "title": "Remind Approver"}},
                                                            {"type": "reply", "reply": {"id": "Cancelapp", "title": "Cancel Pending App"}},
                                                            {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},

                                                        ]
                                                        approoooover = df_employeesappspendingcheck.iat[0,3].title()
                                                        send_whatsapp_message(
                                                            sender_id, 
                                                            f"Hey {first_name}, your recent `{df_employeesappspendingcheck.iat[0,2]}` Leave Application `[ID - {df_employeesappspendingcheck.iat[0,0]}]` applied on `{df_employeesappspendingcheck.iat[0,4].strftime('%d %B %Y')}` for `{df_employeesappspendingcheck.iat[0,7]} days from {df_employeesappspendingcheck.iat[0,5].strftime('%d %B %Y')} to {df_employeesappspendingcheck.iat[0,6].strftime('%d %B %Y')}` is still pending approval from `{approoooover}`.\n\n" 
                                                            f"Select an option below to either remind `{approoooover}` to approve your pending leave application or you can cancel the pending application to submit a new leave application."         
                                                            , 
                                                            buttons
                                                        )

                                                elif "reminder" in button_id.lower():

                                                    app_id = button_id.split("_")[1]
                                                    print(app_id)

                                                    try:
                                                    
                                                        print ("eissssssssshhhhhhhhhhhhhhhhhhhhhhhhhhhh")

                                                        table_name = company_reg + 'main'
                                                        company_name = company_reg.replace("_", " ").title()
                                                        table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                        table_name_apps_approved = f"{company_reg}appsapproved"


                                                        if not app_id:
                                                            print("none on appid")

                                                            return jsonify({"message": "Application ID is missing."}), 400

                                                        print(table_name_apps_pending_approval)

                                                        query = f"SELECT * FROM {table_name_apps_pending_approval} WHERE appid = %s;"
                                                        cursor.execute(query, (app_id,))
                                                        result = cursor.fetchone()
                                                        app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf, statuspre = result
                                                        print("chiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
                                                        print(employee_number)
                                                        print(approver_name)

                                                        try:

                                                            query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation, department FROM {table_name};"
                                                            cursor.execute(query)
                                                            rows = cursor.fetchall()

                                                            df_employees = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Address", "Role","Leave Approver Name","Leave Approver ID","Leave Approver Email", "Leave Approver WhatsAapp", "Leave Days Balance","Days Accumulated per Month", "Department"])
                                                            print(df_employees)
                                                            userdf = df_employees[df_employees['id'] == int(np.int64(employee_number))].reset_index()
                                                            print("yeaarrrrr")
                                                            print(userdf)
                                                            firstname = userdf.iat[0,2]
                                                            surname = userdf.iat[0,3]
                                                            whatsapp = userdf.iat[0,4]
                                                            address = userdf.iat[0,6]
                                                            email = userdf.iat[0,5]
                                                            fullnamedisp = firstname + ' ' + surname
                                                            leaveapprovername = userdf.iat[0,8]
                                                            leaveapproverid = userdf.iat[0,9]
                                                            leaveapproveremail = userdf.iat[0, 10]
                                                            leaveapproverwhatsapp = userdf.iat[0,11]
                                                            role = userdf.iat[0,7]
                                                            leavedaysbalance = userdf.iat[0,12]
                                                            department = userdf.iat[0,14] 
                                                            print('check')

                                                            departmentdf = df_employees[df_employees['Department'] == department].reset_index()
                                                            numberindepartment = len(departmentdf)
                                                            
                                                            startdatex = pd.Timestamp(start_date)
                                                            enddatex = pd.Timestamp(end_date)

                                                            leave_dates = pd.date_range(startdatex, enddatex)

                                                            query = f"""
                                                                SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate,
                                                                    leaveenddate, leavedaysappliedfor, approvalstatus, statusdate,
                                                                    leavedaysbalancebf, department
                                                                FROM {table_name_apps_approved}
                                                                WHERE department = %s;
                                                            """
                                                            cursor.execute(query, (department,))
                                                            rows = cursor.fetchall()

                                                            df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf","department"]) 
                                                            df_employeesappsapprovedcheck["leavestartdate"] = pd.to_datetime(df_employeesappsapprovedcheck["leavestartdate"])
                                                            df_employeesappsapprovedcheck["leaveenddate"] = pd.to_datetime(df_employeesappsapprovedcheck["leaveenddate"])
            
                                                            df_employeesappsapprovedcheck.dropna(subset=["leavestartdate", "leaveenddate"], inplace=True)
                                                            # Create daily impact report
                                                            impact_report = []

                                                            for date in leave_dates:

                                                                date = pd.Timestamp(date)

                                                                print(type(date))  # Should be pandas._libs.tslibs.timestamps.Timestamp or datetime.datetime
                                                                print(df_employeesappsapprovedcheck.dtypes)  # Check all datetime columns

                                                                on_leave = ((df_employeesappsapprovedcheck["leavestartdate"] <= date) & (df_employeesappsapprovedcheck["leaveenddate"] >= date)).sum()
                                                                remaining = numberindepartment - on_leave - 1  # subtract 1 for the new leave
                                                                impact_report.append({
                                                                    "date": date,  # <=== Keep as datetime, don't convert to string
                                                                    "on leave": on_leave + 1,
                                                                    "employees remaining": remaining
                                                                })

                                                            # Convert to DataFrame for display
                                                            impact_df = pd.DataFrame(impact_report)
                                                            print("IMPAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACT")
                                                            print(impact_df)
                                                            print(numberindepartment)

                                                            impact_df["date"] = pd.to_datetime(impact_df["date"], format="%Y-%m-%d")
                                                            impact_df = impact_df[impact_df["date"].dt.weekday != 6].copy()

                                                            change = (impact_df[["on leave", "employees remaining"]] != impact_df[["on leave", "employees remaining"]].shift()).any(axis=1)
                                                            change.iloc[0] = True  # ensure the first row starts a group
                                                            impact_df["group"] = change.cumsum()

                                                            statements = []
                                                            for _, group_df in impact_df.groupby("group"):
                                                                start = group_df["date"].iloc[0].strftime("%d %B %Y")
                                                                end = group_df["date"].iloc[-1].strftime("%d %B %Y")
                                                                on_leave = group_df["on leave"].iloc[0]
                                                                remaining = group_df["employees remaining"].iloc[0]
                                                                
                                                                if start == end:
                                                                    statements.append(f"On {start}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                                                                else:
                                                                    statements.append(f"From {start} to {end}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                                                                    # Combine all statements into a single variable
                                                            final_summary = "\n".join(statements)
                                                            # Print output
                                                            for s in statements:
                                                                print(s)

                                                        except Exception as e:
                                                            print(e)

                                                        sections = [
                                                            {
                                                                "title": "MY PROFILE",
                                                                "rows": [
                                                                    {"id": "Apply", "title": "Apply for Leave"},
                                                                    {"id": "Track", "title": "Track My Application"},
                                                                    {"id": "Checkbal", "title": "Check Days Balance"},
                                                                    {"id": "myhist", "title": "My Applications History"},
                                                                    {"id": "Myinfo", "title": "My Info"},
                                                                    {"id": "Pending", "title": "Apps Pending My Approval"},
                                                                ]
                                                            },
                                                            {
                                                                "title": "ADMINISTRATION",
                                                                "rows": [
                                                                    {"id": "Empmgt", "title": "Employee Management"},
                                                                    {"id": "Analyticscomp", "title": "Analytics & Insights"},
                                                                    {"id": "Company", "title": "Company Profile"},
                                                                ]
                                                            }
                                                        ]
                                                        
                                                        send_whatsapp_list_message(
                                                            sender_id,
                                                            f"Hey {first_name}. A reminder has been sent to {approver_name} on {approver_whatsapp} to decide on your `{leave_days}-day {leave_type} leave` running from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}` that you applied for on `{date_applied.strftime('%d %B %Y')}`✅! \n Select an option below to continue 👇",
                                                            "Admin/Approver Options",
                                                            sections
                                                        )
                                                        

                                                        if approver_whatsapp:

                                                            try:

                                                                buttons = [
                                                                    {"type": "reply", "reply": {"id": f"Approve5appwa_{app_id}", "title": "Approve"}},
                                                                    {"type": "reply", "reply": {"id": f"Disapproveappwa_{app_id}", "title": "Disapprove"}},
                                                                ]
                                                                send_whatsapp_message(
                                                                    f"263{approver_whatsapp}", 
                                                                    f"Hey {approver_name}! 😊. A gentle reminder, you have a new `{leave_type}` Leave Application from `{first_name} {surname}` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}`.\n\n" 
                                                                    f"If you approve this leave application, {final_summary}\n\n"  
                                                                    f"Select an option below to either approve or disapprove the application."         
                                                                    , 
                                                                    buttons
                                                                )

                                                            except Exception as e:
                                                                print(e)



                                                    except Exception as e:
                                                        print(e)
                                                        return jsonify({"message": "Error approving leave application.", "error": str(e)}), 500
                                                
                                                elif button_id == "myhist" or selected_option == "myhist":

                                                    try:

                                                        print(id_user)
                                                        companyxx = company_reg.replace("_"," ").title()

                                                        table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                        table_name_apps_approved = f"{company_reg}appsapproved"
                                                        table_name_apps_declined = f"{company_reg}appsdeclined"
                                                        table_name_apps_cancelled = f"{company_reg}appscancelled"

                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, approvalstatus, TO_CHAR(statusdate, 'FMDD Month YYYY') AS statusdate  FROM {table_name_apps_approved} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()
                                                        df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid", "id","leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate","approvalstatus","statusdate"]) 

                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, approvalstatus, TO_CHAR(statusdate, 'FMDD Month YYYY') AS statusdate   FROM {table_name_apps_declined} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()
                                                        df_employeesappsdeclinedcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate","approvalstatus","statusdate"])  
                                
                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, approvalstatus, TO_CHAR(statusdate, 'FMDD Month YYYY') AS statusdate  FROM {table_name_apps_cancelled} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()
                                                        df_employeesappscancelledcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate","approvalstatus","statusdate"])

                                                        query = f"SELECT appid, id, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, approvalstatus FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()
                                                        df_employeesappspenpendingcheck = pd.DataFrame(rows, columns=["appid", "id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate" ,"approvalstatus"])
                                                        df_employeesappspenpendingcheck["statusdate"] = ""


                                                        all_approved_declined = df_employeesappsapprovedcheck._append(df_employeesappsdeclinedcheck)
                                                        all_approved_declined_cancelled = all_approved_declined._append(df_employeesappscancelledcheck)
                                                        all_approved_declined_cancelled_pending = all_approved_declined_cancelled._append(df_employeesappspenpendingcheck)

                                                        all_approved_declined_cancelled_pending["dateapplied"] = pd.to_datetime(all_approved_declined_cancelled_pending["dateapplied"], errors='coerce')

                                                        all_approved_declined_cancelled_pending = all_approved_declined_cancelled_pending.sort_values(by="dateapplied", ascending=False)

                                                        print("hist hist hist")
                                                        all_approved_declined_cancelled_pending.drop('id', axis=1, inplace=True)
                                                        all_approved_declined_cancelled_pending["dateapplied"] = all_approved_declined_cancelled_pending["dateapplied"].dt.strftime("%-d %B %Y")

                                                        print(all_approved_declined_cancelled_pending)
                                                    
                                                        def generate_leave_hist_pdf():
                                                            app = {
                                                                'company_name': company_reg.replace("_", " ").title(),
                                                                'employee_name': f"{first_name} {last_name}",
                                                                'generated_on': today_date,
                                                                'power': power,
                                                            }

                                                            table_hist_html = all_approved_declined_cancelled_pending.to_html(index=False, classes='data', border=0, justify='center',escape=False)

                                                            html_out = render_template("leave_applications_history.html", app=app, table_hist_html=table_hist_html)
                                                            pdf_bytes = HTML(string=html_out).write_pdf()
                                                            return pdf_bytes

                                                        def upload_pdf_to_whatsapp(pdf_bytes):
                                                            compxxy = company_reg.replace("_"," ").title()
                                                            filename=f"{first_name}_{last_name}_{compxxy}_leave_applications_history.pdf"
                                                        
                                                            url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/media"
                                                            headers = {
                                                                "Authorization": f"Bearer {ACCESS_TOKEN}"
                                                            }

                                                            files = {
                                                                "file": (filename, io.BytesIO(pdf_bytes), "application/pdf"),
                                                                "type": (None, "application/pdf"),
                                                                "messaging_product": (None, "whatsapp")
                                                            }

                                                            response = requests.post(url, headers=headers, files=files)
                                                            print("📥 Full incoming data:", response.text)  # Good for debugging
                                                            response.raise_for_status()
                                                            return response.json()["id"]

                                                                                                        
                                                        def send_whatsapp_pdf_by_media_id(recipient_number, media_id):
                                                            compxxy = company_reg.replace("_"," ").title()
                                                            filename=f"{first_name}_{last_name}_{compxxy}_leave_applications_history.pdf"
                                                            url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                            headers = {
                                                                "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                                "Content-Type": "application/json"
                                                            }
                                                            payload = {
                                                                "messaging_product": "whatsapp",
                                                                "to": recipient_number,
                                                                "type": "document",
                                                                "document": {
                                                                    "id": media_id,            # Media ID from upload step
                                                                    "filename": filename       # Desired file name on recipient's phone
                                                                }
                                                            }

                                                            response = requests.post(url, headers=headers, json=payload)
                                                            response.raise_for_status()
                                                            return response.json()


                                                        pdf_path = generate_leave_hist_pdf()
                                                        media_id = upload_pdf_to_whatsapp(pdf_path)

                                                        appscountnum = len(all_approved_declined_cancelled_pending)

                                                        if appscountnum > 0:

                                                            send_whatsapp_pdf_by_media_id(sender_id, media_id)

                                                            buttons = [
                                                                {"type": "reply", "reply": {"id": "Analytics", "title": "Analytics & Insights"}},
                                                                {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                            ]
                                                            send_whatsapp_message(
                                                                sender_id, 
                                                                f"Hey {first_name} {last_name} from {companyxx}! You may go ahead and download your leave applications history file attached here 😎.", 
                                                                buttons
                                                            )

                                                        else:

                                                            buttons = [
                                                                {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                                                                {"type": "reply", "reply": {"id": "Checkbal", "title": "Check Days Balance"}},
                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                            ]

                                                            send_whatsapp_message(
                                                                sender_id, 
                                                                f"Hello {first_name} {last_name} from {companyxx}!\n\n You have not applied for any leave days yet.", 
                                                                buttons
                                                            )



                                                    except Exception as e:

                                                        send_whatsapp_message(f"+263710910052", f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully! Error; {e}")                      
                                                    
                                                elif selected_option == "Analyticscomp":
                                                
                                                    companyxx = company_reg.replace("_"," ").title()

                                                    buttons = [
                                                        {"type": "reply", "reply": {"id": "Book", "title": "Extract Leave Book"}},
                                                        {"type": "reply", "reply": {"id": "Summarycomp", "title": "Get Insights"}},
                                                        {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                    ]
                                                    send_whatsapp_message(
                                                        sender_id,
                                                        f"Hey {first_name}, select an option below to proceed.",
                                                        buttons
                                                    )

                                                elif selected_option == "Editemail":
                                                
                                                    companyxx = company_reg.replace("_"," ").title()



                                                    sections = [
                                                        {
                                                            "title": "User Options",
                                                            "rows": [
                                                                {"id": "Editname", "title": "Edit My Name"},
                                                                {"id": "Editwhatsapp", "title": "Change My WhatsApp #"},
                                                                {"id": "Editemail", "title": "Change My Email"},
                                                                {"id": "Editwebpass", "title": "Change Web Password"},
                                                                {"id": "Editaddress", "title": "Edit My Address"},
                                                                {"id": "MyInfo", "title": "My Info"},
                                                                {"id": "Menu", "title": "Main Menu"}
                                                            ]
                                                        }
                                                    ]

                                                    send_whatsapp_list_message(
                                                        sender_id, 
                                                        f"Hey {first_name}!\n Enter your new email address starting with the word `email` as shown below 👇. \n\n `email epsilon@gmail.com`", 
                                                    "User Options",
                                                    sections)
                            
                                                elif selected_option == "Myinfo":

                                                    companyxx = company_reg.replace("_"," ").title()

                                                    try:

                                                        table_name = f"{company_reg}main"

                                                        query = f"SELECT id, firstname, surname, whatsapp, address, email, role, department, currentleavedaysbalance, monthlyaccumulation, leaveapprovername, leaveapproverwhatsapp, leaveapproveremail FROM {table_name} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        row = cursor.fetchone()

                                                        if row:

                                                            columns = ["ID", "First Name", "Surname", "WhatsApp", "Address", "Email", 
                                                                    "Role", "Department", "Leave Days", "Monthly Accrual", 
                                                                    "Approver", "Approver WhatsApp", "Approver Email"]

                                                            message_text = "*📄 Employee Details:*\n\n"
                                                            for col, val in zip(columns, row):
                                                                message_text += f"*{col}:* {val}\n"

                                                            sections = [
                                                                {
                                                                    "title": "User Options",
                                                                    "rows": [
                                                                        {"id": "Editname", "title": "Edit My Name"},
                                                                        {"id": "Editwhatsapp", "title": "Change My WhatsApp #"},
                                                                        {"id": "Editemail", "title": "Change My Email"},
                                                                        {"id": "Editwebpass", "title": "Change Web Password"},
                                                                        {"id": "Editaddress", "title": "Edit My Address"},
                                                                        {"id": "MyInfo", "title": "My Info"},
                                                                        {"id": "Menu", "title": "Main Menu"}
                                                                    ]
                                                                }
                                                            ]

                                                            send_whatsapp_list_message(
                                                                sender_id, 
                                                                f"Hey there {first_name}!\n Your information in {companyxx}'s Leave Management System is as follows;\n\n {message_text}", 
                                                            "User Options",
                                                            sections)



                                                    except Exception as e:

                                                        print(e)

                                                        send_whatsapp_message(f"+263710910052", f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully! Error; {e}")                      

                                                elif button_id == "Apply" or selected_option == "Apply":

                                                    table_name_apps_pending_approval = f"{company_reg}appspendingapproval"

                                                    query = f"SELECT id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor  FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                    cursor.execute(query)
                                                    rows = cursor.fetchall()

                                                    df_employeesappspendingcheck = pd.DataFrame(rows, columns=["id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor"])    

                                                    if len(df_employeesappspendingcheck) == 0:

                                                        sections = [
                                                            {
                                                                "title": "Leave Type Options",
                                                                "rows": [
                                                                    {"id": "Annual", "title": "Annual Leave"},
                                                                    {"id": "Sick", "title": "Sick Leave"},
                                                                    {"id": "Study", "title": "Study Leave"},
                                                                    {"id": "Bereavement", "title": "Bereavement Leave"},
                                                                    {"id": "Parental", "title": "Parental Leave"},
                                                                    {"id": "Other", "title": "Other"},
                                                                ]
                                                            }
                                                        ]

                                                        send_whatsapp_list_message(
                                                            sender_id, 
                                                            f"{first_name}, kindly select the type of Leave that you are applying for.", 
                                                            "Leave Type Options",
                                                            sections) 

                                                    elif len(df_employeesappspendingcheck) > 0:

                                                        app_idx = df_employeesappspendingcheck.iat[0,0]

                                                        buttons = [
                                                            {"type": "reply", "reply": {"id": f"Reminder_{app_idx}", "title": "Remind Approver"}},
                                                            {"type": "reply", "reply": {"id": "Cancelapp", "title": "Cancel Pending App"}},
                                                            {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                        ]
                                                        send_whatsapp_message(
                                                            sender_id, 
                                                            f"Oops! 🥲. Sorry {first_name}, you cannot apply for leave whilst you have another leave application which is still pending approval.\n\n" 
                                                            f"Your `{df_employeesappspendingcheck.iat[0,2]}` Leave Application `[ID - {df_employeesappspendingcheck.iat[0,0]}]` applied on `{df_employeesappspendingcheck.iat[0,4].strftime('%d %B %Y')}` for `{df_employeesappspendingcheck.iat[0,7]} days from {df_employeesappspendingcheck.iat[0,5].strftime('%d %B %Y')} to {df_employeesappspendingcheck.iat[0,6].strftime('%d %B %Y')}` is still pending approval from {df_employeesappspendingcheck.iat[0,3]}.\n\n" 
                                                            f"Select an option below to either remind the approver to approved your pending application or you can cancel the pending application to submit a new leave application."         
                                                            , 
                                                            buttons
                                                        )

                                                elif button_id == "Submitapp":
                                        
                                                    try:

                                                        table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                        table_name_apps_approved = f"{company_reg}appsapproved"
                                                        companyxx = company_reg.replace("_", " ").title()

                                                        cursor.execute("""
                                                            SELECT id ,empidwa, leavetypewa, startdate, enddate FROM whatsapptempapplication
                                                            WHERE empidwa = %s
                                                        """, (id_user,))
                                                
                                                        result = cursor.fetchone()

                                                        appid = result[0]
                                                        id_user = result[1]
                                                        leavetype = result[2]
                                                        startdate = result[3]
                                                        enddate = result[4]
                                                        table_name = f"{company_reg}main"
                                                        
                                                        query = f"SELECT id FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                        cursor.execute(query)
                                                        rows = cursor.fetchall()

                                                        df_employeesappspendingcheck = pd.DataFrame(rows, columns=["id"])    

                                                        if len(df_employeesappspendingcheck) == 0:

                                                            query = f"""SELECT appid, id, leavestartdate, leaveenddate FROM {table_name_apps_approved} WHERE id = %s AND leavestartdate <= %s AND leaveenddate >= %s"""

                                                            cursor.execute(query, (id_user, enddate, startdate))
                                                            results = cursor.fetchall()

                                                            # Process results
                                                            if results:
                                                                print("Overlapping records found:")

                                                                try:

                                                                    overlap_messages = []

                                                                    for row in results:

                                                                        formatted_date_start = row[2].strftime("%d %B %Y")
                                                                        formatted_date_end = row[3].strftime("%d %B %Y")

                                                                        overlap_messages.append(f"appID: {row[0]}, Starting Date: {formatted_date_start}, Ending Date: {formatted_date_end}")

                                                                    # Combine into one single string (newline-separated)
                                                                    overlap_info = "\n".join(overlap_messages)

                                                                    buttons = [
                                                                        {"type": "reply", "reply": {"id": f"Apply", "title": "Restart Application"}},
                                                                        {"type": "reply", "reply": {"id": f"ApplyRevoke", "title": "Revoke Conflictn App"}},
                                                                        {"type": "reply", "reply": {"id": f"Menu", "title": "Main Menu"}},
                                                                    ]

                                                                    send_whatsapp_message(sender_id, f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully!\n\n"
                                                                        f"One of your previously approved leave applications include days within the period that you are currently applying for.\n\n Leave App; {overlap_info}.\n\n Either restart your application with different dates from these, or apply that this conflicting approved Leave application be Revoked.",
                                                                        buttons
                                                                        )
                                                                
                                                                except Exception as e:

                                                                    send_whatsapp_message(f"+263710910052", f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully! Error; {e}")                      
                                                            
                                                            else:

                                                                print("No Overlapping records found:")

                                                                table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                                table_name_apps_approved = f"{company_reg}appsapproved"

                                                                query = f"SELECT id FROM {table_name_apps_pending_approval} WHERE id = {str(id_user)};"
                                                                cursor.execute(query)
                                                                rows = cursor.fetchall()

                                                                df_employeesappspendingcheck = pd.DataFrame(rows, columns=["id"])    

                                                                if len(df_employeesappspendingcheck) == 0:

                                                                    cursor.execute("""
                                                                        SELECT id ,empidwa, leavetypewa, startdate, enddate FROM whatsapptempapplication
                                                                        WHERE empidwa = %s
                                                                    """, (id_user,))
                                                            
                                                                    result = cursor.fetchone()

                                                                    appid = result[0]
                                                                    leavetype = result[2]
                                                                    startdate = result[3]
                                                                    enddate = result[4]
                                                                    table_name = f"{company_reg}main"

                                                                    if isinstance(startdate, str):
                                                                        startdate = datetime.datetime.strptime(startdate, "%Y-%m-%d").date()
                                                                    if isinstance(enddate, str):
                                                                        enddate = datetime.datetime.strptime(enddate, "%Y-%m-%d").date()

                                                                    business_days = 0
                                                                    current_date = startdate

                                                                    while current_date <= enddate:
                                                                        if current_date.weekday() != 6:  # 0=Mon, 1=Tue, ..., 4=Fri
                                                                            business_days += 1
                                                                        current_date += timedelta(days=1)  # Use timedelta directly

                                                                    query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation, department FROM {table_name};"
                                                                    cursor.execute(query)
                                                                    rows = cursor.fetchall()

                                                                    df_employees = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Address", "Role","Leave Approver Name","Leave Approver ID","Leave Approver Email", "Leave Approver WhatsAapp", "Leave Days Balance","Days Accumulated per Month", "Department"])
                                                                    print(df_employees)
                                                                    userdf = df_employees[df_employees['id'] == int(np.int64(id_user))].reset_index()
                                                                    print("yeaarrrrr")
                                                                    print(userdf)
                                                                    firstname = userdf.iat[0,2]
                                                                    surname = userdf.iat[0,3]
                                                                    whatsapp = userdf.iat[0,4]
                                                                    address = userdf.iat[0,6]
                                                                    email = userdf.iat[0,5]
                                                                    fullnamedisp = firstname + ' ' + surname
                                                                    leaveapprovername = userdf.iat[0,8]
                                                                    leaveapproverid = userdf.iat[0,9]
                                                                    leaveapproveremail = userdf.iat[0, 10]
                                                                    leaveapproverwhatsapp = userdf.iat[0,11]
                                                                    role = userdf.iat[0,7]
                                                                    leavedaysbalance = userdf.iat[0,12]
                                                                    department = userdf.iat[0,14] 
                                                                    print('check')

                                                                    departmentdf = df_employees[df_employees['Department'] == department].reset_index()
                                                                    numberindepartment = len(departmentdf)
                                                                    
                                                                    startdatex = pd.Timestamp(startdate)
                                                                    enddatex = pd.Timestamp(enddate)

                                                                    leave_dates = pd.date_range(startdatex, enddatex)

                                                                    query = f"""
                                                                        SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate,
                                                                            leaveenddate, leavedaysappliedfor, approvalstatus, statusdate,
                                                                            leavedaysbalancebf, department
                                                                        FROM {table_name_apps_approved}
                                                                        WHERE department = %s;
                                                                    """
                                                                    cursor.execute(query, (department,))
                                                                    rows = cursor.fetchall()

                                                                    df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf","department"]) 
                                                                    df_employeesappsapprovedcheck["leavestartdate"] = pd.to_datetime(df_employeesappsapprovedcheck["leavestartdate"])
                                                                    df_employeesappsapprovedcheck["leaveenddate"] = pd.to_datetime(df_employeesappsapprovedcheck["leaveenddate"])
                    
                                                                    df_employeesappsapprovedcheck.dropna(subset=["leavestartdate", "leaveenddate"], inplace=True)
                                                                    # Create daily impact report
                                                                    impact_report = []

                                                                    for date in leave_dates:

                                                                        date = pd.Timestamp(date)

                                                                        print(type(date))  # Should be pandas._libs.tslibs.timestamps.Timestamp or datetime.datetime
                                                                        print(df_employeesappsapprovedcheck.dtypes)  # Check all datetime columns

                                                                        on_leave = ((df_employeesappsapprovedcheck["leavestartdate"] <= date) & (df_employeesappsapprovedcheck["leaveenddate"] >= date)).sum()
                                                                        remaining = numberindepartment - on_leave - 1  # subtract 1 for the new leave
                                                                        impact_report.append({
                                                                            "date": date,  # <=== Keep as datetime, don't convert to string
                                                                            "on leave": on_leave + 1,
                                                                            "employees remaining": remaining
                                                                        })

                                                                    # Convert to DataFrame for display
                                                                    impact_df = pd.DataFrame(impact_report)
                                                                    print("IMPAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACT")
                                                                    print(impact_df)
                                                                    print(numberindepartment)

                                                                    impact_df["date"] = pd.to_datetime(impact_df["date"], format="%Y-%m-%d")
                                                                    impact_df = impact_df[impact_df["date"].dt.weekday != 6].copy()

                                                                    change = (impact_df[["on leave", "employees remaining"]] != impact_df[["on leave", "employees remaining"]].shift()).any(axis=1)
                                                                    change.iloc[0] = True  # ensure the first row starts a group
                                                                    impact_df["group"] = change.cumsum()

                                                                    statements = []
                                                                    for _, group_df in impact_df.groupby("group"):
                                                                        start = group_df["date"].iloc[0].strftime("%d %B %Y")
                                                                        end = group_df["date"].iloc[-1].strftime("%d %B %Y")
                                                                        on_leave = group_df["on leave"].iloc[0]
                                                                        remaining = group_df["employees remaining"].iloc[0]
                                                                        
                                                                        if start == end:
                                                                            statements.append(f"On {start}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                                                                        else:
                                                                            statements.append(f"From {start} to {end}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                                                                            # Combine all statements into a single variable
                                                                    final_summary = "\n".join(statements)
                                                                    # Print output
                                                                    for s in statements:
                                                                        print(s)

                                                                    if leavetype == "Annual":

                                                                        leavedaysbalancebf = float(leavedaysbalance) - float(business_days)

                                                                    else:

                                                                        leavedaysbalancebf = float(leavedaysbalance)
                                                                
                                                                    if leavedaysbalancebf >= 0:

                                                                        status = "Pending"

                                                                        insert_query = f"""
                                                                        INSERT INTO {table_name_apps_pending_approval} (id, firstname, surname, department, leavetype, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus)
                                                                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                                                        """
                                                                        cursor.execute(insert_query, (int(np.int64(id_user)), first_name, last_name, department, leavetype, leaveapprovername, int(np.int64(leaveapproverid)), leaveapproveremail, int(np.int64(leaveapproverwhatsapp)), float(np.float64(leavedaysbalance)), today_date, startdate, enddate, float(np.int64(business_days)), float(np.float64(leavedaysbalancebf)), status))
                                                                        connection.commit()

                                                                        query = f"SELECT appid FROM {table_name_apps_pending_approval};"
                                                                        cursor.execute(query)
                                                                        rows = cursor.fetchall()

                                                                        df_employees = pd.DataFrame(rows, columns=["id"])
                                                                        leaveappid = df_employees.iat[0,0]
                                                                        companyxx = company_reg.replace("_"," ").title()
                                                                        approovvver = leaveapprovername.title()

                                                                        buttons = [
                                                                            {"type": "reply", "reply": {"id": "Track", "title": "Track Application"}},
                                                                            {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                                            ]

                                                                        send_whatsapp_message(sender_id, f"✅ Great News {first_name} from {companyxx}! \n\n Your `{leavetype} Leave Application` for `{business_days} days` from `{startdate.strftime('%d %B %Y')}` to `{enddate.strftime('%d %B %Y')}` has been submitted successfully!\n\n"
                                                                            f"Your Leave Application ID is `{leaveappid}`.\n\n"
                                                                            f"A Notification has been sent to `{approovvver}`  on `+263{leaveapproverwhatsapp}` to decide on  your application.",
                                                                            buttons)
                                                                        
                                                                        if leaveapproverwhatsapp:
                            
                                                                            buttons = [
                                                                                {"type": "reply", "reply": {"id": f"Approve5appwa_{leaveappid}", "title": "Approve"}},
                                                                                {"type": "reply", "reply": {"id": f"Disapproveappwa_{leaveappid}", "title": "Disapprove"}},
                                                                            ]
                                                                            send_whatsapp_message(
                                                                                f"263{leaveapproverwhatsapp}", 
                                                                                f"Hey {approovvver}! 😊. New `{leavetype}` Leave Application from `{first_name} {surname}` for `{business_days} days` from `{startdate.strftime('%d %B %Y')}` to `{enddate.strftime('%d %B %Y')}`.\n\n" 
                                                                                f"If you approve this leave application, {final_summary}\n\n"  
                                                                                f"Select an option below to either approve or disapprove the application."         
                                                                                , 
                                                                                buttons
                                                                            )

                                                                    else:


                                                                        buttons = [
                                                                            {"type": "reply", "reply": {"id": f"Apply", "title": "Restart Application"}},
                                                                            {"type": "reply", "reply": {"id": f"Checkbal", "title": "Check Days Balance"}},
                                                                            {"type": "reply", "reply": {"id": f"Menu", "title": "Main Menu"}},
                                                                        ]

                                                                        send_whatsapp_message(sender_id, f"Oops, {first_name} from {companyxx}! \n\n Your Leave Application` has NOT been submitted successfully!\n\n"
                                                                            f"You only have *{leavedaysbalance}* days available for leave but you are applying for *{business_days}*.\n\n You can restart your application and apply for leave such that the days between your leave start date and end date do not exceed your available balance of *{leavedaysbalance}* days.",
                                                                            buttons
                                                                            )

                                                                else:
                                                                    print("leave app submission failed")

                                                    except ValueError as e:
                                                        send_whatsapp_message(
                                                            sender_id,
                                                            f"{e}, ❌ No, incorrect message format. Please use:\n"
                                                            "`end 24 january 2025`\n"
                                                            "Example: `end 15 march 2024`"
                                                        )

                                                elif button_id == "Checkbal" or selected_option == "Checkbal":

                                                    buttons = [
                                                    {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                                                    {"type": "reply", "reply": {"id": "Track", "title": "Track Application"}},
                                                    {"type": "reply", "reply": {"id": f"Menu", "title": "Main Menu"}},
                                                    ]

                                                    send_whatsapp_message(
                                                        sender_id, 
                                                        f"Hey {first_name}, your current available leave days balance is `{days_days_balance} days`.\n\n"
                                                        "Select an option below to continue 👇",
                                                        buttons
                                                    )

                                                elif button_id == "Resubapp" :

                                                    table_name_apps_cancelled = f"{company_reg}appscancelled"
                                                    table_name_apps_pending_approval = f"{company_reg}appspendingapproval"

                                                    query = f"SELECT appid, id, firstname, surname, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail , leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf FROM {table_name_apps_cancelled} WHERE id = %s;"
                                                    cursor.execute(query, (id_user,))
                                                    result = cursor.fetchall()

                                                    if result:

                                                        df_employees = pd.DataFrame(result)
                                                        df_employees = df_employees.sort_values(by=df_employees.columns[0], ascending=False)
                                                        print(df_employees)
                                                                
                                                        try:

                                                            status = "Pending"
                                                            app_id = int(np.int64(df_employees.iat[0,0]))
                                                            employee_number = int(np.int64(df_employees.iat[0,1]))
                                                            first_name = df_employees.iat[0,2]
                                                            surname = df_employees.iat[0,3]
                                                            leave_type = df_employees.iat[0,4]
                                                            leave_specify = df_employees.iat[0,5]
                                                            approver_name = df_employees.iat[0,6]
                                                            approver_id =  int(np.int64(df_employees.iat[0,7]))
                                                            approver_email =  df_employees.iat[0,8]
                                                            approver_whatsapp =  int(np.int64(df_employees.iat[0,9]))
                                                            leave_days_balance =  float(np.float64(df_employees.iat[0,10]))
                                                            date_applied = df_employees.iat[0,11]
                                                            start_date = df_employees.iat[0,12]
                                                            end_date = df_employees.iat[0,13]
                                                            leave_days =  float(np.int64(df_employees.iat[0,14]))
                                                            leavedaysbalancebf =  float(np.float64(df_employees.iat[0,15]))
                                                            insert_query = f"""
                                                            INSERT INTO {table_name_apps_pending_approval} 
                                                            (appid, id, firstname, surname, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus)
                                                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                                            """

                                                            cursor.execute(insert_query, (
                                                                app_id, employee_number, first_name, surname, leave_type, leave_specify, approver_name, 
                                                                approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                                                                end_date, leave_days, leavedaysbalancebf, status
                                                            ))
                                                            
                                                            connection.commit()
                                                            print("Insert successful!")

                                                        except Exception as e:
                                                            print("Error inserting data:", e)

                                                        # SQL query to delete or mark the leave as canceled
                                                        query = f"""DELETE FROM {table_name_apps_cancelled} WHERE appid = %s"""
                                                        cursor.execute(query, (app_id,))
                                                        connection.commit()                                       

                                                        companyxx = company_reg.replace("_", " ").title()

                                                        sections = [
                                                            {
                                                                "title": "MY PROFILE",
                                                                "rows": [
                                                                    {"id": "Apply", "title": "Apply for Leave"},
                                                                    {"id": "Track", "title": "Track My Application"},
                                                                    {"id": "Checkbal", "title": "Check Days Balance"},
                                                                    {"id": "myhist", "title": "My Applications History"},
                                                                    {"id": "Myinfo", "title": "My Info"},
                                                                    {"id": "Pending", "title": "Apps Pending My Approval"},
                                                                ]
                                                            },
                                                            {
                                                                "title": "ADMINISTRATION",
                                                                "rows": [
                                                                    {"id": "Empmgt", "title": "Employee Management"},
                                                                    {"id": "Analyticscomp", "title": "Analytics & Insights"},
                                                                    {"id": "Company", "title": "Company Profile"},
                                                                ]
                                                            }
                                                        ]

                                                        send_whatsapp_list_message(sender_id, f"Hey {first_name} from {companyxx}! \n\n Your `{leave_type} Leave Application` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}` has been Re-Submitted for approval successfully✅!",
                                                        "Administrator Options",
                                                        sections)                                          
                                                    
                                                    else:
                                                        print("No record found for the user.")

                                                elif "appwa" in button_id.lower():

                                                    app_id = button_id.split("_")[1]
                                                    print(app_id)

                                                    if "approve5" in button_id.lower():

                                                        try:
                                                        
                                                            print ("eissssssssshhhhhhhhhhhhhhhhhhhhhhhhhhhh")

                                                            table_name = company_reg + 'main'
                                                            company_name = company_reg.replace("_", " ").title()
                                                            table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                            table_name_apps_approved = f"{company_reg}appsapproved"

                                                            if not app_id:
                                                                print("none on appid")

                                                                return jsonify({"message": "Application ID is missing."}), 400

                                                            status = "Approved"
                                                            statusdate = today_date
                                                            print("bababababababababa")
                                                            print(table_name_apps_pending_approval)

                                                            query = f"SELECT * FROM {table_name_apps_pending_approval} WHERE appid = %s;"
                                                            cursor.execute(query, (app_id,))
                                                            result = cursor.fetchone()
                                                            app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf, statuspre = result
                                                            print("chiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
                                                            print(employee_number)
                                                            print(approver_name)

                                                            try:
                                                                insert_query = f"""
                                                                INSERT INTO {table_name_apps_approved} 
                                                                (appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus, statusdate)
                                                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                                                """
                                                                
                                                                cursor.execute(insert_query, (
                                                                    app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, 
                                                                    approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                                                                    end_date, leave_days, leavedaysbalancebf, status, statusdate
                                                                ))
                                                                
                                                                connection.commit()
                                                                print("Insert successful!")

                                                                query = f"UPDATE {table_name} SET currentleavedaysbalance = %s WHERE id = %s;"
                                                                cursor.execute(query, (leavedaysbalancebf, employee_number))
                                                                connection.commit()

                                                            except Exception as e:
                                                                print("Error inserting data:", e)

                                                            query = f"""DELETE FROM {table_name_apps_pending_approval} WHERE appid = %s"""
                                                            cursor.execute(query, (app_id,))
                                                            connection.commit()

                                                            query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation, department FROM {table_name};"
                                                            cursor.execute(query)
                                                            rows = cursor.fetchall()

                                                            df_employees = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Address", "Role","Leave Approver Name","Leave Approver ID","Leave Approver Email", "Leave Approver WhatsAapp", "Leave Days Balance","Days Accumulated per Month", "Department"])
                                                            print(df_employees)
                                                            userdf = df_employees[df_employees['id'] == int(np.int64(employee_number))].reset_index()
                                                            print("yeaarrrrr")
                                                            print(userdf)
                                                            firstname = userdf.iat[0,2].title()
                                                            surname = userdf.iat[0,3].title()
                                                            whatsappemp = userdf.iat[0,4]
                                                            email = userdf.iat[0,5]
                                                            address = userdf.iat[0,6]
                                                            companyxx = company_name.replace("_", " ").title()
                                                            app_namexx = approver_name.title()

                                                            query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate, leavedaysbalancebf, department  FROM {table_name_apps_approved} WHERE id = {str(employee_number)};"
                                                            cursor.execute(query)
                                                            rows = cursor.fetchall()
                                                            df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf", "department"]) 

                                                            df_employeesappsapprovedcheck = df_employeesappsapprovedcheck.sort_values(by="appid", ascending=False)  

                                                            def generate_leave_pdf():
                                                                app = {
                                                                    'company_logo': 44,
                                                                    'company_name': companyxx,
                                                                    'employee_name': f"{first_name} {surname}",
                                                                    'leave_type': leave_type,
                                                                    'generated_on': today_date,
                                                                    'date_applied': df_employeesappsapprovedcheck.iat[0,4].strftime('%d %B %Y'),
                                                                    'approver_name': df_employeesappsapprovedcheck.iat[0,3].title(),
                                                                    'reference_number': df_employeesappsapprovedcheck.iat[0,0],
                                                                    'approved_date': df_employeesappsapprovedcheck.iat[0,9].strftime('%d %B %Y'),
                                                                    'new_balance': df_employeesappsapprovedcheck.iat[0,10],
                                                                    'start_date':  df_employeesappsapprovedcheck.iat[0,5].strftime('%d %B %Y'),
                                                                    'end_date':  df_employeesappsapprovedcheck.iat[0,6].strftime('%d %B %Y'),
                                                                    'days_requested':  df_employeesappsapprovedcheck.iat[0,7], 
                                                                    'department':  department, 
                                                                    'address': address, 
                                                                    'whatsapp': f"+263{whatsappemp}", 
                                                                    'email': email, 
                                                                    'status': 'Approved',
                                                                    'power': power,
                                                                }

                                                                html_out = render_template("leave_pdf_template.html", app=app)
                                                                
                                                                # ✅ Return as bytes instead of saving to file
                                                                pdf_bytes = HTML(string=html_out).write_pdf()
                                                                return pdf_bytes

                                                            

                                                            def upload_pdf_to_whatsapp(pdf_bytes):
                                                                filename=f"leave_application_{df_employeesappsapprovedcheck.iat[0,0]}_{first_name}_{surname}_{companyxx}.pdf"
                                                            
                                                                url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/media"
                                                                headers = {
                                                                    "Authorization": f"Bearer {ACCESS_TOKEN}"
                                                                }

                                                                files = {
                                                                    "file": (filename, io.BytesIO(pdf_bytes), "application/pdf"),
                                                                    "type": (None, "application/pdf"),
                                                                    "messaging_product": (None, "whatsapp")
                                                                }

                                                                response = requests.post(url, headers=headers, files=files)
                                                                print("📥 Full incoming data:", response.text)  # Good for debugging
                                                                response.raise_for_status()
                                                                return response.json()["id"]

                                                                                                            
                                                            def send_whatsapp_pdf_by_media_id(recipient_number, media_id):
                                                                filename=f"leave_application_{df_employeesappsapprovedcheck.iat[0,0]}_{first_name}_{surname}_{companyxx}.pdf"
                                                                url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                                headers = {
                                                                    "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                                    "Content-Type": "application/json"
                                                                }
                                                                payload = {
                                                                    "messaging_product": "whatsapp",
                                                                    "to": recipient_number,
                                                                    "type": "document",
                                                                    "document": {
                                                                        "id": media_id,            # Media ID from upload step
                                                                        "filename": filename       # Desired file name on recipient's phone
                                                                    }
                                                                }

                                                                response = requests.post(url, headers=headers, json=payload)
                                                                response.raise_for_status()
                                                                return response.json()


                                                            pdf_path = generate_leave_pdf()
                                                            media_id = upload_pdf_to_whatsapp(pdf_path)

                                                            buttonsapproval = [
                                                                {"type": "reply", "reply": {"id": "Revoke", "title": "Revoke Approval"}},
                                                                {"type": "reply", "reply": {"id": "Pending", "title": "Pending My Approval"}},
                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                            ]

                                                            send_whatsapp_message(sender_id, f"✅ Great News {approver_name} from {companyxx}! \n\n You have successfully approved `{first_name} {surname}`'s  `{leave_days} day` `{leave_type} Leave Application` running from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}`✅!")
                                                            send_whatsapp_pdf_by_media_id(sender_id, media_id)
                                                            send_whatsapp_message(
                                                                sender_id,
                                                                "Select an option below to continue 👇, or Type `Hello` to view all Administrator/Approver Options",
                                                                buttonsapproval
                                                            )

                                                            if whatsappemp:

                                                                buttons = [
                                                                    {"type": "reply", "reply": {"id": "Revoke", "title": "Revoke Application"}},
                                                                    {"type": "reply", "reply": {"id": "myhist", "title": "Download My History"}},
                                                                    {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                                ]

                                                                send_whatsapp_message(f"263{whatsappemp}", f"✅ Great News {first_name} {surname} from {companyxx}! \n\n Your `{leave_type} Leave Application` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}`, has been Approved ✅ by `{app_namexx}`!")
                                                                send_whatsapp_pdf_by_media_id(f"263{whatsappemp}", media_id)
                                                                send_whatsapp_message(
                                                                    f"263{whatsappemp}",
                                                                    "Select an option below to continue 👇, or Type `Hello` to view all User Options",
                                                                    buttons
                                                                )
                                                        
                                                        except Exception as e:
                                                            return jsonify({"message": "Error approving leave application.", "error": str(e)}), 500

                                                    elif "disapprove" in button_id.lower():

                                                        print("disapproved")

                                                        try:
                                                        
                                                            print ("eissssssssshhhhhhhhhhhhhhhhhhhhhhhhhhhh")

                                                            table_name = company_reg + 'main'
                                                            company_name = company_reg.replace("_", " ").title()
                                                            table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                            table_name_apps_approved = f"{company_reg}appsapproved"
                                                            table_name_apps_declined = f"{company_reg}appsdeclined"


                                                            if not app_id:
                                                                print("none on appid")

                                                                return jsonify({"message": "Application ID is missing."}), 400

                                                            status = "Disapproved"
                                                            statusdate = today_date
                                                            print("bababababababababa")
                                                            print(table_name_apps_pending_approval)

                                                            query = f"SELECT * FROM {table_name_apps_pending_approval} WHERE appid = %s;"
                                                            cursor.execute(query, (app_id,))
                                                            result = cursor.fetchone()
                                                            app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf, statuspre = result
                                                            print("chiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
                                                            print(employee_number)
                                                            print(approver_name)

                                                            try:
                                                                insert_query = f"""
                                                                INSERT INTO {table_name_apps_declined} 
                                                                (appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus, statusdate)
                                                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                                                """
                                                                
                                                                cursor.execute(insert_query, (
                                                                    app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, 
                                                                    approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                                                                    end_date, leave_days, leavedaysbalancebf, status, statusdate
                                                                ))
                                                                
                                                                connection.commit()
                                                                print("Insert successful!")

                                                            except Exception as e:
                                                                print("Error inserting data:", e)

                                                            query = f"""DELETE FROM {table_name_apps_pending_approval} WHERE appid = %s"""
                                                            cursor.execute(query, (app_id,))
                                                            connection.commit()

                                                            query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation, department FROM {table_name};"
                                                            cursor.execute(query)
                                                            rows = cursor.fetchall()

                                                            df_employees = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Address", "Role","Leave Approver Name","Leave Approver ID","Leave Approver Email", "Leave Approver WhatsAapp", "Leave Days Balance","Days Accumulated per Month", "Department"])
                                                            print(df_employees)
                                                            userdf = df_employees[df_employees['id'] == int(np.int64(employee_number))].reset_index()
                                                            print("yeaarrrrr")
                                                            print(userdf)
                                                            firstname = userdf.iat[0,2].title()
                                                            surname = userdf.iat[0,3].title()
                                                            whatsappemp = userdf.iat[0,4]
                                                            email = userdf.iat[0,5]
                                                            address = userdf.iat[0,6]
                                                            companyxx = company_name.replace("_", " ").title()
                                                            app_namexx = approver_name.title()

                                                            query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate, leavedaysbalancebf  FROM {table_name_apps_approved} WHERE id = {str(employee_number)};"
                                                            cursor.execute(query)
                                                            rows = cursor.fetchall()
                                                            df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf"]) 

                                                            df_employeesappsapprovedcheck["dateapplied"] = pd.to_datetime(df_employeesappsapprovedcheck["dateapplied"], errors='coerce')
                                                            df_employeesappsapprovedcheck = df_employeesappsapprovedcheck.sort_values(by="dateapplied", ascending=False)
                                                            

                                                            buttonsapproval = [
                                                                {"type": "reply", "reply": {"id": "Revokedis", "title": "Revoke Disapproval"}},
                                                                {"type": "reply", "reply": {"id": "Pending", "title": "Apps Pending My Approval"}},
                                                                {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                            ]

                                                            send_whatsapp_message(sender_id, f"✅ Hey {approver_name} from {companyxx}! \n\n You have successfully disapproved `{first_name} {surname}`'s  `{leave_days} day` `{leave_type} Leave Application` running from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}`✅!")
                                                            send_whatsapp_message(
                                                                sender_id,
                                                                "Select an option below to continue 👇y, or Type `Hello` to view all Approver options",
                                                                buttonsapproval
                                                            )

                                                            if whatsappemp:

                                                                buttons = [
                                                                    {"type": "reply", "reply": {"id": "Reapply", "title": "Resubmit Application"}},
                                                                    {"type": "reply", "reply": {"id": "myhist", "title": "Download My History"}},
                                                                    {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}}
                                                                ]

                                                                send_whatsapp_message(f"263{whatsappemp}", f"✅ Oops, {first_name} {surname} from {companyxx}! \n\n Your `{leave_type} Leave Application` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}`, has been disapproved ❌ by `{app_namexx}`!")
                                                                send_whatsapp_message(
                                                                    f"263{whatsappemp}",
                                                                    "Select an option below to continue 👇",
                                                                    buttons
                                                                )


                                                        except Exception as e:
                                                            print(e)
                                                            return jsonify({"message": "Error approving leave application.", "error": str(e)}), 500


                                                    else:
                                                        pass

                                                elif button_id == "Menu" or selected_option == "Menu":

                                                    companyxx = company_reg.replace("_"," ").title()
                                                    
                                                    sections = [
                                                        {
                                                            "title": "MY PROFILE",
                                                            "rows": [
                                                                {"id": "Apply", "title": "Apply for Leave"},
                                                                {"id": "Track", "title": "Track My Application"},
                                                                {"id": "Checkbal", "title": "Check Days Balance"},
                                                                {"id": "myhist", "title": "My Applications History"},
                                                                {"id": "Myinfo", "title": "My Info"},
                                                                {"id": "Pending", "title": "Apps Pending My Approval"},
                                                            ]
                                                        },
                                                        {
                                                            "title": "ADMINISTRATION",
                                                            "rows": [
                                                                {"id": "Empmgt", "title": "Employee Management"},
                                                                {"id": "Analyticscomp", "title": "Analytics & Insights"},
                                                                {"id": "Company", "title": "Company Profile"},
                                                            ]
                                                        }
                                                    ]
                                                    
                                                    send_whatsapp_list_message(
                                                        sender_id,
                                                        f"Hello {first_name} {last_name}, LMS Administrator & Leave Applications Approver from {companyxx}!\n\n {bot} LMS Bot Here 😎. How can I assist you?",
                                                        "Admin/Approver Options",
                                                        sections
                                                    )

                                                elif button_id == "Cancelapp" :

                                                    table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                    table_name_apps_cancelled = f"{company_reg}appscancelled"

                                                    query = f"SELECT appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail , leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf FROM {table_name_apps_pending_approval} WHERE id = %s;"
                                                    cursor.execute(query, (id_user,))
                                                    result = cursor.fetchone()
                                                    if result:
                                                        (app_id, employee_number, first_name, surname, department, leave_type,  leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf) = result
                                                    
                                                        try:
                                                                status = "Cancelled"
                                                                statusdate = today_date
                                                            
                                                                insert_query = f"""
                                                                INSERT INTO {table_name_apps_cancelled} 
                                                                (appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus, statusdate)
                                                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                                                """

                                                                cursor.execute(insert_query, (
                                                                    app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, 
                                                                    approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                                                                    end_date, leave_days, leavedaysbalancebf, status, statusdate
                                                                ))
                                                                
                                                                connection.commit()
                                                                print("Insert successful!")

                                                        except Exception as e:
                                                            print("Error inserting data:", e)

                                                        # SQL query to delete or mark the leave as canceled
                                                        query = f"""DELETE FROM {table_name_apps_pending_approval} WHERE appid = %s"""
                                                        cursor.execute(query, (app_id,))
                                                        connection.commit()                                       

                                                        companyxx = company_reg.replace("_", " ").title()
                                                        buttons = [
                                                            {"type": "reply", "reply": {"id": f"Resubapp_{app_id}", "title": "ReSubmit Application"}},
                                                            {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                                                            {"type": "reply", "reply": {"id": "Main", "title": "Main Menu"}},
                                                        ]

                                                        send_whatsapp_message(sender_id, f"Hey {first_name} from {companyxx}! \n\n Your `{leave_type} Leave Application` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}` has been Cancelled successfully✅!\n\n"
                                                            "Select an option below to continue 👇",
                                                            buttons
                                                        )                                          
                                                    
                                                    else:
                                                        print("No record found for the user.")

                                                elif button_id == "Addemp":

                                                    buttons = [
                                                    {"type": "reply", "reply": {"id": "Addone", "title": "Manual Addition"}},
                                                    {"type": "reply", "reply": {"id": "Bulkadd", "title": "Bulk Addition"}},
                                                    {"type": "reply", "reply": {"id": f"Menu", "title": "Main Menu"}},
                                                    ]

                                                    send_whatsapp_message(
                                                        sender_id, 
                                                        f"Hey {first_name}, select an option below on how you want to add employees to your company's Leave Management System.",
                                                        buttons
                                                    )
                                            
                                                elif button_id == "Bulkadd":

                                                    buttons = [
                                                    {"type": "reply", "reply": {"id": "Uptemp", "title": "Upload Template"}},
                                                    {"type": "reply", "reply": {"id": "Downtemp", "title": "Download Template"}},
                                                    {"type": "reply", "reply": {"id": f"Menu", "title": "Main Menu"}},
                                                    ]

                                                    send_whatsapp_message(
                                                        sender_id, 
                                                        f"Hey {first_name}, select an option below on whether you want to upload an Excel temaplate already filled with employee details or you want to download the template to fill with Employee details.",
                                                        buttons
                                                    )

                                                elif button_id == "Uptemp":

                                                    buttons = [
                                                    {"type": "reply", "reply": {"id": "Downtemp", "title": "Download Template"}},
                                                    {"type": "reply", "reply": {"id": f"Menu", "title": "Main Menu"}},
                                                    ]

                                                    send_whatsapp_message(
                                                        sender_id,
                                                        "📁 Please attach 📎 the filled Excel template file now by sending it directly here.",
                                                        buttons
                                                    )

                                                elif button_id == "Downtemp":

                                                    companyxx = company_reg.replace("_", " ").title()

                                                    wb = openpyxl.Workbook()
                                                    ws = wb.active
                                                    ws.title = "Employee Details"

                                                    # Add headers
                                                    headers = [
                                                        "FirstName", "Surname", "WhatsApp", "Email", 
                                                        "Role", "Department", 
                                                        "Current Leave Days Balance", "Monthly Leave Days Accumulation"
                                                    ]
                                                    ws.append(headers)

                                                    # Style headers
                                                    dark_blue = "003366"
                                                    white = "FFFFFF"
                                                    for col in range(1, len(headers) + 1):
                                                        cell = ws.cell(row=1, column=col)
                                                        cell.fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
                                                        cell.font = Font(color=white, bold=True)

                                                    # Role dropdown (inline, short list)
                                                    role_dv = DataValidation(type="list", formula1='"Administrator,Ordinary User"', allow_blank=False)
                                                    ws.add_data_validation(role_dv)
                                                    for row in range(2, 500):
                                                        role_dv.add(ws[f"E{row}"])

                                                    # Department options (long list) - write to column Z
                                                    departments = [
                                                        "Human Resources and Administration",
                                                        "Finance and Accounting",
                                                        "Sales and Marketing",
                                                        "Sales and Distribution",
                                                        "Workshop and Maintenance",
                                                        "Operations and Production",
                                                        "Procurement and Purchasing",
                                                        "Customer Service and Support",
                                                        "IT and Digital Infrastructure",
                                                        "Risk Management",
                                                        "Legal and Compliance",
                                                        "Health, Safety and Environment",
                                                        "Research, Analytics and Reporting"
                                                    ]

                                                    for i, dept in enumerate(departments, start=1):
                                                        ws[f"R{i}"] = dept

                                                    # Department dropdown using cell range
                                                    dept_dv = DataValidation(type="list", formula1="=$R$1:$R$13", allow_blank=False)
                                                    ws.add_data_validation(dept_dv)
                                                    for row in range(2, 500):
                                                        dept_dv.add(ws[f"F{row}"])

                                                    # Hide the reference column
                                                    ws.column_dimensions['R'].hidden = True

                                                    # Save workbook to memory stream
                                                    output = io.BytesIO()
                                                    wb.save(output)
                                                    output.seek(0)

                                                    def send_whatsapp_excel_by_media_id(recipient_number, media_id, company_reg, reference_number=None, caption=None):
                                                        """Sends an Excel file via WhatsApp using the uploaded media ID"""
                                                        
                                                        filename = f"LMS Employee Addition Template {companyxx}.xlsx"
                                                        
                                                        url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                        headers = {
                                                            "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                            "Content-Type": "application/json"
                                                        }
                                                        
                                                        payload = {
                                                            "messaging_product": "whatsapp",
                                                            "to": recipient_number,
                                                            "type": "document",
                                                            "document": {
                                                                "id": media_id,
                                                                "filename": filename
                                                            }
                                                        }
                                                        
                                                        if caption:
                                                            payload["document"]["caption"] = caption

                                                        response = requests.post(url, headers=headers, json=payload)
                                                        response.raise_for_status()
                                                        return response.json()

                                                    def upload_excel_to_whatsapp(excel_bytes, company_reg, reference_number=None):
                                                        compxxy = company_reg.replace("_"," ").title()
                                                        filename = f"LMS Employee Addition Template {companyxx}.xlsx"

                                                        url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/media"
                                                        headers = {
                                                            "Authorization": f"Bearer {ACCESS_TOKEN}"
                                                        }

                                                        files = {
                                                            "file": (filename, io.BytesIO(excel_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                                                            "type": (None, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                                                            "messaging_product": (None, "whatsapp")
                                                        }

                                                        response = requests.post(url, headers=headers, files=files)
                                                        print("📊 Excel upload response:", response.text)
                                                        response.raise_for_status()

                                                        return response.json()["id"]

                                                    try:
                                                        # Get the Excel bytes
                                                        excel_bytes = output.getvalue()

                                                        # Upload Excel to WhatsApp and get media ID
                                                        media_id = upload_excel_to_whatsapp(
                                                            excel_bytes=excel_bytes,
                                                            company_reg=table_name
                                                        )

                                                        # Send Excel to user
                                                        send_whatsapp_excel_by_media_id(
                                                            recipient_number=sender_id,
                                                            media_id=media_id,
                                                            company_reg=table_name,
                                                            caption=f"Employee Registration Template - {companyxx}"
                                                        )

                                                        # Confirmation message with button
                                                        buttons = [
                                                            {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                        ]
                                                        send_whatsapp_message(
                                                            sender_id,
                                                            f"Your *Employee Registration Template* for *{companyxx}* is attached 📎.\n\nYou may fill it in (*Kindly use Microsoft Excel, NOT Google Sheets*) and upload it when ready.",
                                                            buttons
                                                        )

                                                    except Exception as e:
                                                        print("Error sending employee template:", str(e))
                                                        send_whatsapp_message(
                                                            sender_id,
                                                            f"Sorry, we couldn't send the Employee Template right now.\nError: {e}"
                                                        )
                                                                                                        
                                                elif selected_option in ["Annual","Sick","Study","Parental", "Bereavement","Other"] :
                                                    button_id_leave_type = str(selected_option)

                                                    cursor.execute("""
                                                        DELETE FROM whatsapptempapplication
                                                        WHERE empidwa = %s
                                                    """, (id_user,))  
                                                    
                                                    connection.commit()

                                                    cursor.execute(f"""
                                                        INSERT INTO whatsapptempapplication (empidwa, leavetypewa, companynamewa)
                                                        VALUES (%s, %s, %s)
                                                    """, (id_user, button_id_leave_type, company_reg))

                                                    connection.commit()

                                                    send_whatsapp_message(
                                                        sender_id, 
                                                        f"Ok. When would you like your {selected_option} Leave to start {first_name}?\n\n"
                                                        "Please enter your response using the format: 👇🏻\n"
                                                        "`start 24 january 2025`"
                                                    )

                                                    continue

                                                elif selected_option == "Book" or button_id == "Book":
                                                    
                                                    table_name = f"{company_reg}main"
                                                    appsapproved = f"{company_reg}appsapproved"
                                                    companyxx = company_reg.replace("_", " ").title()

                                                    query = f"SELECT id, firstname, surname, whatsapp, email, address ,role,currentleavedaysbalance, monthlyaccumulation, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp  FROM {table_name};"
                                                    cursor.execute(query)
                                                    rows = cursor.fetchall()

                                                    df_employees = pd.DataFrame(rows, columns=["ID","First Name", "Surname", "WhatsApp","Email", "Address", "Role","Leave Days Balance","Days Accumulated per Month","Leave Approver Name", "Leave Approver ID", "Leave Approver Email", "Leave Approver WhatsaApp"])
                                                    df_employees = df_employees.sort_values(by="ID", ascending=True)

                                                    query = f"SELECT appid, id, firstname, surname, leavetype, leaveapprovername, TO_CHAR(dateapplied, 'FMDD-Month-YYYY') AS dateapplied,  TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate,   TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate, leavedaysappliedfor,   TO_CHAR(statusdate, 'FMDD Month YYYY') AS statusdate, leavedaysbalancebf  FROM {appsapproved};"
                                                    cursor.execute(query)
                                                    rows2 = cursor.fetchall()

                                                    df_apps = pd.DataFrame(rows2, columns=["AppID","Emp ID", "First Name", "Surname", "Leave Type","Leave Approver Name", "Date Applied", "Leave Start Date", "Leave End Date","Leave Days Applied for","Date Approved","Leave Days Balance"])
                                                    df_apps = df_apps.sort_values(by="AppID", ascending=False)

                                                    print(df_employees)

                                                    df_apps['Leave Start Date'] = pd.to_datetime(df_apps['Leave Start Date'])
                                                    df_apps['Leave End Date'] = pd.to_datetime(df_apps['Leave End Date'])

                                                    # Function to expand dates and exclude Sundays
                                                    def expand_leave_days(row):
                                                        dates = pd.date_range(row['Leave Start Date'], row['Leave End Date'], freq='D')
                                                        # Exclude Sundays (weekday=6)
                                                        dates = [d for d in dates if d.weekday() != 6]
                                                        return dates

                                                    # Apply the function and explode the DataFrame
                                                    df_apps['Leave Dates'] = df_apps.apply(expand_leave_days, axis=1)
                                                    df_exploded = df_apps.explode('Leave Dates')

                                                    # Extract month and year for grouping
                                                    df_exploded['Month'] = df_exploded['Leave Dates'].dt.to_period('M')

                                                    # Group by Employee and Month
                                                    result = df_exploded.groupby(['Emp ID', 'First Name', 'Surname', 'Month']).size().reset_index(name='Leave Days Taken')

                                                    # Pivot to MoM format (months as columns)
                                                    mom_leave = result.pivot_table(
                                                        index=['Emp ID', 'First Name', 'Surname'],
                                                        columns='Month',
                                                        values='Leave Days Taken',
                                                        fill_value=0
                                                    ).reset_index()

                                                    # Rename columns for clarity
                                                    mom_leave.columns.name = None
                                                    mom_leave.columns = ['Emp ID', 'First Name', 'Surname'] + [f"{col.strftime('%b-%Y')}" for col in mom_leave.columns[3:]]

                                                    print(mom_leave)

                                                    def upload_excel_to_whatsapp(excel_bytes, company_reg, first_name, last_name, reference_number=None):
                                                        """Uploads an Excel file to WhatsApp servers and returns the media ID"""
                                                        compxxy = company_reg.replace("_"," ").title()
                                                        
                                                        filename = f"leave_records_{compxxy}.xlsx"
                                                        
                                                        url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/media"
                                                        headers = {
                                                            "Authorization": f"Bearer {ACCESS_TOKEN}"
                                                        }

                                                        files = {
                                                            "file": (filename, io.BytesIO(excel_bytes), 
                                                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                                                            "type": (None, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                                                            "messaging_product": (None, "whatsapp")
                                                        }

                                                        response = requests.post(url, headers=headers, files=files)
                                                        print("📊 Excel upload response:", response.text)  # Debugging
                                                        response.raise_for_status()
                                                        return response.json()["id"]

                                                    def send_whatsapp_excel_by_media_id(recipient_number, media_id, company_reg, first_name, last_name, reference_number=None, caption=None):
                                                        """Sends an Excel file via WhatsApp using the uploaded media ID"""
                                                        compxxy = company_reg.replace("_"," ").title()
                                                        
                                                        filename = f"leave_records_{compxxy}.xlsx"
                                                        
                                                        url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                                                        headers = {
                                                            "Authorization": f"Bearer {ACCESS_TOKEN}",
                                                            "Content-Type": "application/json"
                                                        }
                                                        
                                                        payload = {
                                                            "messaging_product": "whatsapp",
                                                            "to": recipient_number,
                                                            "type": "document",
                                                            "document": {
                                                                "id": media_id,
                                                                "filename": filename
                                                            }
                                                        }
                                                        
                                                        if caption:
                                                            payload["document"]["caption"] = caption

                                                        response = requests.post(url, headers=headers, json=payload)
                                                        response.raise_for_status()
                                                        return response.json()

                                                    output = io.BytesIO()
                                                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                                        df_employees.to_excel(writer, index=False, sheet_name=f'LMS Book {today_date}')
                                                        df_apps.to_excel(writer, index=False, sheet_name=f'All Approved')
                                                        mom_leave.to_excel(writer, index=False, sheet_name=f'Month on Month')

                                                    output.seek(0)
                                                    excel_bytes = output.getvalue()
                                                    
                                                    try:
                                                        media_id = upload_excel_to_whatsapp(
                                                            excel_bytes=excel_bytes,
                                                            company_reg=company_reg,
                                                            first_name=first_name,
                                                            last_name=last_name
                                                        )
                                                        
                                                        send_whatsapp_excel_by_media_id(
                                                            recipient_number=sender_id,
                                                            media_id=media_id,
                                                            company_reg=company_reg,
                                                            first_name=first_name,
                                                            last_name=last_name,
                                                            caption=f"{companyxx} Employee Leave Records as of {today_date}"
                                                        )
                                                        
                                                        buttons = [
                                                            {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                        ]
                                                        send_whatsapp_message(
                                                            sender_id, 
                                                            f"Hey there {first_name} {last_name}! You may go ahead and download Leave Book for {companyxx} attached here 😎.", 
                                                            buttons
                                                        )

                                                    except Exception as e:
                                                        print(f"Error sending Excel file: {str(e)}")
                                                        send_whatsapp_message(
                                                            sender_id,
                                                            f"Sorry {first_name}, we encountered an error preparing your document -- {e}. Please try again later."
                                                        )

                                                elif selected_option == "Pending" or button_id == "Pending":

                                                    table_name_apps_pending_approval = f"{company_reg}appspendingapproval"

                                                    query = f"SELECT id, leavetype, firstname, surname, leaveapprovername, leaveapproverid, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, appid  FROM {table_name_apps_pending_approval} WHERE leaveapproverid = {str(id_user)};"
                                                    cursor.execute(query)
                                                    rows = cursor.fetchall()

                                                    df_employeesappspendingcheck = pd.DataFrame(rows, columns=["id", "leavetype", "firstname", "surname", "leaveapprovername", "leaveapproverid", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor", "appid"])    
                                                    df_employeesappspendingcheck = df_employeesappspendingcheck.sort_values(by=df_employeesappspendingcheck.columns[10], ascending=False)

                                                    if len(df_employeesappspendingcheck) == 0:

                                                        companyxx = company_reg.replace("_", " ").title()
                                                        
                                                        sections = [
                                                            {
                                                                "title": "MY PROFILE",
                                                                "rows": [
                                                                    {"id": "Apply", "title": "Apply for Leave"},
                                                                    {"id": "Track", "title": "Track My Application"},
                                                                    {"id": "Checkbal", "title": "Check Days Balance"},
                                                                    {"id": "myhist", "title": "My Applications History"},
                                                                    {"id": "Myinfo", "title": "My Info"},
                                                                    {"id": "Pending", "title": "Apps Pending My Approval"},
                                                                ]
                                                            },
                                                            {
                                                                "title": "ADMINISTRATION",
                                                                "rows": [
                                                                    {"id": "Empmgt", "title": "Employee Management"},
                                                                    {"id": "Analyticscomp", "title": "Analytics & Insights"},
                                                                    {"id": "Company", "title": "Company Profile"},
                                                                ]
                                                            }
                                                        ]
        
                                                        send_whatsapp_list_message(
                                                            sender_id, 
                                                            f"{first_name}, there are currently no leave applications that are pending your approval.", 
                                                        "Admin/Approver Options",
                                                        sections) 

                                                    elif len(df_employeesappspendingcheck) > 0:

                                                        firstnameemp2 = df_employeesappspendingcheck.iat[0,2]
                                                        appid = df_employeesappspendingcheck.iat[0,10]
                                                        surnameemp2 = df_employeesappspendingcheck.iat[0,3]
                                                        leave_type2 = df_employeesappspendingcheck.iat[0,1]
                                                        days = df_employeesappspendingcheck.iat[0,9]
                                                        date_applied2 = df_employeesappspendingcheck.iat[0,6]
                                                        start_date2 = df_employeesappspendingcheck.iat[0,7]
                                                        end_date2 = df_employeesappspendingcheck.iat[0,8]

                                                        buttons = [
                                                            {"type": "reply", "reply": {"id": f"Approve5appwa_{appid}", "title": "Approve"}},
                                                            {"type": "reply", "reply": {"id": f"Disapproveappwa_{appid}", "title": "Disapprove"}},
                                                        ]

                                                        send_whatsapp_message(
                                                            sender_id, 
                                                            f"{firstnameemp2} {surnameemp2}'s {days} day {leave_type2} Leave Application, applied on {date_applied2.strftime('%d %B %Y')} and running from {start_date2.strftime('%d %B %Y')} to {end_date2.strftime('%d %B %Y')} is pending your Approval.\n\n" 
                                                            "Select an option below to either approve or disapprove this leave application.", 
                                                            buttons
                                                        )
                                                    
                                                elif selected_option == "RoleApprover":

                                                    sections = [
                                                        {
                                                            "title": "Role & Approver Options",
                                                            "rows": [
                                                                {"id": "Changerole", "title": "Edit Employee Role"},
                                                                {"id": "Changeappr", "title": "Edit Employee Approver"},
                                                                {"id": "RoleApprover", "title": "Role & Approver Schedule"},
                                                                {"id": "Menu", "title": "Main Menu"}
                                                            ]
                                                        }
                                                    ]

                                                    send_whatsapp_list_message(
                                                        sender_id, 
                                                        f"Hey {first_name}, kindly select an option below.",
                                                        "Role & Approver Options",
                                                        sections
                                                    )
                                                            
                                                elif selected_option == "Addrememp":

                                                    buttons = [
                                                    {"type": "reply", "reply": {"id": "Addemp", "title": "Add Employees"}},
                                                    {"type": "reply", "reply": {"id": "Rememp", "title": "Remove Employees"}},
                                                    {"type": "reply", "reply": {"id": "Menu", "title": "Main Menu"}},
                                                    ]

                                                    send_whatsapp_message(
                                                        sender_id, 
                                                        f"Hey {first_name}, select an option below on how you want to add or remove employees to or from your company's Leave Management System.",
                                                        buttons
                                                    )

                                            elif message.get("type") == "text":

                                                text = message.get("text", {}).get("body", "").lower()
                                                print(f"📨 Message from {sender_id}: {text}")
                                                
                                                if "hello" in text.lower():
                                                    companyxx = company_reg.replace("_"," ").title()
                                                    
                                                    sections = [
                                                        {
                                                            "title": "MY PROFILE",
                                                            "rows": [
                                                                {"id": "Apply", "title": "Apply for Leave"},
                                                                {"id": "Track", "title": "Track My Application"},
                                                                {"id": "Checkbal", "title": "Check Days Balance"},
                                                                {"id": "myhist", "title": "My Applications History"},
                                                                {"id": "Myinfo", "title": "My Info"},
                                                                {"id": "Pending", "title": "Apps Pending My Approval"},
                                                            ]
                                                        },
                                                        {
                                                            "title": "ADMINISTRATION",
                                                            "rows": [
                                                                {"id": "Empmgt", "title": "Employee Management"},
                                                                {"id": "Analyticscomp", "title": "Analytics & Insights"},
                                                                {"id": "Company", "title": "Company Profile"},
                                                            ]
                                                        }
                                                    ]

                                                    
                                                    send_whatsapp_list_message(
                                                        sender_id,
                                                        f"Hello {first_name} {last_name}, LMS Administrator & Leave Applications Approver from {companyxx}!\n\n {bot} LMS Bot Here 😎. How can I assist you?",
                                                        "Admin/Approver Options",
                                                        sections
                                                    )

                                                elif "email" in text.lower():

                                                    table_name = company_reg + "main"
                                                    table_name_apps_pending_approval = f"{company_reg}appspendingapproval"
                                                    table_name_apps_cancelled = f"{company_reg}appscancelled"
                                                    table_name_apps_approved = f"{company_reg}appsapproved"
                                                    table_name_apps_declined = f"{company_reg}appsdeclined"
                                                    table_name_apps_revoked = f"{company_reg}appsrevoked"

                                                    # Regex to extract email after the word "email"
                                                    match = re.search(r"email\s+([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)", text.strip(), re.IGNORECASE)

                                                    if not match:
                                                        raise ValueError("Invalid Email format")

                                                        sections = [
                                                            {
                                                                "title": "User Options",
                                                                "rows": [
                                                                    {"id": "Editname", "title": "Edit My Name"},
                                                                    {"id": "Editwhatsapp", "title": "Change My WhatsApp #"},
                                                                    {"id": "Editemail", "title": "Change My Email"},
                                                                    {"id": "Editwebpass", "title": "Change Web Password"},
                                                                    {"id": "Editaddress", "title": "Edit My Address"},
                                                                    {"id": "MyInfo", "title": "My Info"},
                                                                    {"id": "Menu", "title": "Main Menu"}
                                                                ]
                                                            }
                                                        ]

                                                        send_whatsapp_list_message(
                                                            sender_id, 
                                                            f"Hey {first_name}!\n You have provided an invalid Email Address. Enter your new email address starting with the word `email` as shown below 👇. \n\n `email epsilon@gmail.com`", 
                                                        "User Options",
                                                        sections)                


                                                    email = match.group(1)
                                                    print("Extracted email:", email)

                                                    try:

                                                        query = f"UPDATE {table_name} SET email = %s WHERE id = %s;"
                                                        cursor.execute(query, (email, id_user))
                                                        connection.commit()

                                                        query = f"UPDATE {table_name_apps_pending_approval} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                        cursor.execute(query, (email, id_user))
                                                        connection.commit()
                                            
                                                        query = f"UPDATE {table_name_apps_cancelled} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                        cursor.execute(query, (email, id_user))
                                                        connection.commit()

                                                        query = f"UPDATE {table_name_apps_approved} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                        cursor.execute(query, (email, id_user))
                                                        connection.commit()

                                                        query = f"UPDATE {table_name_apps_declined} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                        cursor.execute(query, (email, id_user))
                                                        connection.commit()

                                                        query = f"UPDATE {table_name_apps_revoked} SET leaveapproveremail = %s WHERE leaveapproverid = %s;"
                                                        cursor.execute(query, (email, id_user))
                                                        connection.commit()

                                                        sections = [
                                                            {
                                                                "title": "User Options",
                                                                "rows": [
                                                                    {"id": "Editname", "title": "Edit My Name"},
                                                                    {"id": "Editwhatsapp", "title": "Change My WhatsApp #"},
                                                                    {"id": "Editemail", "title": "Change My Email"},
                                                                    {"id": "Editwebpass", "title": "Change Web Password"},
                                                                    {"id": "Editaddress", "title": "Edit My Address"},
                                                                    {"id": "MyInfo", "title": "My Info"},
                                                                    {"id": "Menu", "title": "Main Menu"}
                                                                ]
                                                            }
                                                        ]

                                                        send_whatsapp_list_message(
                                                            sender_id, 
                                                            f"Hey {first_name}!\n Email Address Successfully Changed to `{email}`. Select an option below to proceed 👇.", 
                                                        "User Options",
                                                        sections)  

                                                    except Exception as e:
                                                        print(e)

                                                        

                                                elif "start" in text.lower():

                                                    try:
                                                        # Match: "start 20 july 2025"
                                                        match = re.match(r"start\s+(\d{1,2}\s+[a-zA-Z]+\s+\d{4})", text.strip(), re.IGNORECASE)
                                                        if not match:
                                                            raise ValueError("Invalid format")

                                                        date_part = match.group(1)
                                                        parsed_date = datetime.strptime(date_part, "%d %B %Y")  # Will raise ValueError if invalid

                                                        # ✅ Now it's safe to update the DB
                                                        cursor.execute("""
                                                            UPDATE whatsapptempapplication
                                                            SET startdate = %s
                                                            WHERE empidwa = %s
                                                        """, (date_part, id_user))
                                                        connection.commit()

                                                        cursor.execute("""
                                                            SELECT empidwa, leavetypewa FROM whatsapptempapplication
                                                            WHERE empidwa = %s
                                                        """, (id_user,))
                                                        result = cursor.fetchone()
                                                        leavetypewa = result[1] if result else "your"

                                                        send_whatsapp_message(sender_id,
                                                            f"✅ Got it! Start date saved.\n\nNow enter your last day on {leavetypewa} leave like this:\n"
                                                            "`end 28 July 2025`"
                                                        )

                                                    except ValueError:
                                                        send_whatsapp_message(
                                                            sender_id,
                                                            f"❌ Invalid start date message format, {first_name}. Please use the date format givem below 👇:\n"
                                                            "`start 24 january 2025`\n\n"
                                                            "Example: `start 15 march 2024`"
                                                        )

                                                    except Exception as e:
                                                        import traceback
                                                        print("🔴 Unexpected error:", e)
                                                        traceback.print_exc()

                                                        try:
                                                            send_whatsapp_message(
                                                                sender_id,
                                                                "⚠️ Something went wrong while processing your start date. Please try again or contact support."
                                                            )
                                                        except Exception as send_err:
                                                            print("🔴 Failed to send WhatsApp error message:", send_err)


                                                elif "end" in text.lower():

                                                    try:
                                                        # ✅ Match "end 24 january 2025"
                                                        match = re.match(r"end\s+(\d{1,2}\s+[a-zA-Z]+\s+\d{4})", text.strip(), re.IGNORECASE)
                                                        if not match:
                                                            raise ValueError("Invalid end date format.")

                                                        date_part = match.group(1)
                                                        parsed_end_date = datetime.strptime(date_part, "%d %B %Y").date()  # Will raise ValueError if invalid

                                                        # ✅ Update DB now that it's valid
                                                        cursor.execute("""
                                                            UPDATE whatsapptempapplication
                                                            SET enddate = %s
                                                            WHERE empidwa = %s
                                                        """, (date_part, id_user))
                                                        connection.commit()

                                                        # ✅ Fetch full leave application
                                                        cursor.execute("""
                                                            SELECT id, empidwa, leavetypewa, startdate, enddate FROM whatsapptempapplication
                                                            WHERE empidwa = %s
                                                        """, (id_user,))
                                                        result = cursor.fetchone()

                                                        if not result:
                                                            raise Exception("No leave record found.")

                                                        appid = result[0]
                                                        leavetype = result[2]
                                                        startdate = result[3]
                                                        enddate = result[4]

                                                        # ✅ Ensure both dates are datetime.date objects
                                                        if isinstance(startdate, str):
                                                            startdate = datetime.strptime(startdate, "%Y-%m-%d").date()
                                                        if isinstance(enddate, str):
                                                            enddate = datetime.strptime(enddate, "%Y-%m-%d").date()

                                                        # ✅ Calculate business days
                                                        business_days = 0
                                                        current_date = startdate
                                                        while current_date <= enddate:
                                                            if current_date.weekday() != 6:  # Weekday: Mon-Fri
                                                                business_days += 1
                                                            current_date += timedelta(days=1)

                                                        # ✅ Ask user to confirm submission
                                                        buttons = [
                                                            {"type": "reply", "reply": {"id": "Submitapp", "title": "Yes, Submit"}},
                                                            {"type": "reply", "reply": {"id": "Dontsubmit", "title": "No"}}
                                                        ]
                                                        send_whatsapp_message(
                                                            sender_id,
                                                            f"📝 Do you wish to submit your `{business_days}-day {leavetype} Leave Application` from "
                                                            f"`{startdate.strftime('%d %B %Y')}` to `{enddate.strftime('%d %B %Y')}`, {first_name}?",
                                                            buttons
                                                        )

                                                    except ValueError:
                                                        send_whatsapp_message(
                                                            sender_id,
                                                            f"❌ Invalid end date message format, {first_name}. Please use the date format givem below 👇:\n"
                                                            "`end 24 january 2025`\n\n"
                                                            "Example: `end 28 march 2024`"
                                                        )

                                                    except Exception as e:
                                                        import traceback
                                                        print("🔴 ERROR during end date processing:", e)
                                                        traceback.print_exc()
                                                        try:
                                                            send_whatsapp_message(
                                                                sender_id,
                                                                "⚠️ Something went wrong while processing your end date. Please try again or contact support."
                                                            )
                                                        except Exception as send_err:
                                                            print("🔴 Failed to send error message via WhatsApp:", send_err)
                                                            
                                                else:
                                                    send_whatsapp_message(
                                                        sender_id, 
                                                        f"{bot} LMS Bot Here 😎. Say 'hello' to start!"
                                                    )

                                            elif message.get("type") == "document":
                                                mime_type = message["document"]["mime_type"]
                                                filename = message["document"]["filename"]
                                                file_id = message["document"]["id"]

                                                def download_whatsapp_media(media_id):
                                                    media_url = f"https://graph.facebook.com/v19.0/{media_id}"
                                                    headers = {
                                                        "Authorization": f"Bearer {VERIFY_TOKEN}"
                                                    }

                                                    # Get the actual download URL
                                                    res = requests.get(media_url, headers=headers)
                                                    res.raise_for_status()
                                                    download_url = res.json()["url"]

                                                    # Download file content
                                                    file_response = requests.get(download_url, headers=headers)
                                                    file_response.raise_for_status()
                                                    return file_response.content


                                                if mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or filename.endswith(".xlsx"):
                                                    try:
                                                        file_bytes = download_whatsapp_media(file_id)

                                                        # Use BytesIO to read Excel from memory
                                                        excel_file = io.BytesIO(file_bytes)

                                                        df = pd.read_excel(excel_file)

                                                        print("yoooooooooooooooooooh upload!!!")
                                                        print(df)

                                                        send_whatsapp_message(sender_id, f"✅ Excel received. It contains {len(df)} employees to add.")

                                                    except Exception as e:
                                                        send_whatsapp_message(sender_id, f"❌ Error reading Excel file: {str(e)}")
                                                else:
                                                    send_whatsapp_message(sender_id, "⚠️ Unsupported file type. Please upload a valid `.xlsx` file.")


                                    except Exception as e:
                                        print(e)

                                        send_whatsapp_message(
                                            sender_id,
                                            f"{e}"
                                        )

                                        send_whatsapp_message(
                                            "+263774822568",
                                            f"{e}"
                                        )



            return jsonify({"status": "received"}), 200
            
        except Exception as e:
            print(f"error; {e}")


@app.route('/paynow/return')
def paynow_return():
    return "<h1>Thank you! Your payment is being verified.</h1>"


@app.route('/paynow/result/update', methods=['POST'])
def paynow_result():

    VERIFY_TOKENcc = "1412803596375322"
    ACCESS_TOKEN = "EAAUppTRo5q4BPATlxuMt4ZANFhgbyrtQI7iB1bR5FAI7K5Rv9yolg1OEwgt5J8xRJKKkTc2F9lHutvNcDXyHPEZAoGEuMQlv1THfAGRuTtZBEzmwbJG04f1sLxEAUFze09rHvmtuqa50ccT6ik2nm7cfcMOI8vn6id1PZBId5fMDf2WNZASQFIBIZBX6UIyTr3vVkaaTvIwO1ZB1ZAnQS6LUMtC6b14MZBeisR6XvHIvZBSSooWwZDZD"
    PHONE_NUMBER_IDcc = "618334968023252"

    data = request.form.to_dict()
    print("Paynow Result Webhook:", data)

    pollurlex = data.get('pollurl')
    status = data.get('status')
    ticketref = data.get('paynowreference')
    #today_date = datetime.now()

    cursor.execute("""
        SELECT id FROM cagwatick
        WHERE pollurl = %s
        ORDER BY id DESC
        LIMIT 1
    """, (pollurlex,))
    result = cursor.fetchone()

    if result:
        highest_id = result[0]
        cursor.execute("""
            UPDATE cagwatick
            SET status = %s, datebought = %s
            WHERE id = %s
        """, (status, today_date, highest_id))

        connection.commit()

    else:
        print("No row found for this sender_id.")

    cursor.execute("""
        SELECT id, idwanumber, route, time, paymethod, fare, ecocashnum, pollurl, status, datebought FROM cagwatick
        WHERE pollurl = %s
    """, (pollurlex,))
    result = cursor.fetchone()

    if result:

        status = result[8]
        print(status)

        if status.lower() == "paid":

            fare = result[5]
            route = result[2]
            time = result[3]
            sender_id = result[1]
            seat = random.randint(1, 65)

            url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_IDcc}/messages"
            headers = {
                "Authorization": f"Bearer {ACCESS_TOKEN}",
                "Content-Type": "application/json"
            }

            payload = {
                "messaging_product": "whatsapp",
                "to": f"263{sender_id}",
                "type": "interactive",
                "interactive": {
                    "type": "list",
                    "header": {
                        "type": "text",
                        "text": "🚍 CAG TOURS MAIN MENU"
                    },
                    "body": {
                        "text": (
                            f"Great News. You have successfully purchased a `USD {fare}` bus ticket for the `{route}` route. Your bus departs at `{time}` from Harare and you have been allocated seat number `{seat}`.\n Attached is you ticket [ticketref `{ticketref}`]\n\n Thank you!"
                        )
                    },
                    "action": {
                        "button": "📋 CAG TOURS MENU",
                        "sections": [
                            {
                                "title": "📦 CAG TOURS SERVICES",
                                "rows": [
                                    {
                                        "id": "book_ticket",
                                        "title": "Book a Ticket",
                                        "description": "Reserve your seat instantly"
                                    },
                                    {
                                        "id": "routes",
                                        "title": "View Routes",
                                        "description": "Get info regarding our travel routes"
                                    },
                                    {
                                        "id": "parcel_delivery",
                                        "title": "Parcel Delivery",
                                        "description": "Send or collect packages"
                                    },
                                    {
                                        "id": "find_stop",
                                        "title": "Find Bus Stop",
                                        "description": "Locate nearest pick-up point"
                                    },
                                    {
                                        "id": "promotions",
                                        "title": "Promotions & Offers",
                                        "description": "Current discounts & deals"
                                    }
                                ]
                            },
                            {
                                "title": "🚌 CAG TOURS",
                                "rows": [
                                    {
                                        "id": "know_more",
                                        "title": "Know More",
                                        "description": "Our story, mission & travel experience"
                                    },
                                    {
                                        "id": "why_choose",
                                        "title": "Why Choose Us",
                                        "description": "Luxury, safety & comfort explained"
                                    }
                                ]
                            },
                            {
                                "title": "🛎 CUSTOMER SERVICE",
                                "rows": [
                                    {
                                        "id": "faqs",
                                        "title": "❓ FAQs",
                                        "description": "Get answers to common questions"
                                    },
                                    {
                                        "id": "policies",
                                        "title": "Travel Policies",
                                        "description": "Baggage rules, safety, refunds"
                                    },
                                    {
                                        "id": "get_help",
                                        "title": "Get Help",
                                        "description": "Talk to a support agent now"
                                    }
                                ]
                            }
                        ]
                    }
                }
            }



            # Send the request to WhatsApp
            response = requests.post(url, headers=headers, json=payload)

            # Optional: Print result for debugging
            print(response.status_code)
            print(response.text)

            return "OK", 200


        else:

            fare = result[5]
            route = result[2]
            time = result[3]
            sender_id = result[1]


            url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_IDcc}/messages"
            headers = {
                "Authorization": f"Bearer {ACCESS_TOKEN}",
                "Content-Type": "application/json"
            }

            payload = {
                "messaging_product": "whatsapp",
                "to": f"263{sender_id}",
                "type": "interactive",
                "interactive": {
                    "type": "list",
                    "header": {
                        "type": "text",
                        "text": "🚍 CAG TOURS MAIN MENU"
                    },
                    "body": {
                        "text": (
                            f"Your attempt to purchase USD {fare} bus ticket for the {route} route failed, try again"
                        )
                    },
                    "action": {
                        "button": "📋 CAG TOURS MENU",
                        "sections": [
                            {
                                "title": "📦 CAG TOURS SERVICES",
                                "rows": [
                                    {
                                        "id": "book_ticket",
                                        "title": "Book a Ticket",
                                        "description": "Reserve your seat instantly"
                                    },
                                    {
                                        "id": "routes",
                                        "title": "View Routes",
                                        "description": "Get info regarding our travel routes"
                                    },
                                    {
                                        "id": "parcel_delivery",
                                        "title": "Parcel Delivery",
                                        "description": "Send or collect packages"
                                    },
                                    {
                                        "id": "find_stop",
                                        "title": "Find Bus Stop",
                                        "description": "Locate nearest pick-up point"
                                    },
                                    {
                                        "id": "promotions",
                                        "title": "Promotions & Offers",
                                        "description": "Current discounts & deals"
                                    }
                                ]
                            },
                            {
                                "title": "🚌 CAG TOURS",
                                "rows": [
                                    {
                                        "id": "know_more",
                                        "title": "Know More",
                                        "description": "Our story, mission & travel experience"
                                    },
                                    {
                                        "id": "why_choose",
                                        "title": "Why Choose Us",
                                        "description": "Luxury, safety & comfort explained"
                                    }
                                ]
                            },
                            {
                                "title": "🛎 CUSTOMER SERVICE",
                                "rows": [
                                    {
                                        "id": "faqs",
                                        "title": "❓ FAQs",
                                        "description": "Get answers to common questions"
                                    },
                                    {
                                        "id": "policies",
                                        "title": "Travel Policies",
                                        "description": "Baggage rules, safety, refunds"
                                    },
                                    {
                                        "id": "get_help",
                                        "title": "Get Help",
                                        "description": "Talk to a support agent now"
                                    }
                                ]
                            }
                        ]
                    }
                }
            }



            # Send the request to WhatsApp
            response = requests.post(url, headers=headers, json=payload)

            # Optional: Print result for debugging
            print(response.status_code)
            print(response.text)













    else:
        print("no result")



    return "OK", 200


def delete_all_tables():
    try:
        
        # Get all table names in the public schema
        cursor.execute("""
            SELECT tablename FROM pg_tables WHERE schemaname = 'public';
        """)
        tables = cursor.fetchall()

        # Loop through and drop each table
        for table in tables:
            table_name = table[0]
            cursor.execute(f"DROP TABLE IF EXISTS {table_name} CASCADE;")
            print(f"Dropped table: {table_name}")
            connection.commit()

        # Close the cursor and connection
        print("All tables deleted successfully.")

    except Exception as e:
        print("Error:", e)

# Run the function


def find_credentials(email, password):
    connection.reconnect()
    cursor = connection.cursor()
    
    try:
        cursor.execute("SHOW TABLES")
        tables = cursor.fetchall()
        
        for table in tables:
            table_name = table[0]
            query = f"""
            SELECT COUNT(*)
            FROM information_schema.columns
            WHERE table_schema = 'LMSuniversal_guarddoing'
            AND table_name = '{table_name}'
            AND column_name IN ('email', 'password')
            """
            cursor.execute(query)
            if cursor.fetchone()[0] == 2:  

                query = f"""
                SELECT * FROM {table_name}
                WHERE email = %s AND password = %s
                """
                cursor.execute(query, (email, password))
                result = cursor.fetchone()
                if result:
                    print(f"Credentials found in table: {table_name}")
                    return result  

    except: 
        print("No matching credentials found in any table.")
        return None


def generate_employees_remaining_chart(df_employees_lineg, df_apps_approved_lineg):
    df_employees = df_employees_lineg
    df_leaves = df_apps_approved_lineg

    today = datetime.today().date()
    next_30_days = today + timedelta(days=30)

    # Total employees per department
    total_by_dept = df_employees.groupby('department').size().to_dict()

    # Ensure leave dates are datetime.date
    df_leaves['leavestartdate'] = pd.to_datetime(df_leaves['leavestartdate']).dt.date
    df_leaves['leaveenddate'] = pd.to_datetime(df_leaves['leaveenddate']).dt.date

    result = {}
    all_dates = pd.date_range(start=today, end=next_30_days).date

    for dept, total_employees in total_by_dept.items():
        result[dept] = []
        df_dept_leaves = df_leaves[df_leaves['department'] == dept]

        for date in all_dates:
            # Count employees on leave on this date
            on_leave = df_dept_leaves[
                (df_dept_leaves['leavestartdate'] <= date) & (df_dept_leaves['leaveenddate'] >= date)
            ].shape[0]

            remaining = total_employees - on_leave

            result[dept].append({
                "date": date.strftime("%Y-%m-%d"),
                "remaining": remaining
            })
    print("CURRENT FIIIIIIIXXX")
    print(result)

    return result

def generate_employees_remaining_bar_chart(df_employees_lineg, df_apps_approved_lineg):
    df_employees = df_employees_lineg
    df_leaves = df_apps_approved_lineg

    today = datetime.today().date()
    next_30_days = today + timedelta(days=30)

    # Total employees per department
    total_by_dept = df_employees.groupby('department').size().to_dict()

    # Ensure leave dates are datetime.date
    df_leaves['leavestartdate'] = pd.to_datetime(df_leaves['leavestartdate']).dt.date
    df_leaves['leaveenddate'] = pd.to_datetime(df_leaves['leaveenddate']).dt.date

    result = {}
    all_dates = pd.date_range(start=today, end=next_30_days).date

    for dept, total_employees in total_by_dept.items():
        result[dept] = []
        df_dept_leaves = df_leaves[df_leaves['department'] == dept]

        for date in all_dates:
            # Count employees on leave on this date
            on_leave = df_dept_leaves[
                (df_dept_leaves['leavestartdate'] <= date) & (df_dept_leaves['leaveenddate'] >= date)
            ].shape[0]

            remaining = ((total_employees - on_leave)/total_employees) * 100

            result[dept].append({
                "date": date.strftime("%Y-%m-%d"),
                "remaining": remaining
            })
    print("CURRENT FIIIIIIIXXX")
    print(result)

    return result




def generate_leave_by_department_data(df_filtered_for_bar_chart):
    # Ensure 'Date Applied' is a datetime object
    df_filtered_for_bar_chart['Leave Start Date'] = pd.to_datetime(df_filtered_for_bar_chart['Leave Start Date'], format='%d %B %Y', errors='coerce')

    grouped = (
        df_filtered_for_bar_chart
        .groupby(['Department', 'Approval Status', 'Leave Start Date'])
        .size()
        .reset_index(name='count')
    )

    result = {}
    for _, row in grouped.iterrows():
        type = row['Department']
        status = row['Approval Status']
        date = row['Leave Start Date'].strftime('%Y-%m-%d')
        count = row['count']

        if type not in result:
            result[type] = {}
        if status not in result[type]:
            result[type][status] = []

        result[type][status].append({'date': date, 'count': count})

    return result


def generate_leave_by_type_data(df_filtered_for_bar_chart_type):
    # Ensure 'Date Applied' is a datetime object
    df_filtered_for_bar_chart_type['Leave Start Date'] = pd.to_datetime(df_filtered_for_bar_chart_type['Leave Start Date'], format='%d %B %Y', errors='coerce')

    grouped = (
        df_filtered_for_bar_chart_type
        .groupby(['Leave Type', 'Approval Status', 'Leave Start Date'])
        .size()
        .reset_index(name='count')
    )

    result = {}
    for _, row in grouped.iterrows():
        dept = row['Leave Type']
        status = row['Approval Status']
        date = row['Leave Start Date'].strftime('%Y-%m-%d')
        count = row['count']

        if dept not in result:
            result[dept] = {}
        if status not in result[dept]:
            result[dept][status] = []

        result[dept][status].append({'date': date, 'count': count})

    return result


def run1(table_name, empid):
    print(empid)

    query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation, department FROM {table_name};"
    cursor.execute(query)
    rows = cursor.fetchall()

    df_employees = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Address", "Role","Leave Approver Name","Leave Approver ID","Leave Approver Email", "Leave Approver WhatsAapp", "Leave Days Balance","Days Accumulated per Month","Department"])
    print(df_employees)
    employee_personal_details = df_employees[["id","firstname", "surname", "whatsapp","Email","Address"]]

    employee_personal_details['Action'] = employee_personal_details.apply(
        lambda row: f'''<div style="display: flex; gap: 10px;font-size: 12px;"><button class="btn btn-primary3 edit-emp-details-comp-btn" data-id="{row['id']}" data-firstname="{row['firstname']}" data-surname="{row['surname']}" data-whatsapp="{row['whatsapp']}" data-email="{row['Email']}" data-address="{row['Address']}">Edit Information</button></div>''', axis=1
    )

    employee_personal_details.columns = ["ID","FIRST NAME","SURNAME","WHATSAPP","EMAIL","ADDRESS","ACTION"]
    employee_personal_details_html = employee_personal_details.to_html(classes="table table-bordered table-theme", table_id="employeespersonalTable", index=False,  escape=False,)

    total_days_available = df_employees["Leave Days Balance"].sum()

    total_employees = len(df_employees)
    userdf = df_employees[df_employees['id'] == empid].reset_index()
    print("yeaarrrrr")
    print(userdf)
    firstname = userdf.iat[0,2]
    surname = userdf.iat[0,3]
    whatsapp = userdf.iat[0,4]
    address = userdf.iat[0,6]
    email = userdf.iat[0,5]
    fullnamedisp = firstname + ' ' + surname
    leaveapprovername = userdf.iat[0,8]
    leaveapproverid = userdf.iat[0,9]
    leaveapproveremail = userdf.iat[0, 10]
    leaveapproverwhatsapp = userdf.iat[0,11]
    role = userdf.iat[0,7]
    leavedaysbalance = userdf.iat[0,12]
    department = userdf.iat[0,14]

    print('check')

    df_employees['Employee Name'] = df_employees['firstname'] + ' ' + df_employees['surname']

    df_employees['Action'] = df_employees['id'].apply(
        lambda x: f'''<div style="display: flex; gap: 10px;font-size: 12px;"> <button class="btn btn-primary3 edit-priv-btn" data-bs-toggle="modal" data-bs-target="#editModalpriv" data-name="{x}"  data-ID="{x}">Edit Role</button> <button class="btn btn-primary3 edit-department-btn" data-bs-toggle="modal" data-bs-target="#editModaldepartment" data-name="{x}"  data-ID="{x}">Change Department</button>  <button class="btn btn-primary3 change-approver-btn" data-bs-toggle="modal" data-bs-target="#editModalapprover" data-name="{x}" data-ID="{x}">Change Approver</button>  <button class="btn btn-primary3 edit-balance-btn" data-bs-toggle="modal" data-bs-target="#editModalbalance" data-name="{x}" data-ID="{x}">Edit Balance</button> </div>'''
    )

    selected_columns = df_employees[['id','Employee Name', "Role", "Department", "Leave Approver Name", "Leave Days Balance", "Action"]]
    selected_columns.columns = ['ID','EMPLOYEE NAME','ROLE','DEPARTMENT','APPROVER','DAYS BALANCE','ACTION']

    table_employees_html = selected_columns.to_html(classes="table table-bordered table-theme", table_id="employeesTable", index=False,  escape=False,)

    selected_columns['Combined'] = selected_columns.apply(
        lambda row: f"{row['ID']}--{row['EMPLOYEE NAME']}", axis=1
    )

    employees_list = selected_columns['Combined'].tolist()

    selected_columns_accumulators = df_employees[['id','Employee Name', "Days Accumulated per Month"]]
    selected_columns_accumulators.columns = ['ID','EMPLOYEE NAME','DAYS ACCUMULATED PER MONTH']
    selected_columns_accumulators.loc[:, 'LEAVE DAYS ACCUMULATED PER MONTH'] = selected_columns_accumulators.apply(
        lambda row: f'<input type="number" step="0.5" class="editable-field" value="{row["DAYS ACCUMULATED PER MONTH"]:.1f}" data-id="{row["ID"]}" style="width: 100%;"/>'
        if row["DAYS ACCUMULATED PER MONTH"] is not None
        else f'<input type="number" step="0.5" class="editable-field" value="0.0" data-id="{row["ID"]}" style="width: 100%;"/>',
        axis=1
    )


    seacc = selected_columns_accumulators[['ID','EMPLOYEE NAME','LEAVE DAYS ACCUMULATED PER MONTH']]

    table_employees_accumulators_html = seacc.to_html(classes="table table-bordered table-theme", table_id="employeesaccumulatorsTable", index=False,  escape=False,)

    rememployees = selected_columns_accumulators[['ID','EMPLOYEE NAME']]
    rememployees.loc[:, 'SELECTION'] = rememployees.apply(
        lambda row: f'<input type="checkbox" class="custom-checkbox employee-checkbox" name="employee_ids" value="{row["ID"]}" data-employee-name="{row["EMPLOYEE NAME"]}">',
        axis=1
    )

    table_rememployees_html = rememployees.to_html(classes="table table-bordered table-theme", table_id="removeemployeesTable", index=False,  escape=False,)

    rememployees1 = selected_columns_accumulators[['ID','EMPLOYEE NAME']]
    rememployees1.loc[:, 'SELECTION'] = rememployees1.apply(
        lambda row: f'<input type="checkbox" class="custom-checkbox employee-checkbox-bulk-approver" name="employee_ids" value="{row["ID"]}" data-employee-name="{row["EMPLOYEE NAME"]}">',
        axis=1
    )
    table_rememployees_bulk1_html = rememployees1.to_html(classes="table table-bordered table-theme", table_id="employeesbulk1Table", index=False,  escape=False,)

    rememployees2 = selected_columns_accumulators[['ID','EMPLOYEE NAME']]
    rememployees2.loc[:, 'SELECTION'] = rememployees2.apply(
        lambda row: f'<input type="checkbox" class="custom-checkbox employee-checkbox-bulk-balances" name="employee_ids" value="{row["ID"]}" data-employee-name="{row["EMPLOYEE NAME"]}">',
        axis=1
    )
    table_rememployees_bulk_balances_html = rememployees2.to_html(classes="table table-bordered table-theme", table_id="employeesbulkbalancesTable", index=False,  escape=False,)

    rememployees3 = selected_columns_accumulators[['ID','EMPLOYEE NAME']]
    rememployees3.loc[:, 'SELECTION'] = rememployees3.apply(
        lambda row: f'<input type="checkbox" class="custom-checkbox employee-checkbox-bulk-accumulators" name="employee_ids" value="{row["ID"]}" data-employee-name="{row["EMPLOYEE NAME"]}">',
        axis=1
    )
    table_rememployees_bulk_accumulators_html = rememployees3.to_html(classes="table table-bordered table-theme", table_id="bulkemployeesbulkaccumulatorsTable", index=False,  escape=False,)

    company_name = table_name.replace("main", "")
    table_name_apps_pending_approval = f"{company_name}appspendingapproval"
    table_name_apps_approved = f"{company_name}appsapproved"
    table_name_apps_declined = f"{company_name}appsdeclined"
    table_name_apps_cancelled = f"{company_name}appscancelled"

    query = f"""SELECT appid, id, firstname, surname, leavetype, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate,  leavedaysappliedfor, leaveapprovername, approvalstatus FROM {table_name_apps_pending_approval};"""
    cursor.execute(query)
    rows = cursor.fetchall()
    df_leave_appsmain_pending_approval = pd.DataFrame(rows, columns=["App ID","ID","First Name", "Surname", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver","Approval Status"])
    df_leave_appsmain_pending_approval['Approval Status'] = '<p style="color: #ffab00; border: 3px solid #ffab00;border-radius: 9px;display: inline-block; margin: 0;padding: 0px 8px;">Pending</p>'
    df_leave_appsmain_pending_approvalcomb = df_leave_appsmain_pending_approval[["App ID","First Name", "Surname", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver","Approval Status"]]
    open_requests = len(df_leave_appsmain_pending_approval)
    print(df_leave_appsmain_pending_approval)

    query = f"""SELECT appid, id, firstname, surname, leavetype, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate,  leavedaysappliedfor, leaveapprovername, approvalstatus FROM {table_name_apps_approved};"""
    cursor.execute(query)
    rows = cursor.fetchall()
    df_leave_appsmain_approved = pd.DataFrame(rows, columns=["App ID","ID","First Name", "Surname", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver","Approval Status"])
    df_leave_appsmain_approved['Approval Status'] = '<p style="color: #28a745; border: 3px solid #28a745;border-radius: 9px;display: inline-block; margin: 0;padding: 0px 8px;">Approved</p>'
    df_leave_appsmain_approved['ACTION'] = df_leave_appsmain_approved['App ID'].apply(lambda x: f'''<div style="display: flex; gap: 10px;"><button class="btn btn-primary3 download-app-btn" data-ID="{x}" onclick="downloadLeaveApp('{x}')">Download</button></div>''')
    df_leave_appsmain_approvedcomb = df_leave_appsmain_approved[["App ID","First Name", "Surname", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver","Approval Status","ACTION"]]
    approved_requests = len(df_leave_appsmain_approved)

    if approved_requests > 0:

        total_leave_days = df_leave_appsmain_approved["Leave Days"].sum()
        top_leave_type = df_leave_appsmain_approved['Leave Type'].value_counts().idxmax()
        longest_leave_days = df_leave_appsmain_approved['Leave Days'].max()

        leave_utilization_rate = round((total_leave_days/ total_days_available) * 100,0)
        avg_leave_days = round(total_leave_days/total_employees,0)


    query = f"""SELECT dateapplied, statusdate FROM {table_name_apps_approved};"""
    cursor.execute(query)
    rows = cursor.fetchall()
    df_leave_appsmain_approved2 = pd.DataFrame(rows, columns=["Date Applied", "Status Date"])

    if len(df_leave_appsmain_approved2) > 0:

        df_leave_appsmain_approved2["Date Applied"] = pd.to_datetime(df_leave_appsmain_approved2["Date Applied"])
        df_leave_appsmain_approved2["Status Date"] = pd.to_datetime(df_leave_appsmain_approved2["Status Date"])

        # Function to count days excluding Sundays
        def count_days_excluding_sundays(start, end):
            return sum((start + pd.to_timedelta(i, unit='D')).weekday() != 6  # 6 = Sunday
                    for i in range((end - start).days))

        df_leave_appsmain_approved2["Days (No Sundays)"] = df_leave_appsmain_approved2.apply(
            lambda row: count_days_excluding_sundays(row["Date Applied"], row["Status Date"]),
            axis=1
        )

        avg_approval_time = round(df_leave_appsmain_approved2["Days (No Sundays)"].mean(),0)
        df_leave_appsmain_approved2['Date Applied'] = pd.to_datetime(df_leave_appsmain_approved2['Date Applied'])

        df_leave_appsmain_approved2['Month-Year'] = df_leave_appsmain_approved2['Date Applied'].dt.to_period('M')

        peak_leave_monthpp = df_leave_appsmain_approved2['Month-Year'].mode()[0]
        peak_leave_month = peak_leave_monthpp.to_timestamp().strftime('%B %Y')

    query = f"""SELECT appid, id, firstname, surname, department, leavetype, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate,  leavedaysappliedfor, leaveapprovername, approvalstatus FROM {table_name_apps_declined};"""
    cursor.execute(query)
    rows = cursor.fetchall()
    df_leave_appsmain_declined = pd.DataFrame(rows, columns=["App ID","ID","First Name", "Surname", "Department", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver","Approval Status"])
    df_leave_appsmain_declined_chart = df_leave_appsmain_declined
    df_leave_appsmain_declined_chart['Approval Status'] = "Declined"
    df_leave_appsmain_declined['Approval Status'] = '<p style="color: #E30022; border: 3px solid #E30022;border-radius: 9px;display: inline-block; margin: 0;padding: 0px 8px;">Declined</p>'
    df_leave_appsmain_declinedcomb = df_leave_appsmain_declined[["App ID","First Name", "Surname", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver","Approval Status"]]
    disapproved_requests = len(df_leave_appsmain_declined)

    approval_rate = ""

    if approved_requests>0 and disapproved_requests>0:

        approval_rate = round((approved_requests/(approved_requests + disapproved_requests)) * 100,0)

    elif approved_requests == 0:
        
        leave_utilization_rate = 0
        avg_leave_days = 0
        avg_approval_time = ""
        top_leave_type = ""
        longest_leave_days = 0
        peak_leave_month = ""

    query = f"""SELECT appid, id, firstname, surname, department, leavetype, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate,  leavedaysappliedfor, leaveapprovername, approvalstatus FROM {table_name_apps_cancelled};"""
    cursor.execute(query)
    rows = cursor.fetchall()
    df_leave_appsmain_cancelled = pd.DataFrame(rows, columns=["App ID","ID","First Name", "Surname", "Department", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver","Approval Status"])
    df_leave_appsmain_cancelled['Approval Status'] = '<p style="color: #E30022; border: 3px solid #E30022;border-radius: 9px;display: inline-block; margin: 0;padding: 0px 8px;">Cancelled</p>'
    df_leave_appsmain_cancelledcomb = df_leave_appsmain_cancelled[["App ID","First Name", "Surname", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver","Approval Status"]]


    queryxx = f"""SELECT appid, id, firstname, surname, department, leavetype, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate,  leavedaysappliedfor, leaveapprovername, approvalstatus FROM {table_name_apps_declined};"""
    cursor.execute(queryxx)
    rows = cursor.fetchall()
    df_leave_appsmain_declinedxx = pd.DataFrame(rows, columns=["App ID","ID","First Name", "Surname", "Department", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver","Approval Status"])
    df_leave_appsmain_declinedxx['Approval Status'] = "Declined"

    queryxy = f"""SELECT appid, id, firstname, surname, department, leavetype, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate,  leavedaysappliedfor, leaveapprovername, approvalstatus FROM {table_name_apps_pending_approval};"""
    cursor.execute(queryxy)
    rows = cursor.fetchall()
    df_leave_appsmain_pendingxx = pd.DataFrame(rows, columns=["App ID","ID","First Name", "Surname", "Department", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver","Approval Status"])
    df_leave_appsmain_pendingxx['Approval Status'] = "Pending"

    queryxyx = f"""SELECT appid, id, firstname, surname, department, leavetype, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate,  leavedaysappliedfor, leaveapprovername, approvalstatus FROM {table_name_apps_cancelled};"""
    cursor.execute(queryxyx)
    rows = cursor.fetchall()
    df_leave_appsmain_cancelledxx = pd.DataFrame(rows, columns=["App ID","ID","First Name", "Surname", "Department", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver","Approval Status"])
    df_leave_appsmain_cancelledxx['Approval Status'] = "Cancelled"

    queryyy = f"""SELECT appid, id, firstname, surname, department, leavetype, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate,  leavedaysappliedfor, leaveapprovername, approvalstatus FROM {table_name_apps_approved};"""
    cursor.execute(queryyy)
    rows = cursor.fetchall()
    df_leave_appsmain_approvedxx = pd.DataFrame(rows, columns=["App ID","ID","First Name", "Surname", "Department", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver","Approval Status"])
    df_leave_appsmain_approvedxx['Approval Status'] = "Approved"



    df_leave_appsmain_analysis1 = df_leave_appsmain_declinedxx._append(df_leave_appsmain_approvedxx)
    df_leave_appsmain_analysis2 = df_leave_appsmain_analysis1._append(df_leave_appsmain_pendingxx)
    df_leave_appsmain_analysis = df_leave_appsmain_analysis2._append(df_leave_appsmain_cancelledxx)


    df_filtered_for_bar_chart = df_leave_appsmain_analysis[['Department', 'Approval Status', 'Leave Start Date']]

    df_filtered_for_bar_chart_type = df_leave_appsmain_analysis[['Leave Type', 'Approval Status', 'Leave Start Date']]


    df_leave_appsmain1 = df_leave_appsmain_pending_approvalcomb._append(df_leave_appsmain_approvedcomb)
    df_leave_appsmain3 = df_leave_appsmain1._append(df_leave_appsmain_declinedcomb)
    df_leave_appsmain = df_leave_appsmain3._append(df_leave_appsmain_cancelledcomb)
    df_leave_appsmain['Employee Name'] = df_leave_appsmain['First Name'] + ' ' + df_leave_appsmain['Surname']
    df_leave_appsmain = df_leave_appsmain[["App ID","Employee Name", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver","Approval Status","ACTION"]].fillna('')

    table_leave_apps_html = df_leave_appsmain.to_html(classes="table table-bordered table-theme", table_id="leaveappsTable", index=False,  escape=False,)

    query = f"""SELECT appid, id, firstname, surname, leavetype, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate,  leavedaysappliedfor FROM {table_name_apps_pending_approval} WHERE leaveapproverid = {empid};"""
    cursor.execute(query)
    rows = cursor.fetchall()
    df_leave_apps_pending_my_approval = pd.DataFrame(rows, columns=["App ID","ID","First Name", "Surname", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days"])
    num_pending_my_approval = len(df_leave_apps_pending_my_approval)
    df_leave_apps_pending_my_approval['ACTION'] = df_leave_apps_pending_my_approval['App ID'].apply(lambda x: f'''<div style="display: flex; gap: 10px;"> <button class="btn btn-primary3 approve-app-btn" data-bs-toggle="modal" data-bs-target="#approveappModal" data-name="{x}" data-ID="{x}">Approve</button> <button class="btn btn-primary3 disapprove-app-btn" data-bs-toggle="modal" data-bs-target="#disapproveappModal" data-name="{x}" data-ID="{x}">Disapprove</button> </div>''') 
    df_leave_apps_pending_my_approval['Approval Status'] = '<p style="color: #ffab00; border: 3px solid #ffab00;border-radius: 9px;display: inline-block; margin: 0;padding: 0px 8px;">Pending</p>'    
    df_leave_apps_pending_my_approval_fin = df_leave_apps_pending_my_approval[["App ID","First Name", "Surname", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days", "Approval Status","ACTION"]]
    

    query = f"""SELECT appid, id, firstname, surname, leavetype, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate,  leavedaysappliedfor FROM {table_name_apps_approved} WHERE leaveapproverid = {empid};"""
    cursor.execute(query)
    rows = cursor.fetchall()
    df_leave_apps_approved_by_me= pd.DataFrame(rows, columns=["App ID","ID","First Name", "Surname", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days"])
    df_leave_apps_approved_by_me['ACTION'] = df_leave_apps_approved_by_me['App ID'].apply(lambda x: f'''<div style="display: flex; gap: 10px;"> <button class="btn btn-primary3 download-app-btn" data-name="{x}" data-ID="{x}" onclick="downloadLeaveApp('{x}')">Download</button><button class="btn btn-primary3 revokeapprover-app-btn" data-bs-toggle="modal" data-bs-target="#revokeapproverappModal" data-name="{x}" data-ID="{x}">Revoke</button></div>''') 
    df_leave_apps_approved_by_me['Approval Status'] = '<p style="color: #28a745; border: 3px solid #28a745;border-radius: 9px;display: inline-block; margin: 0;padding: 0px 8px;">Approved</p>'
    df_leave_apps_approved_by_me = df_leave_apps_approved_by_me[["App ID","First Name", "Surname", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days", "Approval Status","ACTION"]]

    query = f"""SELECT appid, id, firstname, surname, leavetype, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate,  leavedaysappliedfor FROM {table_name_apps_declined} WHERE leaveapproverid = {empid};"""
    cursor.execute(query)
    rows = cursor.fetchall()
    df_leave_apps_declined_by_me= pd.DataFrame(rows, columns=["App ID","ID","First Name", "Surname", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days"])
    df_leave_apps_declined_by_me['ACTION'] = df_leave_apps_declined_by_me['App ID'].apply(lambda x: f'''<div style="display: flex; gap: 10px;"> <button class="btn btn-primary3 revive-app-btn" data-bs-toggle="modal" data-bs-target="#reviveappModal" data-name="{x}" data-ID="{x}">Revive</button></div>''') 
    df_leave_apps_declined_by_me['Approval Status'] = '<p style="color: #E30022; border: 3px solid #E30022;border-radius: 9px;display: inline-block; margin: 0;padding: 0px 8px;">Declined</p>'
    df_leave_apps_declined_by_me = df_leave_apps_declined_by_me[["App ID","First Name", "Surname", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days", "Approval Status","ACTION"]]

    df_leave_apps_approved_declined_by_me = df_leave_apps_approved_by_me._append(df_leave_apps_declined_by_me)
    df_leave_apps_approved_declined_pending_by_me = df_leave_apps_approved_declined_by_me._append(df_leave_apps_pending_my_approval_fin).fillna('')   
    df_leave_apps_approved_declined_pending_by_me['Employee Name'] = df_leave_apps_approved_declined_pending_by_me['First Name'] + ' ' + df_leave_apps_approved_declined_pending_by_me['Surname']
    df_leave_apps_approved_declined_pending_by_me = df_leave_apps_approved_declined_pending_by_me[["App ID","Employee Name", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days", "Approval Status","ACTION"]]

    table_leave_apps_approved_by_me_html = df_leave_apps_approved_declined_pending_by_me.to_html(classes="table table-bordered table-theme", table_id="leaveappsTableapprovedbyme", index=False,  escape=False,)
 
    query = f"""SELECT id, appid, leavetype, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate,  leavedaysappliedfor, leaveapprovername FROM {table_name_apps_approved};"""
    cursor.execute(query)
    rows = cursor.fetchall()
    df_my_leave_apps_approved = pd.DataFrame(rows, columns=["ID","App ID", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver"])
    df_my_leave_apps_approved = df_my_leave_apps_approved[df_my_leave_apps_approved['ID'] == empid]
    print("ahhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhh")
    print(df_my_leave_apps_approved)
    df_my_leave_apps_approved['Approval Status'] = '<p style="color: #28a745; border: 3px solid #28a745;border-radius: 9px;display: inline-block; margin: 0;padding: 0px 8px;">Approved</p>'
    df_my_leave_apps_approved['ACTION'] = df_my_leave_apps_approved['App ID'].apply(lambda x: f'''<div style="display: flex; gap: 10px;"><button class="btn btn-primary3 download-app-btn" data-ID="{x}" onclick="downloadLeaveApp('{x}')">Download</button><button class="btn btn-primary3 revoke-app-btn" data-bs-toggle="modal" data-bs-target="#revokeappModal" data-name="{x}" data-ID="{x}">Revoke</button></div>''')
    df_my_leave_apps_approved_approved_fin =  df_my_leave_apps_approved[["App ID", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Approval Status","Leave Approver","ACTION"]]

    query = f"""SELECT id, appid, leavetype, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate,  leavedaysappliedfor, leaveapprovername FROM {table_name_apps_declined};"""
    cursor.execute(query)
    rows = cursor.fetchall()
    df_my_leave_apps_declined = pd.DataFrame(rows, columns=["ID","App ID", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver"])
    df_my_leave_apps_declined = df_my_leave_apps_declined[df_my_leave_apps_declined['ID'] == empid]
    print(df_my_leave_apps_declined)
    df_my_leave_apps_declined['Approval Status'] = '<p style="color: #E30022; border: 3px solid #E30022;border-radius: 9px;display: inline-block; margin: 0;padding: 0px 8px;">Declined</p>'
    df_my_leave_apps_declined['ACTION'] =  df_my_leave_apps_declined['App ID'].apply(lambda x: f'''<div style="display: flex; gap: 10px;"> <button class="btn btn-primary3 reapply-app-btn" data-bs-toggle="modal" data-bs-target="#reapplyappModal" data-name="{x}" data-ID="{x}">Re-Apply</button>''') 
    df_my_leave_apps_declined_fin_declined =  df_my_leave_apps_declined[["App ID", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Approval Status","Leave Approver","ACTION"]]

    query = f"""SELECT id, appid, leavetype, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate,  leavedaysappliedfor, leaveapprovername FROM {table_name_apps_cancelled};"""
    cursor.execute(query)
    rows = cursor.fetchall()
    df_my_leave_apps_cancelled = pd.DataFrame(rows, columns=["ID","App ID", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver"])
    df_my_leave_apps_cancelled = df_my_leave_apps_cancelled[df_my_leave_apps_cancelled['ID'] == empid]
    print("ahhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhh")
    print(df_my_leave_apps_cancelled)
    df_my_leave_apps_cancelled['Approval Status'] = '<p style="color: #E30022; border: 3px solid #E30022;border-radius: 9px;display: inline-block; margin: 0;padding: 0px 8px;">Cancelled</p>'
    df_my_leave_apps_cancelled['ACTION'] =  df_my_leave_apps_cancelled['App ID'].apply(lambda x: f'''<div style="display: flex; gap: 10px;"> <button class="btn btn-primary3 reapply-app-btn" data-bs-toggle="modal" data-bs-target="#reapplyappModal" data-name="{x}" data-ID="{x}">Re-Apply</button>''') 
    df_my_leave_apps_cancelled_fin =  df_my_leave_apps_cancelled[["App ID", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Approval Status","Leave Approver","ACTION"]]

    df_my_leave_apps_approved_declined_fin1 = df_my_leave_apps_approved_approved_fin._append(df_my_leave_apps_declined_fin_declined)
    df_my_leave_apps_approved_declined_fin = df_my_leave_apps_approved_declined_fin1._append(df_my_leave_apps_cancelled_fin)

    userleaveapppending = df_leave_appsmain_pending_approval[df_leave_appsmain_pending_approval['ID'] == empid].reset_index()
    userleaveapppending['ACTION'] = userleaveapppending['App ID'].apply(lambda x: f'''<div style="display: flex; gap: 10px;"> <button class="btn btn-primary3 edit-priv-btn" data-bs-toggle="modal" data-bs-target="#remindapproverModal" data-name="{x}" data-ID="{x}">Remind Approver</button> <button class="btn btn-primary3 cancel-app-btn" data-bs-toggle="modal" data-bs-target="#cancelappModal" data-name="{x}" data-ID="{x}">Cancel Application</button> </div>''') 
    userleaveapppending = userleaveapppending[["App ID","Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver","Approval Status", "ACTION"]]
    df_my_leave_apps_approved_declined_pending_fin = df_my_leave_apps_approved_declined_fin._append(userleaveapppending)

    table_my_leave_apps_html = df_my_leave_apps_approved_declined_pending_fin.to_html(classes="table table-bordered table-theme", table_id="myleaveappsTable", index=False,  escape=False,)

    def generate_leave_status_chart():

        def categorize_status(status):
            if "pending" in status.lower():
                return "Pending"
            elif "declined" in status.lower():
                return "Declined"
            elif "approved" in status.lower():
                return "Approved"
            elif "cancelled" in status.lower():
                return "Cancelled"
            return status  # Return unchanged if no match

        df_my_leave_apps_approved_declined_pending_fin["Approval Status"] = df_my_leave_apps_approved_declined_pending_fin["Approval Status"].apply(categorize_status)
        print("ragaaaaaaaaaaaaaaaa")
        print(df_my_leave_apps_approved_declined_pending_fin)
        
        status_counts = df_my_leave_apps_approved_declined_pending_fin["Approval Status"].value_counts().to_dict()
        return status_counts  # Return as dictionary

 
    if not approval_rate:
        approval_rate = ""









    query = f"""SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate,leavedaysbalancebf, department FROM {table_name_apps_approved}"""
    cursor.execute(query)
    rowsxxyy = cursor.fetchall()

    df_apps_approved_lineg = pd.DataFrame(rowsxxyy, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate","leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf","department"])

    startdate_lineg = datetime.today().date()
    enddate_lineg = startdate_lineg + timedelta(days=30)


    query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation, department FROM {table_name};"
    cursor.execute(query)
    rowsxxzz = cursor.fetchall()

    df_employees_lineg = pd.DataFrame(rowsxxzz, columns=["id","firstname", "surname", "whatsapp","email", "address", "role","leaveapprovername","leaveapproverid","leaveapproveremail", "leaveapproverwhatsapp", "currentleavedaysbalance","monthlyaccumulation","department"])
   











    return {
        "table_my_leave_apps_html": table_my_leave_apps_html,
        "table_leave_apps_approved_by_me_html": table_leave_apps_approved_by_me_html,
        "num_pending_my_approval": num_pending_my_approval,
        "table_leave_apps_html": table_leave_apps_html,
        "table_employees_accumulators_html": table_employees_accumulators_html,
        "table_rememployees_bulk_balances_html": table_rememployees_bulk_balances_html,
        "table_rememployees_bulk_accumulators_html":  table_rememployees_bulk_accumulators_html,
        "table_rememployees_html": table_rememployees_html,
        "table_rememployees_bulk1_html": table_rememployees_bulk1_html,
        "employees_list": employees_list,
        "role": role,
        "total_employees": total_employees,
        "open_requests": open_requests,
        "approved_requests":  approved_requests,
        "leave_utilization_rate": leave_utilization_rate,
        "avg_leave_days": avg_leave_days,
        "avg_approval_time": avg_approval_time,
        "approval_rate": approval_rate,
        "top_leave_type": top_leave_type,
        "longest_leave_days": longest_leave_days,
        "peak_leave_month": peak_leave_month,
        "department": department,
        "firstname": firstname,
        "surname": surname,
        "fullnamedisp": fullnamedisp,
        "email": email,
        "whatsapp": whatsapp,
        "address": address,
        "table_employees_html": table_employees_html,
        "employee_personal_details_html": employee_personal_details_html,
        "today_date": today_date,
        "leaveapprovername": leaveapprovername,
        "leaveapproverid": leaveapproverid,
        "leavedaysbalance": leavedaysbalance,
        "leaveapproveremail": leaveapproveremail,
        "leaveapproverwhatsapp": leaveapproverwhatsapp,
        "leave_status_chart": generate_leave_status_chart(),  
        "leave_by_department_data": generate_leave_by_department_data(df_filtered_for_bar_chart),
        "leave_by_type_data": generate_leave_by_type_data(df_filtered_for_bar_chart_type),
        "empsremainingbydpt": generate_employees_remaining_chart(df_employees_lineg, df_apps_approved_lineg),
        "empsremainingbydptbar": generate_employees_remaining_bar_chart(df_employees_lineg, df_apps_approved_lineg),
    }



def check_existing_data(df, table_name):
    cursor.execute(f"SELECT whatsapp, email FROM {table_name}")
    existing_data = cursor.fetchall()

    existing_whatsapps = [data[0] for data in existing_data]
    existing_emails = [data[1] for data in existing_data]

    existing_whatsapps = [data[0] for data in existing_data]
    existing_emails = [data[1] for data in existing_data]

    df = df[~df['WhatsApp'].isin(existing_whatsapps)]  
    df = df[~df['Email'].isin(existing_emails)]  

    return df

app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}

if connection.status == psycopg2.extensions.STATUS_READY:
    cursor = connection.cursor()
        
    def allowed_file(filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


    @app.route('/upload-excel', methods=['POST'])
    def upload_excel():

        user_uuid = session.get('user_uuid')
        if user_uuid:

            table_name = session.get('table_name')
            empid = session.get('empid')

            if 'file' not in request.files:
                return jsonify({"status": "error", "message": "No file part"}), 400
            
            file = request.files['file']
            
            if file.filename == '':
                return jsonify({"status": "error", "message": "No selected file"}), 400
            
            if file and allowed_file(file.filename):

                df = pd.read_excel(file, usecols=range(8))
                df = check_existing_data(df, table_name)
                df = df.dropna(subset=['FirstName'])

                print(df)

                if len(df) > 0:

                    for index, row in df.iterrows():
                        first_name = row['FirstName']
                        surname = row['Surname']

                        if pd.notna(row['WhatsApp']):
                            whatsapp_raw = str(int(float(row['WhatsApp']))).replace(" ", "")
                        else:
                            print("NaN skipped")

                        whatsapp = whatsapp_raw[-9:] if len(whatsapp_raw) >= 9 else whatsapp_raw
                        email = row['Email']
                        role = row['Role']
                        department = row['Department']
                        current_leave_days_balance = row['Current Leave Days Balance']
                        monthly_accumulation = row['Monthly Leave Days Accumulation']

                        cursor.execute(f"""
                            INSERT INTO {table_name} (firstname, surname, whatsapp, email, role, department, currentleavedaysbalance, monthlyaccumulation)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """, (first_name, surname, whatsapp, email, role, department, current_leave_days_balance, monthly_accumulation))
                    
                    connection.commit()

                return redirect(url_for('Dashboard'))
    
        else:
            return redirect(url_for('landingpage')) 



    @app.route('/download-excel-template-add-employees')
    def download_excel_template_add_employees():

        user_uuid = session.get('user_uuid')
        if not user_uuid:
            return redirect(url_for('landingpage'))
        

        try:

            table_name = session.get('table_name')
            company_name = table_name.replace("main", "")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Employee Details"

            # Add headers
            headers = [
                "FirstName", "Surname", "WhatsApp", "Email", 
                "Role", "Department", 
                "Current Leave Days Balance", "Monthly Leave Days Accumulation"
            ]
            ws.append(headers)

            # Style headers
            dark_blue = "003366"
            white = "FFFFFF"
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
                cell.font = Font(color=white, bold=True)

            # Role dropdown (inline, short list)
            role_dv = DataValidation(type="list", formula1='"Administrator,Ordinary User"', allow_blank=False)
            ws.add_data_validation(role_dv)
            for row in range(2, 500):
                role_dv.add(ws[f"E{row}"])

            # Department options (long list) - write to column Z
            departments = [
                "Human Resources and Administration",
                "Finance and Accounting",
                "Sales and Marketing",
                "Sales and Distribution",
                "Workshop and Maintenance",
                "Operations and Production",
                "Procurement and Purchasing",
                "Customer Service and Support",
                "IT and Digital Infrastructure",
                "Risk Management",
                "Legal and Compliance",
                "Health, Safety and Environment",
                "Research, Analytics and Reporting"
            ]

            for i, dept in enumerate(departments, start=1):
                ws[f"Z{i}"] = dept

            # Department dropdown using cell range
            dept_dv = DataValidation(type="list", formula1="=$Z$1:$Z$13", allow_blank=False)
            ws.add_data_validation(dept_dv)
            for row in range(2, 500):
                dept_dv.add(ws[f"F{row}"])

            # Hide the reference column
            ws.column_dimensions['Z'].hidden = True

            # Save workbook to memory stream
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"{company_name.strip()}_Employee_Details_Template.xlsx"
            return send_file(output, download_name=filename, as_attachment=True)
        
        except Error as e:
            print(e)

            return redirect(url_for('Dashboard'))



    @app.route('/admin_sign_up', methods=['POST'])
    def submit_form():
        global password
        try:

            external_database_url = "postgresql://lmsdatabase_8ag3_user:6WD9lOnHkiU7utlUUjT88m4XgEYQMTLb@dpg-ctp9h0aj1k6c739h9di0-a.oregon-postgres.render.com/lmsdatabase_8ag3"
            database = 'lmsdatabase_8ag3'

            connection = psycopg2.connect(external_database_url)

            cursor = connection.cursor()

            company_name_w_space = request.form.get('company_name')
            company_name = company_name_w_space.replace(' ', '_')
            firstname = request.form.get('firstname')
            surname = request.form.get('surname')

            whatsapp_raw = str(int(float(request.form['whatsapp']))).replace(" ", "")
            whatsapp = whatsapp_raw[-9:] if len(whatsapp_raw) >= 9 else whatsapp_raw
            email = request.form.get('email')
            password = request.form.get('password')

            table_name = f"{company_name}main"
            table_name_apps_pending_approval = f"{company_name}appspendingapproval"
            table_name_apps_cancelled = f"{company_name}appscancelled"
            table_name_apps_approved = f"{company_name}appsapproved"
            table_name_apps_declined = f"{company_name}appsdeclined"
            table_name_apps_revoked = f"{company_name}appsrevoked"

            check_table_query = f"""
            SELECT COUNT(*) 
            FROM information_schema.tables 
            WHERE table_schema = %s AND table_name = %s;
            """
            cursor.execute(check_table_query, (database, table_name))
            table_exists = cursor.fetchone()[0]

            if table_exists:
                print(f"Table `{table_name}` already exists. Skipping creation.")

                return render_template('index.html')  

            else:

                try:

                    create_table_query = f"""
                    CREATE TABLE {table_name} (
                        id SERIAL PRIMARY KEY,
                        firstname VARCHAR(100),
                        surname VARCHAR(100),
                        whatsapp INT,
                        address VARCHAR(100),
                        email VARCHAR(255),
                        password VARCHAR(255),
                        department VARCHAR(255),
                        role VARCHAR(255),
                        leaveapprovername VARCHAR(255),
                        leaveapproverid INT,
                        leaveapproveremail VARCHAR(255),
                        leaveapproverwhatsapp INT,
                        currentleavedaysbalance NUMERIC(5, 1),
                        monthlyaccumulation NUMERIC(5, 1)
                    );
                    """
                    cursor.execute(create_table_query)
                    connection.commit()
                    print(f"Table `{table_name}` created successfully!")


                    create_table_query = f"""
                    CREATE TABLE {table_name_apps_pending_approval} (
                        appid SERIAL PRIMARY KEY,
                        id INT,
                        firstname VARCHAR(100),
                        surname VARCHAR(100),
                        department VARCHAR(255),
                        leavetype VARCHAR(255),
                        reasonifother VARCHAR(300),
                        leaveapprovername VARCHAR(255),
                        leaveapproverid INT,
                        leaveapproveremail VARCHAR(255),
                        leaveapproverwhatsapp INT,
                        currentleavedaysbalance NUMERIC(5, 1),
                        dateapplied date,
                        leavestartdate date,
                        leaveenddate date,
                        leavedaysappliedfor INT,
                        leavedaysbalancebf NUMERIC(5, 1),
                        approvalstatus VARCHAR(255)
                    );
                    """
                    cursor.execute(create_table_query)
                    connection.commit()
                    print(f"Table `{table_name_apps_pending_approval}` created successfully!")


                    create_table_query = f"""
                    CREATE TABLE {table_name_apps_cancelled} (
                        appid INT,
                        id INT,
                        firstname VARCHAR(100),
                        surname VARCHAR(100),
                        department VARCHAR(255),
                        leavetype VARCHAR(255),
                        reasonifother VARCHAR(300),
                        leaveapprovername VARCHAR(255),
                        leaveapproverid INT,
                        leaveapproveremail VARCHAR(255),
                        leaveapproverwhatsapp INT,
                        currentleavedaysbalance NUMERIC(5, 1),
                        dateapplied date,
                        leavestartdate date,
                        leaveenddate date,
                        leavedaysappliedfor INT,
                        leavedaysbalancebf NUMERIC(5, 1),
                        approvalstatus VARCHAR(255),
                        statusdate date
                    );
                    """
                    cursor.execute(create_table_query)
                    connection.commit()
                    print(f"Table `{table_name_apps_cancelled}` created successfully!")




                    create_table_query = f"""
                    CREATE TABLE {table_name_apps_approved} (
                        appid INT,
                        id INT,
                        firstname VARCHAR(100),
                        surname VARCHAR(100),
                        department VARCHAR(255),
                        leavetype VARCHAR(255),
                        reasonifother VARCHAR(300),
                        leaveapprovername VARCHAR(255),
                        leaveapproverid INT,
                        leaveapproveremail VARCHAR(255),
                        leaveapproverwhatsapp INT,
                        currentleavedaysbalance NUMERIC(5, 1),
                        dateapplied date,
                        leavestartdate date,
                        leaveenddate date,
                        leavedaysappliedfor INT,
                        leavedaysbalancebf NUMERIC(5, 1),
                        approvalstatus VARCHAR(255),
                        statusdate date         
                    );
                    """
                    cursor.execute(create_table_query)
                    connection.commit()
                    print(f"Table `{table_name_apps_approved}` created successfully!")


                    create_table_query = f"""
                    CREATE TABLE {table_name_apps_declined} (
                        appid INT,
                        id INT,
                        firstname VARCHAR(100),
                        surname VARCHAR(100),
                        department VARCHAR(255),
                        leavetype VARCHAR(255),
                        reasonifother VARCHAR(300),
                        leaveapprovername VARCHAR(255),
                        leaveapproverid INT,
                        leaveapproveremail VARCHAR(255),
                        leaveapproverwhatsapp INT,
                        currentleavedaysbalance NUMERIC(5, 1),
                        dateapplied date,
                        leavestartdate date,
                        leaveenddate date,
                        leavedaysappliedfor INT,
                        leavedaysbalancebf NUMERIC(5, 1),
                        approvalstatus VARCHAR(255),
                        statusdate date         
                    );
                    """
                    cursor.execute(create_table_query)
                    connection.commit()
                    print(f"Table `{table_name_apps_declined}` created successfully!")


                    create_table_query = f"""
                    CREATE TABLE {table_name_apps_revoked} (
                        appid INT,
                        id INT,
                        firstname VARCHAR(100),
                        surname VARCHAR(100),
                        department VARCHAR(255),
                        leavetype VARCHAR(255),
                        reasonifother VARCHAR(300),
                        leaveapprovername VARCHAR(255),
                        leaveapproverid INT,
                        leaveapproveremail VARCHAR(255),
                        leaveapproverwhatsapp INT,
                        currentleavedaysbalance NUMERIC(5, 1),
                        dateapplied date,
                        leavestartdate date,
                        leaveenddate date,
                        leavedaysappliedfor INT,
                        leavedaysbalancebf NUMERIC(5, 1),
                        approvalstatus VARCHAR(255),
                        statusdate date         
                    );
                    """
                    cursor.execute(create_table_query)
                    connection.commit()
                    print(f"Table `{table_name_apps_revoked}` created successfully!")


                except psycopg2.errors.DuplicateTable:
                    connection.rollback()  # Roll back the failed CREATE
                    print("Table already exists, redirecting...")
                    return render_template('index.html')  
                    # Call your other function here



                admin = "Administrator"
                company_name_compreg = table_name.lower()


                insert_query_compreg = f"""
                INSERT INTO companyreg (companyname, datecreated)
                VALUES (%s, %s);
                """
                cursor.execute(insert_query_compreg, (company_name_compreg, today_date))
                connection.commit()

                insert_query = f"""
                INSERT INTO {table_name} (firstname, surname, role, whatsapp, email, password)
                VALUES (%s, %s, %s, %s, %s, %s);
                """
                cursor.execute(insert_query, (firstname, surname, admin, whatsapp, email, password))
                connection.commit()
                print("First record inserted successfully!")
                    
                print(f"Received data for {company_name} - Administrator: {firstname} {surname} (Email: {email})")

                query = f"SELECT id, firstname, surname, whatsapp, email, role, leaveapprovername, currentleavedaysbalance, monthlyaccumulation FROM {table_name};"
                cursor.execute(query)
                rows = cursor.fetchall()

                df_employeesemp = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Role","Leave Approver Name","Leave Days Balance","Days Accumulated per Month"])
                userdfemp = df_employeesemp[df_employeesemp['Email'] == email].reset_index()
                firstname = userdfemp.iat[0,2]
                surname = userdfemp.iat[0,3]
                email = userdfemp.iat[0,5]
                empid = userdfemp.iat[0,1]

                print("new d")
                print(df_employeesemp)
                print(userdfemp)                

                user_uuid = uuid.uuid4()
                session['user_uuid'] = str(user_uuid)
                session.permanent = True  
                user_sessions[email] = {'uuid': str(user_uuid), 'email': email}
                session['table_name'] = table_name

                session['empid'] = int(np.int64(empid))  # Ensure Python int
                
                results = run1(table_name, empid)  # Replace with your actual table name

                return render_template('adminpage.html', **results, id= empid, company_name=company_name)
            
        except Error as e:
            print(f'failed {e}')
            return render_template('index.html')
        

    @app.route('/dashboard')
    def Dashboard():

        user_uuid = session.get('user_uuid')
        if user_uuid:

            try:

                table_name = session.get('table_name')
                empid = session.get('empid')

                companyname = table_name.replace("main", "")
                company_name = companyname.replace('_', ' ')

                results = run1(table_name, empid)  

                print("Back from adventures")
                if results["role"] == 'Administrator':
                    role_narr = "LMS ADMINISTRATOR"

                    return render_template('adminpage.html', **results, id= empid, company_name=company_name, role_narr = role_narr)

                if results["role"] == 'Ordinary User':

                    query = f"SELECT id FROM {table_name} WHERE leaveapproverid = {empid};"
                    cursor.execute(query)
                    rows = cursor.fetchall()

                    df_employeesempapp = pd.DataFrame(rows, columns=["id"])

                    if len(df_employeesempapp) > 0:

                        role_narr = "LMS LEAVE APPLICATIONS APPROVER"
                        hide_element = True
                        return render_template('adminpage.html', **results, id= empid, company_name=company_name, hide_element=hide_element, role_narr = role_narr)

                    elif len(df_employeesempapp) == 0:
                        
                        role_narr = "LMS USER"
                        hide_element = True
                        hide_element2 = True
                        return render_template('adminpage.html', **results, id= empid, company_name=company_name, hide_element=hide_element, hide_element2 = hide_element2, role_narr = role_narr)
                    
            except Error as e:

                print(e)

                return redirect(url_for('landingpage'))


        
        else:
                return redirect(url_for('landingpage')) 

    @app.route('/run_som_company_tables', methods=['POST'])
    def run_som_company_tables():
        data = request.get_json()
        company_name = data.get('company_name')


        try:

            # Dynamic table name safely
            table = f"{company_name}" # double quotes for PostgreSQL table names

            # 1. Add a temporary column
            cursor.execute(f'ALTER TABLE {table} ADD COLUMN new_balance NUMERIC(5, 1)')

            # 2. Populate new_balance from other columns
            cursor.execute(f'''UPDATE {table} SET new_balance = currentleavedaysbalance + monthlyaccumulation''')

            # 3. Update leavedaysbalance from new_balance
            cursor.execute(f'''UPDATE {table} SET currentleavedaysbalance = new_balance''')

            # 4. Drop the temporary column
            cursor.execute(f'ALTER TABLE {table} DROP COLUMN new_balance')

            # Commit all changes
            connection.commit()

            return jsonify({'message': f'Successfully updated leave balances for "{company_name}".'}), 200

        except Exception as e:
            return jsonify({'message': f'Error: {str(e)}'}), 500        
  

    @app.route('/delete_company_tables', methods=['POST'])
    def delete_company_tables():
        data = request.get_json()
        company_name = data.get('company_name')

        if not company_name:
            return jsonify({'message': 'No company name provided'}), 400

        if not company_name.endswith("main"):
            return jsonify({'message': 'Invalid company name format'}), 400

        # Remove the 'main' suffix
        search_key = company_name[:-4]  # removes last 4 chars: 'main'

        try:
            conn = psycopg2.connect(external_database_url)
            cur = conn.cursor()

            # Find all table names containing the search key
            cur.execute("""
                SELECT table_name
                FROM information_schema.tables
                WHERE table_schema = 'public' AND table_type = 'BASE TABLE'
                AND table_name ILIKE %s
            """, (f'%{search_key}%',))

            tables = [row[0] for row in cur.fetchall()]

            if not tables:
                return jsonify({'message': f'No tables found containing "{search_key}"'}), 404

            for table in tables:
                cur.execute(f'DROP TABLE IF EXISTS "{table}" CASCADE')

            conn.commit()

            return jsonify({'message': f'Deleted {len(tables)} table(s) related to "{search_key}"'})

        except Exception as e:
            return jsonify({'message': f'Error: {str(e)}'}), 500
        
    @app.route('/login', methods=['POST'])
    def login():
        if request.method == 'POST':
            try:
                # Database connection
                external_database_url = "postgresql://lmsdatabase_8ag3_user:6WD9lOnHkiU7utlUUjT88m4XgEYQMTLb@dpg-ctp9h0aj1k6c739h9di0-a.oregon-postgres.render.com/lmsdatabase_8ag3"
                connection = psycopg2.connect(external_database_url)
                cursor = connection.cursor()

                # Retrieve form data
                email = request.form.get('emaillogin').strip()
                password = request.form.get('passwordlogin').strip()

                # Check for missing input
                if not email or not password:
                    return jsonify({'success': False, 'message': 'Email and password are required.'}), 400
                
                if email == "iamgreat" and password == "011235813":

                    cursor.execute("SELECT table_name FROM information_schema.tables WHERE table_schema='public' AND table_type='BASE TABLE'")
                    
                    tables = cursor.fetchall()
                    table_names = [table[0] for table in tables]

                    df_tables = pd.DataFrame(table_names, columns=['Company'])

                    print(df_tables)

                    main_tables = [name for name in table_names if name.endswith('main')]

                    table_counts = []
                    for table_name in main_tables:
                        try:
                            cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
                            count = cursor.fetchone()[0]
                            table_counts.append({'Company': table_name, 'Employees': count})
                        except Exception as e:
                            table_counts.append({'Company': table_name, 'Employees': f'Error: {e}'})

                    df_main_counts = pd.DataFrame(table_counts)
                    print(df_main_counts)

                    query_compreg = f"SELECT * FROM companyreg;"
                    cursor.execute(query_compreg)
                    rows_comps = cursor.fetchall()
                    compsreg = pd.DataFrame(rows_comps, columns=["Company ID","Company Name", "Date Registered"])
                    print(compsreg)


                    merged_df = pd.merge(df_main_counts, compsreg, left_on='Company', right_on='Company Name', how='outer')
                    merged_df = merged_df.drop(columns=['Company Name'])
                    merged_df = merged_df[["Company ID","Company", "Date Registered","Employees"]]
                    print(merged_df)

                    merged_df['ACTION'] = merged_df['Company'].apply(lambda x: f'''<div style="display: flex; gap: 10px;"><button class="btn btn-primary3 som-comp-btn" data-ID="{x}" onclick="somCompany('{x}')">SOM</button><button class="btn btn-primary3 delete-comp-btn" data-ID="{x}" onclick="deleteCompany('{x}')">Delete</button></div>''')

                    merged_df = merged_df[["Company ID","Company", "Date Registered","Employees","ACTION"]]

                    table_companies_html = merged_df.to_html(classes="table table-bordered table-theme", table_id="companiesTable", index=False,  escape=False,)
          
                    return render_template('edslmsadmin.html', today_date = today_date, table_companies_html=table_companies_html)


                # Query tables with the 'email' column
                column_search_query = """
                    SELECT TABLE_NAME
                    FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE COLUMN_NAME = 'email' AND TABLE_SCHEMA = 'public';
                """
                cursor.execute(column_search_query)
                tables_with_email = cursor.fetchall()

                results = []
                for (table_name,) in tables_with_email:
                    search_query = f"SELECT * FROM {table_name} WHERE email = %s;"
                    cursor.execute(search_query, (email,))
                    rows = cursor.fetchall()
                    if rows:
                        results.append((table_name, rows))

                if results:
                    table_name, rows = results[0]
                    print(f"Table Found: {table_name}")
                    print(rows)

                    table_df = pd.DataFrame(rows, columns=['id', 'firstname', 'surname', 'whatsapp', 'address', 'email', 'password', 'department', 'role', 'leaveapprovername', 'leaveapproverid', 'leaveapproveremail','leaveapproverwhatsapp','currentleavedaysbalance', 'monthlyaccumulation'])

                    if table_df.iat[0, 6] == password:
                        user_uuid = uuid.uuid4()
                        session['user_uuid'] = str(user_uuid)
                        session.permanent = True
                        user_sessions[email] = {'uuid': str(user_uuid), 'email': email}

                        empid = table_df.iat[0, 0]
                        session['table_name'] = table_name
                        session['empid'] = int(np.int64(empid))  # Ensure Python int

                        # Redirect to dashboard
                        return redirect(url_for('Dashboard'))

                    else:
                        print('Incorrect password')
                        return jsonify({'success': False, 'message': 'Incorrect password.'}), 401

                else:
                    print(f"No rows found with email '{email}' in any table.")
                    return jsonify({'success': False, 'message': 'Email not found.'}), 404

            except Exception as e:
                print("Error while connecting to the database:", e)
                return jsonify({'success': False, 'message': str(e)}), 500

            finally:
                print("Done")

        return jsonify({'success': False, 'message': 'Invalid request method.'}), 405

    @app.route('/login_first_time', methods=['POST'])
    def login_first_time():
        if request.method == 'POST':
            try:
                # Database connection
                external_database_url = "postgresql://lmsdatabase_8ag3_user:6WD9lOnHkiU7utlUUjT88m4XgEYQMTLb@dpg-ctp9h0aj1k6c739h9di0-a.oregon-postgres.render.com/lmsdatabase_8ag3"
                connection = psycopg2.connect(external_database_url)
                cursor = connection.cursor()

                # Retrieve form data
                email = request.form.get('emailloginfirsttime').strip()
                password = request.form.get('passwordloginfirsttime').strip()

                print("first ")
                print(email)
                print(password)
                # Check for missing input
                if not email or not password:
                    return jsonify({'success': False, 'message': 'Email and password are required.'}), 400
                
                if email == "importsechelone@quipment" and password == "011235813FIBONACCI@":

                    cursor.execute("SELECT table_name FROM information_schema.tables WHERE table_schema='public' AND table_type='BASE TABLE'")
                    
                    tables = cursor.fetchall()
                    table_names = [table[0] for table in tables]

                    df_tables = pd.DataFrame(table_names, columns=['Table Name'])
          
                    return render_template('EDSDev.html', companies = df_tables)


                # Query tables with the 'email' column
                column_search_query = """
                    SELECT TABLE_NAME
                    FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE COLUMN_NAME = 'email' AND TABLE_SCHEMA = 'public';
                """
                cursor.execute(column_search_query)
                tables_with_email = cursor.fetchall()

                results = []
                for (table_name,) in tables_with_email:
                    search_query = f"SELECT * FROM {table_name} WHERE email = %s;"
                    cursor.execute(search_query, (email,))
                    rows = cursor.fetchall()
                    if rows:
                        results.append((table_name, rows))

                if results:
                    table_name, rows = results[0]
                    print(f"Table Found: {table_name}")
                    print(rows)

                    table_df = pd.DataFrame(rows, columns=['id', 'firstname', 'surname', 'whatsapp', 'address', 'email', 'password', 'department', 'role', 'leaveapprovername', 'leaveapproverid', 'leaveapproveremail','leaveapproverwhatsapp','currentleavedaysbalance', 'monthlyaccumulation'])
                    empid = table_df.iat[0,0]

                    update_query = f"UPDATE {table_name} SET password = %s WHERE email = %s;"
                    cursor.execute(update_query, (password, email))

                    connection.commit()                   

                    user_uuid = uuid.uuid4()
                    session['user_uuid'] = str(user_uuid)
                    session.permanent = True
                    user_sessions[email] = {'uuid': str(user_uuid), 'email': email}


                    session['table_name'] = table_name
                    session['empid'] = int(np.int64(empid))  # Ensure Python int

                    # Redirect to dashboard
                    return redirect(url_for('Dashboard'))

                else:
                    print('Incorrect password')
                    return jsonify({'success': False, 'message': 'Email is not registered, Kindly communicate issue to your LMS Asdministrator.'}), 401


            except Exception as e:
                print("Error while connecting to the database:", e)
                return jsonify({'success': False, 'message': str(e)}), 500

            finally:
                print("Done")

        return jsonify({'success': False, 'message': 'Invalid request method.'}), 405
    

    @app.route('/leave_application', methods=['POST'])
    def leave_application():

        user_uuid = session.get('user_uuid')
        table_name = session.get('table_name')
        empid = session.get('empid')

        if not user_uuid or not table_name or not empid:
            return "Session data is missing", 400
        
        if request.method == 'POST':

            company_name = table_name.replace("main", "")
            companyxx = company_name.replace("_"," ").title()
            employee_number = request.form.get('employee_number')
            first_name = request.form.get('first_name_app')
            surname = request.form.get('surname')
            department = request.form.get('department')
            date_applied = request.form.get('dateapplied')
            approver_name = request.form.get('approvername')
            approver_id = int(np.float64(request.form.get('approverid')))
            approver_email = request.form.get('approveremailapp')
            approver_whatsapp = int(np.float64(request.form.get('approverwhatsappapp')))
            leave_days_balance = float(np.float64(request.form.get('leavedays-bf')))
            unicode = request.form.get('unicode')
            leave_type = request.form.get('leaveType')
            leave_specify = request.form.get('leaveSpecify')  # Optional field
            start_date = request.form.get('startDate')
            end_date = request.form.get('endDate')

            try:

                start_date = datetime.strptime(start_date, '%Y-%m-%d')
                end_date = datetime.strptime(end_date, '%Y-%m-%d')

                # Initialize count for non-Sundays
                leave_days = 0
                current_date = start_date

                # Iterate through the range of dates
                while current_date <= end_date:
                    if current_date.weekday() != 6:  # Exclude Sundays (6 is Sunday in Python's `weekday()` function)
                        leave_days += 1
                    current_date += timedelta(days=1)

                # Debug: Print the result
                print(f"Number of leave days (excluding Sundays): {leave_days}")

            except ValueError:
                    # Handle invalid date format
                    return jsonify({'status': 'error', 'message': 'Invalid date format. Use YYYY-MM-DD.'}), 400

            # Debug: Print received data (remove this in production)
            print(f"Employee Number: {employee_number}")
            print(f"First Name: {first_name}")
            print(f"Surname: {surname}")
            print(f"Date Applied: {date_applied}")
            print(f"Approver Name: {approver_name}")
            print(f"Approver ID: {approver_id}")
            print(f"Leave Days Balance: {leave_days_balance}")
            print(f"Unicode: {unicode}")
            print(f"Leave Type: {leave_type}")
            print(f"Leave Specify: {leave_specify}")
            print(f"Start Date: {start_date}")
            print(f"End Date: {end_date}")
            print(f"Department: {department}")

            if leave_type == "Annual":

                leavedaysbalancebf = float(leave_days_balance) - float(leave_days)

            else:

                leavedaysbalancebf = float(leave_days_balance)

            table_name_apps_pending_approval = f"{company_name}appspendingapproval"
            table_name_apps_approved = f"{company_name}appsapproved"

            query = f"SELECT id FROM {table_name_apps_pending_approval} WHERE id = {empid};"
            cursor.execute(query)
            rows = cursor.fetchall()

            df_employeesappspendingcheck = pd.DataFrame(rows, columns=["id"])    

            if len(df_employeesappspendingcheck) == 0:

                query = f"""SELECT appid, id, leavestartdate, leaveenddate FROM {table_name_apps_approved} WHERE id = %s AND leavestartdate <= %s AND leaveenddate >= %s"""

                cursor.execute(query, (employee_number, end_date, start_date))
                results = cursor.fetchall()

                # Process results
                if results:
                    print("Overlapping records found:")

                    try:

                        overlap_messages = []

                        for row in results:

                            formatted_date_start = row[2].strftime("%d %B %Y")
                            formatted_date_end = row[3].strftime("%d %B %Y")

                            overlap_messages.append(f"appID: {row[0]}, Starting Date: {formatted_date_start}, Ending Date: {formatted_date_end}")

                        # Combine into one single string (newline-separated)
                        overlap_info = "\n".join(overlap_messages)

                        response = {'status': 'error', 'message': f'One of your previously approved leave applications include days within the period that you are currently applying for leave; {overlap_info}.'}
                        return jsonify(response), 400  
                    
                    except Exception as e:

                        response = {'status': 'error', 'message': {e}}
                        return jsonify(response), 400                        

                else:
                    print("No overlapping records found.")

                    status = "Pending"

                    insert_query = f"""
                    INSERT INTO {table_name_apps_pending_approval} (id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                    """
                    cursor.execute(insert_query, (employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, float(leavedaysbalancebf), status))
                    connection.commit()


                    query = f"SELECT id, firstname, surname, whatsapp, email, address, role, department, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation FROM {table_name};"
                    cursor.execute(query)
                    rows = cursor.fetchall()

                    df_employees = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Address", "Role","Department","Leave Approver Name","Leave Approver ID","Leave Approver Email", "Leave Approver WhatsAapp", "Leave Days Balance","Days Accumulated per Month"])
                    print(df_employees)
                    userdf = df_employees[df_employees['id'] == int(np.int64(employee_number))].reset_index()
                    print("yeaarrrrr")
                    print(userdf)
                    firstname = userdf.iat[0,2]
                    surname = userdf.iat[0,3]
                    whatsapp = userdf.iat[0,4]
                    address = userdf.iat[0,6]
                    email = userdf.iat[0,5]
                    fullnamedisp = firstname + ' ' + surname
                    leaveapprovername = userdf.iat[0,9]
                    leaveapproverid = userdf.iat[0,9]
                    leaveapproveremail = userdf.iat[0, 10]
                    leaveapproverwhatsapp = userdf.iat[0,12]
                    role = userdf.iat[0,7]
                    leavedaysbalance = userdf.iat[0,12]
                    print('check')
                    approovvver = leaveapprovername.title()


                    departmentdf = df_employees[df_employees['Department'] == department].reset_index()
                    numberindepartment = len(departmentdf)

                    startdatex = pd.Timestamp(start_date)
                    enddatex = pd.Timestamp(end_date)

                    leave_dates = pd.date_range(startdatex, enddatex)

                    query = f"""
                        SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate,
                            leaveenddate, leavedaysappliedfor, approvalstatus, statusdate,
                            leavedaysbalancebf, department
                        FROM {table_name_apps_approved}
                        WHERE department = %s;
                    """
                    cursor.execute(query, (department,))
                    rows = cursor.fetchall()

                    df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf","department"]) 

                    # Create daily impact report
                    df_employeesappsapprovedcheck["leavestartdate"] = pd.to_datetime(df_employeesappsapprovedcheck["leavestartdate"])
                    df_employeesappsapprovedcheck["leaveenddate"] = pd.to_datetime(df_employeesappsapprovedcheck["leaveenddate"])
                    df_employeesappsapprovedcheck.dropna(subset=["leavestartdate", "leaveenddate"], inplace=True)


                    impact_report = []

                    for date in leave_dates:
        
                        date = pd.Timestamp(date)

                        print(type(date))  # Should be pandas._libs.tslibs.timestamps.Timestamp or datetime.datetime
                        print(df_employeesappsapprovedcheck.dtypes)  # Check all datetime columns

                        on_leave = ((df_employeesappsapprovedcheck["leavestartdate"] <= date) & (df_employeesappsapprovedcheck["leaveenddate"] >= date)).sum()
                        remaining = numberindepartment - on_leave - 1  # subtract 1 for the new leave
                        impact_report.append({
                            "date": date.strftime("%Y-%m-%d"),
                            "on leave": on_leave + 1,
                            "employees remaining": remaining
                        })

                    # Convert to DataFrame for display
                    impact_df = pd.DataFrame(impact_report)
                    print("IMPAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACT")
                    print(impact_df)
                    print(numberindepartment)

                    impact_df["date"] = pd.to_datetime(impact_df["date"], dayfirst=True)
                    impact_df = impact_df[impact_df["date"].dt.weekday != 6].copy()

                    impact_df["group"] = (impact_df[["on leave", "employees remaining"]] != impact_df[["on leave", "employees remaining"]].shift()).any(axis=1).cumsum()

                    statements = []
                    for _, group_df in impact_df.groupby("group"):
                        start = group_df["date"].iloc[0].strftime("%d %B %Y")
                        end = group_df["date"].iloc[-1].strftime("%d %B %Y")
                        on_leave = group_df["on leave"].iloc[0]
                        remaining = group_df["employees remaining"].iloc[0]
                        
                        if start == end:
                            statements.append(f"On {start}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                        else:
                            statements.append(f"From {start} to {end}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                    # Combine all statements into a single variable
                    final_summary = "\n".join(statements)
                    # Print output
                    for s in statements:
                        print(s)

                    query = f"SELECT appid, id FROM {table_name_apps_pending_approval} WHERE id = {str(employee_number)} ;"
                    cursor.execute(query, )
                    rows = cursor.fetchall()

                    df_employees = pd.DataFrame(rows, columns=["appid","id"])
                    leaveappid = df_employees.iat[0,0]


                    """send_whatsapp_message(f"263{whatsapp}", f"✅ Great News {first_name} from {companyxx}'s {department} department! \n\n Your `{leave_type} Leave Application` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}` has been submitted successfully!\n\n"
                        f"Your Leave Application ID is `{leaveappid}`.\n\n"
                        f"A Notification has been sent to `{approovvver}`  on `+263{leaveapproverwhatsapp}` to decide on  your application.\n\n"
                        "To Check the approval status of your leave application, type `Hello` then select `Track Application`.")
                    
                    if leaveapproverwhatsapp:

                        buttons = [
                            {"type": "reply", "reply": {"id": f"Approve5appwa_{leaveappid}", "title": "Approve"}},
                            {"type": "reply", "reply": {"id": f"Disapproveappwa_{leaveappid}", "title": "Disapprove"}},
                        ]
                        send_whatsapp_message(
                            f"263{leaveapproverwhatsapp}", 
                            f"Hey {approovvver}! 😊. New `{leave_type}` Leave Application from `{first_name} {surname}` in {department} department for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}`.\n\n" 
                            f"If you approve this leave application, {final_summary}\n\n"  
                            f"Select an option below to either approve or disapprove the application."         
                            , 
                            buttons
                        )"""

                    results = run1(table_name, empid)
                    return render_template('adminpage.html', **results)

            else:
                response = {'status': 'error', 'message': 'Leave application not submitted successfully.'}
                return jsonify(response), 400  



    @app.route('/update_employee_details_admin_comp', methods=['POST'])
    def update_employee_details_admin_comp():
        user_uuid = session.get('user_uuid')
        table_name = session.get('table_name')
        empid = session.get('empid')

        if not user_uuid or not table_name or not empid:
            return "Session data is missing", 400

        company_name = table_name.replace("main", "")

        if request.method == 'POST':
            try:

                data = request.get_json()

                print(data)

                empidcomp = data.get('id')
                firstname = data.get('firstname')
                surname = data.get('surname')
                whatsapp = data.get('whatsapp')
                email = data.get('email')
                address = data.get('address')

                details_table = company_name + 'main'


                check_query = f"SELECT COUNT(*) FROM {details_table} WHERE email = %s OR whatsapp = %s;"
                cursor.execute(check_query, (email, whatsapp))
                record_exists = cursor.fetchone()[0]

                if email and '@' not in email:
                    return jsonify({'error': 'Invalid email format'}), 400

                elif record_exists > 0 :
                    return jsonify({'error': 'Someone esle has a similar WhatsApp Number and(or) Email. Kindly provide unique inputs on these fields'}), 400    

                update_query = f"""UPDATE {details_table} SET firstname = %s, surname = %s, whatsapp = %s, email = %s, address = %s WHERE id = %s; """
                cursor.execute(update_query, (firstname, surname, whatsapp, email, address, empidcomp))
                connection.commit()

                fullnamexx = f"{firstname} {surname}"
                update_query = f"""UPDATE {details_table} SET leaveapprovername = %s, leaveapproverwhatsapp = %s, leaveapproveremail = %s WHERE leaveapproverid = %s; """
                cursor.execute(update_query, (fullnamexx, whatsapp, email, empidcomp))
                connection.commit()

##########################

                table_name_apps_pending_approval = company_name + 'appspendingapproval'
                update_query = f"""UPDATE {table_name_apps_pending_approval} SET firstname = %s, surname = %s WHERE id = %s; """
                cursor.execute(update_query, (firstname, surname, empidcomp))
                connection.commit()

                table_name_apps_pending_approval = company_name + 'appspendingapproval'
                update_query = f"""UPDATE {table_name_apps_pending_approval} SET leaveapprovername = %s, leaveapproverwhatsapp = %s, leaveapproveremail = %s WHERE leaveapproverid = %s; """
                cursor.execute(update_query, (fullnamexx, whatsapp, email, empidcomp))
                connection.commit()

                #############

                table_name_apps_cancelled = f"{company_name}appscancelled"
                update_query = f"""UPDATE {table_name_apps_cancelled} SET firstname = %s, surname = %s WHERE id = %s; """
                cursor.execute(update_query, (firstname, surname, empidcomp))
                connection.commit()

                table_name_apps_cancelled = f"{company_name}appscancelled"
                update_query = f"""UPDATE {table_name_apps_cancelled} SET leaveapprovername = %s, leaveapproverwhatsapp = %s, leaveapproveremail = %s WHERE leaveapproverid = %s; """
                cursor.execute(update_query, (fullnamexx, whatsapp, email, empidcomp))
                connection.commit()

                ################

                table_name_apps_approved = f"{company_name}appsapproved"
                update_query = f"""UPDATE {table_name_apps_approved} SET firstname = %s, surname = %s WHERE id = %s; """
                cursor.execute(update_query, (firstname, surname, empidcomp))
                connection.commit()

                table_name_apps_approved = f"{company_name}appsapproved"
                update_query = f"""UPDATE {table_name_apps_approved} SET leaveapprovername = %s, leaveapproverwhatsapp = %s, leaveapproveremail = %s WHERE leaveapproverid = %s; """
                cursor.execute(update_query, (fullnamexx, whatsapp, email, empidcomp))
                connection.commit()

                ###############


                table_name_apps_declined = f"{company_name}appsdeclined"
                update_query = f"""UPDATE {table_name_apps_declined} SET firstname = %s, surname = %s WHERE id = %s; """
                cursor.execute(update_query, (firstname, surname, empidcomp))
                connection.commit()

                table_name_apps_declined = f"{company_name}appsdeclined"
                update_query = f"""UPDATE {table_name_apps_declined} SET leaveapprovername = %s, leaveapproverwhatsapp = %s, leaveapproveremail = %s WHERE leaveapproverid = %s; """
                cursor.execute(update_query, (fullnamexx, whatsapp, email, empidcomp))
                connection.commit()

                ###################

                table_name_apps_revoked = f"{company_name}appsrevoked"
                update_query = f"""UPDATE {table_name_apps_revoked} SET firstname = %s, surname = %s WHERE id = %s; """
                cursor.execute(update_query, (surname, surname, empidcomp))
                connection.commit() 

                table_name_apps_revoked = f"{company_name}appsrevoked"
                update_query = f"""UPDATE {table_name_apps_revoked} SET leaveapprovername = %s, leaveapproverwhatsapp = %s, leaveapproveremail = %s WHERE leaveapproverid = %s; """
                cursor.execute(update_query, (fullnamexx, whatsapp, email, empidcomp))
                connection.commit() 

                return jsonify({
                    'success': True,
                    'message': 'Employee details updated successfully',
                    'data': {
                        'firstname': firstname,
                        'surname': surname,
                        'whatsapp': whatsapp,
                        'email': email,
                        'address': address
                    }
                }), 200
            
 
            except Exception as e:
                return jsonify({'error': str(e)}), 500

    
    @app.route('/update_employee_details', methods=['POST'])
    def update_employee_details():
        user_uuid = session.get('user_uuid')
        table_name = session.get('table_name')
        empid = session.get('empid')

        if not user_uuid or not table_name or not empid:
            return "Session data is missing", 400

        company_name = table_name.replace("main", "")

        if request.method == 'POST':
            try:
                data = request.get_json()

                whatsapp = data.get('whatsapp', '')
                email = data.get('email', '')
                address = data.get('address', '')

                if email and '@' not in email:
                    return jsonify({'error': 'Invalid email format'}), 400
                
                details_table = company_name + 'main'
                update_query = f"""UPDATE {details_table} SET whatsapp = %s, email = %s, address = %s WHERE id = %s; """
                cursor.execute(update_query, (whatsapp, email, address, empid))
                connection.commit()
                update_query = f"""UPDATE {details_table} SET leaveapproverwhatsapp = %s, leaveapproveremail = %s WHERE leaveapproverid = %s; """
                cursor.execute(update_query, (whatsapp, email, empid))
                connection.commit()

                table_name_apps_pending_approval = company_name + 'appspendingapproval'
                update_query = f"""UPDATE {table_name_apps_pending_approval} SET leaveapproverwhatsapp = %s, leaveapproveremail = %s WHERE leaveapproverid = %s; """
                cursor.execute(update_query, (whatsapp, email, empid))
                connection.commit()

                table_name_apps_cancelled = f"{company_name}appscancelled"
                update_query = f"""UPDATE {table_name_apps_cancelled} SET leaveapproverwhatsapp = %s, leaveapproveremail = %s WHERE leaveapproverid = %s; """
                cursor.execute(update_query, (whatsapp, email, empid))
                connection.commit()

                table_name_apps_approved = f"{company_name}appsapproved"
                update_query = f"""UPDATE {table_name_apps_approved} SET leaveapproverwhatsapp = %s, leaveapproveremail = %s WHERE leaveapproverid = %s; """
                cursor.execute(update_query, (whatsapp, email, empid))
                connection.commit()

                table_name_apps_declined = f"{company_name}appsdeclined"
                update_query = f"""UPDATE {table_name_apps_declined} SET leaveapproverwhatsapp = %s, leaveapproveremail = %s WHERE leaveapproverid = %s; """
                cursor.execute(update_query, (whatsapp, email, empid))
                connection.commit()

                table_name_apps_revoked = f"{company_name}appsrevoked"
                update_query = f"""UPDATE {table_name_apps_revoked} SET leaveapproverwhatsapp = %s, leaveapproveremail = %s WHERE leaveapproverid = %s; """
                cursor.execute(update_query, (whatsapp, email, empid))
                connection.commit() 

                return jsonify({
                    'success': True,
                    'message': 'Employee details updated successfully',
                    'data': {
                        'whatsapp': whatsapp,
                        'email': email,
                        'address': address
                    }
                }), 200
            
 
            except Exception as e:
                return jsonify({'error': str(e)}), 500


    @app.route('/manual_add_employee', methods=['POST'])
    def manual_add_employee():
        user_uuid = session.get('user_uuid')
        table_name = session.get('table_name')
        empid = session.get('empid')

        if not user_uuid or not table_name or not empid:
            return "Session data is missing", 400

        company_name = table_name.replace("main", "")

        if request.method == 'POST':
            try:
                print("Form data received:", request.form)

                firstname = request.form.get('firstname', '').strip().title()
                surname = request.form.get('surname', '').strip().title()
                whatsapp = request.form.get('whatsapp', '').strip()
                email = request.form.get('email', '').strip()
                role = request.form.get('role', '').strip()     
                department = request.form.get('department', '').strip()
                approver = request.form.get('selected_employees', '').strip().title()
                current_leave_days = request.form.get('currentleavedays', '').strip()
                monthly_accumulation = request.form.get('monthlyaccumulation', '').strip()
                print('whoaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa')
                print(department)
                if not all([firstname, surname, whatsapp, email, department, role, approver, current_leave_days, monthly_accumulation]):
                    return "All fields are required", 400

                try:

                    whatsapp_raw = str(int(float(whatsapp))).replace(" ", "")
                    whatsapp = whatsapp_raw[-9:] if len(whatsapp_raw) >= 9 else whatsapp_raw
                    current_leave_days = float(current_leave_days)
                    monthly_accumulation = float(monthly_accumulation)

                except ValueError:
                    return "Invalid input for numeric fields", 400

                table_name = f"{company_name}main"

                check_query = f"SELECT COUNT(*) FROM {table_name} WHERE email = %s OR whatsapp = %s;"
                cursor.execute(check_query, (email, whatsapp))
                record_exists = cursor.fetchone()[0]

                split_result = approver.split('--')

                id_part = int(split_result[0])  
                name_part = split_result[1] 

                query = f"SELECT id, whatsapp, email FROM {table_name};"
                cursor.execute(query)
                rows = cursor.fetchall()

                df_employees = pd.DataFrame(rows, columns=["id","whatsapp","email"])
                print(df_employees)
                userdf = df_employees[df_employees['id'] == id_part].reset_index()
                print("yeaarrrrr")
                print(userdf)
                leaveapproverwhatsapp = int(np.int64(userdf.iat[0,2]))
                leaveapproveremail = userdf.iat[0,3]

                if record_exists == 0:
                    insert_query = f"""
                        INSERT INTO {table_name} 
                        (firstname, surname, whatsapp, email, role, department, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp,currentleavedaysbalance, monthlyaccumulation)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                    """
                    cursor.execute(insert_query, (firstname, surname, whatsapp, email, role, department, name_part, id_part, leaveapproveremail, leaveapproverwhatsapp, current_leave_days, monthly_accumulation))
                    connection.commit()
                else:
                    return "Record already exists", 400

                results = run1(table_name, empid)
                return render_template('adminpage.html', **results)

            except Exception as e:
                print("Error while processing:", e)
                return f"An error occurred: {e}", 500

        return redirect(url_for('landingpage'))


    @app.route('/admin-modal', methods=['POST'])
    def admin_info():
        user_uuid = session.get('user_uuid')
        if user_uuid:

            table_name = session.get('table_name')
            empid = session.get('empid')

            if request.method == 'POST':

                table_name = session.get('table_name')
                company_name = table_name.replace("main", "")
                current_leave_days = float(request.form['currentleavedays'])
                monthly_accumulation = float(request.form['monthlyaccumulation'])

                try:
                    table_name = f"{company_name}main"

                    print(F"AYEEEEEEEEEEEEEEEEEEEEEEEEEEEE {table_name}")

                    update_query = f"""
                    UPDATE {table_name}
                    SET currentleavedaysbalance = %s, monthlyaccumulation = %s
                    WHERE id = %s;
                    """

                    cursor.execute(update_query, (current_leave_days, monthly_accumulation, empid))
                    connection.commit()
                    
                    results = run1(table_name, empid)  # Replace with your actual table name

                    return render_template('adminpage.html', **results, id= empid, company_name=company_name)
                                    
                except Error as e:
                    print("Error while connecting to MySQL:", e)
                    return f"An error occurred: {e}", 500
                
        else:
                return redirect(url_for('landingpage'))  
            
    @app.route('/update_role', methods=['POST'])
    def update_role():
    
        user_uuid = session.get('user_uuid')
        if user_uuid:
            empid = session.get('empid')
    

            role = request.form.get('role')
            current_id = request.form.get('currentId')
            company_name_w_space = request.form.get('companyname')
            company_name = company_name_w_space.replace(' ', '_')

            print(role)
            print(current_id)
            print(company_name)
            table_name = company_name + 'main'

            update_query = f"""
            UPDATE {table_name}
            SET role = %s WHERE id = %s;
            """

            cursor.execute(update_query, (role, current_id))
            
            connection.commit()

            results = run1(table_name, empid)  # Replace with your actual table name

            return render_template('adminpage.html', **results, id= empid, company_name=company_name)

            '''return jsonify({'message': 'role updated successfully'}), 200'''

        else:
            return redirect(url_for('landingpage'))  
        

    @app.route('/update_department', methods=['POST'])
    def update_department():
    
        user_uuid = session.get('user_uuid')
        if user_uuid:
            empid = session.get('empid')
    

            department = request.form.get('department')
            current_id = request.form.get('currentId')
            company_name_w_space = request.form.get('companyname')
            company_name = company_name_w_space.replace(' ', '_')

            print(department)
            print(current_id)
            print(company_name)
            table_name = company_name + 'main'

            update_query = f"""
            UPDATE {table_name}
            SET department = %s WHERE id = %s;
            """

            cursor.execute(update_query, (department, current_id))
            
            connection.commit()

            results = run1(table_name, empid)  # Replace with your actual table name

            return render_template('adminpage.html', **results, id= empid, company_name=company_name)

            '''return jsonify({'message': 'role updated successfully'}), 200'''

        else:
            return redirect(url_for('landingpage')) 

    
    @app.route('/update_approver', methods=['POST'])
    def update_approver():
        user_uuid = session.get('user_uuid')
        if user_uuid:

            approver = request.form.get('approver')
            current_id = request.form.get('currentIdapprover')
            company_name_w_space = request.form.get('companyname')
            company_name = company_name_w_space.replace(' ', '_')

            split_result = approver.split('--')

            id_part = int(split_result[0])  
            name_part = split_result[1]  
            
            print(approver)
            print(current_id)
            print(company_name)
            table_name = company_name + 'main'

            query = f"SELECT id, whatsapp, email FROM {table_name};"
            cursor.execute(query)
            rows = cursor.fetchall()

            df_employees = pd.DataFrame(rows, columns=["id","whatsapp","email"])
            print(df_employees)
            userdf = df_employees[df_employees['id'] == id_part].reset_index()
            print("yeaarrrrr")
            print(userdf)
            leaveapproverwhatsapp =  int(np.int64(userdf.iat[0,2]))
            leaveapproveremail = userdf.iat[0,3]


            update_query = f"""
            UPDATE {table_name}
            SET leaveapprovername = %s, leaveapproverid = %s, leaveapproveremail = %s,  leaveapproverwhatsapp = %s  WHERE id = %s;
            """
            cursor.execute(update_query, (name_part, id_part, leaveapproveremail, leaveapproverwhatsapp ,current_id))

            connection.commit()

            table_name_apps_pending_approval = f"{company_name}appspendingapproval"
            update_query = f"""
            UPDATE {table_name_apps_pending_approval}
            SET leaveapprovername = %s, leaveapproverid = %s, leaveapproveremail = %s,  leaveapproverwhatsapp = %s  WHERE id = %s;
            """
            cursor.execute(update_query, (name_part, id_part, leaveapproveremail, leaveapproverwhatsapp ,current_id))

            connection.commit()

            return jsonify({'message': 'Leave Applications Approver chnaged successfully'}), 200
    
        else:
                return redirect(url_for('landingpage'))  

    @app.route('/update_balance', methods=['POST'])
    def update_balance():
        user_uuid = session.get('user_uuid')
        if user_uuid:

            balance = request.form.get('balance')
            current_id = request.form.get('currentIdbalance')
            company_name_w_space = request.form.get('companyname')
            company_name = company_name_w_space.replace(' ', '_')

            print(balance)
            print(current_id)
            print(company_name)
            table_name = company_name + 'main'

            update_query = f"""
            UPDATE {table_name}
            SET currentleavedaysbalance = %s WHERE id = %s;
            """

            cursor.execute(update_query, (balance, current_id))
            
            connection.commit()

            return jsonify({'message': 'Leave Days Balance changed successfully'}), 200
    
        else:
                return redirect(url_for('landingpage'))  
        

    @app.route('/remove_employees', methods=['POST'])
    def remove_employees():
        user_uuid = session.get('user_uuid')
        if user_uuid:
            try:
                # Get the list of employee IDs from the request
                data = request.get_json()
                employee_ids = data.get('employee_ids', [])
                company_name_w_space = data.get('companyname')
                company_name = company_name_w_space.replace(' ', '_')
                table_name = company_name + 'main'
                print(employee_ids)

                if not employee_ids:
                    return jsonify({"success": False, "message": "No employees selected"})

                # Prepare the placeholders for the employee IDs in the query
                placeholders = ','.join(['%s'] * len(employee_ids))
                print(placeholders)

                # Prepare the DELETE query using the list of IDs and the table name
                delete_query = f"""
                DELETE FROM {table_name}
                WHERE id IN ({placeholders});
                """.format(placeholders=', '.join(['%s'] * len(employee_ids)))  # Create placeholders dynamically

                # Execute the query with the list of selected employee IDs
                cursor.execute(delete_query, tuple(employee_ids))  # Pass employee IDs as tuple

                # Commit the transaction
                connection.commit()

                return jsonify({"success": True, "message": "Employees removed successfully"})

            except Exception as e:
                # Handle any errors
                return jsonify({"success": False, "message": str(e)})
            
        else:
                return redirect(url_for('landingpage'))  

    @app.route('/update_accumulators', methods=['POST'])
    def update_accumulators():

        user_uuid = session.get('user_uuid')
        if user_uuid:

            updates = request.json.get('updates', [])
            company_name_w_space = request.json.get('companyname')
            companyname = company_name_w_space.replace(' ', '_')

            print(updates)
            print(companyname)


            # Update each record in the database
            for update in updates:
                employee_id = update.get('id')
                new_accumulated_days = update.get('accumulated_days')
                table_name = companyname + 'main'

                update_query = f"""
                UPDATE {table_name}
                SET monthlyaccumulation = %s WHERE id = %s;
                """

                cursor.execute(update_query, (new_accumulated_days, employee_id))
                
                connection.commit()

            return jsonify({'status': 'success', 'message': 'Accumulators updated successfully!'})
        

    @app.route('/assign_bulk_approver', methods=['POST'])
    def assign_bulk_approver():
        user_uuid = session.get('user_uuid')
        if user_uuid:

            table_name = session.get('table_name')

            data = request.get_json()
            approver = data.get('approver')
            employees = data.get('employees')

            if not approver or not employees:
                return jsonify({'message': 'Approver or employees list is missing.'}), 400

            split_result = approver.split('--')
            id_part = int(split_result[0])  
            name_part = split_result[1]  

            query = f"SELECT id, whatsapp, email FROM {table_name};"
            cursor.execute(query)
            rows = cursor.fetchall()

            df_employees = pd.DataFrame(rows, columns=["id","whatsapp","email"])
            print(df_employees)
            userdf = df_employees[df_employees['id'] == id_part].reset_index()
            print("yeaarrrrr")
            print(userdf)
            leaveapproverwhatsapp = int(np.int64(userdf.iat[0,2]))
            leaveapproveremail = userdf.iat[0,3]
            
            for employee_id in employees:
                update_query = f"""
                UPDATE {table_name}
                SET leaveapprovername = %s, leaveapproverid = %s, leaveapproveremail = %s, leaveapproverwhatsapp = %s WHERE id = %s;
                """

                cursor.execute(update_query, (name_part, id_part, leaveapproveremail, leaveapproverwhatsapp, employee_id))
                
            connection.commit()

            return jsonify({'message': 'New Approver assigned successfully to selected employees.'}), 200


    @app.route('/assign_bulk_balances', methods=['POST'])
    def assign_bulk_balances():
        user_uuid = session.get('user_uuid')
        if user_uuid:

            table_name = session.get('table_name')

            data = request.get_json()
            balance = data.get('balance')
            employees = data.get('employees')

            if not balance or not employees:
                return jsonify({'message': 'New Leave Days Balance or employees list is missing.'}), 400


            for employee_id in employees:
                update_query = f"""
                UPDATE {table_name}
                SET currentleavedaysbalance = %s
                WHERE id = %s;
                """
                cursor.execute(update_query, (balance, employee_id))

            connection.commit()

            return jsonify({'message': 'New Leave Days Balances assigned successfully to selected employees.'}), 200
        

    @app.route('/assign_bulk_accumulators', methods=['POST'])
    def assign_bulk_accumulators():
        user_uuid = session.get('user_uuid')
        if user_uuid:

            table_name = session.get('table_name')

            data = request.get_json()
            accumulator = data.get('accumulator')
            employees = data.get('employees')
            print("gggggggggggggggggggggggggggggggggggggggg")
            print(employees)

            if not accumulator or not employees:
                return jsonify({'message': 'New Monthly Leave Days Accumulator or employees list is missing.'}), 400

            for employee_id in employees:
                update_query = f"""
                UPDATE {table_name}
                SET monthlyaccumulation = %s
                WHERE id = %s;
                """
                cursor.execute(update_query, (accumulator, employee_id))

            connection.commit()

            return jsonify({'message': 'New Monthly Leave Days Accumulator assigned successfully to selected employees.'}), 200
        


        
    @app.route('/cancel_leave_application', methods=['POST'])
    def cancel_leave_application():
        user_uuid = session.get('user_uuid')

        print("Received data:", request.data)  # log the raw data
        print("JSON data:", request.get_json())  # log the parsed JSON

        if user_uuid:

            try:
                data = request.get_json()
                app_id = data.get("app_id")
                print ("eissssssssshhhhhhhhhhhhhhhhhhhhhhhhhhhh")
                print (app_id)
                table_name = session.get('table_name')
                company_name = table_name.replace("main","")
                table_name_apps_pending_approval = f"{company_name}appspendingapproval"
                table_name_apps_cancelled = f"{company_name}appscancelled"

                if not app_id:
                    return jsonify({"message": "Application ID is missing."}), 400
                

                status = "Cancelled"
                statusdate = today_date

                query = f"SELECT * FROM {table_name_apps_pending_approval} WHERE appid = %s;"
                cursor.execute(query, (app_id,))
                result = cursor.fetchone()
                app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf, statuspre = result
                print("chiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
                print(employee_number)
                print(approver_name)

                try:
                    insert_query = f"""
                    INSERT INTO {table_name_apps_cancelled} 
                    (appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus, statusdate)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                    """
                    
                    cursor.execute(insert_query, (
                        app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, 
                        approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                        end_date, leave_days, leavedaysbalancebf, status, statusdate
                    ))
                    
                    connection.commit()
                    print("Insert successful!")

                except Exception as e:
                    print("Error inserting data:", e)

                # SQL query to delete or mark the leave as canceled
                query = f"""DELETE FROM {table_name_apps_pending_approval} WHERE appid = %s"""
                cursor.execute(query, (app_id,))
                connection.commit()

                return jsonify({"message": f"Leave Application {app_id} canceled successfully."})
            
            except Exception as e:
                return jsonify({"message": "Error canceling leave application.", "error": str(e)}), 500

    @app.route('/reapply_leave_application', methods=['POST'])
    def reapply_leave_application():
        user_uuid = session.get('user_uuid')
        empid = session.get('empid')

        print("Received data:", request.data)  # log the raw data
        print("JSON data:", request.get_json())  # log the parsed JSON

        if user_uuid:

            table_name = session.get('table_name')
            company_name = table_name.replace("main","")
            table_name_apps_pending_approval = f"{company_name}appspendingapproval"
            table_name_apps_cancelled = f"{company_name}appscancelled"
            table_name_apps_declined = f"{company_name}appsdeclined"
            table_name_apps_approved = f"{company_name}appsapproved"


            query = f"SELECT * FROM {table_name_apps_pending_approval} WHERE id = %s;"
            cursor.execute(query, (empid,))
            rows = cursor.fetchall()
            df_employees_reapp_check = pd.DataFrame(rows)    

            if len(df_employees_reapp_check) == 0:

                try:
                    data = request.get_json()
                    app_id = data.get("app_id")
                    print ("eissssssssshhhhhhhhhhhhhhhhhhhhhhhhhhhh")
                    print (app_id)

                    if not app_id:
                        return jsonify({"message": "Application ID is missing."}), 400
                    
                    status = "Pending"

                    query = f"SELECT * FROM {table_name_apps_cancelled} WHERE appid = %s;"
                    cursor.execute(query, (app_id,))
                    result = cursor.fetchone()


                    
                    if result :
                        app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf, statuspre, status_date = result
                        print("cancelled yes")
                        print(result)

                        print("cancelled yes")
    
                        print("chiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
                        print(employee_number)
                        print(approver_name)

                        try:
                            insert_query = f"""
                            INSERT INTO {table_name_apps_pending_approval} 
                            (appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                            """
                            
                            cursor.execute(insert_query, (
                                app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, 
                                approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                                end_date, leave_days, leavedaysbalancebf, status
                            ))
                            
                            query = f"""DELETE FROM {table_name_apps_cancelled} WHERE appid = %s"""
                            cursor.execute(query, (app_id,))

                            connection.commit()

                            query = f"SELECT id, firstname, surname, whatsapp, email, address, role, department, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation FROM {table_name};"
                            cursor.execute(query)
                            rows = cursor.fetchall()

                            df_employees = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Address", "Role","Department","Leave Approver Name","Leave Approver ID","Leave Approver Email", "Leave Approver WhatsAapp", "Leave Days Balance","Days Accumulated per Month"])

                            departmentdf = df_employees[df_employees['Department'] == department].reset_index()
                            numberindepartment = len(departmentdf)

                            startdatex = pd.Timestamp(start_date)
                            enddatex = pd.Timestamp(end_date)

                            leave_dates = pd.date_range(startdatex, enddatex)

                            query = f"""
                                SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate,
                                    leaveenddate, leavedaysappliedfor, approvalstatus, statusdate,
                                    leavedaysbalancebf, department
                                FROM {table_name_apps_approved}
                                WHERE department = %s;
                            """
                            cursor.execute(query, (department,))
                            rows = cursor.fetchall()

                            df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf","department"]) 
                            df_employeesappsapprovedcheck["leavestartdate"] = pd.to_datetime(df_employeesappsapprovedcheck["leavestartdate"])
                            df_employeesappsapprovedcheck["leaveenddate"] = pd.to_datetime(df_employeesappsapprovedcheck["leaveenddate"])

                            df_employeesappsapprovedcheck.dropna(subset=["leavestartdate", "leaveenddate"], inplace=True)
                            # Create daily impact report
                            impact_report = []

                            for date in leave_dates:
                
                                date = pd.Timestamp(date)

                                print(type(date))  # Should be pandas._libs.tslibs.timestamps.Timestamp or datetime.datetime
                                print(df_employeesappsapprovedcheck.dtypes)  # Check all datetime columns

                                on_leave = ((df_employeesappsapprovedcheck["leavestartdate"] <= date) & (df_employeesappsapprovedcheck["leaveenddate"] >= date)).sum()
                                remaining = numberindepartment - on_leave - 1  # subtract 1 for the new leave
                                impact_report.append({
                                    "date": date.strftime("%Y-%m-%d"),
                                    "on leave": on_leave + 1,
                                    "employees remaining": remaining
                                })

                            # Convert to DataFrame for display
                            impact_df = pd.DataFrame(impact_report)
                            print("IMPAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACT")
                            print(impact_df)
                            print(numberindepartment)

                            impact_df["date"] = pd.to_datetime(impact_df["date"], dayfirst=True)
                            impact_df = impact_df[impact_df["date"].dt.weekday != 6].copy()

                            impact_df["group"] = (impact_df[["on leave", "employees remaining"]] != impact_df[["on leave", "employees remaining"]].shift()).any(axis=1).cumsum()

                            statements = []
                            for _, group_df in impact_df.groupby("group"):
                                start = group_df["date"].iloc[0].strftime("%d %B %Y")
                                end = group_df["date"].iloc[-1].strftime("%d %B %Y")
                                on_leave = group_df["on leave"].iloc[0]
                                remaining = group_df["employees remaining"].iloc[0]
                                
                                if start == end:
                                    statements.append(f"On {start}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                                else:
                                    statements.append(f"From {start} to {end}, the {department} department will have {remaining} employee(s) remaining at work and {on_leave} employee(s) on leave.")
                            # Combine all statements into a single variable
                            final_summary = "\n".join(statements)
                            # Print output
                            for s in statements:
                                print(s)

                            query = f"SELECT appid, id FROM {table_name_apps_pending_approval} WHERE id = {str(employee_number)} ;"
                            cursor.execute(query, )
                            rows = cursor.fetchall()

                            df_employees = pd.DataFrame(rows, columns=["appid","id"])

                            query = f"SELECT id, whatsapp FROM {table_name} WHERE id = {str(employee_number)} ;"
                            cursor.execute(query, )
                            rows = cursor.fetchall()

                            df_employees = pd.DataFrame(rows, columns=["id", "whatsapp"])
                            whatsapp = df_employees.iat[0, 1]
                            approovvver = approver_name.title()
                            companyxx = table_name.replace("main", "").replace("_", " ").title()

                            """send_whatsapp_message(f"263{whatsapp}", f"✅ Great News {first_name} from {companyxx}'s {department} department! \n\n Your `{leave_type} Leave Application` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}` has been submitted successfully!\n\n"
                                f"Your Leave Application ID is `{app_id}`.\n\n"
                                f"A Notification has been sent to `{approovvver}`  on `+263{approver_whatsapp}` to decide on  your application.\n\n"
                                "To Check the approval status of your leave application, type `Hello` then select `Track Application`.")
                            
                            if approver_whatsapp:

                                buttons = [
                                    {"type": "reply", "reply": {"id": f"Approve5appwa_{app_id}", "title": "Approve"}},
                                    {"type": "reply", "reply": {"id": f"Disapproveappwa_{app_id}", "title": "Disapprove"}},
                                ]
                                send_whatsapp_message(
                                    f"263{approver_whatsapp}", 
                                    f"Hey {approovvver}! 😊. New `{leave_type}` Leave Application from `{first_name} {surname}` in {department} department for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}`.\n\n" 
                                    f"If you approve this leave application, {final_summary}\n\n"  
                                    f"Select an option below to either approve or disapprove the application."         
                                    , 
                                    buttons
                                )"""

                            print("Insert successful!")

                        except Exception as e:
                            print("Error inserting data:", e)
                            return jsonify({"message": "Error re-applying leave application.", "error": str(e)}), 500

                    else:

                        query = f"SELECT * FROM {table_name_apps_declined} WHERE appid = %s;"
                        cursor.execute(query, (app_id,))
                        result = cursor.fetchone()
                        app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf, statuspre, status_date = result
        
                        print("chiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
                        print(employee_number)
                        print(approver_name)

                        try:
                            insert_query = f"""
                            INSERT INTO {table_name_apps_pending_approval} 
                            (appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                            """
                            
                            cursor.execute(insert_query, (
                                app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, 
                                approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                                end_date, leave_days, leavedaysbalancebf, status
                            ))

                            # SQL query to delete or mark the leave as canceled
                            query = f"""DELETE FROM {table_name_apps_declined} WHERE appid = %s"""
                            cursor.execute(query, (app_id,))
                            
                            connection.commit()


                            query = f"SELECT id, whatsapp FROM {table_name} WHERE id = {str(employee_number)} ;"
                            cursor.execute(query, )
                            rows = cursor.fetchall()

                            df_employees = pd.DataFrame(rows, columns=["id", "whatsapp"])
                            whatsapp = df_employees.iat[0, 1]
                            approovvver = approver_name.title()
                            companyxx = table_name.replace("main", "").replace("_", " ").title()

                            """send_whatsapp_message(f"263{whatsapp}", f"✅ Great News {first_name} from {companyxx}'s {department} department! \n\n Your `{leave_type} Leave Application` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}` has been submitted successfully!\n\n"
                                f"Your Leave Application ID is `{app_id}`.\n\n"
                                f"A Notification has been sent to `{approovvver}`  on `+263{approver_whatsapp}` to decide on  your application.\n\n"
                                "To Check the approval status of your leave application, type `Hello` then select `Track Application`.")
                            
                            if approver_whatsapp:

                                buttons = [
                                    {"type": "reply", "reply": {"id": f"Approve5appwa_{app_id}", "title": "Approve"}},
                                    {"type": "reply", "reply": {"id": f"Disapproveappwa_{app_id}", "title": "Disapprove"}},
                                ]
                                send_whatsapp_message(
                                    f"263{approver_whatsapp}", 
                                    f"Hey {approovvver}! 😊. New `{leave_type}` Leave Application from `{first_name} {surname}` in {department} department for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}`.\n\n" 
                                    f"Select an option below to either approve or disapprove the application."         
                                    , 
                                    buttons
                                )"""

                            print("Insert successful!")

                        except Exception as e:

                            print("Error inserting data:", e)       

                            return jsonify({"message": "Error re-applying leave application.", "error": str(e)}), 500

                    return jsonify({"message": f"Leave Application {app_id} re-applied successfully."})
                
                except Exception as e:
                    return jsonify({"message": "Error re-applying leave application.", "error": str(e)}), 500
                
            else: 

                response = {'status': 'error', 'message': 'Leave re-application not submitted successfully.'}
                return jsonify(response), 400  
            
    @app.route('/delete-all-tables', methods=['POST'])
    def handle_delete_all_tables():
        try:
            delete_all_tables()  # Your function
            return jsonify({"message": "All tables deleted successfully"}), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 500
        

    @app.route('/export_all_tables')
    def export_all_tables():
        file_path = "exported_tables.xlsx"
        try:
            # Connect to PostgreSQL
            connection = psycopg2.connect(external_database_url)
            cursor = connection.cursor()

            # Get all table names in 'public' schema
            cursor.execute("""
                SELECT DISTINCT table_name
                FROM information_schema.columns
                WHERE table_schema = 'public';
            """)
            tables = [row[0] for row in cursor.fetchall()]


            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for table in tables:
                    df = pd.read_sql(f'SELECT * FROM "{table}"', connection)
                    df.to_excel(writer, sheet_name=table[:31], index=False)


        except Exception as e:
            return f"Error: {str(e)}"

        @after_this_request
        def cleanup(response):
            try:
                os.remove(file_path)
            except Exception as e:
                print("Could not export file:", e)
            return response

        return send_file(file_path, as_attachment=True)


    @app.route('/revoke_leave_application', methods=['POST'])
    def revoke_leave_application():
        user_uuid = session.get('user_uuid')
        empid = session.get('empid')

        if user_uuid:
            print("Received data:", request.data)  # log the raw data
            print("JSON data:", request.get_json())  # log the parsed JSON
            table_name = session.get('table_name')
            company_name = table_name.replace("main","")
            table_name_apps_pending_approval = f"{company_name}appspendingapproval"
            table_name_apps_approved = f"{company_name}appsapproved"


            query = f"SELECT * FROM {table_name_apps_pending_approval} WHERE id = %s;"
            cursor.execute(query, (empid,))
            rows = cursor.fetchall()
            df_employees_reapp_check = pd.DataFrame(rows)    

            if len(df_employees_reapp_check) == 0:

                try:
                    data = request.get_json()
                    app_id = data.get("app_id")
                    print ("eissssssssshhhhhhhhhhhhhhhhhhhhhhhhhhhh")
                    print (app_id)

                    if not app_id:
                        return jsonify({"message": "Application ID is missing."}), 400
                    
                    status = "Pending"

                    query = f"SELECT * FROM {table_name_apps_approved} WHERE appid = %s;"
                    cursor.execute(query, (app_id,))
                    result = cursor.fetchone()
                    
                    if result :
                        app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf, statuspre, status_date = result
                        print("cancelled yes")
                        print(result)

                        print("cancelled yes")
    
                        print("chiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
                        print(employee_number)
                        print(approver_name)

                        try:
                            insert_query = f"""
                            INSERT INTO {table_name_apps_pending_approval} 
                            (appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                            """
                            
                            cursor.execute(insert_query, (
                                app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, 
                                approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                                end_date, leave_days, leavedaysbalancebf, status
                            ))
                            
                            query = f"""DELETE FROM {table_name_apps_approved} WHERE appid = %s"""
                            cursor.execute(query, (app_id,))

                            query = f"""SELECT currentleavedaysbalance FROM {table_name} WHERE id = %s;"""
                            cursor.execute(query, (employee_number,))
                            leavedayscf = cursor.fetchone()
                            leavedayscf = float(leavedayscf[0])
                            print(leavedayscf)

                            new_current_balance = leavedayscf + leave_days
                            query = f"UPDATE {table_name} SET currentleavedaysbalance = %s WHERE id = %s;"
                            cursor.execute(query, (new_current_balance, employee_number))

                            connection.commit()
                            print("Insert successful!")

                        except Exception as e:
                            print("Error inserting data:", e)
                            return jsonify({"message": "Error revoking leave application.", "error": str(e)}), 500

                    return jsonify({"message": f"Leave Application {app_id} Revocation applied successfully."})
                
                except Exception as e:
                    return jsonify({"message": "Error applying for revocation of leave application.", "error": str(e)}), 500
                
            else: 

                response = {'status': 'error', 'message': 'Leave revocation application not submitted successfully.'}
                return jsonify(response), 400  
            

            

    @app.route('/revoke_leave_application_approver', methods=['POST'])
    def revoke_leave_application_approver():
        user_uuid = session.get('user_uuid')

        print("Received data:", request.data)  # log the raw data
        print("JSON data:", request.get_json())  # log the parsed JSON

        if user_uuid:
            data = request.get_json()
            app_id = data.get("app_id")

            table_name = session.get('table_name')
            company_name = table_name.replace("main","")
            table_name_apps_approved = f"{company_name}appsapproved"
            table_name_apps_declined = f"{company_name}appsdeclined"

            try:

                print ("eissssssssshhhhhhhhhhhhhhhhhhhhhhhhhhhh")
                print (app_id)

                if not app_id:
                    return jsonify({"message": "Application ID is missing."}), 400
                
                status = "Declined"

                query = f"SELECT * FROM {table_name_apps_approved} WHERE appid = %s;"
                cursor.execute(query, (app_id,))
                result = cursor.fetchone()
                
                app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf, statuspre, status_date = result
                print("approved yes")
                print(result)

                print("approved yes")

                print("chiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
                print(employee_number)
                print(approver_name)

                try:
                    insert_query = f"""
                    INSERT INTO {table_name_apps_declined} 
                    (appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus, statusdate)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                    """
                    
                    cursor.execute(insert_query, (
                        app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, 
                        approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                        end_date, leave_days, leavedaysbalancebf, status, today_date
                    ))

                    query = f"""SELECT currentleavedaysbalance FROM {table_name} WHERE id = %s;"""
                    cursor.execute(query, (employee_number,))
                    leavedayscf = cursor.fetchone()
                    leavedayscf = float(leavedayscf[0])
                    print(leavedayscf)

                    new_current_balance = leavedayscf + leave_days
                    query = f"UPDATE {table_name} SET currentleavedaysbalance = %s WHERE id = %s;"
                    cursor.execute(query, (new_current_balance, employee_number))
                    connection.commit()

                    query = f"""DELETE FROM {table_name_apps_approved} WHERE appid = %s"""
                    cursor.execute(query, (app_id,))

                    connection.commit()
                    print("Insert successful!")


                except Exception as e:
                    print("Error inserting data:", e)
                    return jsonify({"message": "Error revoking leave application.", "error": str(e)}), 500

                return jsonify({"message": f"Leave Application {app_id} revoked successfully."})
            
            except Exception as e:
                return jsonify({"message": "Error re-applying leave application.", "error": str(e)}), 500
                




    @app.route('/download_leave_app/<app_id>')
    def download_pdf(app_id):
        global today_date
        user_uuid = session.get('user_uuid')
        empid = session.get('empid')

        if user_uuid:
            table_name = session.get('table_name')
            company_name = table_name.replace("main","")
            try:
                table_name_apps_approved = company_name + 'appsapproved'
                query = f"""SELECT appid, id, firstname, surname, leavetype, TO_CHAR(dateapplied, 'FMDD Month YYYY') AS dateapplied, TO_CHAR(leavestartdate, 'FMDD Month YYYY') AS leavestartdate, TO_CHAR(leaveenddate, 'FMDD Month YYYY') AS leaveenddate,  leavedaysappliedfor, leaveapprovername, TO_CHAR(statusdate, 'FMDD Month YYYY') AS statusdate, leavedaysbalancebf, leaveapproverid, department FROM {table_name_apps_approved} WHERE appid = %s;"""
                cursor.execute(query, (app_id,))  
                rows = cursor.fetchall()
                df_leave_appsmain_approved = pd.DataFrame(rows, columns=["App ID","ID","First Name", "Surname", "Leave Type","Date Applied", "Leave Start Date", "Leave End Date", "Leave Days","Leave Approver", "Status Date","New Leave Days Balance", "Leave Approver ID", "Department"])

                query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation FROM {table_name};"
                cursor.execute(query)
                rows = cursor.fetchall()

                df_employees = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Address", "Role","Leave Approver Name","Leave Approver ID","Leave Approver Email", "Leave Approver WhatsAapp", "Leave Days Balance","Days Accumulated per Month"])
                print(df_employees)
                userdf = df_employees[df_employees['id'] == empid].reset_index()
                print("yeaarrrrr")

                employee_name = f"{df_leave_appsmain_approved.iat[0,2].title()} {df_leave_appsmain_approved.iat[0,3].title()}"
                leave_type = df_leave_appsmain_approved.iat[0,4].title()
                company_name_doc = company_name.replace("_"," ").title()

                def get_logo_base64():
                    logo_path = os.path.join(app.static_folder, 'images', 'eds logo blue.png')
                    with open(logo_path, "rb") as img_file:
                        return "data:image/png;base64," + base64.b64encode(img_file.read()).decode('utf-8')

                application = {
                    'company_logo': get_logo_base64() ,
                    'company_name': company_name_doc,
                    'employee_name': employee_name,
                    'employee_id': df_leave_appsmain_approved.iat[0,1],
                    'leave_type': leave_type,
                    'generated_on': today_date,
                    'date_applied': df_leave_appsmain_approved.iat[0,5],
                    'approver_name': df_leave_appsmain_approved.iat[0,9].title(),
                    'approver_id': df_leave_appsmain_approved.iat[0,12],
                    'department': df_leave_appsmain_approved.iat[0,13],
                    'reference_number': app_id,
                    'approved_date': df_leave_appsmain_approved.iat[0,10],
                    'new_balance': df_leave_appsmain_approved.iat[0,11],
                    'start_date': df_leave_appsmain_approved.iat[0,6],
                    'end_date': df_leave_appsmain_approved.iat[0,7],
                    'days_requested': df_leave_appsmain_approved.iat[0,8], 
                    'address': userdf.iat[0,6], 
                    'whatsapp': f"0{userdf.iat[0,4]}", 
                    'email': userdf.iat[0,5], 
                    'status': 'Approved'
                }
                
                # 2. Render HTML
                html = render_template('leave_pdf_template.html', app=application)
                
                # 3. Generate PDF
                pdf = HTML(string=html).write_pdf()
                
                # 4. Create response
                response = make_response(pdf)
                response.headers['Content-Type'] = 'application/pdf'
                response.headers['Content-Disposition'] = \
                    f'attachment; filename=leave_application_{company_name}_{app_id}.pdf'
                
                return response
                
            except Exception as e:
                return str(e), 500
            
    @app.route('/revive_leave_application', methods=['POST'])
    def revive_leave_application():
        user_uuid = session.get('user_uuid')

        print("Received data:", request.data)  # log the raw data
        print("JSON data:", request.get_json())  # log the parsed JSON

        if user_uuid:
            data = request.get_json()
            app_id = data.get("app_id")

            table_name = session.get('table_name')
            company_name = table_name.replace("main","")
            table_name_apps_pending_approval = f"{company_name}appspendingapproval"
            table_name_apps_cancelled = f"{company_name}appscancelled"
            table_name_apps_declined = f"{company_name}appsdeclined"

            query = f"SELECT id FROM {table_name_apps_declined} WHERE appid = %s;"
            cursor.execute(query, (app_id,))
            res = cursor.fetchone()
            employee_number_check = res

            query = f"SELECT * FROM {table_name_apps_pending_approval} WHERE id = %s;"
            cursor.execute(query, (employee_number_check,))
            rows = cursor.fetchall()
            df_employees_reapp_check = pd.DataFrame(rows)    

            if len(df_employees_reapp_check) == 0:

                try:

                    print ("eissssssssshhhhhhhhhhhhhhhhhhhhhhhhhhhh")
                    print (app_id)

                    if not app_id:
                        return jsonify({"message": "Application ID is missing."}), 400
                    
                    status = "Pending"

                    query = f"SELECT * FROM {table_name_apps_declined} WHERE appid = %s;"
                    cursor.execute(query, (app_id,))
                    result = cursor.fetchone()
                    
                    app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf, statuspre, status_date = result
                    print("declined yes")
                    print(result)

                    print("declined yes")

                    print("chiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
                    print(employee_number)
                    print(approver_name)

                    try:
                        insert_query = f"""
                        INSERT INTO {table_name_apps_pending_approval} 
                        (appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                        """
                        
                        cursor.execute(insert_query, (
                            app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, 
                            approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                            end_date, leave_days, leavedaysbalancebf, status
                        ))
                        
                        query = f"""DELETE FROM {table_name_apps_declined} WHERE appid = %s"""
                        cursor.execute(query, (app_id,))

                        connection.commit()
                        print("Insert successful!")

                    except Exception as e:
                        print("Error inserting data:", e)
                        return jsonify({"message": "Error re-applying leave application.", "error": str(e)}), 500


                    return jsonify({"message": f"Leave Application {app_id} revived successfully."})
                
                except Exception as e:
                    return jsonify({"message": "Error re-applying leave application.", "error": str(e)}), 500
                
            else: 

                response = {'status': 'error', 'message': 'Leave application not submitted successfully.'}
                return jsonify(response), 400  

    @app.route('/approve_leave_application', methods=['POST'])
    def approve_leave_application():
        user_uuid = session.get('user_uuid')

        print("Received data:", request.data)  # log the raw data
        print("JSON data:", request.get_json())  # log the parsed JSON

        if user_uuid:

            ACCESS_TOKEN = "EAATESj1oB5YBPIzFCv7ulvosr2S2ZAiWBJrFp7bti6L0ZCWS2AOz5dUABlJ6q16a4hRwEXdq5vZAP5tp4rGXfOQ2sx0hg1EOwMpL002eqUrygbPc3jkY8FPOzR7c6tMvKJxT3XxXP8Qp9U1n30MIMVcNy9JUCZB8UyIwaAZBAjf2U32TVTwSBJlSeHoNYrGH0dwZDZD"
            PHONE_NUMBER_ID = "756962384159644"
            VERIFY_TOKEN = "2644686099068373"
            WHATSAPP_API_URL = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"

            try:
                data = request.get_json()
                app_id = data.get("app_id")
                print ("eissssssssshhhhhhhhhhhhhhhhhhhhhhhhhhhh")
                print (app_id)
                table_name = session.get('table_name')
                company_name = table_name.replace("main","")
                table_name_apps_pending_approval = f"{company_name}appspendingapproval"
                table_name_apps_approved = f"{company_name}appsapproved"

                if not app_id:
                    return jsonify({"message": "Application ID is missing."}), 400

                status = "Approved"
                statusdate = today_date

                query = f"SELECT * FROM {table_name_apps_pending_approval} WHERE appid = %s;"
                cursor.execute(query, (app_id,))
                result = cursor.fetchone()
                app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf, statuspre = result
                print("chiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
                print(employee_number)
                print(approver_name)

                try:
                    insert_query = f"""
                    INSERT INTO {table_name_apps_approved} 
                    (appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus, statusdate)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                    """
                    
                    cursor.execute(insert_query, (
                        app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, 
                        approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                        end_date, leave_days, leavedaysbalancebf, status, statusdate
                    ))
                    
                    connection.commit()
                    print("Insert successful!")

                    query = f"UPDATE {table_name} SET currentleavedaysbalance = %s WHERE id = %s;"
                    cursor.execute(query, (leavedaysbalancebf, employee_number))
                    connection.commit()

                except Exception as e:
                    print("Error inserting data:", e)

                query = f"""DELETE FROM {table_name_apps_pending_approval} WHERE appid = %s"""
                cursor.execute(query, (app_id,))
                connection.commit()

                query = f"SELECT id, firstname, surname, whatsapp, email, address, role, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, monthlyaccumulation, department FROM {table_name};"
                cursor.execute(query)
                rows = cursor.fetchall()

                df_employees = pd.DataFrame(rows, columns=["id","firstname", "surname", "whatsapp","Email", "Address", "Role","Leave Approver Name","Leave Approver ID","Leave Approver Email", "Leave Approver WhatsAapp", "Leave Days Balance","Days Accumulated per Month", "Department"])
                print(df_employees)
                userdf = df_employees[df_employees['id'] == int(np.int64(employee_number))].reset_index()
                print("yeaarrrrr")
                print(userdf)
                firstname = userdf.iat[0,2].title()
                surname = userdf.iat[0,3].title()
                whatsappemp = userdf.iat[0,4]
                email = userdf.iat[0,5]
                address = userdf.iat[0,6]
                companyxx = company_name.replace("_", " ").title()
                app_namexx = approver_name.title()

                query = f"SELECT appid, id, leavetype, leaveapprovername, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, approvalstatus, statusdate, leavedaysbalancebf, department  FROM {table_name_apps_approved} WHERE id = {str(employee_number)};"
                cursor.execute(query)
                rows = cursor.fetchall()
                df_employeesappsapprovedcheck = pd.DataFrame(rows, columns=["appid","id", "leavetype", "leaveapprovername", "dateapplied", "leavestartdate", "leaveenddate", "leavedaysappliedfor","approvalstatus","statusdate", "leavedaysbalancebf", "department"]) 

                df_employeesappsapprovedcheck = df_employeesappsapprovedcheck.sort_values(by="appid", ascending=False)  

                """send_whatsapp_message(f"263{whatsappemp}", f"✅ Great News {firstname} {surname} from {companyxx}! \n\n Your `{leave_type} Leave Application` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}` has been Approved ✅ by `{app_namexx}`!") """

                def generate_leave_pdf():
                    app = {
                        'company_logo': 44,
                        'company_name': companyxx,
                        'employee_name': f"{first_name} {surname}",
                        'leave_type': leave_type,
                        'generated_on': today_date,
                        'date_applied': df_employeesappsapprovedcheck.iat[0,4].strftime('%d %B %Y'),
                        'approver_name': df_employeesappsapprovedcheck.iat[0,3].title(),
                        'reference_number': df_employeesappsapprovedcheck.iat[0,0],
                        'approved_date': df_employeesappsapprovedcheck.iat[0,9].strftime('%d %B %Y'),
                        'new_balance': df_employeesappsapprovedcheck.iat[0,10],
                        'start_date':  df_employeesappsapprovedcheck.iat[0,5].strftime('%d %B %Y'),
                        'end_date':  df_employeesappsapprovedcheck.iat[0,6].strftime('%d %B %Y'),
                        'days_requested':  df_employeesappsapprovedcheck.iat[0,7], 
                        'department': df_employeesappsapprovedcheck.iat[0,11],
                        'address': address, 
                        'whatsapp': f"+263{whatsappemp}", 
                        'email': email, 
                        'status': 'Approved'
                    }

                    html_out = render_template("leave_pdf_template.html", app=app)
                    
                    # ✅ Return as bytes instead of saving to file
                    pdf_bytes = HTML(string=html_out).write_pdf()
                    return pdf_bytes


                def upload_pdf_to_whatsapp(pdf_bytes):
                    filename=f"leave_application_{df_employeesappsapprovedcheck.iat[0,0]}_{first_name}_{surname}_{companyxx}.pdf"
                
                    url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/media"
                    headers = {
                        "Authorization": f"Bearer {ACCESS_TOKEN}"
                    }

                    files = {
                        "file": (filename, io.BytesIO(pdf_bytes), "application/pdf"),
                        "type": (None, "application/pdf"),
                        "messaging_product": (None, "whatsapp")
                    }

                    response = requests.post(url, headers=headers, files=files)
                    print("📥 Full incoming data:", response.text)  # Good for debugging
                    response.raise_for_status()
                    return response.json()["id"]

                                                                
                def send_whatsapp_pdf_by_media_id(recipient_number, media_id):
                    filename=f"leave_application_{df_employeesappsapprovedcheck.iat[0,0]}_{first_name}_{surname}_{companyxx}.pdf"
                    url = f"https://graph.facebook.com/v19.0/{PHONE_NUMBER_ID}/messages"
                    headers = {
                        "Authorization": f"Bearer {ACCESS_TOKEN}",
                        "Content-Type": "application/json"
                    }
                    payload = {
                        "messaging_product": "whatsapp",
                        "to": recipient_number,
                        "type": "document",
                        "document": {
                            "id": media_id,            # Media ID from upload step
                            "filename": filename       # Desired file name on recipient's phone
                        }
                    }

                    response = requests.post(url, headers=headers, json=payload)
                    response.raise_for_status()
                    return response.json()


                pdf_path = generate_leave_pdf()
                media_id = upload_pdf_to_whatsapp(pdf_path)
                send_whatsapp_pdf_by_media_id(f"263{whatsappemp}", media_id)

                return jsonify({"message": f"Leave Application {app_id} approved successfully."})
            
            except Exception as e:
                return jsonify({"message": "Error approving leave application.", "error": str(e)}), 500
            

    @app.route('/disapprove_leave_application', methods=['POST'])
    def disapprove_leave_application():
        user_uuid = session.get('user_uuid')

        print("Received data:", request.data)  # log the raw data
        print("JSON data:", request.get_json())  # log the parsed JSON

        if user_uuid:

            ACCESS_TOKEN = "EAATESj1oB5YBPIzFCv7ulvosr2S2ZAiWBJrFp7bti6L0ZCWS2AOz5dUABlJ6q16a4hRwEXdq5vZAP5tp4rGXfOQ2sx0hg1EOwMpL002eqUrygbPc3jkY8FPOzR7c6tMvKJxT3XxXP8Qp9U1n30MIMVcNy9JUCZB8UyIwaAZBAjf2U32TVTwSBJlSeHoNYrGH0dwZDZD"
            PHONE_NUMBER_ID = "756962384159644"
            VERIFY_TOKEN = "2644686099068373"
            WHATSAPP_API_URL = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"

            try:
                data = request.get_json()
                app_id = data.get("app_id")
                print ("eissssssssshhhhhhhhhhhhhhhhhhhhhhhhhhhh")
                print (app_id)
                table_name = session.get('table_name')
                company_name = table_name.replace("main","")
                companyxx = company_name.replace("_", " ").title()
                table_name_apps_pending_approval = f"{company_name}appspendingapproval"
                table_name_apps_declined = f"{company_name}appsdeclined"

                if not app_id:
                    return jsonify({"message": "Application ID is missing."}), 400

                status = "Declined"
                statusdate = today_date

                query = f"SELECT * FROM {table_name_apps_pending_approval} WHERE appid = %s;"
                cursor.execute(query, (app_id,))
                result = cursor.fetchone()
                app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, end_date, leave_days, leavedaysbalancebf, statuspre = result
                print("chiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
                print(employee_number)
                print(approver_name)
                try:
                    insert_query = f"""
                    INSERT INTO {table_name_apps_declined} 
                    (appid, id, firstname, surname, department, leavetype, reasonifother, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp, currentleavedaysbalance, dateapplied, leavestartdate, leaveenddate, leavedaysappliedfor, leavedaysbalancebf, approvalstatus, statusdate)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                    """
                    
                    cursor.execute(insert_query, (
                        app_id, employee_number, first_name, surname, department, leave_type, leave_specify, approver_name, 
                        approver_id, approver_email, approver_whatsapp, leave_days_balance, date_applied, start_date, 
                        end_date, leave_days, leavedaysbalancebf, status, statusdate
                    ))
                    
                    connection.commit()
                    print("Insert successful!")

                except Exception as e:
                    print("Error inserting data:", e)

                query = f"""DELETE FROM {table_name_apps_pending_approval} WHERE appid = %s"""
                cursor.execute(query, (app_id,))
                connection.commit()

                query = f"SELECT id, firstname, surname, whatsapp, email, address ,role, department,currentleavedaysbalance, monthlyaccumulation, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp  FROM {table_name};"
                cursor.execute(query)
                rows = cursor.fetchall()

                df_employeesx = pd.DataFrame(rows, columns=["ID","First Name", "Surname", "WhatsApp","Email", "Address", "Role", "Department","Leave Days Balance","Days Accumulated per Month","Leave Approver Name", "Leave Approver ID", "Leave Approver Email", "Leave Approver WhatsaApp"])
                userdff = df_employeesx[df_employeesx['id'] == int(np.int64(employee_number))].reset_index()

                whatsappapprover = userdff.iat[0, 13]
                whatsappemp = userdff.iat[0, 3]

                buttonsapproval = [
                    {"type": "reply", "reply": {"id": "Revokedis", "title": "Revoke Disapproval"}},
                    {"type": "reply", "reply": {"id": "Pending", "title": "Apps Pending My Approval"}},
                ]

                """send_whatsapp_message(f"263{whatsappapprover}", f"✅ Hey {approver_name} from {companyxx}! \n\n You have successfully disapproved `{first_name} {surname}`'s  `{leave_days} day` `{leave_type} Leave Application` running from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}`✅!")
                send_whatsapp_message(
                    f"263{whatsappapprover}",
                    "Select an option below to continue 👇y, or Type `Hello` to view all Approver options",
                    buttonsapproval
                )

                if whatsappemp:

                    buttons = [
                        {"type": "reply", "reply": {"id": "Reapply", "title": "Resubmit Application"}},
                        {"type": "reply", "reply": {"id": "Apply", "title": "Apply for Leave"}},
                        {"type": "reply", "reply": {"id": "Checkbal", "title": "Check Days Balance"}},
                    ]

                    send_whatsapp_message(f"263{whatsappemp}", f"✅ Oops, {first_name} {surname} from {companyxx}! \n\n Your `{leave_type} Leave Application` for `{leave_days} days` from `{start_date.strftime('%d %B %Y')}` to `{end_date.strftime('%d %B %Y')}`, has been disapproved ❌ by `{companyxx}`!")
                    send_whatsapp_message(
                        f"263{whatsappemp}",
                        "Select an option below to continue 👇",
                        buttons
                    )"""

                return jsonify({"message": f"Leave Application {app_id} declined successfully."})
            
            except Exception as e:
                return jsonify({"message": "Error approving leave application.", "error": str(e)}), 500


    @app.route('/export_lms_book_excel')
    def export_excel():
        user_uuid = session.get('user_uuid')
        if user_uuid:

            table_name = session.get('table_name')
            company_name = table_name.replace("main","")


            query = f"SELECT id, firstname, surname, whatsapp, email, address ,role, department,currentleavedaysbalance, monthlyaccumulation, leaveapprovername, leaveapproverid, leaveapproveremail, leaveapproverwhatsapp  FROM {table_name};"
            cursor.execute(query)
            rows = cursor.fetchall()

            df_employees = pd.DataFrame(rows, columns=["ID","First Name", "Surname", "WhatsApp","Email", "Address", "Role", "Department","Leave Days Balance","Days Accumulated per Month","Leave Approver Name", "Leave Approver ID", "Leave Approver Email", "Leave Approver WhatsaApp"])
            df_employees = df_employees.sort_values(by="ID", ascending=True)

            print(df_employees)

            # Create an in-memory Excel file
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_employees.to_excel(writer, index=False, sheet_name=f'LMS Book {today_date}')

            output.seek(0)

            # Send the file to the client
            return send_file(
                output,
                as_attachment=True,
                download_name=f'{company_name} LMS Book as at {today_date}.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

    @app.route('/export_lms_book_pdf')
    def export_pdf():
        user_uuid = session.get('user_uuid')
        if user_uuid:
            table_name = session.get('table_name')
            if not table_name:
                return "Table name not found in session", 400

            company_name = table_name.replace("main", "").strip()

            query = f"SELECT id, firstname, surname, whatsapp, email, role, department, leaveapprovername, leaveapproverid, currentleavedaysbalance, monthlyaccumulation FROM {table_name};"
            cursor.execute(query)
            rows = cursor.fetchall()

            df_employees = pd.DataFrame(rows, columns=["ID", "First Name", "Surname", "WhatsApp", "Email", "Role", "Department", "Leave Approver Name","Leave Approver ID", "Leave Days Balance", "Days Accumulated per Month" ])
            df_employees = df_employees.sort_values(by="ID", ascending=True)

            df_html = df_employees.to_html(index=False, classes='table table-bordered', escape=False)

            html_content = f"""
            <html>
            <head>
            <style>
                .table {{
                    width: 100%;
                    border-collapse: collapse;
                }}
                .table th, .table td {{
                    border: 1px solid #ddd;
                    padding: 4px;
                    text-align: left;
                    max-width: 500px;  /* Limit cell width */
                    word-wrap: break-word;  /* Enables wrapping */
                    overflow-wrap: break-word;  /* Ensures wrapping in older browsers */
                    white-space: normal;  /* Forces the text to wrap */
                }}
                .table th {{
                    background-color: #003366;
                    color: white;
                }}
                tr {{
                    word-wrap: break-word;
                }}
                /* Specific styles for the email column */
                td.Email {{
                    max-width: 250px;  /* Limit email column width */
                    word-wrap: break-word;
                    overflow-wrap: break-word;
                    white-space: normal;
                }}
                th.Email {{
                    word-wrap: break-word;
                    overflow-wrap: break-word;
                }}
                h1 {{
                    text-align: center;
                    margin-bottom: 20px;
                }}
            </style>

            </head>
            <body>
                <h1>{company_name} LMS Book</h1>
                <p>As of {today_date}</p>
                {df_html}
            </body>
            </html>
            """

            output = io.BytesIO()
            pisa_status = pisa.CreatePDF(html_content, dest=output)

            if pisa_status.err:
                return "Failed to generate PDF", 500

            output.seek(0)

            return send_file(
                output,
                as_attachment=True,
                download_name=f'{company_name} LMS Book as at {today_date}.pdf',
                mimetype='application/pdf'
            )
        else:
            return "Unauthorized access", 401

    @app.route('/logout')
    def logout():
        # Clear the session data to log the user out
        session.clear()

        # Redirect to the landing page or login page after logout
        return redirect(url_for('explore_lms'))



            
else:
    print('Connection to SQL failed')

@app.route('/')
def landingpage():
    return render_template('main.html') 

@app.route('/echelon-digital-solutions-privacy-policy')
def privacypolicy():
    return render_template('privacypolicy.html') 


@app.route('/explore_lms')
def explore_lms():
    return render_template('index.html')  
    

if __name__ == "__main__":
    app.run(host='0.0.0.0', port= 55, debug=True)
